import json
import logging
import os
import re
import secrets
import string
import threading
import time
import unicodedata
from datetime import date as date_class
from datetime import datetime
from datetime import timedelta
from functools import wraps
from pathlib import Path
from urllib import error as urllib_error
from urllib import request as urllib_request
import io

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from dotenv import load_dotenv
from flask import Flask, jsonify, make_response, request, send_from_directory
from flask_cors import CORS
from supabase import Client, create_client


load_dotenv(Path(__file__).with_name('.env'))


BACKEND_DIR = Path(__file__).resolve().parent
FRONTEND_DIST_DIR = BACKEND_DIR.parent / 'frontend' / 'dist'
SUPER_ADMIN_EMAIL = 'admin@admin'
# AVISO: nunca adicionar aliases por local-part de e-mail ou por cargo — gera escalada de privilégio.


RESOURCE_DEFINITIONS = {
    'filiais': {
        'table': 'filiais',
        'order': 'cidade',
        'required_fields': ['cidade', 'uf', 'parceira'],
        'allowed_fields': ['cidade', 'uf', 'parceira', 'endereco', 'cnpj', 'telefone_filial', 'email_filial', 'ativo'],
        'nullable_fields': ['endereco', 'cnpj', 'telefone_filial', 'email_filial'],
        'partial_match_fields': ['cidade', 'uf', 'parceira'],
        'view_scope': 'menu.filiais',
        'create_scope': 'create.filiais',
    },
    'colaboradores': {
        'table': 'colaboradores',
        'order': 'nome_completo',
        'required_fields': ['filial_id', 'nome_completo', 'cpf', 'cargo', 'data_admissao'],
        'allowed_fields': [
            'user_id',
            'filial_id',
            'nome_completo',
            'cpf',
            'telefone',
            'cargo',
            'turno',
            'escala_servico',
            'horario_padrao_inicio',
            'horario_padrao_fim',
            'carga_horaria_semanal',
            'intervalo_almoco_minutos',
            'salario_base_mensal',
            'percentual_periculosidade',
            'percentual_adicional_clt',
            'adicional_noturno_sobre_periculosidade',
            'beneficios_tipos',
            'beneficios_mensais',
            'tipo_acesso',
            'permissao_app',
            'permissao_desktop',
            'permissao_editar',
            'permissao_excluir',
            'permissao_aprovar_he',
            'foto_url',
            'data_admissao',
            'data_desligamento',
            'ativo',
        ],
        'nullable_fields': ['data_desligamento', 'horario_padrao_inicio', 'horario_padrao_fim', 'telefone'],
        'partial_match_fields': ['nome_completo', 'cargo', 'cpf'],
        'view_scope': 'menu.colaboradores',
        'view_scope_any': [
            'menu.colaboradores',
            'menu.colaborador_documentos',
            'menu.colaborador_contratos',
            'menu.diarias',
            'menu.eventos_rh',
            'menu.contratos_operacionais',
            'menu.quadro_funcionarios',
            'menu.horas_extras',
            'menu.bonificacao',
            'menu.bonificacao_metricas',
            'menu.pedidos_compra',
            'menu.manutencoes',
            'menu.acompanhamento',
            'menu.presenca',
            'menu.carregamento',
        ],
        'create_scope': 'create.colaboradores',
        'filial_scope_field': 'filial_id',
    },
    'contratos_operacionais': {
        'table': 'contratos_operacionais',
        'order': 'nome_contrato',
        'required_fields': ['filial_id', 'codigo_contrato', 'nome_contrato', 'valor_mensal_contrato'],
        'allowed_fields': [
            'filial_id',
            'codigo_contrato',
            'nome_contrato',
            'cliente_nome',
            'tipo_contrato',
            'valor_mensal_contrato',
            'horas_50_cobradas_contrato',
            'horas_100_cobradas_contrato',
            'qtd_colaboradores_contratados',
            'cargos_contrato',
            'valor_por_colaborador',
            'custos_extras_gold_mensais',
            'inicio_vigencia',
            'fim_vigencia',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['codigo_contrato', 'nome_contrato', 'cliente_nome', 'cargos_contrato', 'tipo_contrato', 'observacoes'],
        'nullable_fields': ['cliente_nome', 'qtd_colaboradores_contratados', 'cargos_contrato', 'valor_por_colaborador', 'inicio_vigencia', 'fim_vigencia', 'observacoes'],
        'view_scope': 'menu.contratos_operacionais',
        'create_scope': 'create.contratos_operacionais',
        'filial_scope_field': 'filial_id',
    },
    'contratos_colaboradores': {
        'table': 'contratos_colaboradores',
        'order': 'id',
        'required_fields': ['filial_id', 'contrato_operacional_id', 'tipo_item'],
        'allowed_fields': [
            'filial_id',
            'contrato_operacional_id',
            'tipo_item',
            'colaborador_id',
            'veiculo_carregamento_id',
            'veiculo_proprio_id',
            'nome_item',
            'percentual_alocacao',
            'valor_cobrado_colaborador',
            'horas_50_cobradas',
            'horas_100_cobradas',
            'inicio_vigencia',
            'fim_vigencia',
            'data_inicio_vinculo',
            'data_fim_vinculo',
            'tipo_periodo',
            'dias_duracao_periodo',
            'dias_alerta_antes',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['observacoes'],
        'nullable_fields': [
            'observacoes', 'inicio_vigencia', 'fim_vigencia',
            'data_inicio_vinculo', 'data_fim_vinculo',
            'tipo_periodo', 'dias_duracao_periodo', 'dias_alerta_antes',
            'colaborador_id', 'veiculo_carregamento_id', 'veiculo_proprio_id', 'nome_item',
        ],
        'view_scope': 'menu.contratos_operacionais',
        'create_scope': 'create.contratos_operacionais',
        'filial_scope_field': 'filial_id',
    },
    'contratos_gastos_extras': {
        'table': 'contratos_gastos_extras',
        'order': 'nome_gasto',
        'required_fields': ['filial_id', 'contrato_operacional_id', 'nome_gasto'],
        'allowed_fields': [
            'filial_id',
            'contrato_operacional_id',
            'colaborador_id',
            'percentual_alocacao',
            'nome_gasto',
            'valor_mensal',
            'inicio_vigencia',
            'fim_vigencia',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['nome_gasto', 'observacoes'],
        'nullable_fields': ['colaborador_id', 'percentual_alocacao', 'valor_mensal', 'inicio_vigencia', 'fim_vigencia', 'observacoes'],
        'view_scope': 'menu.contratos_operacionais',
        'create_scope': 'create.contratos_operacionais',
        'filial_scope_field': 'filial_id',
    },
    'colaborador_beneficios': {
        'table': 'colaborador_beneficios',
        'order': 'tipo_beneficio',
        'required_fields': ['filial_id', 'colaborador_id', 'tipo_beneficio', 'valor_mensal'],
        'allowed_fields': [
            'filial_id',
            'colaborador_id',
            'tipo_beneficio',
            'valor_mensal',
            'modo_calculo',
            'valor_unitario',
            'base_dias',
            'desconta_faltas',
            'desconta_eventos',
            'teto_mensal',
            'codigo_rubrica',
            'ordem',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['tipo_beneficio', 'codigo_rubrica', 'observacoes'],
        'nullable_fields': ['observacoes', 'teto_mensal', 'codigo_rubrica'],
        'view_scope': 'menu.colaboradores',
        'create_scope': 'create.colaboradores',
        'filial_scope_field': 'filial_id',
    },
    'colaborador_documentos': {
        'table': 'colaborador_documentos',
        'order': 'data_validade',
        'required_fields': ['colaborador_id', 'filial_id', 'tipo_documento'],
        'allowed_fields': [
            'colaborador_id',
            'filial_id',
            'categoria',
            'tipo_documento',
            'numero_documento',
            'orgao_emissor',
            'data_emissao',
            'data_validade',
            'dias_alerta',
            'arquivo_url',
            'arquivos_extras',
            'status',
            'obrigatorio',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['categoria', 'tipo_documento', 'numero_documento', 'orgao_emissor', 'status'],
        'nullable_fields': ['numero_documento', 'orgao_emissor', 'data_emissao', 'data_validade', 'dias_alerta', 'arquivo_url', 'arquivos_extras', 'status', 'observacoes'],
        'view_scope': 'menu.colaborador_documentos',
        'create_scope': 'create.colaborador_documentos',
        'filial_scope_field': 'filial_id',
    },
    'diarias_valores': {
        'table': 'diarias_valores',
        'order': 'cidade',
        'required_fields': ['cidade'],
        'allowed_fields': ['cidade','uf','cafe','almoco','jantar','pernoite','ativo','observacoes'],
        'partial_match_fields': ['cidade','uf','observacoes'],
        'nullable_fields': ['uf','observacoes'],
        'view_scope': 'menu.diarias',
        'create_scope': 'create.diarias',
    },
    'diarias_solicitacoes': {
        'table': 'diarias_solicitacoes',
        'order': 'data_solicitacao',
        'required_fields': ['filial_id','cidade_destino','data_inicio','data_fim'],
        'allowed_fields': [
            'filial_id','numero_solicitacao','tipo','cidade_destino','uf_destino',
            'data_solicitacao','data_inicio','data_fim','rota','status','valor_total',
            'criado_por','aprovado_por','aprovado_em','motivo_reprovacao','reprovado_por',
            'reprovado_em','banco','observacoes','ativo',
        ],
        'partial_match_fields': ['cidade_destino','rota','status','numero_solicitacao','tipo','observacoes'],
        'nullable_fields': ['numero_solicitacao','uf_destino','rota','criado_por','aprovado_por','aprovado_em','motivo_reprovacao','reprovado_por','reprovado_em','banco','observacoes'],
        'view_scope': 'menu.diarias',
        'create_scope': 'create.diarias',
        'filial_scope_field': 'filial_id',
    },
    'diarias_itens': {
        'table': 'diarias_itens',
        'order': 'id',
        'required_fields': ['solicitacao_id','filial_id','motorista_nome'],
        'allowed_fields': [
            'solicitacao_id','filial_id','colaborador_id','motorista_nome','placa',
            'data_inicio','data_fim','qtd_diarias','qtd_pernoites',
            'inclui_cafe','inclui_almoco','inclui_jantar',
            'valor_cafe','valor_almoco','valor_jantar','valor_pernoite','valor_total',
            'observacoes','hotel_nome','hotel_fornecedor_id','ativo',
        ],
        'partial_match_fields': ['motorista_nome','placa','hotel_nome'],
        'nullable_fields': ['colaborador_id','placa','data_inicio','data_fim','observacoes','hotel_nome','hotel_fornecedor_id'],
        'view_scope': 'menu.diarias',
        'create_scope': 'create.diarias',
        'filial_scope_field': 'filial_id',
    },
    'colaborador_contratos': {
        'table': 'colaborador_contratos',
        'order': 'data_inicio',
        'required_fields': ['colaborador_id', 'tipo_vinculo', 'data_inicio'],
        'allowed_fields': [
            'colaborador_id',
            'filial_id',
            'vinculo_id',
            'tipo_vinculo',
            'fase',
            'data_inicio',
            'data_fim',
            'data_desligamento',
            'motivo_desligamento',
            'cargo',
            'salario',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['tipo_vinculo', 'fase', 'cargo'],
        'nullable_fields': ['filial_id', 'fase', 'data_fim', 'data_desligamento', 'motivo_desligamento', 'cargo', 'salario', 'observacoes'],
        'view_scope': 'menu.colaborador_contratos',
        'create_scope': 'create.colaborador_contratos',
        'filial_scope_field': 'filial_id',
    },
    'eventos_rh': {
        'table': 'eventos_rh',
        'order': 'data_inicio',
        'required_fields': ['colaborador_id', 'filial_id', 'tipo_evento', 'data_inicio', 'data_fim'],
        'allowed_fields': [
            'colaborador_id',
            'filial_id',
            'tipo_evento',
            'data_inicio',
            'data_fim',
            'status',
            'impacta_presenca',
            'observacoes',
            'cobertura_colaborador_id',
            'dias_uteis',
            'abono_pecuniario',
            'parcela',
            'ativo',
        ],
        'partial_match_fields': ['tipo_evento', 'status', 'observacoes'],
        'nullable_fields': ['observacoes', 'cobertura_colaborador_id', 'dias_uteis', 'abono_pecuniario', 'parcela'],
        'view_scope': 'menu.eventos_rh',
        'create_scope': 'create.eventos_rh',
        'filial_scope_field': 'filial_id',
    },
    'veiculos': {
        'table': 'veiculos',
        'order': 'placa',
        'required_fields': ['filial_id', 'placa', 'marca', 'modelo'],
        'allowed_fields': [
            'filial_id',
            'placa',
            'chassi',
            'marca',
            'modelo',
            'ano_modelo',
            'cor',
            'status',
            'odometro_atual',
            'descricao',
            'tipo_veiculo',
            'tipo',
            'combustivel',
            'capacidade_tanque',
            'data_vencimento_crlv',
            'data_vencimento_seguro',
            'data_ultima_revisao',
            'km_proxima_revisao',
        ],
        'partial_match_fields': ['placa', 'marca', 'modelo', 'chassi', 'cor', 'status', 'tipo_veiculo', 'tipo'],
        'nullable_fields': [
            'chassi', 'ano_modelo', 'cor', 'odometro_atual', 'descricao',
            'tipo_veiculo', 'tipo', 'combustivel', 'capacidade_tanque',
            'data_vencimento_crlv', 'data_vencimento_seguro',
            'data_ultima_revisao', 'km_proxima_revisao',
        ],
        'view_scope': 'menu.veiculos',
        'create_scope': 'create.veiculos',
        'filial_scope_field': 'filial_id',
    },
    'rotas_carregamento': {
        'table': 'rotas_carregamento',
        'order': 'nome',
        'required_fields': ['filial_id', 'nome'],
        'allowed_fields': ['filial_id', 'codigo', 'nome', 'origem', 'destino', 'ativo', 'observacoes'],
        'partial_match_fields': ['codigo', 'nome', 'origem', 'destino'],
        'nullable_fields': ['codigo', 'origem', 'destino', 'observacoes'],
        'view_scope': 'menu.rotas_carregamento',
        'create_scope': 'create.rotas_carregamento',
        'filial_scope_field': 'filial_id',
    },
    'veiculos_carregamento': {
        'table': 'veiculos_carregamento',
        'order': 'placa',
        'required_fields': ['filial_id', 'placa', 'transportadora'],
        'allowed_fields': [
            'filial_id',
            'rota_id',
            'placa',
            'transportadora',
            'tipo_veiculo',
            'capacidade_cilindros',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['placa', 'transportadora', 'tipo_veiculo'],
        'nullable_fields': ['rota_id', 'tipo_veiculo', 'capacidade_cilindros', 'observacoes'],
        'view_scope': 'menu.veiculos_carregamento',
        'create_scope': 'create.veiculos_carregamento',
        'filial_scope_field': 'filial_id',
    },
    'motivos_parada_carregamento': {
        'table': 'motivos_parada_carregamento',
        'order': 'ordem',
        'required_fields': ['descricao'],
        'allowed_fields': ['descricao', 'ordem', 'exige_observacao', 'ativo'],
        'partial_match_fields': ['descricao'],
        'nullable_fields': ['ordem'],
        'view_scope': 'menu.motivos_parada_carregamento',
        'create_scope': 'create.motivos_parada_carregamento',
    },
    'bonificacao_metricas': {
        'table': 'bonificacao_metricas',
        'order': 'ordem',
        'required_fields': ['nome', 'categoria', 'valor'],
        'allowed_fields': ['nome', 'categoria', 'valor', 'ordem', 'descricao', 'ativo'],
        'partial_match_fields': ['nome', 'categoria', 'descricao'],
        'nullable_fields': ['ordem', 'descricao'],
        'view_scope': 'menu.bonificacao_metricas',
        'create_scope': 'create.bonificacao_metricas',
    },
    'presencas_diarias': {
        'table': 'presencas_diarias',
        'order': 'data_referencia',
        'required_fields': ['data_referencia', 'colaborador_id', 'filial_id', 'status'],
        'allowed_fields': ['data_referencia', 'colaborador_id', 'filial_id', 'status', 'observacoes', 'origem', 'alterado_por'],
        'partial_match_fields': ['status', 'observacoes'],
        'filial_scope_field': 'filial_id',
        'view_scope': 'menu.presenca',
    },
    'pedidos_compra': {
        'table': 'pedidos_compra',
        'order': 'data_pedido',
        'required_fields': ['filial_id', 'data_pedido'],
        'allowed_fields': [
            'filial_id', 'numero_pedido', 'data_pedido', 'data_necessidade',
            'status', 'fornecedor', 'forma_pagamento', 'prazo_pagamento',
            'centro_custo', 'criado_por', 'observacoes', 'valor_total', 'ativo',
            'tipo_reembolso', 'chave_pix', 'dados_bancarios', 'favorecido',
            'aprovado_por', 'aprovado_em', 'em_analise_por', 'em_analise_em',
            'motivo_reprovacao', 'reprovado_por', 'reprovado_em',
            'numero_solicitacao', 'contas_pagar_id', 'data_vencimento',
        ],
        'partial_match_fields': ['numero_pedido', 'fornecedor', 'centro_custo', 'observacoes', 'numero_solicitacao'],
        'nullable_fields': [
            'numero_pedido', 'data_necessidade', 'fornecedor', 'forma_pagamento',
            'prazo_pagamento', 'centro_custo', 'criado_por', 'observacoes',
            'tipo_reembolso', 'chave_pix', 'dados_bancarios', 'favorecido',
            'aprovado_por', 'aprovado_em', 'em_analise_por', 'em_analise_em',
            'motivo_reprovacao', 'reprovado_por', 'reprovado_em',
            'numero_solicitacao', 'contas_pagar_id', 'data_vencimento',
        ],
        'view_scope': 'menu.pedidos_compra',
        'create_scope': 'create.pedidos_compra',
        'filial_scope_field': 'filial_id',
    },
    'itens_catalogo': {
        'table': 'itens_catalogo',
        'order': 'nome',
        'required_fields': ['nome', 'categoria', 'unidade'],
        'allowed_fields': [
            'filial_id', 'nome', 'categoria', 'unidade',
            'valor_referencia', 'fornecedor_habitual', 'observacoes', 'ativo',
        ],
        'partial_match_fields': ['nome', 'categoria', 'fornecedor_habitual', 'observacoes'],
        'nullable_fields': ['filial_id', 'valor_referencia', 'fornecedor_habitual', 'observacoes'],
        'view_scope': 'menu.pedidos_compra',
        'create_scope': 'create.pedidos_compra',
        'filial_scope_field': 'filial_id',
        'filial_scope_include_null': True,
    },
    'fornecedores': {
        'table': 'fornecedores',
        'order': 'nome',
        'required_fields': ['nome'],
        'allowed_fields': ['filial_id', 'nome', 'cnpj', 'telefone', 'email', 'contato_nome', 'categoria', 'endereco', 'observacoes', 'ativo'],
        'partial_match_fields': ['nome', 'cnpj', 'contato_nome', 'categoria', 'observacoes'],
        'nullable_fields': ['filial_id', 'cnpj', 'telefone', 'email', 'contato_nome', 'categoria', 'endereco', 'observacoes'],
        'create_scope': 'create.fornecedores',
        'filial_scope_field': 'filial_id',
        'filial_scope_include_null': True,
    },
    'clientes': {
        'table': 'clientes',
        'order': 'nome',
        'required_fields': ['nome'],
        'allowed_fields': ['filial_id', 'nome', 'cnpj', 'telefone', 'email', 'contato_nome', 'endereco', 'observacoes', 'ativo'],
        'partial_match_fields': ['nome', 'cnpj', 'contato_nome', 'observacoes'],
        'nullable_fields': ['filial_id', 'cnpj', 'telefone', 'email', 'contato_nome', 'endereco', 'observacoes'],
        'create_scope': 'create.clientes',
        'filial_scope_field': 'filial_id',
        'filial_scope_include_null': True,
    },
    'pedidos_compra_itens': {
        'table': 'pedidos_compra_itens',
        'order': 'id',
        'required_fields': ['filial_id', 'pedido_id', 'descricao', 'categoria', 'unidade', 'quantidade', 'valor_unitario'],
        'allowed_fields': [
            'filial_id',
            'pedido_id',
            'descricao',
            'categoria',
            'unidade',
            'quantidade',
            'valor_unitario',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['descricao', 'categoria', 'observacoes'],
        'nullable_fields': ['observacoes'],
        'view_scope': 'menu.pedidos_compra',
        'create_scope': 'create.pedidos_compra',
        'filial_scope_field': 'filial_id',
    },
    'feriados': {
        'table': 'feriados',
        'order': 'data',
        'required_fields': ['nome', 'data', 'tipo'],
        'allowed_fields': [
            'nome',
            'data',
            'tipo',
            'uf',
            'municipio',
            'filial_id',
            'recorrente',
            'ativo',
            'observacoes',
            'tem_expediente',
            'horario_expediente',
        ],
        'partial_match_fields': ['nome', 'uf', 'municipio', 'observacoes'],
        'nullable_fields': ['uf', 'municipio', 'filial_id', 'observacoes', 'horario_expediente'],
        'view_scope': 'menu.feriados',
        'create_scope': 'create.feriados',
    },
    'notas_cte': {
        'table': 'notas_cte',
        'order': 'data_vencimento',
        'required_fields': ['filial_id', 'tipo', 'valor_total'],
        'allowed_fields': [
            'filial_id',
            'tipo',
            'numero_documento',
            'chave_acesso',
            'emitente',
            'destinatario',
            'data_emissao',
            'data_vencimento',
            'data_pagamento',
            'valor_total',
            'descricao',
            'status',
            'categoria',
            'pedido_compra_id',
            'comprovante_url',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['numero_documento', 'chave_acesso', 'emitente', 'destinatario', 'descricao'],
        'nullable_fields': [
            'numero_documento', 'chave_acesso', 'emitente', 'destinatario',
            'data_emissao', 'data_vencimento', 'data_pagamento', 'descricao',
            'categoria', 'pedido_compra_id', 'comprovante_url', 'observacoes',
        ],
        'view_scope': 'menu.notas_cte',
        'create_scope': 'create.notas_cte',
        'filial_scope_field': 'filial_id',
    },
    'cargos': {
        'table': 'cargos',
        'order': 'ordem',
        'required_fields': ['nome'],
        'allowed_fields': ['nome', 'descricao', 'permissoes_padrao', 'ordem', 'ativo'],
        'partial_match_fields': ['nome', 'descricao'],
        'nullable_fields': ['descricao', 'permissoes_padrao', 'ordem'],
        'view_scope': 'menu.cargos',
        'create_scope': 'create.cargos',
    },
    'estoque_itens': {
        'table': 'estoque_itens',
        'order': 'nome',
        'required_fields': ['filial_id', 'nome', 'categoria', 'unidade'],
        'allowed_fields': [
            'filial_id',
            'codigo',
            'nome',
            'descricao',
            'categoria',
            'unidade',
            'estoque_atual',
            'estoque_minimo',
            'localizacao',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['codigo', 'nome', 'descricao', 'localizacao', 'categoria'],
        'nullable_fields': ['codigo', 'descricao', 'localizacao', 'observacoes'],
        'view_scope': 'menu.estoque',
        'create_scope': 'create.estoque',
        'filial_scope_field': 'filial_id',
    },
    'estoque_movimentos': {
        'table': 'estoque_movimentos',
        'order': 'data_movimento',
        'required_fields': ['filial_id', 'item_id', 'tipo', 'quantidade', 'data_movimento'],
        'allowed_fields': [
            'filial_id',
            'item_id',
            'tipo',
            'quantidade',
            'saldo_apos',
            'colaborador_id',
            'fornecedor',
            'numero_nota',
            'motivo',
            'data_movimento',
            'registrado_por',
            'ativo',
            'observacoes',
        ],
        'partial_match_fields': ['tipo', 'fornecedor', 'numero_nota', 'motivo', 'observacoes'],
        'nullable_fields': ['saldo_apos', 'colaborador_id', 'fornecedor', 'numero_nota', 'motivo', 'registrado_por', 'observacoes'],
        'view_scope': 'menu.estoque',
        'create_scope': 'create.estoque_movimentos',
        'filial_scope_field': 'filial_id',
    },
    'veiculos_abastecimentos': {
        'table': 'veiculos_abastecimentos',
        'order': 'data_abastecimento',
        'required_fields': ['filial_id', 'veiculo_id', 'data_abastecimento', 'odometro_km', 'litros', 'valor_litro'],
        'allowed_fields': [
            'filial_id',
            'veiculo_id',
            'data_abastecimento',
            'odometro_km',
            'litros',
            'valor_litro',
            'tipo_combustivel',
            'fornecedor',
            'numero_nota',
            'motorista_id',
            'registrado_por',
            'observacoes',
            'status',
            'ativo',
        ],
        'partial_match_fields': ['fornecedor', 'numero_nota', 'observacoes'],
        'nullable_fields': ['fornecedor', 'numero_nota', 'motorista_id', 'registrado_por', 'observacoes'],
        'view_scope': 'menu.abastecimentos',
        'create_scope': 'create.abastecimentos',
        'filial_scope_field': 'filial_id',
    },
    'veiculos_pneus': {
        'table': 'veiculos_pneus',
        'order': 'veiculo_id',
        'required_fields': ['filial_id', 'veiculo_id', 'posicao', 'status'],
        'allowed_fields': [
            'filial_id',
            'veiculo_id',
            'posicao',
            'marca',
            'modelo',
            'medida',
            'numero_serie',
            'dot',
            'odometro_instalacao',
            'odometro_atual',
            'data_instalacao',
            'vida',
            'status',
            'status_aprovacao',
            'pedido_compra_id',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['posicao', 'marca', 'modelo', 'medida', 'numero_serie', 'observacoes'],
        'nullable_fields': ['marca', 'modelo', 'medida', 'numero_serie', 'dot', 'odometro_instalacao', 'odometro_atual', 'data_instalacao', 'pedido_compra_id', 'observacoes'],
        'view_scope': 'menu.pneus',
        'create_scope': 'create.pneus',
        'filial_scope_field': 'filial_id',
    },
    'manutencoes': {
        'table': 'manutencoes',
        'order': 'data_abertura',
        'required_fields': ['filial_id', 'veiculo_id', 'tipo', 'titulo'],
        'allowed_fields': [
            'filial_id',
            'veiculo_id',
            'tipo',
            'status',
            'titulo',
            'descricao',
            'odometro_entrada',
            'data_abertura',
            'data_previsao',
            'data_inicio',
            'data_conclusao',
            'oficina',
            'responsavel_id',
            'solicitado_por',
            'aprovado_por',
            'aprovado_em',
            'reprovado_por',
            'reprovado_em',
            'motivo_reprovacao',
            'valor_estimado',
            'valor_final',
            'pedido_compra_id',
            'prioridade',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['titulo', 'descricao', 'oficina', 'motivo_reprovacao', 'observacoes'],
        'nullable_fields': [
            'descricao', 'odometro_entrada', 'data_previsao', 'data_inicio', 'data_conclusao',
            'oficina', 'responsavel_id', 'solicitado_por', 'aprovado_por', 'aprovado_em',
            'reprovado_por', 'reprovado_em', 'motivo_reprovacao', 'valor_estimado', 'valor_final',
            'pedido_compra_id', 'observacoes',
        ],
        'view_scope': 'menu.manutencoes',
        'create_scope': 'create.manutencoes',
        'filial_scope_field': 'filial_id',
    },
    'manutencao_itens': {
        'table': 'manutencao_itens',
        'order': 'id',
        'required_fields': ['filial_id', 'manutencao_id', 'tipo', 'descricao', 'quantidade', 'valor_unitario'],
        'allowed_fields': [
            'filial_id',
            'manutencao_id',
            'tipo',
            'descricao',
            'quantidade',
            'unidade',
            'valor_unitario',
            'numero_nota',
            'fornecedor',
            'estoque_item_id',
            'observacoes',
            'ativo',
        ],
        'partial_match_fields': ['descricao', 'fornecedor', 'numero_nota', 'observacoes'],
        'nullable_fields': ['numero_nota', 'fornecedor', 'estoque_item_id', 'observacoes'],
        'view_scope': 'menu.manutencoes',
        'create_scope': 'create.manutencoes',
        'filial_scope_field': 'filial_id',
    },
    'veiculos_documentos': {
        'table': 'veiculos_documentos',
        'order': 'data_validade',
        'required_fields': ['veiculo_id', 'tipo_documento', 'data_validade'],
        'allowed_fields': [
            'veiculo_id',
            'tipo_documento',
            'numero_documento',
            'orgao_emissor',
            'data_emissao',
            'data_validade',
            'prazo_renovacao_dias',
            'arquivo_url',
            'status',
            'observacoes',
        ],
        'partial_match_fields': ['tipo_documento', 'numero_documento', 'orgao_emissor', 'status'],
        'nullable_fields': ['numero_documento', 'orgao_emissor', 'data_emissao', 'prazo_renovacao_dias', 'arquivo_url', 'status', 'observacoes'],
        'view_scope': 'menu.veiculos_documentos',
        'create_scope': 'create.veiculos_documentos',
    },
    'horas_extras': {
        'table': 'horas_extras',
        'order': 'data_solicitacao',
        'required_fields': ['colaborador_id', 'filial_id', 'motivo', 'qtd_horas', 'data_solicitacao'],
        'allowed_fields': [
            'colaborador_id',
            'filial_id',
            'servico_id',
            'jornada_id',
            'motivo',
            'qtd_horas',
            'data_solicitacao',
            'status',
            'justificativa_gestor',
            'data_aprovacao',
            'aprovado_por',
        ],
        'partial_match_fields': ['motivo', 'status', 'justificativa_gestor'],
        'nullable_fields': ['servico_id', 'jornada_id', 'status', 'justificativa_gestor', 'data_aprovacao', 'aprovado_por'],
        'view_scope': 'menu.horas_extras',
        'create_scope': 'create.horas_extras',
        'filial_scope_field': 'filial_id',
    },
}

PERMISSION_SCOPE_GROUPS = [
    # ── Acompanhamento / Aprovações ───────────────────────────────────────────
    {
        'key': 'acompanhamento',
        'title': 'Acompanhamento / Aprovações',
        'items': [
            {'name': 'menu.acompanhamento',     'label': 'Ver tela de acompanhamento',   'platforms': ['web', 'app'], 'auto_enable': [],                                             'description': 'Mostra a tela centralizada de aprovação de manutenções, pedidos de compra e horas extras.'},
            {'name': 'aprovar.pedidos_compra',  'label': 'Aprovar pedidos de compra',    'platforms': ['web', 'app'], 'auto_enable': ['menu.acompanhamento', 'menu.pedidos_compra'], 'description': 'Permite aprovar ou reprovar pedidos de compra.'},
            {'name': 'analisar.pedidos_compra', 'label': 'Analisar pedidos de compra',   'platforms': ['web', 'app'], 'auto_enable': ['menu.acompanhamento', 'menu.pedidos_compra'], 'description': 'Permite colocar um pedido em análise (pendente → analise) e rejeitá-lo nesta etapa.'},
            {'name': 'aprovar.horas_extras',    'label': 'Aprovar horas extras',         'platforms': ['web', 'app'], 'auto_enable': ['menu.acompanhamento', 'menu.horas_extras'],   'description': 'Permite aprovar ou reprovar solicitações de horas extras.'},
            {'name': 'aprovar.manutencoes',     'label': 'Aprovar OS de manutenção',     'platforms': ['web'],        'auto_enable': ['menu.acompanhamento', 'menu.manutencoes'],    'description': 'Permite aprovar ou reprovar ordens de serviço de manutenção.'},
            {'name': 'aprovar.abastecimentos',  'label': 'Aprovar abastecimentos',       'platforms': ['web'],        'auto_enable': ['menu.acompanhamento', 'menu.abastecimentos'], 'description': 'Permite aprovar ou reprovar lançamentos de abastecimento.'},
            {'name': 'aprovar.pneus',           'label': 'Aprovar controle de pneus',    'platforms': ['web'],        'auto_enable': ['menu.acompanhamento', 'menu.pneus'],          'description': 'Permite aprovar ou reprovar lançamentos de controle de pneus.'},
        ],
    },
    # ── RH ────────────────────────────────────────────────────────────────────
    {
        'key': 'rh',
        'title': 'RH',
        'items': [
            {'name': 'menu.colaboradores',           'label': 'Ver colaboradores',                   'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o cadastro de colaboradores no menu.'},
            {'name': 'create.colaboradores',         'label': 'Cadastrar colaboradores',             'platforms': ['web'],        'auto_enable': ['menu.colaboradores'],       'description': 'Permite criar novos colaboradores.'},
            {'name': 'menu.custos_rh',               'label': 'Ver custos RH',                       'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o painel de custos de mão de obra, contratos e acuracidade financeira por base.'},
            {'name': 'menu.contratos_operacionais',  'label': 'Ver contratos operacionais',          'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o cadastro de contratos por base, com valor, headcount e cargos previstos.'},
            {'name': 'create.contratos_operacionais','label': 'Cadastrar contratos operacionais',    'platforms': ['web'],        'auto_enable': ['menu.contratos_operacionais'], 'description': 'Permite criar e manter contratos por base para análise de acuracidade e custos.'},
            {'name': 'menu.colaborador_documentos',  'label': 'Ver documentos RH',                   'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o controle documental de colaboradores, como CNH, ASO e certificados.'},
            {'name': 'create.colaborador_documentos','label': 'Cadastrar documentos RH',             'platforms': ['web'],        'auto_enable': ['menu.colaborador_documentos'], 'description': 'Permite manter vencimentos, arquivos e dados documentais dos colaboradores.'},
            {'name': 'menu.diarias',                 'label': 'Ver diárias / hotelaria',             'platforms': ['web', 'app'], 'auto_enable': [],                          'description': 'Mostra a tela de solicitações de diárias e hotelaria para motoristas em rota.'},
            {'name': 'create.diarias',               'label': 'Solicitar diárias / hotelaria',       'platforms': ['web', 'app'], 'auto_enable': ['menu.diarias'],            'description': 'Permite criar solicitações de diárias/hotelaria. Gestores de base usam pra abrir o pedido.'},
            {'name': 'aprovar.diarias',              'label': 'Aprovar diárias / hotelaria',         'platforms': ['web'],        'auto_enable': ['menu.acompanhamento', 'menu.diarias'], 'description': 'Permite aprovar ou reprovar pedidos de diárias e ajustar valores na tela de Acompanhamento — geralmente quem atende o financeiro.'},
            {'name': 'menu.colaborador_contratos',   'label': 'Ver contratos de colaboradores',      'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra os vínculos contratuais (CLT, estágio, PJ, temporário, aprendiz) e o histórico 45+45→indeterminado.'},
            {'name': 'create.colaborador_contratos', 'label': 'Cadastrar contratos de colaboradores','platforms': ['web'],        'auto_enable': ['menu.colaborador_contratos'], 'description': 'Permite registrar e prorrogar vínculos contratuais e desligamentos.'},
            {'name': 'menu.eventos_rh',              'label': 'Ver planejamento RH',                 'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra férias, afastamentos, licenças e demais eventos planejados do RH.'},
            {'name': 'create.eventos_rh',            'label': 'Cadastrar eventos RH',                'platforms': ['web'],        'auto_enable': ['menu.eventos_rh'],          'description': 'Permite planejar férias, afastamentos, licenças e folgas programadas.'},
            {'name': 'menu.horas_extras',            'label': 'Ver horas extras',                    'platforms': ['web', 'app'], 'auto_enable': [],                          'description': 'Mostra o registro e aprovação de horas extras por colaborador e filial.'},
            {'name': 'create.horas_extras',          'label': 'Lançar horas extras',                 'platforms': ['web', 'app'], 'auto_enable': ['menu.horas_extras'],        'description': 'Permite registrar solicitações de horas extras para colaboradores.'},
            {'name': 'menu.horas_extras_rtm',        'label': 'Calc. Horas Extras (RTM)',            'platforms': ['web'],        'auto_enable': [],                          'description': 'Calculadora RTM: cola dados da planilha e calcula totais de horas extras. Disponível apenas no web (usa colagem de planilha).'},
            {'name': 'menu.quadro_funcionarios',     'label': 'Ver quadro de funcionários',          'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o quadro por base com totais e status da equipe.'},
            {'name': 'menu.bonificacao',             'label': 'Ver bonificação',                     'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o controle mensal de bonificação por colaborador e os totais pagos.'},
            {'name': 'manage.bonificacao',           'label': 'Modificar bonificação',               'platforms': ['web'],        'auto_enable': ['menu.bonificacao'],         'description': 'Permite alterar lançamentos mensais de bonificação por colaborador.'},
            {'name': 'menu.bonificacao_metricas',    'label': 'Ver métricas de bonificação',         'platforms': ['web'],        'auto_enable': [],                          'description': 'Mostra o cadastro das métricas e valores que compõem a bonificação.'},
            {'name': 'create.bonificacao_metricas',  'label': 'Cadastrar métricas de bonificação',   'platforms': ['web'],        'auto_enable': ['menu.bonificacao_metricas'], 'description': 'Permite criar e editar métricas usadas no cálculo da bonificação.'},
        ],
    },
    # ── Frota ─────────────────────────────────────────────────────────────────
    {
        'key': 'frota',
        'title': 'Frota',
        'items': [
            {'name': 'menu.veiculos',              'label': 'Ver veículos',                  'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o cadastro de veículos no menu.'},
            {'name': 'create.veiculos',            'label': 'Cadastrar veículos',            'platforms': ['web'], 'auto_enable': ['menu.veiculos'],           'description': 'Permite criar novos veículos.'},
            {'name': 'menu.frota_dashboard',       'label': 'Ver dashboard de frota',        'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o painel consolidado de frota com custos, manutenções e consumo.'},
            {'name': 'menu.manutencoes',           'label': 'Ver manutenções',               'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o módulo de ordens de serviço de manutenção de veículos.'},
            {'name': 'create.manutencoes',         'label': 'Abrir OS de manutenção',        'platforms': ['web'], 'auto_enable': ['menu.manutencoes'],        'description': 'Permite criar ordens de serviço de manutenção e adicionar peças/serviços.'},
            {'name': 'menu.abastecimentos',        'label': 'Ver abastecimentos',            'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o histórico de abastecimentos por veículo e filial.'},
            {'name': 'create.abastecimentos',      'label': 'Registrar abastecimento',       'platforms': ['web'], 'auto_enable': ['menu.abastecimentos'],     'description': 'Permite lançar abastecimentos de combustível por veículo.'},
            {'name': 'menu.pneus',                 'label': 'Ver controle de pneus',         'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o controle de pneus por posição e veículo.'},
            {'name': 'create.pneus',               'label': 'Gerenciar pneus',               'platforms': ['web'], 'auto_enable': ['menu.pneus'],              'description': 'Permite cadastrar e atualizar o controle de pneus dos veículos.'},
            {'name': 'menu.veiculos_documentos',   'label': 'Ver documentos de frota',       'platforms': ['web'], 'auto_enable': [],                         'description': 'Mostra o controle documental dos veículos: CRLV, seguro, licenças e outros.'},
            {'name': 'create.veiculos_documentos', 'label': 'Cadastrar documentos de frota', 'platforms': ['web'], 'auto_enable': ['menu.veiculos_documentos'], 'description': 'Permite cadastrar e manter documentos dos veículos como CRLV, seguro e licenças.'},
        ],
    },
    # ── Compras ───────────────────────────────────────────────────────────────
    {
        'key': 'compras',
        'title': 'Compras',
        'items': [
            {'name': 'menu.pedidos_compra',   'label': 'Ver pedidos de compra',    'platforms': ['web', 'app'], 'auto_enable': [],                     'description': 'Permite visualizar pedidos de compra e acompanhar aprovações. No web também libera o formulário de criação.'},
            {'name': 'create.pedidos_compra', 'label': 'Criar pedidos de compra',  'platforms': ['web'],        'auto_enable': ['menu.pedidos_compra'], 'description': 'Permite criar pedidos de compra, adicionar itens e alterar status.'},
            {'name': 'menu.fornecedores',     'label': 'Ver fornecedores',         'platforms': ['web'],        'auto_enable': [],                     'description': 'Acesso ao cadastro de fornecedores para sugestão nos pedidos de compra.'},
            {'name': 'create.fornecedores',   'label': 'Cadastrar fornecedores',   'platforms': ['web'],        'auto_enable': ['menu.fornecedores'],   'description': 'Permite criar e editar fornecedores no cadastro.'},
            {'name': 'menu.clientes',         'label': 'Ver clientes',             'platforms': ['web'],        'auto_enable': [],                     'description': 'Acesso ao cadastro de clientes para sugestão nos contratos e contas a receber.'},
            {'name': 'create.clientes',       'label': 'Cadastrar clientes',       'platforms': ['web'],        'auto_enable': ['menu.clientes'],       'description': 'Permite criar e editar clientes no cadastro.'},
            {'name': 'menu.notas_cte',        'label': 'Ver notas / CT-e',         'platforms': ['web'],        'auto_enable': [],                     'description': 'Permite lançar e controlar notas fiscais, CT-e e faturas pendentes de pagamento.'},
            {'name': 'create.notas_cte',      'label': 'Lançar notas / CT-e',      'platforms': ['web'],        'auto_enable': ['menu.notas_cte'],      'description': 'Permite criar, editar e quitar notas fiscais, CT-e e faturas.'},
            {'name': 'menu.feriados',         'label': 'Ver feriados',             'platforms': ['web'],        'auto_enable': [],                     'description': 'Permite visualizar e gerenciar feriados por UF, município e filial no calendário.'},
            {'name': 'create.feriados',       'label': 'Cadastrar feriados',       'platforms': ['web'],        'auto_enable': ['menu.feriados'],       'description': 'Permite criar e editar feriados no calendário.'},
        ],
    },
    # ── Financeiro ────────────────────────────────────────────────────────────
    {
        'key': 'financeiro',
        'title': 'Financeiro',
        'items': [
            {'name': 'menu.contas_receber',   'label': 'Ver contas a receber',     'platforms': ['web'], 'auto_enable': [],                      'description': 'Controle de obrigações a receber de clientes, com status, documentos e SLA.'},
            {'name': 'create.contas_receber', 'label': 'Lançar contas a receber',  'platforms': ['web'], 'auto_enable': ['menu.contas_receber'],  'description': 'Permite criar e editar lançamentos de contas a receber.'},
            {'name': 'menu.contas_pagar',     'label': 'Ver contas a pagar',       'platforms': ['web'], 'auto_enable': [],                      'description': 'Controle de obrigações a pagar: fornecedores, horas extras Gold, hospedagens e despesas.'},
            {'name': 'create.contas_pagar',   'label': 'Lançar contas a pagar',    'platforms': ['web'], 'auto_enable': ['menu.contas_pagar'],    'description': 'Permite criar e editar lançamentos de contas a pagar.'},
            {'name': 'menu.banco',            'label': 'Ver banco / conciliação',  'platforms': ['web'], 'auto_enable': [],                      'description': 'Gestão de contas bancárias, lançamentos e conciliação com contas a receber e pagar.'},
            {'name': 'create.banco',          'label': 'Lançar no banco',          'platforms': ['web'], 'auto_enable': ['menu.banco'],           'description': 'Permite cadastrar contas bancárias, lançamentos e marcar conciliações.'},
        ],
    },
    # ── Estoque ───────────────────────────────────────────────────────────────
    {
        'key': 'estoque',
        'title': 'Estoque',
        'items': [
            {'name': 'menu.estoque',             'label': 'Ver estoque',                  'platforms': ['web', 'app'], 'auto_enable': [],                                     'description': 'Mostra o módulo de estoque no menu. No app: aba "Estoque & Trocas".'},
            {'name': 'create.estoque',           'label': 'Cadastrar estoque',            'platforms': ['web'],        'auto_enable': ['menu.estoque'],                       'description': 'Permite criar novos itens de estoque. Disponível apenas no web.'},
            {'name': 'menu.estoque_movimentos',  'label': 'Ver movimentos de estoque',    'platforms': ['web', 'app'], 'auto_enable': ['menu.estoque'],                       'description': 'Mostra a tela de lançamento de movimentos de estoque (entradas, saídas, trocas). No app: usado para registrar trocas de equipamento.'},
            {'name': 'create.estoque_movimentos','label': 'Lançar movimentos de estoque', 'platforms': ['web', 'app'], 'auto_enable': ['menu.estoque', 'menu.estoque_movimentos'], 'description': 'Permite registrar entradas, saídas, trocas, devoluções e ajustes de estoque.'},
        ],
    },
    # ── Operação RTM ──────────────────────────────────────────────────────────
    {
        'key': 'operacao_rtm',
        'title': 'Operação RTM',
        'items': [
            {'name': 'menu.presenca',                    'label': 'Ver presença',                    'platforms': ['web', 'app'], 'auto_enable': [],                      'description': 'Mostra o controle diário de presença.'},
            {'name': 'manage.presenca',                  'label': 'Modificar presença',              'platforms': ['web', 'app'], 'auto_enable': ['menu.presenca'],        'description': 'Permite alterar e salvar o quadro diário de presença.'},
            {'name': 'menu.carregamento',                'label': 'Ver carregamento',                'platforms': ['web', 'app'], 'auto_enable': [],                      'description': 'Mostra o painel de carregamento por turno, caminhão, referência operacional e eventos.'},
            {'name': 'manage.programacao_carregamento',  'label': 'Programar carregamento',          'platforms': ['web', 'app'], 'auto_enable': ['menu.carregamento'],    'description': 'Permite abrir jornadas e definir quais caminhões vão operar no turno.'},
            {'name': 'manage.operacao_carregamento',     'label': 'Operar carregamento',             'platforms': ['web', 'app'], 'auto_enable': ['menu.carregamento'],    'description': 'Permite registrar carga, paradas, ocorrências e fechamento do caminhão.'},
            {'name': 'menu.rotas_carregamento',          'label': 'Ver referências de carregamento', 'platforms': ['web'],        'auto_enable': [],                      'description': 'Mostra o cadastro das referências operacionais usadas no carregamento.'},
            {'name': 'create.rotas_carregamento',        'label': 'Cadastrar referências',           'platforms': ['web'],        'auto_enable': ['menu.rotas_carregamento'], 'description': 'Permite criar referências operacionais usadas na operação de carregamento.'},
            {'name': 'menu.veiculos_carregamento',       'label': 'Ver veículos de carregamento',    'platforms': ['web'],        'auto_enable': [],                      'description': 'Mostra o cadastro dos caminhões terceiros usados na operação.'},
            {'name': 'create.veiculos_carregamento',     'label': 'Cadastrar veículos de carregamento','platforms': ['web'],      'auto_enable': ['menu.veiculos_carregamento'], 'description': 'Permite criar caminhões terceiros usados no carregamento.'},
            {'name': 'menu.motivos_parada_carregamento', 'label': 'Ver motivos de parada',           'platforms': ['web'],        'auto_enable': [],                      'description': 'Mostra o cadastro dos motivos padrão de parada operacional.'},
            {'name': 'create.motivos_parada_carregamento','label': 'Cadastrar motivos de parada',    'platforms': ['web'],        'auto_enable': ['menu.motivos_parada_carregamento'], 'description': 'Permite manter os motivos padrão de parada operacional.'},
        ],
    },
    # ── Administração ─────────────────────────────────────────────────────────
    {
        'key': 'administracao',
        'title': 'Administração',
        'items': [
            {'name': 'menu.dashboard',      'label': 'Ver dashboard',          'platforms': ['web'], 'auto_enable': [],              'description': 'Permite abrir o painel inicial.'},
            {'name': 'menu.filiais',        'label': 'Ver filiais',            'platforms': ['web'], 'auto_enable': [],              'description': 'Mostra o cadastro de filiais no menu.'},
            {'name': 'create.filiais',      'label': 'Cadastrar filiais',      'platforms': ['web'], 'auto_enable': ['menu.filiais'], 'description': 'Permite criar novas filiais.'},
            {'name': 'menu.auditoria',      'label': 'Ver auditoria',          'platforms': ['web'], 'auto_enable': [],              'description': 'Mostra a trilha de movimentações e alterações realizadas no sistema.'},
            {'name': 'menu.gestao_acessos', 'label': 'Ver gestão de acessos',  'platforms': ['web'], 'auto_enable': [],              'description': 'Visualizar colaboradores com acesso ao sistema e redefinir senhas.'},
            {'name': 'menu.permissoes',     'label': 'Ver permissões',         'platforms': ['web'], 'auto_enable': [],              'description': 'Mostra a tela de gestão de permissões.'},
            {'name': 'menu.cargos',         'label': 'Ver cargos / funções',   'platforms': ['web'], 'auto_enable': [],              'description': 'Mostra o cadastro de cargos e modelos de permissões por função.'},
            {'name': 'create.cargos',       'label': 'Cadastrar cargos',       'platforms': ['web'], 'auto_enable': ['menu.cargos'], 'description': 'Permite criar e editar cargos e seus modelos de permissão.'},
        ],
    },
    # ── Ações granulares (botões específicos dentro de cada tela) ─────────────
    # Comportamento: se o usuário NÃO tem nenhum escopo "action.*" configurado,
    # todos os botões aparecem (compat retroativa). Quando o admin marca ao
    # menos UM escopo de ação, apenas os marcados aparecem para esse usuário.
    {
        'key': 'acoes_granulares',
        'title': 'Ações granulares (botões dentro das telas)',
        'items': [
            {'name': 'action.documentos_rh.importar', 'label': 'Importar documentos RH (Excel/PDF lote)', 'platforms': ['web'], 'auto_enable': ['menu.colaborador_documentos'], 'description': 'Mostra os botões "Importar Excel" e "Upload em lote" na tela Documentos RH.'},
            {'name': 'action.documentos_rh.exportar', 'label': 'Exportar documentos RH para Excel',        'platforms': ['web'], 'auto_enable': ['menu.colaborador_documentos'], 'description': 'Mostra o botão "Exportar Excel" e ações de exportação em lote.'},
            {'name': 'action.documentos_rh.renovar',  'label': 'Renovar documentos em lote',               'platforms': ['web'], 'auto_enable': ['menu.colaborador_documentos'], 'description': 'Mostra o botão "♻ Renovar" na barra de ações em lote (cria novo registro com validade calculada).'},
            {'name': 'action.documentos_rh.inativar', 'label': 'Inativar / marcar n/a documentos',         'platforms': ['web'], 'auto_enable': ['menu.colaborador_documentos'], 'description': 'Mostra os botões "Inativar" e "Marcar n/a" na barra de ações em lote.'},
            {'name': 'action.documentos_rh.excluir',  'label': 'Excluir documentos RH (definitivo)',       'platforms': ['web'], 'auto_enable': ['menu.colaborador_documentos'], 'description': 'Mostra o botão "Excluir" na barra de ações em lote. Ação destrutiva e irreversível.'},
            {'name': 'action.pedidos_compra.editar_analise', 'label': 'Editar pedido em análise',         'platforms': ['web'], 'auto_enable': ['menu.pedidos_compra'],         'description': 'Permite alterar campos do pedido enquanto ele está no status "em análise" (antes de aprovar).'},
            {'name': 'action.global.bulk_export',     'label': 'Exportar listas em massa (genérico)',      'platforms': ['web'], 'auto_enable': [],                              'description': 'Mostra botões "Exportar selecionados" e "Exportar tudo" em listas que suportam exportação.'},
            {'name': 'action.global.bulk_delete',     'label': 'Excluir em massa (genérico)',              'platforms': ['web'], 'auto_enable': [],                              'description': 'Permite seleção múltipla + exclusão em lote em listas que suportam (ação destrutiva).'},
        ],
    },
]

ALL_PERMISSION_SCOPES = [
    item['name']
    for group in PERMISSION_SCOPE_GROUPS
    for item in group['items']
]

PERMISSION_SCOPE_MAP = {
    item['name']: item
    for group in PERMISSION_SCOPE_GROUPS
    for item in group['items']
}

MANAGED_COLLABORATOR_PERMISSION_FIELDS = [
    'permissao_app',
    'permissao_desktop',
    'permissao_editar',
    'permissao_excluir',
    'permissao_aprovar_he',
    'ativo',
]

PRESENCE_STATUS_OPTIONS = [
    'pendente',
    'presente',
    'falta',
    'folga',
    'atestado',
    'ferias',
    'afastado',
]

RH_EVENT_STATUS_OPTIONS = ['planejado', 'aprovado', 'em_andamento', 'concluido', 'cancelado']
RH_EVENT_TYPES = {
    'ferias': 'ferias',
    'afastamento': 'afastado',
    'licenca': 'afastado',
    'atestado': 'atestado',
    'folga_programada': 'folga',
}

WEEKDAY_TOKEN_TO_INDEX = {
    'seg': 0,
    'segunda': 0,
    'ter': 1,
    'terca': 1,
    'terça': 1,
    'qua': 2,
    'quarta': 2,
    'qui': 3,
    'quinta': 3,
    'sex': 4,
    'sexta': 4,
    'sab': 5,
    'sabado': 5,
    'sábado': 5,
    'dom': 6,
    'domingo': 6,
}

LOADING_SHIFT_OPTIONS = ['manha', 'tarde', 'noite']
LOADING_JOURNEY_STATUS_OPTIONS = ['planejado', 'em_operacao', 'finalizado', 'cancelado']
BONIFICACAO_CATEGORIAS = ['individual', 'coletivo']
HIDDEN_COLLABORATOR_NAMES = {'mestre'}

# ============================================================================
# ENUMS PARA VALIDAÇÃO (Garante integridade dos dados)
# ============================================================================
TIPO_ITEM_VALID_VALUES = {
    'colaborador',
    'colaborador_fora_contrato',
    'veiculo',
    'veiculo_proprio',
    'pacote_motorista_veiculo',
    'caminhao',
    'outro',
}

PRESENCE_STATUS_ENUM = set(PRESENCE_STATUS_OPTIONS)
RH_EVENT_STATUS_ENUM = set(RH_EVENT_STATUS_OPTIONS)

# ============================================================================
# CACHE SIMPLES EM MEMÓRIA (Performance + Confiabilidade)
# ============================================================================
# Cache com expiração automática para dashboard de custos
DASHBOARD_CACHE = {}
DASHBOARD_CACHE_TTL = 3600  # 1 hora em segundos

def get_cached_dashboard(filial_id, month_reference):
    """Retorna dashboard do cache se válido, senão None."""
    cache_key = f"{filial_id}_{month_reference}"
    if cache_key in DASHBOARD_CACHE:
        cached_data, timestamp = DASHBOARD_CACHE[cache_key]
        # Verifica se cache ainda é válido
        if (datetime.now() - timestamp).total_seconds() < DASHBOARD_CACHE_TTL:
            return cached_data
        else:
            # Remove cache expirado
            del DASHBOARD_CACHE[cache_key]
    return None

def set_cached_dashboard(filial_id, month_reference, data):
    """Armazena dashboard no cache com timestamp."""
    cache_key = f"{filial_id}_{month_reference}"
    DASHBOARD_CACHE[cache_key] = (data, datetime.now())

def invalidate_dashboard_cache(filial_id=None):
    """Invalida cache do dashboard (por filial ou global)."""
    if filial_id:
        # Remove apenas cache da filial específica
        keys_to_remove = [k for k in DASHBOARD_CACHE.keys() if k.startswith(f"{filial_id}_")]
        for key in keys_to_remove:
            del DASHBOARD_CACHE[key]
    else:
        # Remove cache global
        DASHBOARD_CACHE.clear()

# ============================================================================
# FUNÇÕES DE VALIDAÇÃO (Garante dados confiáveis no banco)
# ============================================================================

def validate_tipo_item(tipo_item_value):
    """
    Valida se tipo_item está nos valores aceitos.
    Retorna True se válido, False caso contrário.
    """
    if not tipo_item_value:
        return False
    
    normalized = str(tipo_item_value).strip().lower()
    return normalized in TIPO_ITEM_VALID_VALUES

def validate_presence_status(status_value):
    """Valida se status de presença está nos valores aceitos."""
    if not status_value:
        return False
    return status_value.lower().strip() in PRESENCE_STATUS_ENUM

def sanitize_decimal(value, default=0.0):
    """Sanitiza valor decimal: remove NaN, Inf, garante positivo se necessário."""
    try:
        float_val = float(value) if value is not None else default
        # Garante que não é NaN ou Inf
        if not (-10**10 < float_val < 10**10):  # Limite razoável
            return default
        return float_val
    except (TypeError, ValueError):
        return default

def validate_colaborador_beneficios_entry(entry):
    """
    Valida entrada de benefício do colaborador.
    Retorna (válido, erros) tuple.
    """
    erros = []
    
    # Validar valor_mensal (não negativo)
    if entry.get('valor_mensal') is not None:
        try:
            valor = float(entry.get('valor_mensal', 0))
            if valor < 0:
                erros.append("valor_mensal não pode ser negativo")
        except (TypeError, ValueError):
            erros.append("valor_mensal deve ser um número")
    
    # Validar tipo_beneficio (não vazio)
    if not entry.get('tipo_beneficio', '').strip():
        erros.append("tipo_beneficio é obrigatório")
    
    # Validar que colaborador_id existe
    if not entry.get('colaborador_id'):
        erros.append("colaborador_id é obrigatório")
    
    return len(erros) == 0, erros

def validate_contratos_colaboradores_entry(entry):
    """
    Valida entrada de colaborador em contrato.
    Retorna (válido, erros) tuple.
    """
    erros = []
    
    # Validar tipo_item (enum)
    tipo_item = entry.get('tipo_item', '').strip()
    if not validate_tipo_item(tipo_item):
        erros.append(f"tipo_item '{tipo_item}' inválido. Valores aceitos: {', '.join(sorted(TIPO_ITEM_VALID_VALUES))}")
    
    # Validar que contrato_id existe
    if not entry.get('contrato_operacional_id'):
        erros.append("contrato_operacional_id é obrigatório")
    
    # Validar percentual_alocacao (0-100) se presente
    if entry.get('percentual_alocacao') is not None:
        try:
            percentual = float(entry.get('percentual_alocacao', 0))
            if not (0 <= percentual <= 100):
                erros.append("percentual_alocacao deve estar entre 0 e 100")
        except (TypeError, ValueError):
            erros.append("percentual_alocacao deve ser um número")
    
    # Validar valor_cobrado (não negativo) se presente
    if entry.get('valor_cobrado_colaborador') is not None:
        try:
            valor = float(entry.get('valor_cobrado_colaborador', 0))
            if valor < 0:
                erros.append("valor_cobrado_colaborador não pode ser negativo")
        except (TypeError, ValueError):
            erros.append("valor_cobrado_colaborador deve ser um número")
    
    # Validar datas de vigência
    data_inicio = entry.get('data_inicio_vinculo')
    data_fim = entry.get('data_fim_vinculo')
    if data_inicio and data_fim:
        try:
            # Tenta fazer parse simples (esperando ISO format)
            from datetime import datetime as dt
            inicio = dt.fromisoformat(str(data_inicio).split('T')[0])
            fim = dt.fromisoformat(str(data_fim).split('T')[0])
            if fim < inicio:
                erros.append("data_fim_vinculo não pode ser anterior a data_inicio_vinculo")
        except (ValueError, TypeError):
            erros.append("datas de vigência em formato inválido")
    
    return len(erros) == 0, erros


def create_app():
    app = Flask(__name__)
    app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2 MB por requisição
    CORS(
        app,
        resources={r'/api/*': {'origins': os.getenv('FRONTEND_URL', 'http://localhost:5173')}},
        supports_credentials=True,
    )

    @app.after_request
    def add_security_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
        response.headers['X-Permitted-Cross-Domain-Policies'] = 'none'
        return response

    supabase_url = os.getenv('SUPABASE_URL')
    supabase_anon_key = os.getenv('SUPABASE_ANON_KEY')
    supabase_service_role_key = os.getenv('SUPABASE_SERVICE_ROLE_KEY')
    supabase_secret_key = os.getenv('SUPABASE_SECRET_KEY')
    supabase_server_key = supabase_service_role_key or supabase_secret_key

    if not supabase_url or not supabase_server_key:
        raise RuntimeError('Defina SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY ou SUPABASE_SECRET_KEY no backend/.env')

    supabase: Client = create_client(supabase_url, supabase_server_key)
    auth_cache_ttl = int(os.getenv('AUTH_CACHE_TTL_SECONDS', '30'))
    auth_cache = {}
    _profile_cache = {}          # user_id → {'profile': ..., 'expires_at': float}
    _profile_cache_ttl = 120     # seconds — rebuild profile after 2 min

    # ─── Reconexão automática Supabase (HTTP/2 fecha conexões ociosas) ────────
    _sb_lock = threading.Lock()
    _sb_last_check = [0.0]
    _sb_check_interval = float(os.getenv('SUPABASE_KEEPALIVE_INTERVAL', '20'))

    def _is_disconnect_err(exc):
        msg = str(exc).lower()
        return 'server disconnected' in msg or 'connection reset' in msg or 'remoteprotocol' in msg

    @app.before_request
    def _supabase_keepalive():
        nonlocal supabase
        now = time.time()
        if now - _sb_last_check[0] < _sb_check_interval:
            return
        with _sb_lock:
            if now - _sb_last_check[0] < _sb_check_interval:
                return
            _sb_last_check[0] = time.time()
            try:
                supabase.table('filiais').select('id').limit(1).execute()
            except Exception as ping_exc:
                if _is_disconnect_err(ping_exc):
                    app.logger.warning('Supabase keepalive: desconectado, recriando cliente...')
                    try:
                        supabase = create_client(supabase_url, supabase_server_key)
                    except Exception as rc_exc:
                        app.logger.error('Supabase: falha ao recriar cliente: %s', rc_exc)

    # ─── Rate limiter em memória (janela fixa por IP + endpoint) ─────────────
    _rl_buckets: dict = {}
    _rl_lock = threading.Lock()

    def _is_rate_limited(key: str, max_requests: int = 60, window_seconds: int = 60) -> bool:
        now_val = time.time()
        with _rl_lock:
            bucket = _rl_buckets.get(key)
            if bucket is None or now_val - bucket[0] >= window_seconds:
                _rl_buckets[key] = (now_val, 1)
                # Limpeza periódica para evitar crescimento ilimitado
                if len(_rl_buckets) > 10000:
                    cutoff = now_val - window_seconds
                    stale = [k for k, v in list(_rl_buckets.items()) if v[0] < cutoff]
                    for k in stale:
                        _rl_buckets.pop(k, None)
                return False
            count = bucket[1] + 1
            _rl_buckets[key] = (bucket[0], count)
            return count > max_requests

    def rate_limit_endpoint(max_requests: int = 120, window_seconds: int = 60):
        """Decorador de rate limiting por IP. Aplica ANTES de @require_auth."""
        def decorator(f):
            @wraps(f)
            def wrapper(*args, **kwargs):
                client_ip = (
                    (request.headers.get('X-Forwarded-For', '') or request.remote_addr or '')
                    .split(',')[0].strip()
                ) or 'unknown'
                if _is_rate_limited(f'{f.__name__}:{client_ip}', max_requests, window_seconds):
                    return jsonify({'error': 'Muitas requisições. Aguarde um momento e tente novamente.'}), 429
                return f(*args, **kwargs)
            return wrapper
        return decorator
    # ─────────────────────────────────────────────────────────────────────────

    table_ready_cache_ttl_seconds = max(15, int(os.getenv('TABLE_READY_CACHE_TTL_SECONDS', '120')))
    table_ready_cache = {}
    alerts_interval_minutes = max(1, int(os.getenv('ALERTS_INTERVAL_MINUTES', '30')))
    alerts_horizon_days = max(1, int(os.getenv('ALERTS_HORIZON_DAYS', '7')))
    alerts_enabled = os.getenv('ALERTS_ENGINE_ENABLED', 'true').strip().lower() not in {'0', 'false', 'no', 'off'}
    alert_engine_state = {
        'last_run_at': None,
        'last_error': None,
        'summary': {
            'total': 0,
            'critical': 0,
            'warning': 0,
            'info': 0,
        },
        'items': [],
    }
    alert_engine_lock = threading.Lock()

    def suggest_filial_code(*values):
        for raw_value in values:
            if not isinstance(raw_value, str):
                continue

            normalized_value = unicodedata.normalize('NFKD', raw_value)
            ascii_value = normalized_value.encode('ascii', 'ignore').decode('ascii').upper().strip()
            compact_value = re.sub(r'[^A-Z0-9]+', '', ascii_value)
            if compact_value:
                return compact_value[:4]

        return ''

    def decorate_filial_row(row):
        if not isinstance(row, dict):
            return row

        normalized_row = dict(row)
        codigo = normalized_row.get('codigo')
        if isinstance(codigo, str):
            codigo = codigo.strip().upper()

        normalized_row['codigo'] = codigo or suggest_filial_code(
            normalized_row.get('cidade'),
            normalized_row.get('parceira'),
        )
        return normalized_row

    def fetch_filial_code(filial_id):
        if not filial_id:
            return ''

        response = (
            supabase.table('filiais')
            .select('*')
            .eq('id', filial_id)
            .limit(1)
            .execute()
        )
        filial = response.data[0] if response.data else None
        if not filial:
            return ''

        return decorate_filial_row(filial).get('codigo') or ''

    def auth_token_from_request():
        header = request.headers.get('Authorization', '')
        if not header.startswith('Bearer '):
            return None
        token = header.split(' ', 1)[1].strip()
        if not token or token.lower() in {'undefined', 'null'}:
            return None
        return token

    def get_cached_auth_entry(cache_key):
        cached = auth_cache.get(cache_key)
        if not cached:
            return None
        if cached['expires_at'] <= time.time():
            auth_cache.pop(cache_key, None)
            return None
        return cached['value']

    def set_cached_auth_entry(cache_key, value):
        # Evita crescimento ilimitado: purga vencidos e, se ainda cheio, descarta os mais antigos
        if len(auth_cache) >= 2000:
            now_ts = time.time()
            expired = [k for k, v in list(auth_cache.items()) if v.get('expires_at', 0) <= now_ts]
            for k in expired:
                auth_cache.pop(k, None)
            if len(auth_cache) >= 2000:
                for k in list(auth_cache)[:500]:
                    auth_cache.pop(k, None)
        auth_cache[cache_key] = {
            'value': value,
            'expires_at': time.time() + auth_cache_ttl,
        }

    def translate_database_error(exc):
        raw_message = exc.args[0] if getattr(exc, 'args', None) else exc
        message = raw_message if isinstance(raw_message, str) else str(raw_message)

        if 'duplicate key value' in message.lower() or 'duplicate' in message.lower():
            normalized = message.lower()
            if 'colaboradores_cpf_key' in normalized and '000.000.000.00' in normalized:
                return 'CPF placeholder repetido bloqueado por constraint única no banco. Rode a migration de CPF placeholder para permitir repetição do 000.000.000.00.'
            if 'user_id' in normalized:
                return 'Já existe colaborador vinculado a este usuário/e-mail.'
            if 'cpf' in normalized:
                return 'Já existe colaborador com este CPF.'
            if 'nome_completo' in normalized:
                return 'Já existe colaborador com este nome no contexto atual.'
            return 'Já existe um registro com esses dados únicos.'
        if 'foreign key' in message.lower():
            return 'O registro informado referencia um vínculo inexistente.'
        if 'violates check constraint' in message.lower() or 'check constraint' in message.lower():
            import re as _re
            m = _re.search(r'check constraint "([^"]+)"', message)
            constraint = m.group(1) if m else 'desconhecida'
            if 'categoria' in constraint.lower():
                return 'Categoria inválida. Use um destes valores: pessoal, saude, habilitacao, treinamento, contratual.'
            return f'Valor inválido para a coluna (constraint: {constraint}). Confira os campos.'
        if 'does not exist' in message.lower():
            return 'Estrutura de banco ausente para este módulo. Rode a migration correspondente.'
        if 'jwt' in message.lower() or 'token' in message.lower():
            return 'Sua sessão expirou ou o token é inválido. Faça login novamente.'
        return message

    def is_missing_relation_error(exc):
        message = str(exc).lower()
        return 'does not exist' in message

    def parse_iso_date(value):
        try:
            return date_class.fromisoformat(value)
        except (TypeError, ValueError):
            return None

    def parse_iso_datetime(value):
        try:
            return datetime.fromisoformat(value)
        except (TypeError, ValueError):
            return None

    def normalize_cpf(value):
        placeholder_cpf = '000.000.000.00'
        if value is None:
            return placeholder_cpf

        digits = re.sub(r'\D', '', str(value))
        if not digits:
            return placeholder_cpf

        if len(digits) == 11:
            if digits == '00000000000':
                return placeholder_cpf
            # Mantém CPF no padrão solicitado para exibição/importação em planilha.
            return f'{digits[:3]}.{digits[3:6]}.{digits[6:9]}.{digits[9:]}'

        return str(value).strip()

    def parse_import_bool(value, default=False):
        if value is None:
            return default

        if isinstance(value, bool):
            return value

        normalized = str(value).strip().lower()
        if normalized in {'1', 'true', 'verdadeiro', 'sim', 's'}:
            return True
        if normalized in {'0', 'false', 'falso', 'nao', 'não', 'n'}:
            return False
        return default

    def normalize_import_date(value):
        if value is None:
            return None

        raw_value = str(value).strip()
        if not raw_value:
            return None

        iso_value = parse_iso_date(raw_value)
        if iso_value:
            return iso_value.isoformat()

        br_match = re.fullmatch(r'(\d{2})/(\d{2})/(\d{4})', raw_value)
        if br_match:
            day, month, year = br_match.groups()
            parsed = parse_iso_date(f'{year}-{month}-{day}')
            return parsed.isoformat() if parsed else None

        return raw_value

    def parse_float_or_default(value, default_value=0.0):
        if value in (None, ''):
            return float(default_value)

        if isinstance(value, (int, float)):
            return float(value)

        normalized = str(value).strip().replace('R$', '').replace(' ', '')
        if ',' in normalized and '.' in normalized:
            normalized = normalized.replace('.', '').replace(',', '.')
        elif ',' in normalized:
            normalized = normalized.replace(',', '.')

        try:
            return float(normalized)
        except ValueError:
            return float(default_value)

    def parse_int_or_default(value, default_value=0):
        # Suporta default_value=None para indicar que deve retornar None
        if value in (None, ''):
            return None if default_value is None else int(default_value)

        if isinstance(value, bool):
            return None if default_value is None else int(default_value)

        try:
            return int(float(value))
        except (TypeError, ValueError):
            return None if default_value is None else int(default_value)

    def parse_time_to_minutes(value):
        raw_value = str(value or '').strip()
        if not raw_value:
            return None

        match = re.fullmatch(r'(\d{1,2}):(\d{2})(?::\d{2})?', raw_value)
        if not match:
            return None

        hours = int(match.group(1))
        minutes = int(match.group(2))
        if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
            return None

        return (hours * 60) + minutes

    def estimated_daily_work_hours(collaborator):
        start_minutes = parse_time_to_minutes(collaborator.get('horario_padrao_inicio'))
        end_minutes = parse_time_to_minutes(collaborator.get('horario_padrao_fim'))
        if start_minutes is None or end_minutes is None:
            return None

        if end_minutes <= start_minutes:
            end_minutes += 24 * 60

        lunch_minutes = max(0, parse_int_or_default(collaborator.get('intervalo_almoco_minutos'), 60))
        total_minutes = max(0, end_minutes - start_minutes - lunch_minutes)
        return round(total_minutes / 60.0, 2)

    def estimated_night_eligible_daily_hours(collaborator):
        start_minutes = parse_time_to_minutes(collaborator.get('horario_padrao_inicio'))
        end_minutes = parse_time_to_minutes(collaborator.get('horario_padrao_fim'))
        if start_minutes is None or end_minutes is None:
            return 0.0

        # Turno cruzando meia-noite: ex. 19:00 -> 04:48.
        if end_minutes <= start_minutes:
            end_minutes += 24 * 60

        night_minutes = 0
        current_minute = start_minutes
        while current_minute < end_minutes:
            minute_of_day = current_minute % (24 * 60)
            if minute_of_day >= (22 * 60) or minute_of_day < (5 * 60):
                night_minutes += 1
            current_minute += 1

        return round(night_minutes / 60.0, 2)

    def safe_accuracy_percent(real_value, target_value):
        """Calcula acuracidade. Mantém sinal do desvio para debug."""
        if target_value <= 0:
            return None

        deviation = real_value - target_value  # Com sinal para análise
        absolute_deviation = abs(deviation)
        accuracy = max(0.0, 100.0 - ((absolute_deviation / target_value) * 100.0))
        return round(accuracy, 2)

    def extract_missing_column_name(exc):
        message = str(exc)
        patterns = [
            r'column\s+"([a-zA-Z0-9_]+)"\s+of\s+relation\s+"[a-zA-Z0-9_]+"\s+does\s+not\s+exist',
            r'column\s+\'([a-zA-Z0-9_]+)\'\s+of\s+relation\s+\'[a-zA-Z0-9_]+\'\s+does\s+not\s+exist',
            r'column\s+[a-zA-Z0-9_]+\.([a-zA-Z0-9_]+)\s+does\s+not\s+exist',
            r'Could not find the\s+"([a-zA-Z0-9_]+)"\s+column',
            r"Could not find the\s+'([a-zA-Z0-9_]+)'\s+column",
            r"column\s+([a-zA-Z0-9_]+)\s+does\s+not\s+exist",
        ]

        for pattern in patterns:
            match = re.search(pattern, message, flags=re.IGNORECASE)
            if match:
                return match.group(1)

        return None

    def insert_or_update_collaborator_with_schema_fallback(normalized_payload, existing_id=None, max_attempts=None):
        payload = dict(normalized_payload)
        removed_columns = []
        attempts = 0
        max_allowed_attempts = max_attempts or (len(payload) + 5)

        while attempts < max_allowed_attempts:
            attempts += 1
            try:
                if existing_id:
                    (
                        supabase.table('colaboradores')
                        .update(payload)
                        .eq('id', existing_id)
                        .execute()
                    )
                else:
                    supabase.table('colaboradores').insert(payload).execute()
                return removed_columns
            except Exception as exc:
                missing_column = extract_missing_column_name(exc)
                if not missing_column or missing_column not in payload:
                    raise

                payload.pop(missing_column, None)
                removed_columns.append(missing_column)

                if not payload:
                    raise RuntimeError('Não restaram colunas válidas para inserir/atualizar colaborador no schema atual.')

            raise RuntimeError('Não foi possível salvar colaborador após múltiplos ajustes de schema legado.')

            return removed_columns

    def execute_mutation_with_schema_fallback(table_name, payload, action='insert', item_id=None, max_attempts=None):
        mutable_payload = dict(payload or {})
        removed_columns = []
        attempts = 0
        max_allowed_attempts = max_attempts or (len(mutable_payload) + 3)

        while attempts < max_allowed_attempts:
            attempts += 1
            try:
                query = supabase.table(table_name)
                if action == 'update':
                    return query.update(mutable_payload).eq('id', item_id).execute(), removed_columns
                return query.insert(mutable_payload).execute(), removed_columns
            except Exception as exc:
                missing_column = extract_missing_column_name(exc)
                if not missing_column or missing_column not in mutable_payload:
                    raise

                mutable_payload.pop(missing_column, None)
                removed_columns.append(missing_column)
                if not mutable_payload:
                    raise RuntimeError('Falha de compatibilidade de schema: nenhuma coluna válida restante para salvar o registro.')

        raise RuntimeError('Não foi possível salvar o registro após múltiplos ajustes de compatibilidade de schema.')

    def now_iso():
        return datetime.now().astimezone().isoformat()

    def is_transient_disconnect_error(exc):
        message = str(exc).lower()
        # Errno 11 = EAGAIN (non-blocking socket would block) — transient on Linux
        if getattr(exc, 'errno', None) == 11:
            return True
        return any(kw in message for kw in (
            'server disconnected', 'connection reset', 'remoteprotocolerror',
            'getaddrinfo failed', 'name or service not known', 'temporary failure',
            'connection refused', 'timed out', 'connection aborted',
            'resource temporarily unavailable', 'eagain',
        ))

    def supabase_retry(fn, attempts=3, base_delay=0.5):
        """Call fn() with exponential-backoff retry on transient network errors."""
        last_exc = None
        for i in range(attempts):
            try:
                return fn()
            except Exception as exc:
                last_exc = exc
                if is_transient_disconnect_error(exc):
                    time.sleep(base_delay * (2 ** i))
                    continue
                raise
        raise last_exc

    def table_exists_ready(table_name, retry_attempts=3):
        cache_entry = table_ready_cache.get(table_name)
        now_ts = time.time()
        if cache_entry and cache_entry.get('expires_at', 0) > now_ts:
            return bool(cache_entry.get('value'))

        last_error = None
        for attempt in range(1, retry_attempts + 1):
            try:
                supabase.table(table_name).select('id').limit(1).execute()
                table_ready_cache[table_name] = {
                    'value': True,
                    'expires_at': now_ts + table_ready_cache_ttl_seconds,
                }
                return True
            except Exception as exc:
                last_error = exc
                if is_transient_disconnect_error(exc):
                    time.sleep(0.1 * attempt)
                    continue

                if is_missing_relation_error(exc):
                    table_ready_cache[table_name] = {
                        'value': False,
                        'expires_at': now_ts + min(60, table_ready_cache_ttl_seconds),
                    }
                    return False

                # Para outros erros inesperados, evita falso negativo de "tabela ausente".
                app.logger.warning('Falha inesperada ao verificar tabela %s: %s', table_name, exc)
                return True

        if last_error and is_transient_disconnect_error(last_error):
            app.logger.warning('Instabilidade ao verificar tabela %s, mantendo disponibilidade para evitar falso alerta.', table_name)
            # Cache "assume available" for 30s so we don't hammer DNS while it's down
            table_ready_cache[table_name] = {'value': True, 'expires_at': now_ts + 30}
            return True

        table_ready_cache[table_name] = {
            'value': False,
            'expires_at': now_ts + min(30, table_ready_cache_ttl_seconds),
        }
        return False

    def audit_table_ready():
        return table_exists_ready('auditoria_movimentacoes')

    def write_audit_event(profile, action, resource_name, entity_id=None, status='ok', details=None, filial_id=None):
        if not audit_table_ready():
            return

        try:
            payload = {
                'acao': action,
                'recurso': resource_name,
                'entidade_id': str(entity_id) if entity_id is not None else None,
                'status': status,
                'detalhes': details or {},
                'colaborador_id': profile.get('id') if profile else None,
                'usuario_id': profile.get('user_id') if profile else None,
                'nome_colaborador': profile.get('nome_completo') if profile else None,
                'filial_id': filial_id or (profile.get('filial_id') if profile else None),
                'ip_origem': (request.headers.get('X-Forwarded-For', '') or request.remote_addr).split(',')[0].strip() or request.remote_addr,
                'criado_em': now_iso(),
            }
            supabase.table('auditoria_movimentacoes').insert(payload).execute()
        except Exception as exc:
            app.logger.warning('Falha ao registrar auditoria (%s/%s): %s', action, resource_name, exc)

    def list_audit_events(profile, filters):
        query = supabase.table('auditoria_movimentacoes').select('*').order('criado_em', desc=True)

        if profile_has_filial_scope(profile):
            query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])

        if filters.get('resource'):
            query = query.eq('recurso', filters['resource'])
        if filters.get('action'):
            query = query.eq('acao', filters['action'])
        if filters.get('status'):
            query = query.eq('status', filters['status'])
        if filters.get('filial_id'):
            query = query.eq('filial_id', filters['filial_id'])
        if filters.get('date_from'):
            query = query.gte('criado_em', f"{filters['date_from']}T00:00:00")
        if filters.get('date_to'):
            query = query.lte('criado_em', f"{filters['date_to']}T23:59:59")

        response = query.limit(min(filters.get('limit', 200), 500)).execute()
        return response.data or []

    def find_auth_user_by_email(email):
        """
        Localiza um usuário no Supabase Auth pelo e-mail sem usar list_users
        (list_users pode retornar 500 quando a base de usuários é grande ou
        o endpoint está com instabilidade).

        Estratégia 1 — verificar nossa tabela colaboradores: se já existe um
        colaborador com esse e-mail e user_id vinculado, busca o auth user
        pelo ID usando get_user_by_id (endpoint muito mais leve).

        Estratégia 2 — tentar list_users como fallback silencioso caso a
        estratégia 1 não encontre nada (novo colaborador, e-mail ainda não
        cadastrado no nosso DB).
        """
        if not email:
            return None
        normalized = email.strip().lower()

        # Estratégia 1: checar tabela colaboradores
        try:
            rows = (
                supabase.table('colaboradores')
                .select('user_id')
                .ilike('email', normalized)
                .not_.is_('user_id', 'null')
                .limit(1)
                .execute()
                .data or []
            )
            if rows and rows[0].get('user_id'):
                uid = str(rows[0]['user_id'])
                result = supabase.auth.admin.get_user_by_id(uid)
                user = getattr(result, 'user', None) or result
                if user and getattr(user, 'email', '').lower() == normalized:
                    return user
        except Exception:
            pass

        # Estratégia 2: list_users como fallback (pode falhar — tratamos silenciosamente)
        try:
            page = 1
            while page <= 20:
                users_page = supabase.auth.admin.list_users(page=page, per_page=100)
                users_list = getattr(users_page, 'users', None) or (users_page if isinstance(users_page, list) else [])
                if not users_list:
                    break
                for u in users_list:
                    if getattr(u, 'email', '').lower() == normalized:
                        return u
                if len(users_list) < 100:
                    break
                page += 1
        except Exception:
            pass  # list_users falhou; seguimos para create_user

        return None

    def ensure_auth_user(email, full_name):
        normalized_email = (email or '').strip().lower()
        if not normalized_email:
            raise ValueError('Informe o e-mail do colaborador para vincular ao Supabase Auth.')

        existing_user = find_auth_user_by_email(normalized_email)
        if existing_user:
            return existing_user

        _pw_chars = string.ascii_letters + string.digits + '!@#$&'
        _temp_password = ''.join(secrets.choice(_pw_chars) for _ in range(20))
        try:
            created = supabase.auth.admin.create_user({
                'email': normalized_email,
                'password': _temp_password,
                'email_confirm': True,
                'user_metadata': {
                    'full_name': full_name or normalized_email,
                },
            })
        except Exception as exc:
            err_str = str(exc).lower()
            # Usuário já existe no Auth mas não foi encontrado pelo find_auth_user_by_email
            if any(k in err_str for k in ('already registered', 'already exists', 'duplicate', 'user already')):
                raise ValueError(
                    f'O e-mail "{normalized_email}" já está cadastrado no Supabase Auth mas não '
                    'está vinculado a nenhum colaborador no sistema. Acesse Gestão de Acessos para '
                    'verificar o vínculo ou use um e-mail diferente.'
                )
            raise ValueError(f'Falha ao criar usuário no Supabase Auth: {exc}')

        auth_user = getattr(created, 'user', None)
        if not auth_user:
            raise ValueError('Não foi possível criar o usuário no Supabase Auth.')

        return auth_user

    def parse_payload(resource_name, partial=False):
        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        config = RESOURCE_DEFINITIONS[resource_name]
        sanitized = {
            key: value
            for key, value in payload.items()
            if key in config['allowed_fields']
        }

        if resource_name == 'filiais':
            for _f in ('cidade', 'uf', 'parceira'):
                if _f in sanitized and isinstance(sanitized[_f], str):
                    sanitized[_f] = sanitized[_f].strip().upper()

        if resource_name in {'colaborador_documentos', 'eventos_rh'}:
            if not sanitized.get('filial_id') and sanitized.get('colaborador_id'):
                sanitized = fill_filial_from_collaborator(sanitized)
            for nullable_field in config.get('nullable_fields', []):
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

        if resource_name == 'colaborador_documentos':
            if 'status' in sanitized and isinstance(sanitized['status'], str):
                sanitized['status'] = sanitized['status'].strip().lower()
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True
            if 'obrigatorio' not in sanitized and not partial:
                sanitized['obrigatorio'] = False
            if 'dias_alerta' in sanitized and sanitized.get('dias_alerta') == '':
                sanitized['dias_alerta'] = None

        if resource_name == 'eventos_rh':
            if 'tipo_evento' in sanitized and isinstance(sanitized['tipo_evento'], str):
                sanitized['tipo_evento'] = sanitized['tipo_evento'].strip().lower()
            if 'status' in sanitized and isinstance(sanitized['status'], str):
                sanitized['status'] = sanitized['status'].strip().lower()
            if sanitized.get('status') and sanitized['status'] not in RH_EVENT_STATUS_OPTIONS:
                return None, f"Status de evento RH inválido: {sanitized['status']}."
            if 'status' not in sanitized and not partial:
                sanitized['status'] = 'planejado'
            if 'impacta_presenca' not in sanitized and not partial:
                sanitized['impacta_presenca'] = True
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True
            if sanitized.get('tipo_evento') and sanitized['tipo_evento'] not in RH_EVENT_TYPES:
                return None, f"Tipo de evento RH inválido: {sanitized['tipo_evento']}."

        if resource_name == 'rotas_carregamento':
            if 'codigo' in sanitized and isinstance(sanitized['codigo'], str):
                sanitized['codigo'] = sanitized['codigo'].strip().upper()
            if not sanitized.get('codigo') and sanitized.get('filial_id'):
                sanitized['codigo'] = fetch_filial_code(sanitized.get('filial_id')) or None

        if resource_name == 'veiculos':
            if 'placa' in sanitized and isinstance(sanitized['placa'], str):
                sanitized['placa'] = sanitized['placa'].strip().upper()
            if 'chassi' in sanitized and isinstance(sanitized['chassi'], str):
                sanitized['chassi'] = sanitized['chassi'].strip().upper()
            if 'status' in sanitized and isinstance(sanitized['status'], str):
                sanitized['status'] = sanitized['status'].strip().lower()
            for nullable_field in config.get('nullable_fields', []):
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None
            if not sanitized.get('status'):
                sanitized['status'] = 'ativo'

        if resource_name in {'rotas_carregamento', 'veiculos_carregamento', 'motivos_parada_carregamento', 'bonificacao_metricas'}:
            for nullable_field in config.get('nullable_fields', []):
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

        if resource_name == 'veiculos_carregamento':
            if 'placa' in sanitized and isinstance(sanitized['placa'], str):
                sanitized['placa'] = sanitized['placa'].strip().upper()
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'rotas_carregamento' and 'ativo' not in sanitized and not partial:
            sanitized['ativo'] = True

        if resource_name == 'motivos_parada_carregamento':
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True
            if 'exige_observacao' not in sanitized and not partial:
                sanitized['exige_observacao'] = False

        if resource_name == 'bonificacao_metricas':
            if 'categoria' in sanitized and isinstance(sanitized['categoria'], str):
                sanitized['categoria'] = sanitized['categoria'].strip().lower()
            if sanitized.get('categoria') and sanitized['categoria'] not in BONIFICACAO_CATEGORIAS:
                return None, f"Categoria de bonificação inválida: {sanitized['categoria']}."
            if 'categoria' not in sanitized and not partial:
                sanitized['categoria'] = 'individual'
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True
            if 'ordem' not in sanitized and not partial:
                sanitized['ordem'] = 0

        if resource_name == 'colaboradores' and 'tipo_acesso' not in sanitized and not partial:
            sanitized['tipo_acesso'] = 'app'

        if resource_name == 'colaboradores' and 'cpf' in sanitized:
            sanitized['cpf'] = normalize_cpf(sanitized.get('cpf'))

        if resource_name == 'colaboradores':
            numeric_defaults = {
                'carga_horaria_semanal': 44,
                'intervalo_almoco_minutos': 60,
                'salario_base_mensal': 0,
                'percentual_periculosidade': 0,
                'percentual_adicional_clt': 0,
                'beneficios_mensais': 0,
            }
            for field_name, default_value in numeric_defaults.items():
                if field_name in sanitized:
                    if field_name in {'carga_horaria_semanal', 'intervalo_almoco_minutos'}:
                        sanitized[field_name] = parse_int_or_default(sanitized.get(field_name), default_value)
                    elif field_name in {'percentual_periculosidade', 'percentual_adicional_clt'}:
                        sanitized[field_name] = max(0.0, min(100.0, parse_float_or_default(sanitized.get(field_name), default_value)))
                    else:
                        sanitized[field_name] = parse_float_or_default(sanitized.get(field_name), default_value)
                elif not partial:
                    sanitized[field_name] = default_value

            if 'adicional_noturno_sobre_periculosidade' in sanitized:
                sanitized['adicional_noturno_sobre_periculosidade'] = bool(sanitized.get('adicional_noturno_sobre_periculosidade'))
            elif not partial:
                sanitized['adicional_noturno_sobre_periculosidade'] = False

            # Converter strings vazias em null para campos nullable
            for nullable_field in config.get('nullable_fields', []):
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

        if resource_name == 'contratos_operacionais':
            if 'codigo_contrato' in sanitized and isinstance(sanitized['codigo_contrato'], str):
                sanitized['codigo_contrato'] = sanitized['codigo_contrato'].strip().upper()
            numeric_defaults = {
                'valor_mensal_contrato': 0,
                'horas_50_cobradas_contrato': 0,
                'horas_100_cobradas_contrato': 0,
                'qtd_colaboradores_contratados': 0,
                'valor_por_colaborador': 0,
                'custos_extras_gold_mensais': 0,
            }
            for field_name, default_value in numeric_defaults.items():
                if field_name in sanitized:
                    if field_name == 'qtd_colaboradores_contratados':
                        sanitized[field_name] = parse_int_or_default(sanitized.get(field_name), default_value)
                    else:
                        sanitized[field_name] = parse_float_or_default(sanitized.get(field_name), default_value)
                elif field_name in {'valor_mensal_contrato'} and not partial:
                    sanitized[field_name] = default_value

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'contratos_colaboradores':
            if 'tipo_item' in sanitized and isinstance(sanitized['tipo_item'], str):
                sanitized['tipo_item'] = sanitized['tipo_item'].strip().lower()
            if 'tipo_item' not in sanitized and not partial:
                sanitized['tipo_item'] = 'colaborador'

            item_type = (sanitized.get('tipo_item') or 'colaborador').strip().lower()
            if item_type not in {'colaborador', 'colaborador_fora_contrato', 'caminhao', 'veiculo_proprio', 'pacote_motorista_veiculo', 'outro'}:
                return None, 'Tipo de item inválido para contrato. Use colaborador, colaborador_fora_contrato, caminhao, veiculo_proprio, pacote_motorista_veiculo ou outro.'

            if item_type in {'colaborador', 'colaborador_fora_contrato'} and not sanitized.get('colaborador_id'):
                return None, 'Para tipo colaborador, selecione o colaborador.'
            if item_type == 'caminhao' and not sanitized.get('veiculo_carregamento_id'):
                return None, 'Para tipo caminhão terceiro, selecione o veículo de carregamento.'
            if item_type == 'veiculo_proprio' and not sanitized.get('veiculo_proprio_id'):
                return None, 'Para tipo veículo próprio, selecione o veículo da frota.'
            if item_type == 'pacote_motorista_veiculo':
                if not sanitized.get('colaborador_id'):
                    return None, 'Para pacote motorista+veículo, selecione o motorista (colaborador).'
                if not sanitized.get('veiculo_proprio_id'):
                    return None, 'Para pacote motorista+veículo, selecione o veículo da frota.'
            if item_type == 'outro' and not sanitized.get('nome_item'):
                return None, 'Para tipo outro, informe o nome/descrição do item.'

            if item_type in {'colaborador', 'colaborador_fora_contrato'}:
                sanitized['veiculo_carregamento_id'] = None
                sanitized['veiculo_proprio_id'] = None
                sanitized['nome_item'] = None
            elif item_type == 'caminhao':
                sanitized['colaborador_id'] = None
                sanitized['veiculo_proprio_id'] = None
                sanitized['nome_item'] = None
            elif item_type == 'veiculo_proprio':
                sanitized['colaborador_id'] = None
                sanitized['veiculo_carregamento_id'] = None
                sanitized['nome_item'] = None
            else:
                sanitized['colaborador_id'] = None
                sanitized['veiculo_carregamento_id'] = None
                sanitized['veiculo_proprio_id'] = None

            if 'percentual_alocacao' in sanitized:
                sanitized['percentual_alocacao'] = max(0.0, min(100.0, parse_float_or_default(sanitized.get('percentual_alocacao'), 100)))
            elif not partial:
                sanitized['percentual_alocacao'] = 100.0

            for field_name in ['valor_cobrado_colaborador', 'horas_50_cobradas', 'horas_100_cobradas']:
                if field_name in sanitized:
                    sanitized[field_name] = max(0.0, parse_float_or_default(sanitized.get(field_name), 0))
                elif not partial:
                    sanitized[field_name] = 0.0

            # Processar tipo_periodo
            if 'tipo_periodo' in sanitized and isinstance(sanitized['tipo_periodo'], str):
                sanitized['tipo_periodo'] = sanitized['tipo_periodo'].strip().lower()
                if sanitized['tipo_periodo'] not in {'normal', 'prorrogacao'}:
                    sanitized['tipo_periodo'] = 'normal'
            elif 'tipo_periodo' not in sanitized and not partial:
                sanitized['tipo_periodo'] = 'normal'

            # Processar dias_duracao_periodo
            if 'dias_duracao_periodo' in sanitized:
                sanitized['dias_duracao_periodo'] = parse_int_or_default(sanitized.get('dias_duracao_periodo'), 45)
                if sanitized['dias_duracao_periodo'] < 0:
                    sanitized['dias_duracao_periodo'] = 45
            elif 'dias_duracao_periodo' not in sanitized and not partial:
                sanitized['dias_duracao_periodo'] = 45

            # Processar dias_alerta_antes
            if 'dias_alerta_antes' in sanitized:
                sanitized['dias_alerta_antes'] = parse_int_or_default(sanitized.get('dias_alerta_antes'), 10)
                if sanitized['dias_alerta_antes'] < 0:
                    sanitized['dias_alerta_antes'] = 10

            # Converter strings vazias em null para campos nullable (evita violação de CHECK constraints)
            for nullable_field in config.get('nullable_fields', []):
                if nullable_field in sanitized and sanitized[nullable_field] == '':
                    sanitized[nullable_field] = None

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'contratos_gastos_extras':
            if 'valor_mensal' in sanitized:
                raw_vm = sanitized.get('valor_mensal')
                if raw_vm in ('', None):
                    sanitized['valor_mensal'] = None
                else:
                    sanitized['valor_mensal'] = max(0.0, parse_float_or_default(raw_vm, 0))
            # valor_mensal é opcional quando colaborador_id for informado (custo será calculado automaticamente)

            if 'percentual_alocacao' in sanitized:
                raw_pa = sanitized.get('percentual_alocacao')
                if raw_pa in ('', None):
                    sanitized['percentual_alocacao'] = None
                else:
                    sanitized['percentual_alocacao'] = max(0.0, min(100.0, parse_float_or_default(raw_pa, 100.0)))

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'colaborador_beneficios':
            if 'valor_mensal' in sanitized:
                sanitized['valor_mensal'] = max(0.0, parse_float_or_default(sanitized.get('valor_mensal'), 0))
            elif not partial:
                sanitized['valor_mensal'] = 0.0

            if 'valor_unitario' in sanitized:
                sanitized['valor_unitario'] = max(0.0, parse_float_or_default(sanitized.get('valor_unitario'), 0))
            elif not partial:
                sanitized['valor_unitario'] = 0.0

            if 'ordem' in sanitized:
                sanitized['ordem'] = parse_int_or_default(sanitized.get('ordem'), 0)
            elif not partial:
                sanitized['ordem'] = 0

            if 'modo_calculo' in sanitized and isinstance(sanitized['modo_calculo'], str):
                sanitized['modo_calculo'] = normalize_benefit_mode(sanitized['modo_calculo'])
            if 'base_dias' in sanitized and isinstance(sanitized['base_dias'], str):
                sanitized['base_dias'] = sanitized['base_dias'].strip().lower()

            if 'modo_calculo' not in sanitized and not partial:
                sanitized['modo_calculo'] = 'fixo_mensal'
            if 'base_dias' not in sanitized and not partial:
                sanitized['base_dias'] = 'presenca'

            if sanitized.get('modo_calculo') and sanitized['modo_calculo'] not in {'fixo_mensal', 'por_dia', 'percentual_salario'}:
                return None, 'Modo de cálculo inválido para benefício. Use fixo_mensal, por_dia ou percentual_salario.'
            if sanitized.get('base_dias') and sanitized['base_dias'] not in {'presenca', 'uteis', 'escala', 'calendario'}:
                return None, 'Base de dias inválida para benefício. Use presenca, uteis, escala ou calendario.'

            if 'desconta_faltas' not in sanitized and not partial:
                sanitized['desconta_faltas'] = True
            if 'desconta_eventos' not in sanitized and not partial:
                sanitized['desconta_eventos'] = True

            if 'teto_mensal' in sanitized:
                if sanitized.get('teto_mensal') in ('', None):
                    sanitized['teto_mensal'] = None
                else:
                    sanitized['teto_mensal'] = max(0.0, parse_float_or_default(sanitized.get('teto_mensal'), 0))

            benefit_mode = sanitized.get('modo_calculo')
            if benefit_mode == 'fixo_mensal':
                sanitized['valor_unitario'] = 0.0 if not partial or 'valor_unitario' in sanitized else sanitized.get('valor_unitario')
            elif benefit_mode == 'por_dia':
                if not partial and max(0.0, parse_float_or_default(sanitized.get('valor_unitario'), 0.0)) <= 0:
                    return None, 'Para benefício variável por dia, informe o valor unitário por dia.'
            elif benefit_mode == 'percentual_salario':
                if not partial and max(0.0, parse_float_or_default(sanitized.get('valor_unitario'), 0.0)) <= 0:
                    return None, 'Para benefício percentual, informe o percentual em valor unitário/%.'

            if 'tipo_beneficio' in sanitized and isinstance(sanitized['tipo_beneficio'], str):
                sanitized['tipo_beneficio'] = sanitized['tipo_beneficio'].strip()

            if 'codigo_rubrica' in sanitized and isinstance(sanitized['codigo_rubrica'], str):
                normalized_rubrica = sanitized['codigo_rubrica'].strip().upper()
                sanitized['codigo_rubrica'] = normalized_rubrica or None

            benefit_type = normalize_bonus_token(sanitized.get('tipo_beneficio'))
            if benefit_type in {'vale_transporte', 'ajuda_custo', 'vale_alimentacao', 'vale_refeicao'}:
                rubrica_by_type = {
                    'vale_transporte': 'VT',
                    'ajuda_custo': 'AJUDA',
                    'vale_alimentacao': 'VA',
                    'vale_refeicao': 'VA',
                }
                sanitized['codigo_rubrica'] = rubrica_by_type.get(benefit_type, sanitized.get('codigo_rubrica'))
                sanitized['modo_calculo'] = 'por_dia'
                sanitized['base_dias'] = 'escala'
                sanitized['valor_mensal'] = 0.0
                sanitized['teto_mensal'] = None
                sanitized['ordem'] = 0

            if is_bonus_benefit_item(sanitized):
                sanitized['codigo_rubrica'] = 'BONUS'
                sanitized['modo_calculo'] = 'fixo_mensal'
                sanitized['base_dias'] = 'presenca'
                sanitized['valor_mensal'] = 0.0
                sanitized['valor_unitario'] = 0.0
                sanitized['teto_mensal'] = None
                sanitized['desconta_faltas'] = False
                sanitized['desconta_eventos'] = False
                sanitized['ordem'] = 0
                if not sanitized.get('observacoes'):
                    sanitized['observacoes'] = 'Valor calculado automaticamente na tela de bonificação mensal.'

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'colaboradores' and partial:
            sanitized.pop('user_id', None)
            # Campos de controle de acesso são gerenciados exclusivamente via /api/permissoes/.
            # Removê-los do PATCH genérico impede escalada de privilégio por qualquer usuário
            # com permissao_editar=True alterando seu próprio cargo ou flags de acesso.
            for _restricted in (
                'permissao_app', 'permissao_desktop', 'permissao_editar',
                'permissao_excluir', 'permissao_aprovar_he', 'tipo_acesso',
            ):
                sanitized.pop(_restricted, None)

        if resource_name == 'cargos':
            if 'nome' in sanitized and isinstance(sanitized['nome'], str):
                sanitized['nome'] = sanitized['nome'].strip()
            if not sanitized.get('nome') and not partial:
                return None, 'Nome do cargo é obrigatório.'

            if 'ordem' in sanitized:
                raw_ordem = sanitized.get('ordem')
                sanitized['ordem'] = parse_int_or_default(raw_ordem, None) if raw_ordem not in ('', None) else None

            if 'permissoes_padrao' in sanitized:
                raw_pp = sanitized.get('permissoes_padrao')
                if isinstance(raw_pp, list):
                    sanitized['permissoes_padrao'] = [str(s) for s in raw_pp if isinstance(s, str)]
                elif raw_pp in ('', None):
                    sanitized['permissoes_padrao'] = []
                else:
                    sanitized['permissoes_padrao'] = []
            elif not partial:
                sanitized['permissoes_padrao'] = []

            for nullable_field in ['descricao', 'ordem']:
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'estoque_itens':
            ESTOQUE_CATEGORIAS = {
                'epi', 'uniforme', 'limpeza', 'manutencao', 'escritorio',
                'informatica', 'ferramentas', 'alimentacao', 'combustivel', 'outro',
            }
            ESTOQUE_UNIDADES = {'un', 'kg', 'l', 'cx', 'pct', 'm', 'par', 'rolo'}

            if 'categoria' in sanitized and isinstance(sanitized['categoria'], str):
                sanitized['categoria'] = sanitized['categoria'].strip().lower()
            if 'categoria' not in sanitized and not partial:
                sanitized['categoria'] = 'outro'
            if sanitized.get('categoria') and sanitized['categoria'] not in ESTOQUE_CATEGORIAS:
                return None, f"Categoria inválida: {sanitized['categoria']}."

            if 'unidade' in sanitized and isinstance(sanitized['unidade'], str):
                sanitized['unidade'] = sanitized['unidade'].strip().lower()
            if 'unidade' not in sanitized and not partial:
                sanitized['unidade'] = 'un'
            if sanitized.get('unidade') and sanitized['unidade'] not in ESTOQUE_UNIDADES:
                return None, f"Unidade inválida: {sanitized['unidade']}."

            if 'estoque_atual' in sanitized:
                sanitized['estoque_atual'] = max(0.0, parse_float_or_default(sanitized.get('estoque_atual'), 0))
            elif not partial:
                sanitized['estoque_atual'] = 0.0

            if 'estoque_minimo' in sanitized:
                sanitized['estoque_minimo'] = max(0.0, parse_float_or_default(sanitized.get('estoque_minimo'), 0))
            elif not partial:
                sanitized['estoque_minimo'] = 0.0

            for nullable_field in ['codigo', 'descricao', 'localizacao', 'observacoes']:
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'estoque_movimentos':
            ESTOQUE_TIPOS = {
                'entrada', 'saida_colaborador', 'saida_geral',
                'saida_fornecedor', 'troca', 'ajuste_positivo', 'ajuste_negativo',
            }

            if 'tipo' in sanitized and isinstance(sanitized['tipo'], str):
                sanitized['tipo'] = sanitized['tipo'].strip().lower()
            if sanitized.get('tipo') and sanitized['tipo'] not in ESTOQUE_TIPOS:
                return None, f"Tipo de movimento inválido: {sanitized['tipo']}. Use: {', '.join(sorted(ESTOQUE_TIPOS))}."

            if 'quantidade' in sanitized:
                sanitized['quantidade'] = max(0.001, parse_float_or_default(sanitized.get('quantidade'), 1))
            elif not partial:
                sanitized['quantidade'] = 1.0

            if 'colaborador_id' in sanitized:
                raw_cid = sanitized.get('colaborador_id')
                sanitized['colaborador_id'] = parse_int_or_default(raw_cid, None) if raw_cid not in ('', None) else None

            if 'registrado_por' in sanitized:
                raw_rp = sanitized.get('registrado_por')
                sanitized['registrado_por'] = parse_int_or_default(raw_rp, None) if raw_rp not in ('', None) else None

            if 'item_id' in sanitized:
                raw_iid = sanitized.get('item_id')
                sanitized['item_id'] = parse_int_or_default(raw_iid, None) if raw_iid not in ('', None) else None

            for nullable_field in ['fornecedor', 'numero_nota', 'motivo', 'observacoes']:
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'pedidos_compra':
            PEDIDO_STATUS_OPTIONS = {
                'rascunho', 'pendente', 'analise',
                'pendente_aprovacao', 'em_analise',   # legado
                'aprovado', 'reprovado', 'em_compra', 'recebido', 'cancelado',
            }
            PEDIDO_FORMA_PAGAMENTO_OPTIONS = {
                'dinheiro', 'pix', 'cartao_debito', 'cartao_credito', 'boleto', 'credito_fornecedor',
            }
            PEDIDO_REEMBOLSO_OPTIONS = {'pix', 'dinheiro', 'transferencia', 'cartao', 'nenhum'}
            if 'status' in sanitized and isinstance(sanitized['status'], str):
                sanitized['status'] = sanitized['status'].strip().lower()
            if 'status' not in sanitized and not partial:
                sanitized['status'] = 'rascunho'
            if sanitized.get('status') and sanitized['status'] not in PEDIDO_STATUS_OPTIONS:
                return None, f"Status inválido: {sanitized['status']}."

            if 'forma_pagamento' in sanitized and isinstance(sanitized['forma_pagamento'], str):
                sanitized['forma_pagamento'] = sanitized['forma_pagamento'].strip().lower()
            if sanitized.get('forma_pagamento') and sanitized['forma_pagamento'] not in PEDIDO_FORMA_PAGAMENTO_OPTIONS:
                return None, f"Forma de pagamento inválida: {sanitized['forma_pagamento']}."

            if 'tipo_reembolso' in sanitized and isinstance(sanitized['tipo_reembolso'], str):
                sanitized['tipo_reembolso'] = sanitized['tipo_reembolso'].strip().lower() or None
            if sanitized.get('tipo_reembolso') and sanitized['tipo_reembolso'] not in PEDIDO_REEMBOLSO_OPTIONS:
                sanitized['tipo_reembolso'] = None

            for nullable_field in [
                'data_necessidade', 'fornecedor', 'forma_pagamento', 'prazo_pagamento',
                'centro_custo', 'observacoes', 'chave_pix', 'dados_bancarios',
                'tipo_reembolso', 'motivo_reprovacao', 'numero_solicitacao',
            ]:
                if sanitized.get(nullable_field) == '':
                    sanitized[nullable_field] = None

            for int_field in ['criado_por', 'aprovado_por', 'em_analise_por', 'reprovado_por', 'contas_pagar_id']:
                if int_field in sanitized:
                    raw = sanitized.get(int_field)
                    sanitized[int_field] = parse_int_or_default(raw, None) if raw not in ('', None) else None

            if 'valor_total' not in sanitized and not partial:
                sanitized['valor_total'] = 0.0

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

            # Gera número do pedido automaticamente se não foi informado
            if not sanitized.get('numero_pedido') and not partial:
                filial_id = sanitized.get('filial_id')
                filial_code = fetch_filial_code(filial_id) if filial_id else 'GLD'
                year = datetime.now().year
                try:
                    last_row = (
                        supabase.table('pedidos_compra')
                        .select('numero_pedido')
                        .ilike('numero_pedido', f'PED-{filial_code}-{year}-%')
                        .order('numero_pedido', desc=True)
                        .limit(1)
                        .execute()
                    ).data or []
                    if last_row:
                        last_num_str = (last_row[0].get('numero_pedido') or '').rsplit('-', 1)[-1]
                        last_num = parse_int_or_default(last_num_str, 0)
                    else:
                        last_num = 0
                    sanitized['numero_pedido'] = f'PED-{filial_code}-{year}-{str(last_num + 1).zfill(4)}'
                except Exception:
                    sanitized['numero_pedido'] = f'PED-{filial_code}-{year}-0001'

        if resource_name == 'pedidos_compra_itens':
            PEDIDO_CATEGORIAS = {
                'limpeza', 'manutencao', 'epi', 'escritorio', 'alimentacao',
                'combustivel', 'informatica', 'uniforme', 'ferramentas', 'outro',
            }
            PEDIDO_UNIDADES = {'un', 'kg', 'l', 'cx', 'pct', 'm', 'par', 'rolo'}

            if 'categoria' in sanitized and isinstance(sanitized['categoria'], str):
                sanitized['categoria'] = sanitized['categoria'].strip().lower()
            if 'categoria' not in sanitized and not partial:
                sanitized['categoria'] = 'outro'
            if sanitized.get('categoria') and sanitized['categoria'] not in PEDIDO_CATEGORIAS:
                return None, f"Categoria inválida: {sanitized['categoria']}."

            if 'unidade' in sanitized and isinstance(sanitized['unidade'], str):
                sanitized['unidade'] = sanitized['unidade'].strip().lower()
            if 'unidade' not in sanitized and not partial:
                sanitized['unidade'] = 'un'
            if sanitized.get('unidade') and sanitized['unidade'] not in PEDIDO_UNIDADES:
                return None, f"Unidade inválida: {sanitized['unidade']}."

            if 'quantidade' in sanitized:
                sanitized['quantidade'] = max(0.001, parse_float_or_default(sanitized.get('quantidade'), 1))
            elif not partial:
                sanitized['quantidade'] = 1.0

            if 'valor_unitario' in sanitized:
                sanitized['valor_unitario'] = max(0.0, parse_float_or_default(sanitized.get('valor_unitario'), 0))
            elif not partial:
                sanitized['valor_unitario'] = 0.0

            if sanitized.get('observacoes') == '':
                sanitized['observacoes'] = None

            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        if resource_name == 'manutencoes' and not partial:
            # Manutenções novas entram como aguardando_aprovacao (atende CHECK constraint)
            if sanitized.get('status') in (None, 'aberta'):
                sanitized['status'] = 'aguardando_aprovacao'

        if resource_name == 'veiculos_abastecimentos' and not partial:
            # Novos abastecimentos entram sempre como pendente; o aprovador libera via /approvals
            sanitized['status'] = 'pendente_aprovacao'
            if 'ativo' not in sanitized:
                sanitized['ativo'] = True

        if resource_name == 'veiculos_pneus':
            # Deriva status_aprovacao do status do pneu: 'trocar' exige aprovação
            if 'status' in sanitized:
                sanitized['status_aprovacao'] = (
                    'pendente_aprovacao' if sanitized['status'] == 'trocar' else 'aprovado'
                )
            elif not partial:
                sanitized['status_aprovacao'] = 'aprovado'
            if 'ativo' not in sanitized and not partial:
                sanitized['ativo'] = True

        # Converte string vazia → None para qualquer campo em nullable_fields da config.
        # Garante que campos numéricos opcionais (ex: odometro_entrada, valor_estimado)
        # não cheguem ao banco como "" causando erro 22P02 (invalid input for numeric).
        for _nf in config.get('nullable_fields', []):
            if _nf in sanitized and sanitized[_nf] == '':
                sanitized[_nf] = None

        required_fields = list(config['required_fields'])
        if resource_name == 'colaboradores':
            # Permite cadastro inicial sem CPF definitivo (placeholder 000.000.000.00 vira None).
            required_fields = [field for field in required_fields if field != 'cpf']

        missing_fields = [
            field for field in required_fields if not partial and sanitized.get(field) in (None, '')
        ]
        if missing_fields:
            return None, f'Campos obrigatórios ausentes: {", ".join(missing_fields)}.'

        return sanitized, None

    def fetch_authenticated_user(token):
        cached_user = get_cached_auth_entry(f'user:{token}')
        if cached_user:
            return cached_user

        auth_request = urllib_request.Request(
            f'{supabase_url}/auth/v1/user',
            headers={
                'apikey': supabase_anon_key or supabase_secret_key or supabase_server_key,
                'Authorization': f'Bearer {token}',
            },
        )

        try:
            with urllib_request.urlopen(auth_request, timeout=10) as response:
                user = json.loads(response.read().decode('utf-8'))
                set_cached_auth_entry(f'user:{token}', user)
                return user
        except urllib_error.HTTPError as exc:
            exc.read()  # drena o corpo sem logar conteúdo interno do Supabase
            app.logger.warning('Token rejeitado pelo Supabase Auth. status=%s', exc.code)
            return None
        except Exception as exc:
            app.logger.exception('Falha de infraestrutura ao validar token no Supabase Auth')
            raise RuntimeError('Falha de comunicação com o Supabase Auth.') from exc

    def fetch_permissions(collaborator_id):
        try:
            result = (
                supabase.table('permissoes')
                .select('*')
                .eq('colaborador_id', collaborator_id)
                .eq('ativo', True)
                .order('permissao_nome')
                .execute()
            )
            return result.data or []
        except Exception:
            return []

    def is_hidden_collaborator_row(row):
        if not isinstance(row, dict):
            return False

        normalized_name = (row.get('nome_completo') or '').strip().lower()
        return normalized_name in HIDDEN_COLLABORATOR_NAMES

    def filter_visible_collaborators(rows):
        return [row for row in (rows or []) if not is_hidden_collaborator_row(row)]

    def fetch_collaborator_row(collaborator_id):
        if not collaborator_id:
            return None

        response = (
            supabase.table('colaboradores')
            .select('id, filial_id, nome_completo, cargo, ativo')
            .eq('id', collaborator_id)
            .limit(1)
            .execute()
        )
        return response.data[0] if response.data else None

    def fetch_collaborator_by_user_id(user_id):
        if not user_id:
            return None

        try:
            response = (
                supabase.table('colaboradores')
                .select('id')
                .eq('user_id', str(user_id))
                .limit(1)
                .execute()
            )
            return response.data[0] if response.data else None
        except Exception:
            # Schema legado pode não ter user_id na tabela colaboradores.
            return None

    def fetch_collaborator_by_name_and_filial(nome_completo, filial_id):
        normalized_name = (nome_completo or '').strip()
        if not normalized_name or not filial_id:
            return None

        try:
            response = (
                supabase.table('colaboradores')
                .select('id')
                .eq('filial_id', filial_id)
                .ilike('nome_completo', normalized_name)
                .limit(1)
                .execute()
            )
            return response.data[0] if response.data else None
        except Exception as exc:
            app.logger.warning('fetch_collaborator_by_name_and_filial falhou: %s', exc)
            return None

    def fetch_collaborator_by_cpf_and_filial(cpf, filial_id):
        normalized_cpf = (cpf or '').strip()
        if not normalized_cpf or not filial_id:
            return None
        trivial = {'000.000.000.00', '000000000', '00000000000', '0', '00000', '000000', '0000000', '00000000'}
        if normalized_cpf in trivial:
            return None
        try:
            response = (
                supabase.table('colaboradores')
                .select('id')
                .eq('filial_id', filial_id)
                .eq('cpf', normalized_cpf)
                .limit(1)
                .execute()
            )
            return response.data[0] if response.data else None
        except Exception as exc:
            app.logger.warning('fetch_collaborator_by_cpf_and_filial falhou: %s', exc)
            return None

    def fill_filial_from_collaborator(payload):
        collaborator = fetch_collaborator_row(payload.get('colaborador_id'))
        if collaborator and collaborator.get('filial_id'):
            payload['filial_id'] = collaborator['filial_id']
        return payload

    def rh_events_table_ready():
        return table_exists_ready('eventos_rh')

    def fetch_active_rh_events_for_date(target_date, filial_ids=None):
        if not rh_events_table_ready():
            return []

        query = (
            supabase.table('eventos_rh')
            .select('*')
            .eq('ativo', True)
            .in_('status', ['aprovado', 'em_andamento'])
            .lte('data_inicio', target_date.isoformat())
            .gte('data_fim', target_date.isoformat())
            .order('data_inicio')
        )
        if filial_ids:
            query = query.in_('filial_id', filial_ids)

        response = query.execute()
        return response.data or []

    def normalize_plain_text(value):
        normalized_value = unicodedata.normalize('NFKD', str(value or ''))
        ascii_value = normalized_value.encode('ascii', 'ignore').decode('ascii').lower()
        return re.sub(r'[^a-z0-9]+', ' ', ascii_value).strip()

    def is_projected_daily_benefit_item(item):
        benefit_type = normalize_bonus_token((item or {}).get('tipo_beneficio'))
        rubrica = str((item or {}).get('codigo_rubrica') or '').strip().upper()
        return benefit_type in {'vale_transporte', 'ajuda_custo', 'vale_alimentacao', 'vale_refeicao'} or rubrica in {'VT', 'AJUDA', 'VA'}

    def scale_is_44h_saturday(scale_value):
        """Retorna True se a escala é Seg-Sáb com sábado compensatório (4h) para fechar 44h semanais."""
        normalized = normalize_plain_text(scale_value)
        if not normalized:
            return False
        return any(keyword in normalized for keyword in ['seg sab', 'segunda sabado', 'seg a sab', 'seg ate sab', '6x1'])

    def infer_workdays_from_scale(scale_value):
        normalized_scale = normalize_plain_text(scale_value)
        if not normalized_scale:
            return None

        if any(keyword in normalized_scale for keyword in ['seg sex', 'segunda sexta', 'seg a sex', 'seg ate sex', '5x2']):
            return {0, 1, 2, 3, 4}
        if any(keyword in normalized_scale for keyword in ['seg sab', 'segunda sabado', 'seg a sab', 'seg ate sab', '6x1']):
            return {0, 1, 2, 3, 4, 5}
        if any(keyword in normalized_scale for keyword in ['seg dom', 'segunda domingo', 'seg a dom', 'seg ate dom', '7x0']):
            return {0, 1, 2, 3, 4, 5, 6}

        found_weekdays = set()
        for token, weekday_index in WEEKDAY_TOKEN_TO_INDEX.items():
            if re.search(rf'\b{re.escape(token)}\b', normalized_scale):
                found_weekdays.add(weekday_index)

        if found_weekdays:
            return found_weekdays

        return None

    def collaborator_works_on_date(collaborator, target_date):
        if not target_date:
            return True

        workdays = infer_workdays_from_scale(collaborator.get('escala_servico'))
        if workdays is None:
            return True

        return target_date.weekday() in workdays

    def was_collaborator_active_on_date(collaborator, target_date):
        """
        Verifica se um colaborador estava ativo em uma data específica.
        Retorna True se:
        - É ativo (ativo = True) E (sem data_desligamento OU data_desligamento > target_date)
        - OU tem registro de presença/evento naquela data (é garantido pela chamadora)
        """
        if not target_date:
            return collaborator.get('ativo', False)
        
        is_active = collaborator.get('ativo', False)
        desligamento_date_str = collaborator.get('data_desligamento')
        
        if not is_active:
            # Se inativo, só mostra se tem data de desligamento futura à data alvo
            if not desligamento_date_str:
                return False
            try:
                desligamento_date = datetime.strptime(desligamento_date_str, '%Y-%m-%d').date() if isinstance(desligamento_date_str, str) else desligamento_date_str
                return desligamento_date > target_date
            except (ValueError, TypeError):
                return False
        
        # Se ativo, verifica se não foi desligado ainda naquela data
        if not desligamento_date_str:
            return collaborator_works_on_date(collaborator, target_date)
        
        try:
            desligamento_date = datetime.strptime(desligamento_date_str, '%Y-%m-%d').date() if isinstance(desligamento_date_str, str) else desligamento_date_str
            if desligamento_date <= target_date:
                return False
        except (ValueError, TypeError):
            pass
        
        return collaborator_works_on_date(collaborator, target_date)

    def build_rh_events_lookup(rh_events):
        return {
            int(event['colaborador_id']): event
            for event in rh_events
            if event.get('colaborador_id') is not None and event.get('impacta_presenca', True)
        }

    def event_type_to_presence_status(event_type):
        return RH_EVENT_TYPES.get((event_type or '').strip().lower(), 'pendente')

    def collaborator_documents_table_ready():
        return table_exists_ready('colaborador_documentos')

    def enrich_collaborator_document(row, reference_date=None):
        if not isinstance(row, dict):
            return row

        reference_date = reference_date or date_class.today()
        normalized_row = dict(row)
        validity_date = parse_iso_date(normalized_row.get('data_validade'))
        configured_status = (normalized_row.get('status') or '').strip().lower()
        days_alert = normalized_row.get('dias_alerta')

        try:
            days_alert = int(days_alert) if days_alert is not None else 30
        except (TypeError, ValueError):
            days_alert = 30

        days_until_due = None
        computed_status = configured_status or 'pendente'

        if configured_status == 'nao_se_aplica':
            computed_status = 'nao_se_aplica'
        elif validity_date:
            days_until_due = (validity_date - reference_date).days
            if days_until_due < 0:
                computed_status = 'vencido'
            elif days_until_due <= days_alert:
                computed_status = 'vence_em_breve'
            else:
                computed_status = 'vigente'
        elif configured_status:
            computed_status = configured_status

        normalized_row['status_calculado'] = computed_status
        normalized_row['dias_para_vencer'] = days_until_due
        normalized_row['arquivo_enviado'] = bool(normalized_row.get('arquivo_url'))
        return normalized_row

    def build_collaborator_documents_alert_summary(profile):
        if not collaborator_documents_table_ready():
            return {
                'available': True,
                'database_ready': False,
                'cards': [],
            }

        query = supabase.table('colaborador_documentos').select('*').eq('ativo', True)
        query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['colaborador_documentos'])
        rows = query.execute().data or []
        documents = [enrich_collaborator_document(row) for row in rows]

        overdue_count = sum(1 for item in documents if item.get('status_calculado') == 'vencido')
        due_soon_count = sum(1 for item in documents if item.get('status_calculado') == 'vence_em_breve')
        pending_count = sum(1 for item in documents if item.get('status_calculado') == 'pendente')
        missing_file_count = sum(1 for item in documents if not item.get('arquivo_enviado'))

        return {
            'available': True,
            'database_ready': True,
            'cards': [
                {'label': 'Docs RH vencidos', 'value': overdue_count, 'hint': 'Documentos com validade expirada'},
                {'label': 'Vencem em breve', 'value': due_soon_count, 'hint': 'Dentro da régua de alerta do documento'},
                {'label': 'Pendentes', 'value': pending_count, 'hint': 'Sem validade ou status final definido'},
                {'label': 'Sem arquivo', 'value': missing_file_count, 'hint': 'Cadastro sem PDF, foto ou link do documento'},
            ],
        }

    def list_active_collaborators_for_presence(filial_id=None, target_date=None, ensure_ids=None):
        query = (
            supabase.table('colaboradores')
            .select('id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim, ativo, data_desligamento')
            .order('nome_completo')
        )
        if filial_id:
            query = query.eq('filial_id', filial_id)

        response = query.execute()
        collaborators = filter_visible_collaborators(response.data or [])
        if not target_date:
            return collaborators

        ensured_ids = {
            int(item)
            for item in (ensure_ids or [])
            if str(item).isdigit()
        }
        filtered = []
        for collaborator in collaborators:
            collaborator_id = int(collaborator.get('id')) if collaborator.get('id') is not None else None
            if collaborator_id in ensured_ids or was_collaborator_active_on_date(collaborator, target_date):
                filtered.append(collaborator)

        return filtered

    def fetch_accessible_filiais(profile):
        query = supabase.table('filiais').select('*').order('cidade')
        if profile_has_filial_scope(profile):
            query = query.in_('id', profile.get('allowed_filial_ids') or [])
        response = query.execute()
        return [decorate_filial_row(item) for item in (response.data or [])]

    def fetch_presence_rows_for_date(target_date, filial_ids=None):
        if not attendance_table_ready():
            return []

        query = (
            supabase.table('presencas_diarias')
            .select('*')
            .eq('data_referencia', target_date.isoformat())
        )
        if filial_ids:
            query = query.in_('filial_id', filial_ids)
        response = query.execute()
        return response.data or []

    def build_presence_items(target_date, collaborators, stored_rows=None, rh_events=None):
        stored_rows = stored_rows or []
        rh_events = rh_events or []
        rows_by_collaborator = {
            int(row['colaborador_id']): row
            for row in stored_rows
            if row.get('colaborador_id') is not None
        }
        rh_events_by_collaborator = build_rh_events_lookup(rh_events)

        items = []
        for collaborator in collaborators:
            stored = rows_by_collaborator.get(int(collaborator['id']), {})
            rh_event = rh_events_by_collaborator.get(int(collaborator['id']))
            default_status = event_type_to_presence_status(rh_event.get('tipo_evento')) if rh_event else 'pendente'
            default_observacao = rh_event.get('observacoes') or '' if rh_event else ''
            items.append({
                'id': stored.get('id'),
                'data_referencia': stored.get('data_referencia', target_date.isoformat()),
                'colaborador_id': collaborator['id'],
                'filial_id': collaborator['filial_id'],
                'nome_completo': collaborator['nome_completo'],
                'cargo': collaborator.get('cargo'),
                'turno': collaborator.get('turno') or '',
                'escala_servico': collaborator.get('escala_servico') or '',
                'horario_padrao_inicio': collaborator.get('horario_padrao_inicio') or '',
                'horario_padrao_fim': collaborator.get('horario_padrao_fim') or '',
                'status': stored.get('status') or default_status,
                'observacoes': stored.get('observacoes') or default_observacao,
                'origem': stored.get('origem') or ('rh' if rh_event else 'web'),
                'rh_evento': rh_event,
            })
        return items

    def build_workforce_board(profile, filial_id=None, target_date=None):
        accessible_filiais = fetch_accessible_filiais(profile)
        if filial_id:
            accessible_filiais = [item for item in accessible_filiais if int(item['id']) == int(filial_id)]

        accessible_filial_ids = [int(item['id']) for item in accessible_filiais]
        collaborators_query = (
            supabase.table('colaboradores')
            .select('id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim, ativo')
            .order('nome_completo')
        )
        if accessible_filial_ids:
            collaborators_query = collaborators_query.in_('filial_id', accessible_filial_ids)
        if filial_id:
            collaborators_query = collaborators_query.eq('filial_id', filial_id)
        collaborators = filter_visible_collaborators(collaborators_query.execute().data or [])
        reference_date = target_date or date_class.today()
        presence_rows = fetch_presence_rows_for_date(reference_date, accessible_filial_ids)
        rh_events = fetch_active_rh_events_for_date(reference_date, accessible_filial_ids)
        presence_by_collaborator = {
            int(row['colaborador_id']): row
            for row in presence_rows
            if row.get('colaborador_id') is not None
        }
        rh_events_by_collaborator = build_rh_events_lookup(rh_events)

        summary_by_filial = {
            int(item['id']): {
                'filial_id': item['id'],
                'filial_nome': f"{item['cidade']}/{item['uf']}",
                'parceira': item.get('parceira'),
                'ativos': 0,
                'inativos': 0,
                'presentes': 0,
                'faltas': 0,
                'ferias': 0,
                'afastados': 0,
                'atestados': 0,
                'folgas': 0,
                'pendentes': 0,
            }
            for item in accessible_filiais
        }

        employee_rows = []
        for collaborator in collaborators:
            current_filial_id = int(collaborator['filial_id'])
            if accessible_filial_ids and current_filial_id not in accessible_filial_ids:
                continue
            summary = summary_by_filial.setdefault(current_filial_id, {
                'filial_id': current_filial_id,
                'filial_nome': f'Base {current_filial_id}',
                'parceira': None,
                'ativos': 0,
                'inativos': 0,
            })
            if collaborator.get('ativo', True):
                summary['ativos'] += 1
            else:
                summary['inativos'] += 1

            presence_row = presence_by_collaborator.get(int(collaborator['id']), {})
            rh_event = rh_events_by_collaborator.get(int(collaborator['id']))
            works_on_reference_date = collaborator_works_on_date(collaborator, reference_date)
            day_status = presence_row.get('status') or (
                event_type_to_presence_status(rh_event.get('tipo_evento')) if rh_event else ('pendente' if works_on_reference_date else 'folga')
            )

            if collaborator.get('ativo', True):
                if day_status == 'presente':
                    summary['presentes'] += 1
                elif day_status == 'falta':
                    summary['faltas'] += 1
                elif day_status == 'ferias':
                    summary['ferias'] += 1
                elif day_status == 'afastado':
                    summary['afastados'] += 1
                elif day_status == 'atestado':
                    summary['atestados'] += 1
                elif day_status == 'folga':
                    summary['folgas'] += 1
                else:
                    summary['pendentes'] += 1

            employee_rows.append({
                'colaborador_id': collaborator['id'],
                'nome_completo': collaborator['nome_completo'],
                'cargo': collaborator.get('cargo'),
                'filial_id': current_filial_id,
                'filial_nome': summary['filial_nome'],
                'turno': collaborator.get('turno') or '',
                'escala_servico': collaborator.get('escala_servico') or '',
                'horario_padrao_inicio': collaborator.get('horario_padrao_inicio') or '',
                'horario_padrao_fim': collaborator.get('horario_padrao_fim') or '',
                'status_cadastro': 'Ativo' if collaborator.get('ativo', True) else 'Inativo',
                'status_dia': day_status,
                'origem_status_dia': presence_row.get('origem') or ('rh' if rh_event else 'cadastro'),
                'evento_rh': rh_event,
            })

        return {
            'reference_date': reference_date.isoformat(),
            'filiais': accessible_filiais,
            'summary_by_filial': sorted(summary_by_filial.values(), key=lambda item: item['filial_nome']),
            'employees': employee_rows,
        }

    def contratos_operacionais_table_ready():
        return table_exists_ready('contratos_operacionais')

    def contratos_colaboradores_table_ready():
        return table_exists_ready('contratos_colaboradores')

    def contratos_gastos_extras_table_ready():
        return table_exists_ready('contratos_gastos_extras')

    def colaborador_beneficios_table_ready():
        return table_exists_ready('colaborador_beneficios')

    def month_date_bounds(month_reference):
        month_start = date_class(month_reference.year, month_reference.month, 1)
        if month_reference.month == 12:
            next_month = date_class(month_reference.year + 1, 1, 1)
        else:
            next_month = date_class(month_reference.year, month_reference.month + 1, 1)
        month_end = next_month - timedelta(days=1)
        return month_start, month_end

    def month_business_days_count(month_reference):
        month_start, month_end = month_date_bounds(month_reference)
        current_day = month_start
        total = 0
        while current_day <= month_end:
            if current_day.weekday() < 5:
                total += 1
            current_day += timedelta(days=1)
        return total

    def previous_month_reference(month_reference):
        if month_reference.month == 1:
            return date_class(month_reference.year - 1, 12, 1)
        return date_class(month_reference.year, month_reference.month - 1, 1)

    def scale_workdays_count_for_month(collaborator, month_reference):
        month_start, month_end = month_date_bounds(month_reference)
        workdays = infer_workdays_from_scale((collaborator or {}).get('escala_servico'))
        if workdays is None:
            return month_business_days_count(month_reference)

        current_day = month_start
        total = 0
        while current_day <= month_end:
            if current_day.weekday() in workdays:
                total += 1
            current_day += timedelta(days=1)
        return total

    def scale_benefit_days_count_for_month(collaborator, month_reference):
        """Cômputo de dias para cálculo de benefícios diários (VT/VA/VR).
        Para escalas Seg-Sáb (44h com sábado compensatório de 4h), o sábado é excluído:
        o colaborador trabalha no sábado mas não recebe vale nesse dia."""
        month_start, month_end = month_date_bounds(month_reference)
        escala = (collaborator or {}).get('escala_servico')
        workdays = infer_workdays_from_scale(escala)

        # Sábado compensatório: exclui fins de semana (dia 5) do cômputo de benefícios
        if scale_is_44h_saturday(escala) and workdays is not None:
            workdays = workdays - {5}

        if workdays is None:
            return month_business_days_count(month_reference)

        current_day = month_start
        total = 0
        while current_day <= month_end:
            if current_day.weekday() in workdays:
                total += 1
            current_day += timedelta(days=1)
        return total

    def saturday_count_for_month(month_reference):
        """Retorna quantos sábados há no mês de referência."""
        month_start, month_end = month_date_bounds(month_reference)
        current_day = month_start
        total = 0
        while current_day <= month_end:
            if current_day.weekday() == 5:
                total += 1
            current_day += timedelta(days=1)
        return total

    def projected_daily_benefit_days(benefit_item, month_reference, collaborator, previous_presence_summary):
        # Para escalas Seg-Sáb com sábado compensatório (44h), sábado é excluído do
        # cômputo de dias de benefícios (VT/VA/VR): o colaborador trabalha 4h mas não recebe vale.
        dias_base = scale_benefit_days_count_for_month(collaborator or {}, month_reference)
        desconta_faltas = bool(benefit_item.get('desconta_faltas', True))
        desconta_eventos = bool(benefit_item.get('desconta_eventos', True))
        previous_presence = previous_presence_summary or {}

        faltas_desconto = int(previous_presence.get('falta', 0)) if desconta_faltas else 0
        eventos_desconto = (
            int(previous_presence.get('ferias', 0)) + int(previous_presence.get('afastado', 0))
            if desconta_eventos else 0
        )
        dias_calculados = max(0, dias_base - faltas_desconto - eventos_desconto)
        return {
            'dias_base': dias_base,
            'faltas_desconto': faltas_desconto,
            'eventos_desconto': eventos_desconto,
            'dias_calculados': dias_calculados,
        }

    def normalize_benefit_mode(value):
        normalized = str(value or '').strip().lower()
        aliases = {
            'fixo': 'fixo_mensal',
            'fixo_mensal': 'fixo_mensal',
            'mensal': 'fixo_mensal',
            'variavel': 'por_dia',
            'variável': 'por_dia',
            'por_dia': 'por_dia',
            'dia': 'por_dia',
            'percentual': 'percentual_salario',
            'percentual_salario': 'percentual_salario',
            'percentual_salário': 'percentual_salario',
        }
        return aliases.get(normalized, normalized)

    def summarize_presence_by_collaborator(profile, month_reference, filial_id=None):
        if not attendance_table_ready():
            return {}

        month_start, month_end = month_date_bounds(month_reference)
        query = (
            supabase.table('presencas_diarias')
            .select('colaborador_id, status')
            .gte('data_referencia', month_start.isoformat())
            .lte('data_referencia', month_end.isoformat())
        )
        presence_config = RESOURCE_DEFINITIONS.get('presencas_diarias', {'filial_scope_field': 'filial_id'})
        query = apply_filial_scope(query, profile, presence_config)
        if filial_id:
            query = query.eq('filial_id', filial_id)

        rows = query.execute().data or []
        summary_by_collaborator = {}
        for row in rows:
            collaborator_id = row.get('colaborador_id')
            if collaborator_id is None:
                continue

            collaborator_id = int(collaborator_id)
            status = (row.get('status') or '').strip().lower()
            summary = summary_by_collaborator.setdefault(collaborator_id, {
                'presente': 0,
                'falta': 0,
                'folga': 0,
                'atestado': 0,
                'ferias': 0,
                'afastado': 0,
                'pendente': 0,
                'registros_total': 0,
            })
            if status in summary:
                summary[status] += 1
            summary['registros_total'] += 1

        return summary_by_collaborator

    def calculate_benefit_monthly_value(benefit_item, salary, month_reference, presence_summary, collaborator=None, previous_presence_summary=None):
        mode = normalize_benefit_mode(benefit_item.get('modo_calculo') or 'fixo_mensal')
        base_dias = (benefit_item.get('base_dias') or 'presenca').strip().lower()
        is_projected_daily_benefit = is_projected_daily_benefit_item(benefit_item)
        if is_projected_daily_benefit:
            # Compatibilidade com registros antigos: evita cálculo por percentual/fixo para VT/VA/AJUDA.
            mode = 'por_dia'
            base_dias = 'escala'
        valor_mensal = max(0.0, parse_float_or_default(benefit_item.get('valor_mensal'), 0.0))
        valor_unitario = max(0.0, parse_float_or_default(benefit_item.get('valor_unitario'), 0.0))
        teto_mensal = benefit_item.get('teto_mensal')
        teto_mensal = None if teto_mensal in (None, '') else max(0.0, parse_float_or_default(teto_mensal, 0.0))
        desconta_faltas = bool(benefit_item.get('desconta_faltas', True))
        desconta_eventos = bool(benefit_item.get('desconta_eventos', True))

        presentes = int(presence_summary.get('presente', 0))
        faltas = int(presence_summary.get('falta', 0))
        ferias = int(presence_summary.get('ferias', 0))
        afastados = int(presence_summary.get('afastado', 0))
        registros_total = int(presence_summary.get('registros_total', 0))

        if mode == 'fixo_mensal':
            value = valor_mensal
        elif mode == 'percentual_salario':
            value = max(0.0, salary) * (max(0.0, valor_unitario) / 100.0)
        else:
            if is_projected_daily_benefit:
                projection = projected_daily_benefit_days(
                    benefit_item,
                    month_reference,
                    collaborator,
                    previous_presence_summary,
                )
                dias_base = projection['dias_calculados']
            else:
                if base_dias == 'uteis':
                    dias_base = month_business_days_count(month_reference)
                elif base_dias == 'calendario':
                    _, month_end = month_date_bounds(month_reference)
                    dias_base = month_end.day
                elif base_dias == 'escala':
                    dias_base = registros_total
                else:
                    dias_base = presentes

                if desconta_faltas:
                    dias_base -= faltas
                if desconta_eventos:
                    dias_base -= (ferias + afastados)

            dias_base = max(0, dias_base)
            value = dias_base * valor_unitario

        value = max(0.0, value)
        if teto_mensal is not None:
            value = min(value, teto_mensal)

        return round(value, 2)

    def enrich_collaborator_benefit_rows(profile, rows, month_reference, filial_id=None):
        normalized_rows = [dict(item) for item in (rows or [])]
        collaborator_ids = sorted({
            int(item.get('colaborador_id'))
            for item in normalized_rows
            if item.get('colaborador_id') is not None and str(item.get('colaborador_id')).isdigit()
        })
        if not collaborator_ids:
            return normalized_rows

        collaborators_query = (
            supabase.table('colaboradores')
            .select(
                'id, filial_id, escala_servico, horario_padrao_inicio, horario_padrao_fim, '
                'intervalo_almoco_minutos, salario_base_mensal, ativo'
            )
            .in_('id', collaborator_ids)
        )
        collaborators_query = apply_filial_scope(collaborators_query, profile, RESOURCE_DEFINITIONS['colaboradores'])
        if filial_id:
            collaborators_query = collaborators_query.eq('filial_id', filial_id)

        collaborators = {
            int(item['id']): item
            for item in (collaborators_query.execute().data or [])
            if item.get('id') is not None
        }

        month_presence = summarize_presence_by_collaborator(profile, month_reference, filial_id)
        previous_presence = summarize_presence_by_collaborator(
            profile,
            previous_month_reference(month_reference),
            filial_id,
        )

        for row in normalized_rows:
            collaborator_id = row.get('colaborador_id')
            if collaborator_id is None or not str(collaborator_id).isdigit():
                row['mes_referencia_calculo'] = month_reference.isoformat()
                row['valor_total_projetado'] = 0.0
                continue

            collaborator_id = int(collaborator_id)
            collaborator = collaborators.get(collaborator_id)
            salary = max(
                0.0,
                parse_float_or_default((collaborator or {}).get('salario_base_mensal'), 0.0),
            )
            row['mes_referencia_calculo'] = month_reference.isoformat()
            row['valor_total_projetado'] = calculate_benefit_monthly_value(
                row,
                salary,
                month_reference,
                month_presence.get(collaborator_id, {}),
                collaborator,
                previous_presence.get(collaborator_id, {}),
            )

        return normalized_rows

    def fetch_active_contract_links(profile, month_reference, filial_id=None):
        if not contratos_colaboradores_table_ready():
            return []

        month_start, month_end = month_date_bounds(month_reference)
        query = (
            supabase.table('contratos_colaboradores')
            .select('*')
            .eq('ativo', True)
            .or_(f'inicio_vigencia.is.null,inicio_vigencia.lte.{month_end.isoformat()}')
            .or_(f'fim_vigencia.is.null,fim_vigencia.gte.{month_start.isoformat()}')
            .order('id')
        )
        query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['contratos_colaboradores'])
        if filial_id:
            query = query.eq('filial_id', filial_id)
        return query.execute().data or []

    def fetch_active_contract_expenses(profile, month_reference, filial_id=None):
        if not contratos_gastos_extras_table_ready():
            return []

        month_start, month_end = month_date_bounds(month_reference)
        query = (
            supabase.table('contratos_gastos_extras')
            .select('*')
            .eq('ativo', True)
            .or_(f'inicio_vigencia.is.null,inicio_vigencia.lte.{month_end.isoformat()}')
            .or_(f'fim_vigencia.is.null,fim_vigencia.gte.{month_start.isoformat()}')
            .order('id')
        )
        query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['contratos_gastos_extras'])
        if filial_id:
            query = query.eq('filial_id', filial_id)
        return query.execute().data or []

    def fetch_active_benefit_items(profile, month_reference, filial_id=None):
        del month_reference
        if not colaborador_beneficios_table_ready():
            return []

        query = (
            supabase.table('colaborador_beneficios')
            .select('*')
            .eq('ativo', True)
            .order('tipo_beneficio')
        )
        query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['colaborador_beneficios'])
        if filial_id:
            query = query.eq('filial_id', filial_id)
        return query.execute().data or []

    def build_collaborator_contract_names(profile, target_date=None):
        if not (contratos_operacionais_table_ready() and contratos_colaboradores_table_ready()):
            return {}

        reference_date = target_date or date_class.today()
        month_reference = date_class(reference_date.year, reference_date.month, 1)
        links = fetch_active_contract_links(profile, month_reference)

        # Também busca gastos extras com colaborador vinculado
        expense_links = []
        if contratos_gastos_extras_table_ready():
            month_start, month_end = month_date_bounds(month_reference)
            exp_query = (
                supabase.table('contratos_gastos_extras')
                .select('colaborador_id, contrato_operacional_id')
                .eq('ativo', True)
                .not_.is_('colaborador_id', 'null')
                .or_(f'inicio_vigencia.is.null,inicio_vigencia.lte.{month_end.isoformat()}')
                .or_(f'fim_vigencia.is.null,fim_vigencia.gte.{month_start.isoformat()}')
            )
            exp_query = apply_filial_scope(exp_query, profile, RESOURCE_DEFINITIONS['contratos_gastos_extras'])
            expense_links = exp_query.execute().data or []

        all_sources = list(links) + list(expense_links)
        if not all_sources:
            return {}

        contract_ids = sorted({
            int(item['contrato_operacional_id'])
            for item in all_sources
            if item.get('contrato_operacional_id') is not None
        })
        if not contract_ids:
            return {}

        contracts = (
            supabase.table('contratos_operacionais')
            .select('id, nome_contrato')
            .in_('id', contract_ids)
            .execute()
            .data or []
        )
        name_by_contract_id = {
            int(item['id']): item.get('nome_contrato') or f"Contrato {item['id']}"
            for item in contracts
            if item.get('id') is not None
        }

        names_by_collaborator = {}
        for link in all_sources:
            collaborator_id = link.get('colaborador_id')
            contract_id = link.get('contrato_operacional_id')
            if collaborator_id is None or contract_id is None:
                continue

            contract_name = name_by_contract_id.get(int(contract_id))
            if not contract_name:
                continue

            names_by_collaborator.setdefault(int(collaborator_id), set()).add(contract_name)

        return {
            collaborator_id: sorted(list(contract_names))
            for collaborator_id, contract_names in names_by_collaborator.items()
        }

    def build_costs_dashboard(profile, month_reference, filial_id=None):
        # ======================================================================
        # FIX #5: CACHE SIMPLES PARA PERFORMANCE
        # Evita recalcular 600+ linhas a cada request
        # ======================================================================
        # Determina filial para cache (usa profile se filial_id não especificado)
        filial_cache = filial_id or profile.get('filial_id', 'global')
        month_str = month_reference.isoformat()
        
        # Tenta retornar do cache primeiro
        cached_result = get_cached_dashboard(filial_cache, month_str)
        if cached_result is not None:
            logging.info(f"Dashboard retornado do cache: filial={filial_cache}, mês={month_str}")
            return cached_result
        
        collaborators_select = (
            'id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim, '
            'carga_horaria_semanal, intervalo_almoco_minutos, salario_base_mensal, percentual_periculosidade, percentual_adicional_clt, '
            'adicional_noturno_sobre_periculosidade, '
            'beneficios_tipos, beneficios_mensais, ativo'
        )
        legacy_collaborators_select = (
            'id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim, '
            'carga_horaria_semanal, intervalo_almoco_minutos, salario_base_mensal, beneficios_tipos, beneficios_mensais, ativo'
        )

        def execute_collaborators_query(select_fields):
            query = (
                supabase.table('colaboradores')
                .select(select_fields)
                .eq('ativo', True)
                .order('nome_completo')
            )
            query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['colaboradores'])
            if filial_id:
                query = query.eq('filial_id', filial_id)
            return query.execute().data or []

        try:
            collaborators_rows = execute_collaborators_query(collaborators_select)
        except Exception as exc:
            missing_column = extract_missing_column_name(exc)
            if missing_column not in {'percentual_periculosidade', 'percentual_adicional_clt', 'adicional_noturno_sobre_periculosidade'}:
                raise
            collaborators_rows = execute_collaborators_query(legacy_collaborators_select)
            for collaborator_row in collaborators_rows:
                collaborator_row.setdefault('percentual_periculosidade', 0)
                collaborator_row.setdefault('percentual_adicional_clt', 0)
                collaborator_row.setdefault('adicional_noturno_sobre_periculosidade', False)

        collaborators = filter_visible_collaborators(collaborators_rows)
        benefit_items = fetch_active_benefit_items(profile, month_reference, filial_id)
        monthly_presence_by_collaborator = summarize_presence_by_collaborator(profile, month_reference, filial_id)
        previous_month_presence_by_collaborator = summarize_presence_by_collaborator(
            profile,
            previous_month_reference(month_reference),
            filial_id,
        )
        bonus_eligible_ids = fetch_bonus_eligible_collaborator_ids(profile, filial_id)
        benefit_types_by_collaborator = {}
        for item in benefit_items:
            collaborator_id = item.get('colaborador_id')
            if collaborator_id is None:
                continue
            collaborator_id = int(collaborator_id)
            benefit_type = (item.get('tipo_beneficio') or '').strip()
            if benefit_type:
                benefit_types_by_collaborator.setdefault(collaborator_id, set()).add(benefit_type)

        bonificacao_total_by_collaborator = {}
        if bonificacao_tables_ready():
            for lancamento in fetch_bonificacao_lancamentos(profile, month_reference, filial_id):
                collaborator_id = lancamento.get('colaborador_id')
                if collaborator_id is None:
                    continue
                collaborator_id = int(collaborator_id)
                # Bonificação no custo só entra para colaborador elegível no cadastro de benefícios.
                if bonus_eligible_ids is not None and collaborator_id not in bonus_eligible_ids:
                    continue
                bonificacao_total_by_collaborator[collaborator_id] = (
                    bonificacao_total_by_collaborator.get(collaborator_id, 0.0)
                    + max(0.0, parse_float_or_default(lancamento.get('valor_aplicado'), 0.0))
                )

        filiais = fetch_accessible_filiais(profile)
        filial_labels = {
            int(item['id']): f"{item.get('cidade', '-')}/{item.get('uf', '-') }"
            for item in filiais
        }

        collaborator_rows = []
        collaborator_cost_by_id = {}
        collaborator_financials_by_id = {}
        totals = {
            'salary': 0.0,
            'benefits': 0.0,
            'monthly_cost': 0.0,
            'headcount': 0,
        }
        totals_by_filial = {}
        contract_links = fetch_active_contract_links(profile, month_reference, filial_id)

        for collaborator in collaborators:
            collaborator_id = int(collaborator.get('id')) if collaborator.get('id') is not None else None
            current_filial_id = int(collaborator.get('filial_id')) if collaborator.get('filial_id') is not None else None
            weekly_hours = max(1, parse_int_or_default(collaborator.get('carga_horaria_semanal'), 44))
            monthly_hours = round(weekly_hours * 4.3333, 2)
            lunch_minutes = max(0, parse_int_or_default(collaborator.get('intervalo_almoco_minutos'), 60))
            salary_base = max(0.0, parse_float_or_default(collaborator.get('salario_base_mensal'), 0.0))
            percentual_periculosidade = max(0.0, min(100.0, parse_float_or_default(collaborator.get('percentual_periculosidade'), 0.0)))
            percentual_adicional_clt = max(0.0, min(100.0, parse_float_or_default(collaborator.get('percentual_adicional_clt'), 0.0)))
            adicional_noturno_sobre_periculosidade = bool(collaborator.get('adicional_noturno_sobre_periculosidade', False))
            valor_periculosidade = round(salary_base * (percentual_periculosidade / 100.0), 2)
            salary_base_for_night = salary_base + (valor_periculosidade if adicional_noturno_sobre_periculosidade else 0.0)
            hourly_base_salary = round((salary_base_for_night / monthly_hours), 4) if monthly_hours > 0 else 0.0
            projected_workdays = scale_workdays_count_for_month(collaborator, month_reference)
            night_eligible_daily_hours = estimated_night_eligible_daily_hours(collaborator)
            night_eligible_monthly_hours = round(night_eligible_daily_hours * projected_workdays, 2)
            valor_adicional_clt = round(hourly_base_salary * (percentual_adicional_clt / 100.0) * night_eligible_monthly_hours, 2)
            salary = round(salary_base + valor_periculosidade + valor_adicional_clt, 2)
            fixed_benefits = max(0.0, parse_float_or_default(collaborator.get('beneficios_mensais'), 0.0))
            collaborator_presence = monthly_presence_by_collaborator.get(collaborator_id or -1, {})

            detailed_benefits = 0.0
            variable_benefits = 0.0
            projected_days_base_total = 0
            projected_days_discounts_total = 0
            projected_days_final_total = 0
            if collaborator_id is not None:
                collaborator_benefit_items = [
                    item for item in benefit_items
                    if item.get('colaborador_id') is not None and int(item.get('colaborador_id')) == collaborator_id
                ]
                for benefit_item in collaborator_benefit_items:
                    previous_presence = previous_month_presence_by_collaborator.get(collaborator_id or -1, {})
                    if is_projected_daily_benefit_item(benefit_item):
                        projection = projected_daily_benefit_days(
                            benefit_item,
                            month_reference,
                            collaborator,
                            previous_presence,
                        )
                        projected_days_base_total += int(projection['dias_base'])
                        projected_days_discounts_total += int(projection['faltas_desconto']) + int(projection['eventos_desconto'])
                        projected_days_final_total += int(projection['dias_calculados'])
                    benefit_value = calculate_benefit_monthly_value(
                        benefit_item,
                        salary,
                        month_reference,
                        collaborator_presence,
                        collaborator,
                        previous_presence,
                    )
                    detailed_benefits += benefit_value
                    if (benefit_item.get('modo_calculo') or 'fixo_mensal').strip().lower() != 'fixo_mensal':
                        variable_benefits += benefit_value

            bonificacao_mensal = max(0.0, parse_float_or_default(bonificacao_total_by_collaborator.get(collaborator_id or -1), 0.0))
            benefits = round(fixed_benefits + detailed_benefits + bonificacao_mensal, 2)
            monthly_total_cost = round(salary + benefits, 2)
            hourly_cost = round((monthly_total_cost / monthly_hours), 2) if monthly_hours > 0 else 0.0
            daily_hours = estimated_daily_work_hours(collaborator)
            estimated_daily_cost = round((hourly_cost * daily_hours), 2) if daily_hours is not None else None

            totals['salary'] += salary
            totals['benefits'] += benefits
            totals['monthly_cost'] += monthly_total_cost
            totals['headcount'] += 1

            if current_filial_id is not None:
                filial_summary = totals_by_filial.setdefault(current_filial_id, {
                    'filial_id': current_filial_id,
                    'filial_nome': filial_labels.get(current_filial_id, f'Base {current_filial_id}'),
                    'headcount': 0,
                    'monthly_cost': 0.0,
                })
                filial_summary['headcount'] += 1
                filial_summary['monthly_cost'] += monthly_total_cost

            is_44h_saturday = scale_is_44h_saturday(collaborator.get('escala_servico'))
            sabados_mes = saturday_count_for_month(month_reference) if is_44h_saturday else 0
            collaborator_rows.append({
                'id': collaborator.get('id'),
                'filial_id': current_filial_id,
                'filial_nome': filial_labels.get(current_filial_id, '-') if current_filial_id is not None else '-',
                'nome_completo': collaborator.get('nome_completo'),
                'cargo': collaborator.get('cargo'),
                'turno': collaborator.get('turno') or '',
                'escala_servico': collaborator.get('escala_servico') or '',
                'horario_padrao_inicio': collaborator.get('horario_padrao_inicio') or '',
                'horario_padrao_fim': collaborator.get('horario_padrao_fim') or '',
                'carga_horaria_semanal': weekly_hours,
                'intervalo_almoco_minutos': lunch_minutes,
                'salario_base_mensal': round(salary_base, 2),
                'percentual_periculosidade': round(percentual_periculosidade, 2),
                'percentual_adicional_clt': round(percentual_adicional_clt, 2),
                'adicional_noturno_sobre_periculosidade': adicional_noturno_sobre_periculosidade,
                'valor_periculosidade': round(valor_periculosidade, 2),
                'valor_adicional_clt': round(valor_adicional_clt, 2),
                'horas_noturnas_elegiveis_dia': night_eligible_daily_hours,
                'dias_escala_projetados_mes': projected_workdays,
                'horas_noturnas_elegiveis_mes': night_eligible_monthly_hours,
                'sabado_compensatorio': is_44h_saturday,
                'sabados_excluidos_beneficio_mes': sabados_mes,
                'salario_clt_mensal': round(salary, 2),
                'beneficios_fixo_mensais': round(fixed_benefits, 2),
                'beneficios_itens_mensais': round(detailed_benefits, 2),
                'beneficios_variaveis_mensais': round(variable_benefits, 2),
                'bonificacao_mensal': round(bonificacao_mensal, 2),
                'beneficios_mensais': round(benefits, 2),
                'beneficios_tipos_texto': ', '.join(sorted(list(benefit_types_by_collaborator.get(int(collaborator.get('id')), set())))) if collaborator.get('id') is not None else '',
                'dias_presente_mes': int(collaborator_presence.get('presente', 0)),
                'dias_falta_mes': int(collaborator_presence.get('falta', 0)),
                'dias_eventos_mes': int(collaborator_presence.get('ferias', 0)) + int(collaborator_presence.get('afastado', 0)),
                'beneficios_dias_base_projetados': projected_days_base_total,
                'beneficios_dias_descontos_previstos': projected_days_discounts_total,
                'beneficios_dias_calculados': projected_days_final_total,
                'custo_mensal_total': monthly_total_cost,
                'custo_hora_estimado': hourly_cost,
                'horas_dia_estimadas': daily_hours,
                'custo_dia_estimado': estimated_daily_cost,
                'contratos_vinculados': [],
            })
            if collaborator_id is not None:
                collaborator_cost_by_id[collaborator_id] = monthly_total_cost
                collaborator_financials_by_id[collaborator_id] = {
                    'salario_clt_mensal': round(salary, 2),
                    'beneficios_mensais': round(benefits, 2),
                    'bonificacao_mensal': round(bonificacao_mensal, 2),
                    'beneficios_sem_bonificacao_mensais': round(max(0.0, benefits - bonificacao_mensal), 2),
                }

        contract_ids_in_links = sorted({
            int(item['contrato_operacional_id'])
            for item in contract_links
            if item.get('contrato_operacional_id') is not None
        })
        contract_names_by_id = {}
        if contract_ids_in_links and contratos_operacionais_table_ready():
            contract_name_rows = (
                supabase.table('contratos_operacionais')
                .select('id, nome_contrato')
                .in_('id', contract_ids_in_links)
                .execute()
                .data or []
            )
            contract_names_by_id = {
                int(item['id']): item.get('nome_contrato') or f"Contrato {item['id']}"
                for item in contract_name_rows
                if item.get('id') is not None
            }

        links_by_contract = {}
        links_by_collaborator = {}
        for link in contract_links:
            collaborator_id = link.get('colaborador_id')
            contract_id = link.get('contrato_operacional_id')
            if contract_id is None:
                continue

            contract_id = int(contract_id)
            # Itens sem colaborador (veiculo_proprio, caminhao, outro) ainda
            # entram no contrato pra somar valor_cobrado e listar a frota.
            links_by_contract.setdefault(contract_id, []).append(link)

            if collaborator_id is not None:
                collaborator_id = int(collaborator_id)
                contract_name = contract_names_by_id.get(contract_id)
                if contract_name:
                    links_by_collaborator.setdefault(collaborator_id, set()).add(contract_name)

        for row in collaborator_rows:
            collaborator_id = row.get('id')
            contract_names = sorted(list(links_by_collaborator.get(int(collaborator_id), set()))) if collaborator_id is not None else []
            row['contratos_vinculados'] = contract_names
            row['contratos_vinculados_texto'] = ', '.join(contract_names) if contract_names else '-'

        contracts = []
        contracts_ready = contratos_operacionais_table_ready()
        contract_expenses = fetch_active_contract_expenses(profile, month_reference, filial_id)

        # Pré-carrega custo básico + status ativo de colab referenciados em
        # gastos extras que NÃO estão em collaborator_cost_by_id (típico:
        # colab inativo, mas gasto extra ainda vinculado a ele)
        expense_colab_ids_all = {
            int(e.get('colaborador_id'))
            for e in contract_expenses
            if e.get('colaborador_id') is not None
        }
        expense_colab_ativo_map = {}  # cid -> bool (True/False)
        expense_colab_ids_missing = {
            cid for cid in expense_colab_ids_all
            if cid not in collaborator_cost_by_id
        }
        if expense_colab_ids_all:
            try:
                fallback_resp = supabase.table('colaboradores').select(
                    'id, salario_base_mensal, beneficios_mensais, ativo'
                ).in_('id', list(expense_colab_ids_all)).execute()
                for c in (fallback_resp.data or []):
                    cid_fb = int(c['id'])
                    expense_colab_ativo_map[cid_fb] = bool(c.get('ativo', True))
                    if cid_fb in expense_colab_ids_missing:
                        salario_fb = max(0.0, parse_float_or_default(c.get('salario_base_mensal'), 0.0))
                        beneficios_fb = max(0.0, parse_float_or_default(c.get('beneficios_mensais'), 0.0))
                        collaborator_cost_by_id[cid_fb] = round(salario_fb + beneficios_fb, 2)
            except Exception as exc:
                logging.warning(f'Falha ao carregar custo/ativo fallback colab gasto extra: {exc}')

        expense_rows_by_contract = {}
        expense_total_by_contract = {}
        for expense in contract_expenses:
            contract_id = expense.get('contrato_operacional_id')
            if contract_id is None:
                continue
            contract_id = int(contract_id)
            exp_colaborador_id = expense.get('colaborador_id')
            valor_manual = parse_float_or_default(expense.get('valor_mensal'), 0.0)
            if valor_manual > 0:
                # Valor manual tem prioridade (usuário informou explicitamente)
                expense_value_calc = round(valor_manual, 2)
            elif exp_colaborador_id is not None:
                # Cálculo automático: custo do colaborador proporcional à alocação
                exp_colaborador_id = int(exp_colaborador_id)
                exp_allocation = max(0.0, min(100.0, parse_float_or_default(expense.get('percentual_alocacao'), 100.0)))
                exp_collab_cost = parse_float_or_default(collaborator_cost_by_id.get(exp_colaborador_id), 0.0)
                expense_value_calc = round(exp_collab_cost * (exp_allocation / 100.0), 2)
            else:
                expense_value_calc = 0.0

            # Se o colab vinculado está inativo, o registro permanece mas
            # NÃO conta no total de gastos extras (só informativo)
            colab_ativo_flag = True
            if exp_colaborador_id is not None:
                colab_ativo_flag = expense_colab_ativo_map.get(int(exp_colaborador_id), True)

            expense_value_total = expense_value_calc if colab_ativo_flag else 0.0
            expense_total_by_contract[contract_id] = expense_total_by_contract.get(contract_id, 0.0) + expense_value_total
            expense_rows_by_contract.setdefault(contract_id, []).append({
                'id': expense.get('id'),
                'nome_gasto': expense.get('nome_gasto') or '',
                'colaborador_id': expense.get('colaborador_id'),
                'percentual_alocacao': expense.get('percentual_alocacao'),
                'valor_mensal': round(expense_value_total, 2),
                'valor_mensal_calculado': round(expense_value_calc, 2),
                'colaborador_ativo': colab_ativo_flag,
            })

        if contracts_ready:
            contract_query = (
                supabase.table('contratos_operacionais')
                .select('*')
                .eq('ativo', True)
                .order('nome_contrato')
            )
            contract_query = apply_filial_scope(contract_query, profile, RESOURCE_DEFINITIONS['contratos_operacionais'])
            if filial_id:
                contract_query = contract_query.eq('filial_id', filial_id)
            contracts = contract_query.execute().data or []

        # ────────────────────────────────────────────────────────────────
        # Bulk-load contratos_colaboradores SEM filtro de ativo (somente
        # para aba Variável/RTM — colab inativo que teve horas no mês
        # deve aparecer no agregado RTM, mas não no custo de salário)
        # ────────────────────────────────────────────────────────────────
        rtm_links_by_contract = {}
        try:
            month_start_rtm, month_end_rtm = month_date_bounds(month_reference)
            rtm_links_query = (
                supabase.table('contratos_colaboradores')
                .select('*')
                .or_(f'inicio_vigencia.is.null,inicio_vigencia.lte.{month_end_rtm.isoformat()}')
            )
            rtm_links_query = apply_filial_scope(rtm_links_query, profile, RESOURCE_DEFINITIONS['contratos_colaboradores'])
            if filial_id:
                rtm_links_query = rtm_links_query.eq('filial_id', filial_id)
            for link_row in (rtm_links_query.execute().data or []):
                cid_rtm = link_row.get('contrato_operacional_id')
                if cid_rtm is None:
                    continue
                rtm_links_by_contract.setdefault(int(cid_rtm), []).append(link_row)
        except Exception as exc:
            logging.warning(f'Falha ao carregar contratos_colaboradores (incl. inativos) p/ RTM: {exc}')

        # ────────────────────────────────────────────────────────────────
        # Bulk-load RTM horas extras do mês (para aba Variável do contrato)
        # Também coleta meses disponíveis (qualquer período) por colaborador
        # ────────────────────────────────────────────────────────────────
        rtm_by_colaborador = {}
        rtm_meses_por_colaborador = {}  # colaborador_id -> set(mes_prefix)
        try:
            mes_ref_prefix = month_reference.strftime('%Y-%m')
            rtm_query_all = supabase.table('horas_extras_rtm_registros').select(
                'colaborador_id, funcionario_nome, filial_id, mes_referencia, '
                'total_50, total_100, total_geral, horas_normais, horas_extra_100, '
                'valor_hora_50, valor_hora_100, tipo_hora'
            )
            if filial_id:
                rtm_query_all = rtm_query_all.eq('filial_id', filial_id)
            rtm_resp_all = rtm_query_all.execute()
            for reg in (rtm_resp_all.data or []):
                cid = reg.get('colaborador_id')
                if cid is None:
                    continue
                key = int(cid)
                mes_reg = (reg.get('mes_referencia') or '')[:7]
                if mes_reg:
                    rtm_meses_por_colaborador.setdefault(key, set()).add(mes_reg)
                # Só agrega ao total quando casa com o mês selecionado
                if not mes_reg.startswith(mes_ref_prefix):
                    continue
                cur = rtm_by_colaborador.setdefault(key, {
                    'total_50': 0.0, 'total_100': 0.0, 'total_geral': 0.0,
                    'horas_50': 0.0, 'horas_100': 0.0,
                    'funcionario_nome': reg.get('funcionario_nome', ''),
                    'tipo_hora': reg.get('tipo_hora', ''),
                })
                cur['total_50'] += parse_float_or_default(reg.get('total_50'), 0.0)
                cur['total_100'] += parse_float_or_default(reg.get('total_100'), 0.0)
                cur['total_geral'] += parse_float_or_default(reg.get('total_geral'), 0.0)
                cur['horas_50'] += parse_float_or_default(reg.get('horas_normais'), 0.0)
                cur['horas_100'] += parse_float_or_default(reg.get('horas_extra_100'), 0.0)
        except Exception as exc:
            logging.warning(f'Falha ao carregar horas_extras_rtm_registros: {exc}')

        # Bulk-load nome + ativo de TODOS os colaboradores referenciados nos vínculos
        # (inclusive inativos — dashboard principal filtra ativo=True)
        colab_info_map = {}
        all_linked_cids = set()
        for link in contract_links:
            cid = link.get('colaborador_id')
            if cid is not None:
                all_linked_cids.add(int(cid))
        # Inclui também colab dos vínculos inativos (para RTM)
        for links_list in rtm_links_by_contract.values():
            for link in links_list:
                cid = link.get('colaborador_id')
                if cid is not None:
                    all_linked_cids.add(int(cid))
        if all_linked_cids:
            try:
                info_resp = supabase.table('colaboradores').select(
                    'id, nome_completo, ativo, cargo'
                ).in_('id', list(all_linked_cids)).execute()
                for c in (info_resp.data or []):
                    colab_info_map[int(c['id'])] = {
                        'nome': c.get('nome_completo') or '',
                        'cargo': c.get('cargo') or '',
                        'ativo': bool(c.get('ativo', True)),
                    }
            except Exception as exc:
                logging.warning(f'Falha ao carregar info colaboradores vinculados: {exc}')

        contracts_enriched = []
        acuracidade_headcount_values = []
        acuracidade_valor_values = []
        negative_margin_count = 0
        contracts_summary_totals = {
            'valor_contrato_itens_total': 0.0,
            'valor_contrato_cadastro_total': 0.0,
            'valor_por_fora_total': 0.0,
            'gasto_salario_total': 0.0,
            'gasto_beneficios_total': 0.0,
            'gasto_bonificacao_total': 0.0,
            'custo_real_base_total': 0.0,
            'custo_total_gold_total': 0.0,
            'custo_por_fora_total': 0.0,
            'saldo_por_fora_total': 0.0,
            'margem_total': 0.0,
            'headcount_contratado_total': 0,
            'headcount_real_total': 0,
            'headcount_fora_total': 0,
        }

        for contract in contracts:
            contract_filial_id = contract.get('filial_id')
            contract_filial_id = int(contract_filial_id) if contract_filial_id is not None else None
            filial_totals = totals_by_filial.get(contract_filial_id, {
                'headcount': 0,
                'monthly_cost': 0.0,
                'filial_nome': filial_labels.get(contract_filial_id, '-') if contract_filial_id is not None else '-',
            })

            planned_headcount = max(0, parse_int_or_default(contract.get('qtd_colaboradores_contratados'), 0))
            contract_value_cadastro = max(0.0, parse_float_or_default(contract.get('valor_mensal_contrato'), 0.0))
            planned_value_per_collaborator = max(0.0, parse_float_or_default(contract.get('valor_por_colaborador'), 0.0))
            custos_extras_gold_fixo = max(0.0, parse_float_or_default(contract.get('custos_extras_gold_mensais'), 0.0))
            custos_extras_gold_linhas = round(expense_total_by_contract.get(int(contract.get('id')), 0.0), 2) if contract.get('id') is not None else 0.0
            custos_extras_gold = round(custos_extras_gold_fixo + custos_extras_gold_linhas, 2)

            linked_rows = links_by_contract.get(int(contract.get('id')), [])
            linked_rows_contrato = []
            linked_rows_fora_contrato = []
            for item in linked_rows:
                tipo_item = (item.get('tipo_item') or 'colaborador').strip().lower()
                if tipo_item in {'colaborador_fora_contrato', 'fora_contrato'}:
                    linked_rows_fora_contrato.append(item)
                else:
                    linked_rows_contrato.append(item)

            linked_collaborator_ids = sorted({
                int(item['colaborador_id'])
                for item in linked_rows_contrato
                if item.get('colaborador_id') is not None
            })
            linked_collaborator_ids_fora = sorted({
                int(item['colaborador_id'])
                for item in linked_rows_fora_contrato
                if item.get('colaborador_id') is not None
            })
            linked_item_count = len(linked_rows_contrato)
            linked_item_count_fora = len(linked_rows_fora_contrato)
            real_headcount = len(linked_collaborator_ids)
            real_cost = 0.0
            real_cost_vinculos_contrato = 0.0
            real_cost_fora_contrato = 0.0
            salary_cost = 0.0
            benefits_cost = 0.0
            bonus_cost = 0.0
            total_horas_50_itens = 0.0
            total_horas_100_itens = 0.0
            total_valor_cobrado_colaborador = 0.0
            total_valor_cobrado_colaborador_fora = 0.0
            for link in linked_rows:
                total_horas_50_itens += max(0.0, parse_float_or_default(link.get('horas_50_cobradas'), 0.0))
                total_horas_100_itens += max(0.0, parse_float_or_default(link.get('horas_100_cobradas'), 0.0))
                valor_cobrado_link = max(0.0, parse_float_or_default(link.get('valor_cobrado_colaborador'), 0.0))
                tipo_item_link = (link.get('tipo_item') or 'colaborador').strip().lower()
                is_fora_contrato = tipo_item_link in {'colaborador_fora_contrato', 'fora_contrato'}
                if is_fora_contrato:
                    total_valor_cobrado_colaborador_fora += valor_cobrado_link
                else:
                    total_valor_cobrado_colaborador += valor_cobrado_link

                collaborator_id = link.get('colaborador_id')
                if collaborator_id is None:
                    continue

                collaborator_id = int(collaborator_id)
                collaborator_cost = parse_float_or_default(collaborator_cost_by_id.get(collaborator_id), 0.0)
                collaborator_financials = collaborator_financials_by_id.get(collaborator_id, {})
                allocation = max(0.0, min(100.0, parse_float_or_default(link.get('percentual_alocacao'), 100.0)))
                allocation_factor = allocation / 100.0
                link_cost = collaborator_cost * allocation_factor
                real_cost += link_cost
                if is_fora_contrato:
                    real_cost_fora_contrato += link_cost
                else:
                    real_cost_vinculos_contrato += link_cost
                salary_cost += parse_float_or_default(collaborator_financials.get('salario_clt_mensal'), 0.0) * allocation_factor
                benefits_cost += parse_float_or_default(collaborator_financials.get('beneficios_mensais'), 0.0) * allocation_factor
                bonus_cost += parse_float_or_default(collaborator_financials.get('bonificacao_mensal'), 0.0) * allocation_factor
            real_cost = round(real_cost, 2)
            real_cost_vinculos_contrato = round(real_cost_vinculos_contrato, 2)
            real_cost_fora_contrato = round(real_cost_fora_contrato, 2)
            salary_cost = round(salary_cost, 2)
            benefits_cost = round(benefits_cost, 2)
            bonus_cost = round(bonus_cost, 2)
            benefits_without_bonus_cost = round(max(0.0, benefits_cost - bonus_cost), 2)
            
            # FIX #1: Unificar valor do contrato (dentro + fora) para acuracidade correta
            # Antes: só usava total_valor_cobrado_colaborador (apenas vinculados)
            # Agora: inclui colaboradores "por fora" também
            contract_value_from_items = round(total_valor_cobrado_colaborador, 2)
            contract_value_from_items_fora = round(total_valor_cobrado_colaborador_fora, 2)
            contract_value = round(contract_value_from_items + contract_value_from_items_fora, 2)
            
            real_cost_total_gold = round(real_cost + custos_extras_gold, 2)
            
            # FIX #2: Headcount completo (dentro + fora) para cálculos corretos
            # Antes: real_headcount = len(linked_collaborator_ids) só os vinculados
            # Agora: inclui colaboradores "por fora"
            real_headcount_total = len(linked_collaborator_ids) + len(linked_collaborator_ids_fora)
            real_value_per_collaborator = round((contract_value / real_headcount_total), 2) if real_headcount_total > 0 else 0.0
            margem_contrato = round(contract_value - real_cost_total_gold, 2)
            margem_percentual = round((margem_contrato / contract_value) * 100.0, 2) if contract_value > 0 else None
            if margem_contrato < 0:
                negative_margin_count += 1

            # FIX #3: Acuracidade usa headcount_total (dentro + fora)
            # Para balancear com real_cost_total_gold que também inclui todos
            headcount_accuracy = safe_accuracy_percent(real_headcount_total, planned_headcount)
            value_accuracy = safe_accuracy_percent(real_cost_total_gold, contract_value)
            if headcount_accuracy is not None:
                acuracidade_headcount_values.append(headcount_accuracy)
            if value_accuracy is not None:
                acuracidade_valor_values.append(value_accuracy)

            contracts_summary_totals['valor_contrato_itens_total'] += contract_value
            contracts_summary_totals['valor_contrato_cadastro_total'] += contract_value_cadastro
            contracts_summary_totals['valor_por_fora_total'] += total_valor_cobrado_colaborador_fora
            contracts_summary_totals['gasto_salario_total'] += salary_cost
            contracts_summary_totals['gasto_beneficios_total'] += benefits_cost
            contracts_summary_totals['gasto_bonificacao_total'] += bonus_cost
            contracts_summary_totals['custo_real_base_total'] += real_cost
            contracts_summary_totals['custo_total_gold_total'] += real_cost_total_gold
            contracts_summary_totals['custo_por_fora_total'] += real_cost_fora_contrato
            contracts_summary_totals['saldo_por_fora_total'] += (total_valor_cobrado_colaborador_fora - real_cost_fora_contrato)
            contracts_summary_totals['margem_total'] += margem_contrato
            contracts_summary_totals['headcount_contratado_total'] += planned_headcount
            # FIX: Usar real_headcount_total (dentro + fora) para totalização correta
            contracts_summary_totals['headcount_real_total'] += real_headcount_total
            contracts_summary_totals['headcount_fora_total'] += len(linked_collaborator_ids_fora)

            horas_50_contrato = max(0.0, parse_float_or_default(contract.get('horas_50_cobradas_contrato'), 0.0))
            horas_100_contrato = max(0.0, parse_float_or_default(contract.get('horas_100_cobradas_contrato'), 0.0))
            total_horas_50 = horas_50_contrato if horas_50_contrato > 0 else total_horas_50_itens
            total_horas_100 = horas_100_contrato if horas_100_contrato > 0 else total_horas_100_itens

            # ──────────────────────────────────────────────────────────────
            # Detalhe por colaborador + totais RTM (aba Variável)
            # União de vínculos ATIVOS (linked_rows) + INATIVOS do mês
            # (rtm_links_by_contract) — RTM soma inativos; custo só ativos
            # ──────────────────────────────────────────────────────────────
            colaboradores_detalhe = []
            rtm_total_50_valor = 0.0
            rtm_total_100_valor = 0.0
            rtm_total_geral_valor = 0.0
            rtm_total_horas_50 = 0.0
            rtm_total_horas_100 = 0.0
            colaboradores_inativos_count = 0
            rtm_meses_set = set()

            contract_id_int = int(contract.get('id')) if contract.get('id') is not None else None
            active_link_cids = {int(l.get('colaborador_id')) for l in linked_rows if l.get('colaborador_id') is not None}
            union_links = list(linked_rows)
            for extra_link in rtm_links_by_contract.get(contract_id_int, []):
                ecid = extra_link.get('colaborador_id')
                if ecid is None:
                    continue
                if int(ecid) in active_link_cids:
                    continue
                union_links.append(extra_link)

            seen_cids_detalhe = set()
            for link in union_links:
                cid_link = link.get('colaborador_id')
                if cid_link is None:
                    continue
                cid_int = int(cid_link)
                if cid_int in seen_cids_detalhe:
                    continue
                seen_cids_detalhe.add(cid_int)
                tipo_item_link2 = (link.get('tipo_item') or 'colaborador').strip().lower()
                is_fora2 = tipo_item_link2 in {'colaborador_fora_contrato', 'fora_contrato'}
                info = colab_info_map.get(cid_int, {'nome': '', 'ativo': True, 'cargo': ''})
                rtm = rtm_by_colaborador.get(cid_int)
                allocation2 = max(0.0, min(100.0, parse_float_or_default(link.get('percentual_alocacao'), 100.0)))
                vinculo_ativo = bool(link.get('ativo', True)) and (cid_int in active_link_cids)
                # Custo só conta para vínculos ativos
                custo_col = (parse_float_or_default(collaborator_cost_by_id.get(cid_int), 0.0) * (allocation2 / 100.0)
                             if vinculo_ativo else 0.0)
                valor_cobrado_link2 = max(0.0, parse_float_or_default(link.get('valor_cobrado_colaborador'), 0.0))
                if not info.get('ativo', True):
                    colaboradores_inativos_count += 1
                colaboradores_detalhe.append({
                    'colaborador_id': cid_int,
                    'nome': info.get('nome') or (rtm.get('funcionario_nome') if rtm else ''),
                    'cargo': info.get('cargo') or '',
                    'ativo': info.get('ativo', True),
                    'vinculo_ativo': vinculo_ativo,
                    'tipo_item': tipo_item_link2,
                    'is_fora_contrato': is_fora2,
                    'percentual_alocacao': round(allocation2, 2),
                    'custo_alocado': round(custo_col, 2),
                    'valor_cobrado': round(valor_cobrado_link2, 2),
                    'rtm_total_geral': round(rtm['total_geral'], 2) if rtm else 0.0,
                    'rtm_total_50': round(rtm['total_50'], 2) if rtm else 0.0,
                    'rtm_total_100': round(rtm['total_100'], 2) if rtm else 0.0,
                    'rtm_horas_50': round(rtm['horas_50'], 2) if rtm else 0.0,
                    'rtm_horas_100': round(rtm['horas_100'], 2) if rtm else 0.0,
                })
                # Totais variáveis: colab fixos (não-fora) — ativos OU inativos com RTM no mês
                if not is_fora2:
                    if rtm:
                        rtm_total_50_valor += rtm['total_50']
                        rtm_total_100_valor += rtm['total_100']
                        rtm_total_geral_valor += rtm['total_geral']
                        rtm_total_horas_50 += rtm['horas_50']
                        rtm_total_horas_100 += rtm['horas_100']
                    rtm_meses_set.update(rtm_meses_por_colaborador.get(cid_int, set()))

            contracts_enriched.append({
                'id': contract.get('id'),
                'filial_id': contract_filial_id,
                'filial_nome': filial_totals.get('filial_nome') or filial_labels.get(contract_filial_id, '-'),
                'codigo_contrato': contract.get('codigo_contrato') or '',
                'nome_contrato': contract.get('nome_contrato'),
                'cliente_nome': contract.get('cliente_nome') or '',
                'valor_mensal_contrato': round(contract_value, 2),
                'valor_mensal_contrato_cadastro': round(contract_value_cadastro, 2),
                'valor_mensal_contrato_itens': round(contract_value_from_items, 2),
                'qtd_colaboradores_contratados': planned_headcount,
                'cargos_contrato': contract.get('cargos_contrato') or '',
                'valor_por_colaborador': round(planned_value_per_collaborator, 2),
                'custos_extras_gold_fixo_mensais': round(custos_extras_gold_fixo, 2),
                'custos_extras_gold_linhas_mensais': round(custos_extras_gold_linhas, 2),
                'custos_extras_gold_mensais': round(custos_extras_gold, 2),
                'headcount_real': len(linked_collaborator_ids),  # Apenas vinculados
                'headcount_fora_contrato': len(linked_collaborator_ids_fora),
                'headcount_total': real_headcount_total,  # Total (dentro + fora) para cálculos
                'gasto_salario_mensal': salary_cost,
                'gasto_beneficios_mensal': benefits_cost,
                'gasto_beneficios_sem_bonificacao_mensal': benefits_without_bonus_cost,
                'gasto_bonificacao_mensal': bonus_cost,
                'custo_mensal_vinculos_contrato': real_cost_vinculos_contrato,
                'custo_mensal_fora_contrato': real_cost_fora_contrato,
                'custo_mensal_real': real_cost,
                'custo_total_gold_real': real_cost_total_gold,
                'margem_contrato': margem_contrato,
                'margem_percentual': margem_percentual,
                'valor_por_colaborador_real': real_value_per_collaborator,
                'valor_cobrado_colaboradores_total': round(total_valor_cobrado_colaborador, 2),
                'valor_cobrado_colaboradores_fora_total': round(total_valor_cobrado_colaborador_fora, 2),
                'horas_50_cobradas_contrato': round(horas_50_contrato, 2),
                'horas_100_cobradas_contrato': round(horas_100_contrato, 2),
                'horas_50_cobradas_total': round(total_horas_50, 2),
                'horas_100_cobradas_total': round(total_horas_100, 2),
                'colaboradores_vinculados': linked_collaborator_ids,
                'itens_vinculados_total': linked_item_count,
                'itens_fora_contrato_total': linked_item_count_fora,
                'gastos_extras_linhas': expense_rows_by_contract.get(int(contract.get('id')), []) if contract.get('id') is not None else [],
                'acuracidade_headcount': headcount_accuracy,
                'acuracidade_valor': value_accuracy,
                # Variável: RTM horas extras (somente colaboradores fixos do contrato)
                'rtm_valor_total_50': round(rtm_total_50_valor, 2),
                'rtm_valor_total_100': round(rtm_total_100_valor, 2),
                'rtm_valor_total_geral': round(rtm_total_geral_valor, 2),
                'rtm_horas_50_total': round(rtm_total_horas_50, 2),
                'rtm_horas_100_total': round(rtm_total_horas_100, 2),
                # Detalhe por colaborador (com ativo/inativo + RTM individual)
                'colaboradores_detalhe': colaboradores_detalhe,
                'colaboradores_inativos_count': colaboradores_inativos_count,
                # Meses (YYYY-MM) com dados RTM para colab fixos deste contrato
                # (Excluindo o mês atualmente selecionado — só sugestões)
                'rtm_meses_disponiveis': sorted(
                    [m for m in rtm_meses_set if m != month_reference.strftime('%Y-%m')],
                    reverse=True,
                )[:6],
            })

        avg_headcount_accuracy = round(sum(acuracidade_headcount_values) / len(acuracidade_headcount_values), 2) if acuracidade_headcount_values else None
        avg_value_accuracy = round(sum(acuracidade_valor_values) / len(acuracidade_valor_values), 2) if acuracidade_valor_values else None
        total_contracts_value_accuracy = safe_accuracy_percent(
            totals['monthly_cost'],
            contracts_summary_totals['valor_contrato_itens_total'],
        )
        total_contracts_headcount_accuracy = safe_accuracy_percent(
            contracts_summary_totals['headcount_real_total'],
            contracts_summary_totals['headcount_contratado_total'],
        )
        custo_vinculos_contrato_total = round(
            contracts_summary_totals['custo_real_base_total'] - contracts_summary_totals['custo_por_fora_total'],
            2,
        )
        custo_sem_vinculo_total = round(
            totals['monthly_cost'] - contracts_summary_totals['custo_real_base_total'],
            2,
        )
        margem_total_contratos_base = round(
            contracts_summary_totals['valor_contrato_itens_total'] - totals['monthly_cost'],
            2,
        )
        despesa_total_operacao = round(totals['monthly_cost'], 2)
        lucro_total_operacao = margem_total_contratos_base

        # Prepara resultado
        result = {
            'month_reference': month_reference.isoformat(),
            'month_label': f"{month_reference.month:02d}/{month_reference.year}",
            'filiais': filiais,
            'database_ready': {
                'contracts': contracts_ready,
                'contract_links': contratos_colaboradores_table_ready(),
                'contract_expenses': contratos_gastos_extras_table_ready(),
                'benefit_items': colaborador_beneficios_table_ready(),
                'presence': attendance_table_ready(),
            },
            'summary': {
                'headcount_total': totals['headcount'],
                'salary_total': round(totals['salary'], 2),
                'benefits_total': round(totals['benefits'], 2),
                'monthly_total_cost': round(totals['monthly_cost'], 2),
                'average_headcount_accuracy': avg_headcount_accuracy,
                'average_value_accuracy': avg_value_accuracy,
                'total_contracts_headcount_accuracy': total_contracts_headcount_accuracy,
                'total_contracts_value_accuracy': total_contracts_value_accuracy,
                'contracts_with_negative_margin': negative_margin_count,
                'despesa_total_operacao': despesa_total_operacao,
                'lucro_total_operacao': lucro_total_operacao,
                'valor_contrato_itens_total': round(contracts_summary_totals['valor_contrato_itens_total'], 2),
                'valor_contrato_cadastro_total': round(contracts_summary_totals['valor_contrato_cadastro_total'], 2),
                'valor_por_fora_total': round(contracts_summary_totals['valor_por_fora_total'], 2),
                'gasto_salario_total': round(contracts_summary_totals['gasto_salario_total'], 2),
                'gasto_beneficios_total': round(contracts_summary_totals['gasto_beneficios_total'], 2),
                'gasto_bonificacao_total': round(contracts_summary_totals['gasto_bonificacao_total'], 2),
                'custo_real_base_total': round(contracts_summary_totals['custo_real_base_total'], 2),
                'custo_total_gold_total': round(contracts_summary_totals['custo_total_gold_total'], 2),
                'custo_vinculos_contrato_total': custo_vinculos_contrato_total,
                'custo_por_fora_total': round(contracts_summary_totals['custo_por_fora_total'], 2),
                'custo_sem_vinculo_total': custo_sem_vinculo_total,
                'saldo_por_fora_total': round(contracts_summary_totals['saldo_por_fora_total'], 2),
                'margem_total_contratos': margem_total_contratos_base,
                'headcount_contratado_total': int(contracts_summary_totals['headcount_contratado_total']),
                'headcount_real_total': int(contracts_summary_totals['headcount_real_total']),
                'headcount_fora_total': int(contracts_summary_totals['headcount_fora_total']),
                'colaboradores_sabado_compensatorio': sum(1 for row in collaborator_rows if row.get('sabado_compensatorio')),
            },
            'totals_by_filial': sorted(totals_by_filial.values(), key=lambda item: item['filial_nome']),
            'collaborators': collaborator_rows,
            'contracts': contracts_enriched,
        }
        
        # ======================================================================
        # Armazena resultado no cache para próximas requisições
        # ======================================================================
        try:
            set_cached_dashboard(filial_cache, month_str, result)
            logging.info(f"Dashboard cacheado: filial={filial_cache}, mês={month_str}")
        except Exception as e:
            # Erro no cache não deve afetar a resposta
            logging.warning(f"Erro ao cachear dashboard: {e}")
        
        return result

    def attendance_table_ready():
        return table_exists_ready('presencas_diarias')

    def parse_month_reference(value):
        if not value:
            today = date_class.today()
            return date_class(today.year, today.month, 1)

        normalized = str(value).strip()
        if re.fullmatch(r'\d{4}-\d{2}', normalized):
            normalized = f'{normalized}-01'

        parsed = parse_iso_date(normalized)
        if not parsed:
            return None

        return date_class(parsed.year, parsed.month, 1)

    def bonificacao_tables_ready():
        required_tables = ['bonificacao_metricas', 'bonificacao_lancamentos']
        return all(table_exists_ready(table_name) for table_name in required_tables)

    def ocorrencias_table_ready():
        return table_exists_ready('ocorrencias')

    def fetch_bonificacao_metricas():
        query = (
            supabase.table('bonificacao_metricas')
            .select('*')
            .eq('ativo', True)
            .order('ordem')
            .order('nome')
        )
        return query.execute().data or []

    def normalize_bonus_token(value):
        normalized_value = unicodedata.normalize('NFKD', str(value or ''))
        ascii_value = normalized_value.encode('ascii', 'ignore').decode('ascii').lower().strip()
        return re.sub(r'[^a-z0-9]+', '_', ascii_value)

    def is_bonus_benefit_item(item):
        tipo = normalize_bonus_token(item.get('tipo_beneficio'))
        rubrica = normalize_bonus_token(item.get('codigo_rubrica'))
        known_tokens = {
            'bonus',
            'bonificacao',
            'bonificacao_mensal',
            'premio',
            'premio_mensal',
            'produtividade',
        }
        return tipo in known_tokens or rubrica in known_tokens

    def fetch_bonus_eligible_collaborator_ids(profile, filial_id=None):
        if not colaborador_beneficios_table_ready():
            return None

        query = (
            supabase.table('colaborador_beneficios')
            .select('colaborador_id, tipo_beneficio, codigo_rubrica, ativo')
            .eq('ativo', True)
            .order('colaborador_id')
        )
        query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['colaborador_beneficios'])
        if filial_id:
            query = query.eq('filial_id', filial_id)

        eligible_ids = {
            int(item['colaborador_id'])
            for item in (query.execute().data or [])
            if item.get('colaborador_id') is not None and is_bonus_benefit_item(item)
        }
        return eligible_ids

    def fetch_bonificacao_colaboradores(profile, filial_id=None):
        eligible_ids = fetch_bonus_eligible_collaborator_ids(profile, filial_id)

        base_query = (
            supabase.table('colaboradores')
            .select('id, filial_id, nome_completo, cargo, ativo')
            .eq('ativo', True)
            .order('nome_completo')
        )

        if filial_id:
            base_query = base_query.eq('filial_id', filial_id)
        elif profile_has_filial_scope(profile):
            base_query = base_query.in_('filial_id', profile.get('allowed_filial_ids') or [])

        try:
            if eligible_ids is not None:
                if not eligible_ids:
                    return []
                base_query = base_query.in_('id', sorted(list(eligible_ids)))

            return filter_visible_collaborators(base_query.execute().data or [])
        except Exception as exc:
            if not extract_missing_column_name(exc):
                raise

            fallback_query = (
                supabase.table('colaboradores')
                .select('id, filial_id, nome_completo, cargo, ativo')
                .eq('ativo', True)
                .order('nome_completo')
            )
            if filial_id:
                fallback_query = fallback_query.eq('filial_id', filial_id)
            elif profile_has_filial_scope(profile):
                fallback_query = fallback_query.in_('filial_id', profile.get('allowed_filial_ids') or [])

            if eligible_ids is not None:
                if not eligible_ids:
                    return []
                fallback_query = fallback_query.in_('id', sorted(list(eligible_ids)))

            return filter_visible_collaborators(fallback_query.execute().data or [])

    def fetch_bonificacao_lancamentos(profile, month_reference, filial_id=None):
        query = (
            supabase.table('bonificacao_lancamentos')
            .select('*')
            .eq('mes_referencia', month_reference.isoformat())
            .order('colaborador_id')
        )

        if filial_id:
            query = query.eq('filial_id', filial_id)
        elif profile_has_filial_scope(profile):
            query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])

        return query.execute().data or []

    def fetch_bonificacao_total_ano(profile, year, filial_id=None):
        start_month = date_class(year, 1, 1)
        end_month = date_class(year, 12, 1)

        query = (
            supabase.table('bonificacao_lancamentos')
            .select('valor_aplicado, filial_id')
            .gte('mes_referencia', start_month.isoformat())
            .lte('mes_referencia', end_month.isoformat())
        )

        if filial_id:
            query = query.eq('filial_id', filial_id)
        elif profile_has_filial_scope(profile):
            query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])

        rows = query.execute().data or []
        return float(sum(float(item.get('valor_aplicado') or 0) for item in rows))

    def fetch_monthly_occurrences_by_collaborator(profile, month_reference, collaborator_ids):
        if not collaborator_ids or not ocorrencias_table_ready():
            return {}

        month_start, month_end = month_date_bounds(month_reference)
        query = (
            supabase.table('ocorrencias')
            .select('colaborador_id, tipo_ocorrencia, hora_registro')
            .in_('colaborador_id', collaborator_ids)
            .gte('hora_registro', f'{month_start.isoformat()}T00:00:00')
            .lte('hora_registro', f'{month_end.isoformat()}T23:59:59')
        )

        if profile_has_filial_scope(profile):
            # Restrição indireta por colaborador já respeita perfil/base.
            pass

        rows = query.execute().data or []
        by_collaborator = {}
        for row in rows:
            collaborator_id = row.get('colaborador_id')
            if collaborator_id is None:
                continue
            collaborator_id = int(collaborator_id)
            tipo = (row.get('tipo_ocorrencia') or '').strip().lower()
            bucket = by_collaborator.setdefault(collaborator_id, {
                'ocorrencias_total': 0,
                'ocorrencias_criticas': 0,
            })
            bucket['ocorrencias_total'] += 1
            if any(token in tipo for token in ['cilindro', 'queda', 'derram', 'avaria', 'grave']):
                bucket['ocorrencias_criticas'] += 1

        return by_collaborator

    def build_bonificacao_indicators(profile, month_reference, collaborators, filial_id=None):
        collaborator_ids = [int(item['id']) for item in collaborators if item.get('id') is not None]
        presence_summary = summarize_presence_by_collaborator(profile, month_reference, filial_id)
        occurrences_summary = fetch_monthly_occurrences_by_collaborator(profile, month_reference, collaborator_ids)

        indicators = {}
        for collaborator in collaborators:
            collaborator_id = int(collaborator.get('id')) if collaborator.get('id') is not None else None
            if collaborator_id is None:
                continue

            presence = presence_summary.get(collaborator_id, {})
            occurrences = occurrences_summary.get(collaborator_id, {})

            faltas = int(presence.get('falta', 0))
            ocorrencias_total = int(occurrences.get('ocorrencias_total', 0))
            ocorrencias_criticas = int(occurrences.get('ocorrencias_criticas', 0))

            # Regra simples e transparente para sugestão automática, mantendo edição manual livre.
            penalty_points = (faltas * 15) + (ocorrencias_total * 5) + (ocorrencias_criticas * 20)
            suggested_percent = max(0, min(100, 100 - penalty_points))

            indicators[collaborator_id] = {
                'faltas_mes': faltas,
                'ocorrencias_mes': ocorrencias_total,
                'ocorrencias_criticas_mes': ocorrencias_criticas,
                'penalty_points': penalty_points,
                'suggested_percent': suggested_percent,
            }

        return indicators

    def build_bonificacao_board(profile, month_reference, filial_id=None):
        metrics = fetch_bonificacao_metricas()
        collaborators = fetch_bonificacao_colaboradores(profile, filial_id)
        lancamentos = fetch_bonificacao_lancamentos(profile, month_reference, filial_id)
        allowed_collaborator_ids = {
            int(item['id'])
            for item in collaborators
            if item.get('id') is not None
        }
        lancamentos = [
            item for item in lancamentos
            if item.get('colaborador_id') is not None and int(item.get('colaborador_id')) in allowed_collaborator_ids
        ]
        indicators = build_bonificacao_indicators(profile, month_reference, collaborators, filial_id)
        annual_total_paid = fetch_bonificacao_total_ano(profile, month_reference.year, filial_id)

        monthly_total_paid = float(sum(float(item.get('valor_aplicado') or 0) for item in lancamentos))

        collaborator_totals = {
            int(item['id']): {
                'colaborador_id': item['id'],
                'nome_completo': item.get('nome_completo') or f"Colaborador {item['id']}",
                'filial_id': item.get('filial_id'),
                'total': 0.0,
            }
            for item in collaborators
        }

        metric_totals = {
            int(item['id']): {
                'metrica_id': item['id'],
                'nome': item.get('nome') or f"Métrica {item['id']}",
                'categoria': item.get('categoria') or 'individual',
                'total': 0.0,
            }
            for item in metrics
        }

        for lancamento in lancamentos:
            colaborador_id = int(lancamento['colaborador_id'])
            metrica_id = int(lancamento['metrica_id'])
            valor = float(lancamento.get('valor_aplicado') or 0)

            if colaborador_id in collaborator_totals:
                collaborator_totals[colaborador_id]['total'] += valor

            if metrica_id in metric_totals:
                metric_totals[metrica_id]['total'] += valor

        return {
            'database_ready': bonificacao_tables_ready(),
            'month_reference': month_reference.isoformat(),
            'month': month_reference.month,
            'year': month_reference.year,
            'month_label': f"{month_reference.month:02d}/{month_reference.year}",
            'can_manage': profile_has_scope_permission(profile, 'manage.bonificacao'),
            'filiais': fetch_accessible_filiais(profile),
            'metricas': metrics,
            'colaboradores': collaborators,
            'indicators_by_collaborator': indicators,
            'lancamentos': lancamentos,
            'summary': {
                'monthly_total_paid': monthly_total_paid,
                'annual_total_paid': annual_total_paid,
                'collaborator_totals': sorted(collaborator_totals.values(), key=lambda item: item['nome_completo']),
                'metric_totals': sorted(metric_totals.values(), key=lambda item: (item['categoria'], item['nome'])),
            },
        }

    def loading_tables_ready():
        required_tables = [
            'rotas_carregamento',
            'veiculos_carregamento',
            'motivos_parada_carregamento',
            'jornadas_carregamento',
            'jornada_carregamento_eventos',
            'jornada_carregamento_fechamentos',
        ]
        return all(table_exists_ready(table_name) for table_name in required_tables)

    def fetch_filial_labels(filial_ids):
        if not filial_ids:
            return []
        query = supabase.table('filiais').select('id, cidade, uf').in_('id', filial_ids).order('cidade')
        response = supabase_retry(query.execute)
        return [f"{item['cidade']}/{item['uf']}" for item in (response.data or [])]

    def fetch_loading_support_rows(profile):
        filiais = fetch_accessible_filiais(profile)

        rotas_query = supabase.table('rotas_carregamento').select('*').eq('ativo', True).order('nome')
        veiculos_query = supabase.table('veiculos_carregamento').select('*').eq('ativo', True).order('placa')
        if profile_has_filial_scope(profile):
            allowed_ids = profile.get('allowed_filial_ids') or []
            rotas_query = rotas_query.in_('filial_id', allowed_ids)
            veiculos_query = veiculos_query.in_('filial_id', allowed_ids)

        motivos_query = (
            supabase.table('motivos_parada_carregamento')
            .select('*')
            .eq('ativo', True)
            .order('ordem')
        )

        return {
            'filiais': filiais,
            'rotas': rotas_query.execute().data or [],
            'veiculos': veiculos_query.execute().data or [],
            'motivos': motivos_query.execute().data or [],
        }

    def fetch_loading_journey(journey_id):
        response = (
            supabase.table('jornadas_carregamento')
            .select('*')
            .eq('id', journey_id)
            .limit(1)
            .execute()
        )
        return response.data[0] if response.data else None

    def ensure_loading_journey_allowed(profile, journey_id):
        journey = fetch_loading_journey(journey_id)
        if not journey:
            return None, (jsonify({'error': 'Jornada de carregamento não encontrada.'}), 404)
        if not ensure_profile_can_access_filial(profile, journey.get('filial_id')):
            return None, (jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403)
        return journey, None

    def fetch_loading_events(journey_ids):
        if not journey_ids:
            return {}

        response = (
            supabase.table('jornada_carregamento_eventos')
            .select('*')
            .in_('jornada_id', journey_ids)
            .order('inicio_evento')
            .execute()
        )
        grouped = {int(journey_id): [] for journey_id in journey_ids}
        for row in response.data or []:
            grouped.setdefault(int(row['jornada_id']), []).append(row)
        return grouped

    def fetch_loading_closures(journey_ids):
        if not journey_ids:
            return {}

        response = (
            supabase.table('jornada_carregamento_fechamentos')
            .select('*')
            .in_('jornada_id', journey_ids)
            .execute()
        )
        return {
            int(row['jornada_id']): row
            for row in (response.data or [])
        }

    def fetch_collaborator_labels(collaborator_ids):
        if not collaborator_ids:
            return {}

        response = (
            supabase.table('colaboradores')
            .select('id, nome_completo, cargo')
            .in_('id', collaborator_ids)
            .execute()
        )
        return {
            int(row['id']): row
            for row in (response.data or [])
        }

    def minutes_between(start_value, end_value=None):
        start_dt = parse_iso_datetime(start_value)
        end_dt = parse_iso_datetime(end_value) if end_value else datetime.now().astimezone()
        if not start_dt or not end_dt:
            return 0
        return max(0, int((end_dt - start_dt).total_seconds() // 60))

    def build_loading_event_summary(event_row):
        event_type = event_row.get('tipo_evento')
        start_value = event_row.get('inicio_evento')
        end_value = event_row.get('fim_evento')
        duration_minutes = minutes_between(start_value, end_value) if event_type in {'carga', 'parada'} and start_value else 0
        return {
            'id': event_row['id'],
            'tipo_evento': event_type,
            'inicio_evento': start_value,
            'fim_evento': end_value,
            'motivo_parada_id': event_row.get('motivo_parada_id'),
            'observacao': event_row.get('observacao') or '',
            'registrado_por': event_row.get('registrado_por'),
            'registrado_em': event_row.get('registrado_em'),
            'duration_minutes': duration_minutes,
            'is_open': bool(event_type in {'carga', 'parada'} and start_value and not end_value),
        }

    def summarize_loading_journey(journey, events, closure, maps):
        event_summaries = [build_loading_event_summary(item) for item in events]
        total_loading_minutes = sum(item['duration_minutes'] for item in event_summaries if item['tipo_evento'] == 'carga')
        total_stoppage_minutes = sum(item['duration_minutes'] for item in event_summaries if item['tipo_evento'] == 'parada')
        carga_open = next((item for item in event_summaries if item['tipo_evento'] == 'carga' and item['is_open']), None)
        parada_open = next((item for item in event_summaries if item['tipo_evento'] == 'parada' and item['is_open']), None)
        lider = maps['collaborators'].get(int(journey['lider_colaborador_id'])) if journey.get('lider_colaborador_id') else None
        veiculo = maps['veiculos'].get(int(journey['veiculo_carregamento_id'])) if journey.get('veiculo_carregamento_id') else None
        rota = maps['rotas'].get(int(journey['rota_id'])) if journey.get('rota_id') else None
        filial = maps['filiais'].get(int(journey['filial_id'])) if journey.get('filial_id') else None

        return {
            'id': journey['id'],
            'data_operacao': journey['data_operacao'],
            'filial_id': journey['filial_id'],
            'filial_nome': f"{filial['cidade']}/{filial['uf']}" if filial else f"Base {journey['filial_id']}",
            'turno': journey.get('turno'),
            'status': journey.get('status') or 'planejado',
            'observacao_abertura': journey.get('observacao_abertura') or '',
            'veiculo_carregamento_id': journey.get('veiculo_carregamento_id'),
            'placa': veiculo.get('placa') if veiculo else '',
            'transportadora': veiculo.get('transportadora') if veiculo else '',
            'tipo_veiculo': veiculo.get('tipo_veiculo') if veiculo else '',
            'capacidade_cilindros': veiculo.get('capacidade_cilindros') if veiculo else None,
            'rota_id': journey.get('rota_id'),
            'rota_nome': rota.get('nome') if rota else '',
            'lider_colaborador_id': journey.get('lider_colaborador_id'),
            'lider_nome': lider.get('nome_completo') if lider else 'Não identificado',
            'lider_cargo': lider.get('cargo') if lider else '',
            'aberto_por': journey.get('aberto_por'),
            'iniciado_em': journey.get('iniciado_em'),
            'finalizado_em': journey.get('finalizado_em'),
            'tempo_carregamento_minutos': total_loading_minutes,
            'tempo_parado_minutos': total_stoppage_minutes,
            'tempo_operacao_minutos': total_loading_minutes + total_stoppage_minutes,
            'ocorrencias_count': sum(1 for item in event_summaries if item['tipo_evento'] == 'ocorrencia'),
            'carga_em_aberto': bool(carga_open),
            'parada_em_aberto': bool(parada_open),
            'eventos': event_summaries,
            'fechamento': {
                'quantidade_cilindros': closure.get('quantidade_cilindros') if closure else None,
                'divergencias': closure.get('divergencias') if closure else '',
                'observacao_fechamento': closure.get('observacao_fechamento') if closure else '',
                'finalizado_por': closure.get('finalizado_por') if closure else None,
                'finalizado_em': closure.get('finalizado_em') if closure else None,
            },
        }

    def list_loading_journeys(profile, target_date, filial_id=None, turno=None):
        query = (
            supabase.table('jornadas_carregamento')
            .select('*')
            .eq('data_operacao', target_date.isoformat())
            .order('id')
        )
        if profile_has_filial_scope(profile):
            query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])
        if filial_id:
            query = query.eq('filial_id', filial_id)
        if turno:
            query = query.eq('turno', turno)

        journeys = query.execute().data or []
        journey_ids = [int(item['id']) for item in journeys]
        events_by_journey = fetch_loading_events(journey_ids)
        closures_by_journey = fetch_loading_closures(journey_ids)
        support_rows = fetch_loading_support_rows(profile)

        maps = {
            'filiais': {int(item['id']): item for item in support_rows['filiais']},
            'rotas': {int(item['id']): item for item in support_rows['rotas']},
            'veiculos': {int(item['id']): item for item in support_rows['veiculos']},
            'collaborators': fetch_collaborator_labels([
                int(item['lider_colaborador_id'])
                for item in journeys
                if item.get('lider_colaborador_id')
            ]),
        }

        return [
            summarize_loading_journey(item, events_by_journey.get(int(item['id']), []), closures_by_journey.get(int(item['id'])), maps)
            for item in journeys
        ]

    def find_open_loading_event(events, event_type):
        return next(
            (item for item in events if item.get('tipo_evento') == event_type and item.get('inicio_evento') and not item.get('fim_evento')),
            None,
        )

    def build_loading_dashboard_summary(profile, target_date):
        if not profile_has_scope_permission(profile, 'menu.carregamento'):
            return {
                'available': False,
                'database_ready': False,
                'date': target_date.isoformat(),
                'cards': [],
                'highlights': [],
            }

        database_ready = loading_tables_ready()
        if not database_ready:
            return {
                'available': True,
                'database_ready': False,
                'date': target_date.isoformat(),
                'cards': [],
                'highlights': [],
            }

        journeys = list_loading_journeys(profile, target_date)
        total_trucks = len(journeys)
        total_stoppage_minutes = sum(int(item.get('tempo_parado_minutos') or 0) for item in journeys)
        open_divergences = sum(1 for item in journeys if item.get('ocorrencias_count') and item.get('status') != 'finalizado')
        finalized_with_divergence = sum(
            1
            for item in journeys
            if (item.get('fechamento') or {}).get('divergencias')
        )

        highlight_rows = sorted(
            journeys,
            key=lambda item: (
                int(item.get('tempo_parado_minutos') or 0),
                int(item.get('ocorrencias_count') or 0),
            ),
            reverse=True,
        )[:5]

        return {
            'available': True,
            'database_ready': True,
            'date': target_date.isoformat(),
            'cards': [
                {
                    'label': 'Caminhões do dia',
                    'value': total_trucks,
                    'hint': 'Jornadas abertas para a data atual nas bases visíveis',
                },
                {
                    'label': 'Tempo parado total',
                    'value': f"{total_stoppage_minutes} min",
                    'hint': 'Soma das paradas registradas na operação do dia',
                },
                {
                    'label': 'Divergências abertas',
                    'value': open_divergences,
                    'hint': 'Jornadas em aberto com ocorrência registrada',
                },
                {
                    'label': 'Divergências fechadas',
                    'value': finalized_with_divergence,
                    'hint': 'Jornadas finalizadas com divergência informada no fechamento',
                },
            ],
            'highlights': [
                {
                    'id': item['id'],
                    'placa': item.get('placa') or '-',
                    'rota_nome': item.get('rota_nome') or '-',
                    'filial_nome': item.get('filial_nome') or '-',
                    'status': item.get('status') or '-',
                    'tempo_parado_minutos': item.get('tempo_parado_minutos') or 0,
                    'ocorrencias_count': item.get('ocorrencias_count') or 0,
                }
                for item in highlight_rows
            ],
        }

    def build_document_alert_items(reference_date):
        if not collaborator_documents_table_ready():
            return []

        rows = (
            supabase.table('colaborador_documentos')
            .select('*')
            .eq('ativo', True)
            .order('data_validade')
            .execute()
        ).data or []

        items = []
        for row in rows:
            document = enrich_collaborator_document(row, reference_date)
            status = document.get('status_calculado')
            if status not in {'vencido', 'vence_em_breve'}:
                continue

            days_until_due = document.get('dias_para_vencer')
            if status == 'vencido':
                title = f"Documento vencido: {document.get('tipo_documento') or 'Documento RH'}"
                message = 'Regularize imediatamente para evitar risco operacional.'
                severity = 'critical'
            else:
                title = f"Documento proximo do vencimento: {document.get('tipo_documento') or 'Documento RH'}"
                if days_until_due is None:
                    message = 'Documento dentro da janela de alerta configurada.'
                else:
                    message = f"Vence em {days_until_due} dia(s)."
                severity = 'warning'

            items.append({
                'type': 'documento_rh',
                'severity': severity,
                'title': title,
                'message': message,
                'filial_id': document.get('filial_id'),
                'reference_id': document.get('id'),
                'colaborador_id': document.get('colaborador_id'),
                'status': status,
                'date_reference': reference_date.isoformat(),
                'date_limit': document.get('data_validade'),
                'days_to_due': days_until_due,
            })

        return items

    def build_rh_event_alert_items(reference_date):
        if not rh_events_table_ready():
            return []

        horizon_date = reference_date + timedelta(days=alerts_horizon_days)
        rows = (
            supabase.table('eventos_rh')
            .select('*')
            .eq('ativo', True)
            .in_('status', ['planejado', 'aprovado', 'em_andamento'])
            .gte('data_inicio', reference_date.isoformat())
            .lte('data_inicio', horizon_date.isoformat())
            .order('data_inicio')
            .execute()
        ).data or []

        items = []
        for event in rows:
            start_date = parse_iso_date(event.get('data_inicio'))
            days_to_start = (start_date - reference_date).days if start_date else None
            severity = 'warning' if days_to_start in (0, 1) else 'info'

            if days_to_start is None:
                message = 'Evento RH com inicio proximo na janela de monitoramento.'
            elif days_to_start == 0:
                message = 'Evento RH inicia hoje.'
            elif days_to_start == 1:
                message = 'Evento RH inicia amanha.'
            else:
                message = f"Evento RH inicia em {days_to_start} dia(s)."

            items.append({
                'type': 'evento_rh',
                'severity': severity,
                'title': f"Evento RH programado: {event.get('tipo_evento') or 'evento'}",
                'message': message,
                'filial_id': event.get('filial_id'),
                'reference_id': event.get('id'),
                'colaborador_id': event.get('colaborador_id'),
                'status': event.get('status'),
                'date_reference': reference_date.isoformat(),
                'date_limit': event.get('data_inicio'),
                'days_to_due': days_to_start,
            })

        return items

    def build_loading_alert_items(reference_date):
        if not loading_tables_ready():
            return []

        journeys = (
            supabase.table('jornadas_carregamento')
            .select('id, data_operacao, filial_id, turno, status, veiculo_carregamento_id')
            .in_('status', ['planejado', 'em_operacao'])
            .lte('data_operacao', reference_date.isoformat())
            .order('data_operacao')
            .execute()
        ).data or []

        vehicle_ids = sorted({
            int(item['veiculo_carregamento_id'])
            for item in journeys
            if item.get('veiculo_carregamento_id')
        })
        vehicle_map = {}
        if vehicle_ids:
            vehicle_rows = (
                supabase.table('veiculos_carregamento')
                .select('id, placa')
                .in_('id', vehicle_ids)
                .execute()
            ).data or []
            vehicle_map = {int(row['id']): row for row in vehicle_rows}

        items = []
        for journey in journeys:
            operation_date = parse_iso_date(journey.get('data_operacao'))
            is_overdue = bool(operation_date and operation_date < reference_date)
            vehicle = vehicle_map.get(int(journey['veiculo_carregamento_id'])) if journey.get('veiculo_carregamento_id') else None

            if is_overdue:
                severity = 'critical'
                message = 'Jornada de dia anterior ainda sem fechamento.'
            elif journey.get('status') == 'em_operacao':
                severity = 'warning'
                message = 'Jornada em operacao exige acompanhamento de fechamento.'
            else:
                severity = 'info'
                message = 'Jornada planejada para hoje ainda nao finalizada.'

            items.append({
                'type': 'jornada_carregamento',
                'severity': severity,
                'title': f"Jornada pendente: {vehicle.get('placa') if vehicle else journey.get('veiculo_carregamento_id')}",
                'message': message,
                'filial_id': journey.get('filial_id'),
                'reference_id': journey.get('id'),
                'colaborador_id': None,
                'status': journey.get('status'),
                'date_reference': reference_date.isoformat(),
                'date_limit': journey.get('data_operacao'),
                'days_to_due': (operation_date - reference_date).days if operation_date else None,
            })

        return items

    def summarize_alert_items(items):
        summary = {
            'total': len(items),
            'critical': 0,
            'warning': 0,
            'info': 0,
        }
        for item in items:
            severity = item.get('severity')
            if severity in summary:
                summary[severity] += 1
        return summary

    def refresh_alert_engine_snapshot():
        now_dt = datetime.now().astimezone()
        reference_date = now_dt.date()

        items = []
        items.extend(build_document_alert_items(reference_date))
        items.extend(build_rh_event_alert_items(reference_date))
        items.extend(build_loading_alert_items(reference_date))

        severity_order = {'critical': 0, 'warning': 1, 'info': 2}
        items.sort(key=lambda item: (severity_order.get(item.get('severity'), 99), item.get('date_limit') or '9999-12-31'))

        with alert_engine_lock:
            alert_engine_state['items'] = items
            alert_engine_state['summary'] = summarize_alert_items(items)
            alert_engine_state['last_run_at'] = now_dt.isoformat()
            alert_engine_state['last_error'] = None

    def run_alert_engine_loop():
        app.logger.info('Motor de alertas iniciado. intervalo=%s min horizonte=%s dias', alerts_interval_minutes, alerts_horizon_days)
        while True:
            try:
                refresh_alert_engine_snapshot()
            except Exception as exc:
                with alert_engine_lock:
                    alert_engine_state['last_error'] = str(exc)
                    alert_engine_state['last_run_at'] = datetime.now().astimezone().isoformat()
                if is_transient_disconnect_error(exc):
                    # Network hiccup — retry sooner (60 s) instead of waiting full interval
                    app.logger.warning('Motor de alertas: falha de rede transitória (%s). Tentando novamente em 60s.', exc)
                    time.sleep(60)
                    continue
                app.logger.error('Falha no motor de alertas: %s', exc)
            time.sleep(alerts_interval_minutes * 60)

    def filter_alert_items_for_profile(profile, items):
        module_scope_map = {
            'documento_rh': 'menu.colaborador_documentos',
            'evento_rh': 'menu.eventos_rh',
            'jornada_carregamento': 'menu.carregamento',
        }

        filtered_items = []
        for item in items:
            required_scope = module_scope_map.get(item.get('type'))
            if required_scope and not profile_has_scope_permission(profile, required_scope):
                continue

            filial_id = item.get('filial_id')
            if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
                continue

            filtered_items.append(item)

        return filtered_items

    def extract_known_permission_scopes(permission_rows):
        scopes = []
        for permission_row in permission_rows:
            permission_name = permission_row.get('permissao_nome')
            if permission_name in PERMISSION_SCOPE_MAP:
                scopes.append(permission_name)
        return sorted(set(scopes))

    def extract_allowed_filial_ids(permission_rows):
        filial_ids = []
        for permission_row in permission_rows:
            permission_name = permission_row.get('permissao_nome') or ''
            if not permission_name.startswith('filial.'):
                continue

            _, _, raw_filial_id = permission_name.partition('.')
            if raw_filial_id.isdigit():
                filial_ids.append(int(raw_filial_id))

        return sorted(set(filial_ids))

    def profile_has_filial_scope(profile):
        if profile.get('is_super_admin'):
            return False
        return bool(profile.get('allowed_filial_ids'))

    def ensure_profile_can_access_filial(profile, filial_id):
        if profile.get('is_super_admin'):
            return True
        if not filial_id:
            return True
        if not profile_has_filial_scope(profile):
            return True
        return int(filial_id) in set(profile.get('allowed_filial_ids') or [])

    def apply_filial_scope(query, profile, config):
        filial_scope_field = config.get('filial_scope_field')
        if not filial_scope_field or not profile_has_filial_scope(profile):
            return query

        allowed_filial_ids = profile.get('allowed_filial_ids') or []
        if not allowed_filial_ids:
            return query

        if config.get('filial_scope_include_null'):
            ids_str = ','.join(str(f) for f in allowed_filial_ids)
            return query.or_(f'{filial_scope_field}.in.({ids_str}),{filial_scope_field}.is.null')

        return query.in_(filial_scope_field, allowed_filial_ids)

    def fetch_resource_item_for_scope(resource_name, item_id):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return None

        response = (
            supabase.table(config['table'])
            .select('*')
            .eq('id', item_id)
            .limit(1)
            .execute()
        )
        return response.data[0] if response.data else None

    def ensure_resource_item_allowed(profile, resource_name, item_id):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return None, (jsonify({'error': 'Recurso não encontrado.'}), 404)

        item = fetch_resource_item_for_scope(resource_name, item_id)
        if not item:
            return None, (jsonify({'error': 'Registro não encontrado.'}), 404)

        filial_scope_field = config.get('filial_scope_field')
        if filial_scope_field and not ensure_profile_can_access_filial(profile, item.get(filial_scope_field)):
            return None, (jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403)

        return item, None

    def profile_has_scope_permission(profile, scope_name):
        if profile.get('is_super_admin'):
            return True

        if not scope_name:
            return True

        known_scopes = set(profile.get('permission_scopes') or [])
        if not profile.get('has_scope_permissions'):
            return True

        return scope_name in known_scopes

    def build_profile(collaborator, user):
        detailed_permissions = fetch_permissions(collaborator['id'])
        permission_scopes = extract_known_permission_scopes(detailed_permissions)
        allowed_filial_ids = extract_allowed_filial_ids(detailed_permissions)
        allowed_filial_labels = fetch_filial_labels(allowed_filial_ids)
        return {
            'id': collaborator['id'],
            'user_id': collaborator['user_id'],
            'filial_id': collaborator['filial_id'],
            'nome_completo': collaborator['nome_completo'],
            'cpf': collaborator['cpf'],
            'telefone': collaborator.get('telefone'),
            'cargo': collaborator['cargo'],
            'foto_url': collaborator.get('foto_url') or collaborator.get('foto_perfil_url'),
            'tipo_acesso': collaborator.get('tipo_acesso', 'app'),
            'permissao_app': collaborator.get('permissao_app', True),
            'permissao_desktop': collaborator.get('permissao_desktop', False),
            'permissao_editar': collaborator.get('permissao_editar', False),
            'permissao_excluir': collaborator.get('permissao_excluir', False),
            'permissao_aprovar_he': collaborator.get('permissao_aprovar_he', False),
            'ativo': collaborator.get('ativo', True),
            'database_online': True,
            'email': user.get('email'),
            'permissions': {
                'app': collaborator.get('permissao_app', True),
                'desktop': collaborator.get('permissao_desktop', False),
                'edit': collaborator.get('permissao_editar', False),
                'delete': collaborator.get('permissao_excluir', False),
                'approve_he': collaborator.get('permissao_aprovar_he', False),
            },
            'permission_scopes': permission_scopes,
            'allowed_filial_ids': allowed_filial_ids,
            'allowed_filial_labels': allowed_filial_labels,
            'has_filial_scope': bool(allowed_filial_ids),
            'has_scope_permissions': bool(permission_scopes),
            'detailed_permissions': detailed_permissions,
            'is_super_admin': False,
        }

    def build_super_admin_profile(user):
        return {
            'id': 0,
            'user_id': str(user['id']),
            'filial_id': None,
            'nome_completo': 'Administrador Master',
            'cpf': '-',
            'telefone': None,
            'cargo': 'Superadmin',
            'foto_url': None,
            'tipo_acesso': 'ambos',
            'permissao_app': True,
            'permissao_desktop': True,
            'permissao_editar': True,
            'permissao_excluir': True,
            'permissao_aprovar_he': True,
            'ativo': True,
            'database_online': True,
            'email': user.get('email'),
            'permissions': {
                'app': True,
                'desktop': True,
                'edit': True,
                'delete': True,
                'approve_he': True,
            },
            'permission_scopes': ALL_PERMISSION_SCOPES,
            'allowed_filial_ids': [],
            'allowed_filial_labels': [],
            'has_filial_scope': False,
            'has_scope_permissions': True,
            'detailed_permissions': [
                {
                    'colaborador_id': 0,
                    'permissao_nome': scope_name,
                    'ativo': True,
                    'descricao': PERMISSION_SCOPE_MAP[scope_name]['description'],
                }
                for scope_name in ALL_PERMISSION_SCOPES
            ],
            'is_super_admin': True,
        }

    def is_super_admin_identity(user, collaborator=None):
        # SEGURANÇA: apenas o e-mail exato do administrador master concede acesso irrestrito.
        # Aliases por local-part ('admin@qualquer.com') e verificação por cargo ('master')
        # foram removidos — ambos criam vetores de escalada de privilégio.
        normalized_email = (user.get('email') or '').strip().lower()
        return normalized_email == SUPER_ADMIN_EMAIL

    def fetch_user_profile():
        token = auth_token_from_request()
        if not token:
            return None, (jsonify({'error': 'Token ausente.'}), 401)

        try:
            user = fetch_authenticated_user(token)
        except RuntimeError as exc:
            return None, (jsonify({'error': str(exc)}), 502)

        if not user:
            return None, (jsonify({'error': 'Token inválido ou expirado.'}), 401)

        if is_super_admin_identity(user):
            return build_super_admin_profile(user), None

        user_id = str(user['id'])

        # Serve from cache when fresh — avoids a Supabase roundtrip on every request
        cached = _profile_cache.get(user_id)
        if cached and cached['expires_at'] > time.time():
            return cached['profile'], None

        try:
            profile_response = supabase_retry(
                lambda: supabase.table('colaboradores')
                .select('*')
                .eq('user_id', user_id)
                .limit(1)
                .execute()
            )
        except Exception as exc:
            app.logger.error('Falha ao carregar colaborador para user_id=%s: %s', user_id[:36], exc)
            # Return stale cache on transient error rather than failing the request
            if cached:
                return cached['profile'], None
            return None, (jsonify({'error': 'Falha ao carregar perfil. Tente novamente.'}), 500)

        if not profile_response.data:
            return None, (
                jsonify({
                    'error': 'Usuário autenticado, mas sem cadastro em colaboradores.',
                    'user_id': user_id,
                }),
                403,
            )

        collaborator = profile_response.data[0]
        if is_super_admin_identity(user, collaborator):
            return build_super_admin_profile(user), None

        profile = build_profile(collaborator, user)
        _profile_cache[user_id] = {'profile': profile, 'expires_at': time.time() + _profile_cache_ttl}
        return profile, None

    def require_auth(handler):
        @wraps(handler)
        def wrapper(*args, **kwargs):
            profile, error = fetch_user_profile()
            if error:
                return error
            return handler(profile, *args, **kwargs)

        return wrapper

    def require_edit_permission(profile):
        if profile.get('is_super_admin'):
            return None

        if profile['permissions']['edit'] or profile['permissions']['approve_he']:
            return None
        return jsonify({'error': 'Sem permissão para editar registros.'}), 403

    def require_create_permission(profile, resource_name):
        if profile.get('is_super_admin'):
            return None

        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        create_scope = config.get('create_scope')
        if profile_has_scope_permission(profile, create_scope):
            return None

        return jsonify({'error': 'Sem permissão para cadastrar neste módulo.'}), 403

    def require_delete_permission(profile):
        if profile.get('is_super_admin'):
            return None

        if profile['permissions']['delete']:
            return None
        return jsonify({'error': 'Sem permissão para excluir registros.'}), 403

    def require_scope_permission(profile, scope_name, message='Sem permissão para acessar este módulo.'):
        if profile_has_scope_permission(profile, scope_name):
            return None
        return jsonify({'error': message}), 403

    def serve_frontend_index():
        index_file = FRONTEND_DIST_DIR / 'index.html'
        if index_file.exists():
            response = make_response(send_from_directory(str(FRONTEND_DIST_DIR), 'index.html'))
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        return jsonify({'name': 'SEG API', 'status': 'ok', 'frontend': 'dist ausente'})

    @app.get('/')
    def root():
        return serve_frontend_index()

    @app.get('/assets/<path:path>')
    def frontend_asset(path):
        asset_path = FRONTEND_DIST_DIR / 'assets' / path
        try:
            asset_path.resolve().relative_to((FRONTEND_DIST_DIR / 'assets').resolve())
        except ValueError:
            return jsonify({'error': 'Recurso não encontrado.'}), 404
        if asset_path.exists() and asset_path.is_file():
            response = make_response(send_from_directory(str(FRONTEND_DIST_DIR / 'assets'), path))
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        return jsonify({'error': 'Recurso não encontrado.'}), 404

    @app.get('/<path:path>')
    def frontend_routes(path):
        if path.startswith('api/'):
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        direct_file = FRONTEND_DIST_DIR / path
        try:
            direct_file.resolve().relative_to(FRONTEND_DIST_DIR.resolve())
        except ValueError:
            return serve_frontend_index()
        if direct_file.exists() and direct_file.is_file():
            response = make_response(send_from_directory(str(FRONTEND_DIST_DIR), path))
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response

        return serve_frontend_index()

    @app.get('/api')
    def api_root():
        return jsonify({'name': 'SEG API', 'status': 'ok'})

    @app.get('/api/health')
    def health():
        return jsonify({'status': 'ok'})

    # ─── Recuperação de senha (público) ───────────────────────────────────────
    # Fluxo OTP de 6 dígitos enviado por SMTP para o Gmail informado na hora.
    # Usamos isso porque o e-mail registrado em auth.users pode ser fictício
    # (apenas para entrada no sistema) e não chega na caixa real do colaborador.
    #
    # Endpoints:
    #   POST /api/recuperar-senha/enviar-codigo  { email_login, email_destino }
    #   POST /api/recuperar-senha/validar-codigo { email_login, codigo }     → reset_token
    #   POST /api/recuperar-senha/redefinir      { reset_token, nova_senha }

    import smtplib as _smtplib
    from email.mime.text import MIMEText as _MIMEText
    from email.mime.multipart import MIMEMultipart as _MIMEMultipart
    import secrets as _secrets
    import uuid as _uuid

    def _smtp_config():
        host = os.getenv('SMTP_HOST', 'smtp.gmail.com')
        port = int(os.getenv('SMTP_PORT', '587'))
        user = os.getenv('SMTP_USER', '')
        password = os.getenv('SMTP_PASS', '')
        from_addr = os.getenv('SMTP_FROM', user)
        from_name = os.getenv('SMTP_FROM_NAME', 'SEG - Gold Transportes')
        return host, port, user, password, from_addr, from_name

    def _send_email_otp(destination, codigo, nome_remetente='Gold Transportes'):
        host, port, user, password, from_addr, from_name = _smtp_config()
        if not user or not password or not from_addr:
            raise RuntimeError('SMTP não configurado no servidor. Defina SMTP_HOST/PORT/USER/PASS/FROM no .env do backend.')
        msg = _MIMEMultipart('alternative')
        msg['Subject'] = f'Código de recuperação de senha: {codigo}'
        msg['From'] = f'{from_name} <{from_addr}>'
        msg['To'] = destination
        text = (
            f'Olá,\n\n'
            f'Seu código de recuperação de senha no SEG é:\n\n'
            f'    {codigo}\n\n'
            f'Este código vale por 15 minutos. Se você não solicitou, ignore este e-mail.\n\n'
            f'— {nome_remetente}'
        )
        html = f"""
        <div style="font-family:Arial,Helvetica,sans-serif;max-width:480px;margin:0 auto;padding:20px;color:#1a1a1a">
          <h2 style="color:#c49512;margin:0 0 12px">Recuperação de senha</h2>
          <p>Olá,</p>
          <p>Seu código de recuperação no SEG é:</p>
          <div style="font-size:28px;letter-spacing:6px;font-weight:700;background:#f6f2e6;border:1px solid #e3d8b5;padding:12px 20px;border-radius:6px;display:inline-block;font-family:'Courier New',monospace;margin:8px 0">
            {codigo}
          </div>
          <p style="color:#666;font-size:13px">Vale por <strong>15 minutos</strong>. Se você não solicitou, ignore.</p>
          <hr style="border:none;border-top:1px solid #ddd;margin:20px 0"/>
          <p style="color:#999;font-size:11px">{nome_remetente} — SEG (Sistema de Gestão)</p>
        </div>
        """
        msg.attach(_MIMEText(text, 'plain', 'utf-8'))
        msg.attach(_MIMEText(html, 'html', 'utf-8'))

        with _smtplib.SMTP(host, port, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(user, password)
            smtp.sendmail(from_addr, [destination], msg.as_string())

    def _find_auth_user_by_email(email):
        """Busca um usuário no auth.users pelo e-mail. Retorna o objeto ou None."""
        email_norm = (email or '').strip().lower()
        if not email_norm:
            return None
        try:
            page = 1
            while page <= 50:  # limita varredura (até 5000 usuários)
                resp = supabase.auth.admin.list_users(page=page, per_page=100)
                users = resp if isinstance(resp, list) else getattr(resp, 'users', None) or resp
                if not users:
                    break
                for u in users:
                    u_email = (getattr(u, 'email', None) or (u.get('email') if isinstance(u, dict) else None) or '').lower()
                    if u_email == email_norm:
                        return u
                if len(users) < 100:
                    break
                page += 1
        except Exception as exc:
            app.logger.warning('Falha ao buscar usuário auth por e-mail: %s', exc)
        return None

    def _ip_origem():
        return (request.headers.get('X-Forwarded-For', '') or request.remote_addr or '').split(',')[0].strip()

    @app.post('/api/recuperar-senha/enviar-codigo')
    def recuperar_senha_enviar_codigo():
        data = request.get_json(silent=True) or {}
        email_login = (data.get('email_login') or '').strip().lower()
        email_destino = (data.get('email_destino') or '').strip()
        if not email_login or '@' not in email_login:
            return jsonify({'error': 'Informe o e-mail de login válido.'}), 400
        if not email_destino or '@' not in email_destino:
            return jsonify({'error': 'Informe um Gmail válido para receber o código.'}), 400

        # Rate limit por email_login (máx 3 pedidos em 1h) e por IP (máx 10 em 1h)
        try:
            uma_hora_atras = (datetime.utcnow() - timedelta(hours=1)).isoformat()
            ip = _ip_origem()
            recents_login = supabase.table('password_reset_codes').select('id', count='exact').eq('email_login', email_login).gte('criado_em', uma_hora_atras).execute()
            if (recents_login.count or 0) >= 3:
                return jsonify({'error': 'Limite de 3 pedidos por hora atingido para este login. Aguarde antes de tentar de novo.'}), 429
            if ip:
                recents_ip = supabase.table('password_reset_codes').select('id', count='exact').eq('ip_origem', ip).gte('criado_em', uma_hora_atras).execute()
                if (recents_ip.count or 0) >= 10:
                    return jsonify({'error': 'Muitos pedidos vindos do seu IP. Aguarde alguns minutos.'}), 429
        except Exception as exc:
            app.logger.warning('Falha no rate limit de password reset: %s', exc)

        # Confere se o login existe — mas não revela o resultado pro chamador
        # (anti-enumeração de e-mails). Sempre responde sucesso.
        user = _find_auth_user_by_email(email_login)
        if not user:
            app.logger.info('Pedido de reset para e-mail não cadastrado: %s', email_login)
            return jsonify({'ok': True, 'message': 'Se o e-mail estiver cadastrado, um código será enviado.'})

        codigo = ''.join([_secrets.choice('0123456789') for _ in range(6)])
        expira_em = (datetime.utcnow() + timedelta(minutes=15)).isoformat()
        try:
            supabase.table('password_reset_codes').insert({
                'email_login': email_login,
                'email_destino': email_destino,
                'codigo': codigo,
                'ip_origem': _ip_origem() or None,
                'expira_em': expira_em,
            }).execute()
        except Exception as exc:
            app.logger.error('Falha ao gravar password_reset_codes: %s', exc)
            return jsonify({'error': 'Falha temporária. Tente novamente em alguns minutos.'}), 500

        # Envia o e-mail. Se SMTP falhar, retorna erro com diagnóstico.
        try:
            _send_email_otp(email_destino, codigo)
        except Exception as exc:
            app.logger.error('Falha ao enviar e-mail OTP para %s: %s', email_destino, exc)
            return jsonify({'error': 'Não conseguimos enviar o e-mail agora. Verifique o endereço e tente novamente.'}), 500

        return jsonify({'ok': True, 'message': 'Código enviado.'})

    @app.post('/api/recuperar-senha/validar-codigo')
    def recuperar_senha_validar_codigo():
        data = request.get_json(silent=True) or {}
        email_login = (data.get('email_login') or '').strip().lower()
        codigo = (data.get('codigo') or '').strip()
        if not email_login or not codigo:
            return jsonify({'error': 'Informe o e-mail e o código.'}), 400
        if not codigo.isdigit() or len(codigo) != 6:
            return jsonify({'error': 'Código inválido.'}), 400

        try:
            agora = datetime.utcnow().isoformat()
            rows = (supabase.table('password_reset_codes')
                    .select('id, codigo, tentativas, expira_em, usado')
                    .eq('email_login', email_login)
                    .eq('usado', False)
                    .gte('expira_em', agora)
                    .order('criado_em', desc=True)
                    .limit(1)
                    .execute()).data or []
        except Exception as exc:
            app.logger.error('Falha ao consultar password_reset_codes: %s', exc)
            return jsonify({'error': 'Falha temporária. Tente novamente.'}), 500

        if not rows:
            return jsonify({'error': 'Código inválido ou expirado. Solicite um novo.'}), 400

        row = rows[0]
        if (row.get('tentativas') or 0) >= 5:
            try:
                supabase.table('password_reset_codes').update({'usado': True, 'usado_em': agora}).eq('id', row['id']).execute()
            except Exception:
                pass
            return jsonify({'error': 'Muitas tentativas erradas. Solicite um novo código.'}), 429

        if str(row.get('codigo')) != codigo:
            try:
                supabase.table('password_reset_codes').update({'tentativas': (row.get('tentativas') or 0) + 1}).eq('id', row['id']).execute()
            except Exception:
                pass
            return jsonify({'error': 'Código não confere.'}), 400

        # Código OK: gera reset_token (10 min) e marca como usado
        reset_token = _uuid.uuid4().hex
        expira_token = (datetime.utcnow() + timedelta(minutes=10)).isoformat()
        try:
            supabase.table('password_reset_codes').update({
                'usado': True,
                'usado_em': agora,
                'reset_token': reset_token,
                'expira_em': expira_token,  # reusa o campo: agora marca expiração do token
            }).eq('id', row['id']).execute()
        except Exception as exc:
            app.logger.error('Falha ao gerar reset_token: %s', exc)
            return jsonify({'error': 'Falha temporária.'}), 500

        return jsonify({'ok': True, 'reset_token': reset_token})

    @app.post('/api/recuperar-senha/redefinir')
    def recuperar_senha_redefinir():
        data = request.get_json(silent=True) or {}
        reset_token = (data.get('reset_token') or '').strip()
        nova_senha = data.get('nova_senha') or ''
        if not reset_token:
            return jsonify({'error': 'Sessão inválida. Recomece o processo.'}), 400
        if not isinstance(nova_senha, str) or len(nova_senha) < 6:
            return jsonify({'error': 'A senha precisa ter pelo menos 6 caracteres.'}), 400

        try:
            agora = datetime.utcnow().isoformat()
            rows = (supabase.table('password_reset_codes')
                    .select('id, email_login, expira_em, reset_token')
                    .eq('reset_token', reset_token)
                    .gte('expira_em', agora)
                    .limit(1)
                    .execute()).data or []
        except Exception as exc:
            app.logger.error('Falha ao consultar reset_token: %s', exc)
            return jsonify({'error': 'Falha temporária.'}), 500

        if not rows:
            return jsonify({'error': 'Sessão de redefinição expirou. Recomece o processo.'}), 400

        row = rows[0]
        email_login = row.get('email_login')
        user = _find_auth_user_by_email(email_login)
        if not user:
            return jsonify({'error': 'Usuário não encontrado.'}), 404

        user_id = getattr(user, 'id', None) or (user.get('id') if isinstance(user, dict) else None)
        try:
            supabase.auth.admin.update_user_by_id(str(user_id), {'password': nova_senha})
        except Exception as exc:
            app.logger.error('Falha ao atualizar senha do user %s: %s', user_id, exc)
            return jsonify({'error': 'Não foi possível atualizar a senha. Tente novamente.'}), 500

        # Invalida o token consumindo o registro
        try:
            supabase.table('password_reset_codes').update({
                'reset_token': None,
            }).eq('id', row['id']).execute()
        except Exception:
            pass

        return jsonify({'ok': True})

    # ─── Assinatura SaaS ──────────────────────────────────────────────────────

    _assinatura_cache = {'data': None, 'expires_at': 0.0}
    _assinatura_lock = threading.Lock()
    ASSINATURA_CACHE_TTL = 300  # 5 minutos

    def fetch_assinatura_info(force=False):
        now = time.time()
        with _assinatura_lock:
            if not force and _assinatura_cache['data'] is not None and now < _assinatura_cache['expires_at']:
                return _assinatura_cache['data']

        try:
            rows = (
                supabase.table('assinatura')
                .select('*, planos(nome, max_colaboradores, max_filiais, preco_mensal_brl, stripe_price_id)')
                .limit(1)
                .execute()
            ).data or []
        except Exception:
            return None

        info = rows[0] if rows else None
        if info:
            hoje = date_class.today()
            trial_end = None
            dias_restantes_trial = None
            if info.get('trial_end_date'):
                try:
                    trial_end = date_class.fromisoformat(str(info['trial_end_date']))
                    dias_restantes_trial = (trial_end - hoje).days
                except Exception:
                    pass
            period_end = None
            if info.get('current_period_end'):
                try:
                    period_end = date_class.fromisoformat(str(info['current_period_end']))
                except Exception:
                    pass
            info['dias_restantes_trial'] = dias_restantes_trial
            info['trial_end_date'] = trial_end.isoformat() if trial_end else None
            info['current_period_end'] = period_end.isoformat() if period_end else None
            plano = info.pop('planos', None) or {}
            info['plano_nome'] = plano.get('nome', '')
            info['max_colaboradores'] = plano.get('max_colaboradores')
            info['max_filiais'] = plano.get('max_filiais')
            info['preco_mensal_brl'] = plano.get('preco_mensal_brl')
            info['stripe_price_id'] = plano.get('stripe_price_id')

        with _assinatura_lock:
            _assinatura_cache['data'] = info
            _assinatura_cache['expires_at'] = time.time() + ASSINATURA_CACHE_TTL
        return info

    def invalidate_assinatura_cache():
        with _assinatura_lock:
            _assinatura_cache['data'] = None
            _assinatura_cache['expires_at'] = 0.0

    STRIPE_SECRET_KEY = os.getenv('STRIPE_SECRET_KEY', '')
    STRIPE_WEBHOOK_SECRET = os.getenv('STRIPE_WEBHOOK_SECRET', '')
    STRIPE_PUBLISHABLE_KEY = os.getenv('STRIPE_PUBLISHABLE_KEY', '')
    STRIPE_SUCCESS_URL = os.getenv('STRIPE_SUCCESS_URL', f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/assinatura?checkout=success")
    STRIPE_CANCEL_URL = os.getenv('STRIPE_CANCEL_URL', f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/assinatura?checkout=cancel")

    try:
        import stripe as stripe_lib
        stripe_lib.api_key = STRIPE_SECRET_KEY
        _stripe_available = bool(STRIPE_SECRET_KEY)
    except ImportError:
        stripe_lib = None
        _stripe_available = False

    @app.get('/api/me')
    @require_auth
    def me(profile):
        assinatura = fetch_assinatura_info()
        if assinatura:
            profile['assinatura_status'] = assinatura.get('status')
            profile['assinatura_plano'] = assinatura.get('plano_nome')
            profile['assinatura_dias_trial'] = assinatura.get('dias_restantes_trial')
            profile['assinatura_period_end'] = assinatura.get('current_period_end')
        return jsonify(profile)

    @app.get('/api/assinatura')
    @require_auth
    def get_assinatura(profile):
        info = fetch_assinatura_info()
        if not info:
            return jsonify({'error': 'Assinatura não encontrada.'}), 404
        return jsonify(info)

    @app.get('/api/planos')
    @require_auth
    def get_planos(profile):
        try:
            rows = supabase.table('planos').select('*').eq('ativo', True).order('preco_mensal_brl').execute().data or []
            return jsonify(rows)
        except Exception as exc:
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/assinatura/checkout')
    @require_auth
    def create_checkout(profile):
        if not profile.get('is_super_admin'):
            return jsonify({'error': 'Apenas o administrador pode gerenciar a assinatura.'}), 403
        if not _stripe_available:
            return jsonify({'error': 'Stripe não configurado. Defina STRIPE_SECRET_KEY no servidor.'}), 503

        body = request.get_json(silent=True) or {}
        plano_id = body.get('plano_id')
        if not plano_id:
            return jsonify({'error': 'plano_id obrigatório.'}), 400

        try:
            plano_rows = supabase.table('planos').select('*').eq('id', plano_id).eq('ativo', True).limit(1).execute().data or []
            if not plano_rows:
                return jsonify({'error': 'Plano não encontrado.'}), 404
            plano = plano_rows[0]
            if not plano.get('stripe_price_id'):
                return jsonify({'error': 'Plano sem stripe_price_id configurado.'}), 400

            assinatura = fetch_assinatura_info() or {}
            customer_id = assinatura.get('stripe_customer_id')
            if not customer_id:
                customer = stripe_lib.Customer.create(email=profile.get('email'), metadata={'seg_system': '1'})
                customer_id = customer.id
                sub_rows = supabase.table('assinatura').select('id').limit(1).execute().data or []
                if sub_rows:
                    supabase.table('assinatura').update({'stripe_customer_id': customer_id}).eq('id', sub_rows[0]['id']).execute()
                invalidate_assinatura_cache()

            session = stripe_lib.checkout.Session.create(
                customer=customer_id,
                mode='subscription',
                line_items=[{'price': plano['stripe_price_id'], 'quantity': 1}],
                success_url=STRIPE_SUCCESS_URL,
                cancel_url=STRIPE_CANCEL_URL,
                metadata={'plano_id': str(plano_id)},
            )
            return jsonify({'checkout_url': session.url})
        except Exception as exc:
            app.logger.error('Erro ao criar checkout Stripe: %s', exc)
            return jsonify({'error': str(exc)}), 500

    @app.post('/api/assinatura/portal')
    @require_auth
    def create_portal(profile):
        if not profile.get('is_super_admin'):
            return jsonify({'error': 'Apenas o administrador pode gerenciar a assinatura.'}), 403
        if not _stripe_available:
            return jsonify({'error': 'Stripe não configurado.'}), 503

        assinatura = fetch_assinatura_info() or {}
        customer_id = assinatura.get('stripe_customer_id')
        if not customer_id:
            return jsonify({'error': 'Sem cliente Stripe associado. Faça uma assinatura primeiro.'}), 400

        try:
            session = stripe_lib.billing_portal.Session.create(
                customer=customer_id,
                return_url=f"{os.getenv('FRONTEND_URL', 'http://localhost:5173')}/assinatura",
            )
            return jsonify({'portal_url': session.url})
        except Exception as exc:
            app.logger.error('Erro ao criar portal Stripe: %s', exc)
            return jsonify({'error': str(exc)}), 500

    @app.post('/api/stripe/webhook')
    def stripe_webhook():
        if not _stripe_available:
            return jsonify({'error': 'Stripe não configurado.'}), 503

        payload = request.get_data()
        sig_header = request.headers.get('Stripe-Signature', '')
        try:
            event = stripe_lib.Webhook.construct_event(payload, sig_header, STRIPE_WEBHOOK_SECRET)
        except stripe_lib.error.SignatureVerificationError:
            return jsonify({'error': 'Assinatura inválida.'}), 400
        except Exception as exc:
            return jsonify({'error': str(exc)}), 400

        event_type = event['type']
        obj = event['data']['object']

        try:
            sub_rows = supabase.table('assinatura').select('id').limit(1).execute().data or []
            if not sub_rows:
                return jsonify({'received': True})
            assinatura_id = sub_rows[0]['id']

            STATUS_MAP = {
                'active': 'ativa',
                'past_due': 'inadimplente',
                'canceled': 'cancelada',
                'unpaid': 'inadimplente',
                'trialing': 'trial',
            }

            if event_type in ('customer.subscription.updated', 'customer.subscription.created'):
                novo_status = STATUS_MAP.get(obj.get('status', ''), 'inadimplente')
                period_end = None
                if obj.get('current_period_end'):
                    try:
                        period_end = datetime.fromtimestamp(obj['current_period_end']).date().isoformat()
                    except Exception:
                        pass
                update = {'status': novo_status, 'stripe_subscription_id': obj.get('id')}
                if period_end:
                    update['current_period_end'] = period_end
                plano_id_meta = (obj.get('metadata') or {}).get('plano_id')
                if plano_id_meta and str(plano_id_meta).isdigit():
                    update['plano_id'] = int(plano_id_meta)
                supabase.table('assinatura').update(update).eq('id', assinatura_id).execute()
                invalidate_assinatura_cache()

            elif event_type == 'customer.subscription.deleted':
                supabase.table('assinatura').update({'status': 'cancelada'}).eq('id', assinatura_id).execute()
                invalidate_assinatura_cache()

            elif event_type == 'invoice.payment_failed':
                supabase.table('assinatura').update({'status': 'inadimplente'}).eq('id', assinatura_id).execute()
                invalidate_assinatura_cache()

            elif event_type == 'invoice.payment_succeeded':
                period_end = None
                lines = obj.get('lines', {}).get('data', [])
                if lines and lines[0].get('period', {}).get('end'):
                    try:
                        period_end = datetime.fromtimestamp(lines[0]['period']['end']).date().isoformat()
                    except Exception:
                        pass
                update = {'status': 'ativa'}
                if period_end:
                    update['current_period_end'] = period_end
                supabase.table('assinatura').update(update).eq('id', assinatura_id).execute()
                invalidate_assinatura_cache()

        except Exception as exc:
            app.logger.error('Erro ao processar webhook Stripe %s: %s', event_type, exc)

        return jsonify({'received': True})

    # -------------------------------------------------------------------------
    # FROTA DASHBOARD
    # -------------------------------------------------------------------------
    @app.get('/api/dashboard/frota')
    @require_auth
    def dashboard_frota(profile):
        require_scope(profile, 'menu.frota_dashboard')
        filial_id = request.args.get('filial_id')
        mes = request.args.get('mes')  # YYYY-MM, default = mês atual

        from datetime import date
        import calendar

        if mes:
            try:
                ano, m = int(mes[:4]), int(mes[5:7])
            except (ValueError, IndexError):
                return jsonify({'error': 'Parâmetro mes inválido. Use YYYY-MM.'}), 400
        else:
            hoje = date.today()
            ano, m = hoje.year, hoje.month

        mes_inicio = f'{ano:04d}-{m:02d}-01'
        _, ultimo_dia = calendar.monthrange(ano, m)
        mes_fim = f'{ano:04d}-{m:02d}-{ultimo_dia:02d}'

        def q(table, filters=None, select='*', count_only=False):
            query = supabase.table(table).select(select, count='exact' if count_only else None)
            if filters:
                for k, v in filters.items():
                    if isinstance(v, tuple) and v[0] == 'gte':
                        query = query.gte(k, v[1])
                    elif isinstance(v, tuple) and v[0] == 'lte':
                        query = query.lte(k, v[1])
                    else:
                        query = query.eq(k, v)
            r = query.execute()
            return r

        fil = {'filial_id': filial_id} if filial_id else {}

        # Veículos
        veic_r = q('veiculos', {**fil})
        veiculos_rows = veic_r.data or []
        total_veic = len(veiculos_rows)
        ativos = sum(1 for v in veiculos_rows if v.get('status') == 'ativo')
        em_manut = sum(1 for v in veiculos_rows if v.get('status') == 'manutencao')

        # Manutenções do mês
        manut_fil = {**fil, 'data_abertura': ('gte', mes_inicio)}
        manut_r = q('manutencoes', {**manut_fil, 'ativo': True})
        manut_rows = manut_r.data or []
        os_abertas = sum(1 for m in manut_rows if m.get('status') not in ('concluida', 'cancelada'))
        os_aguardando = sum(1 for m in manut_rows if m.get('status') == 'aguardando_aprovacao')
        custo_manut = sum(float(m.get('valor_final') or m.get('valor_estimado') or 0) for m in manut_rows)

        # Abastecimentos do mês
        abastec_fil = {
            **fil,
            'data_abastecimento': ('gte', mes_inicio),
            'ativo': True,
        }
        abastec_r = q('veiculos_abastecimentos', abastec_fil)
        abastec_rows = abastec_r.data or []
        litros_mes = sum(float(a.get('litros') or 0) for a in abastec_rows)
        gasto_combustivel = sum(float(a.get('valor_total') or 0) for a in abastec_rows)

        # Pneus com status crítico
        pneu_r = q('veiculos_pneus', {**fil, 'ativo': True})
        pneu_rows = pneu_r.data or []
        pneus_trocar = sum(1 for p in pneu_rows if p.get('status') == 'trocar')
        pneus_rodiziar = sum(1 for p in pneu_rows if p.get('status') == 'rodiziar')

        return jsonify({
            'mes': f'{ano:04d}-{m:02d}',
            'veiculos': {
                'total': total_veic,
                'ativos': ativos,
                'em_manutencao': em_manut,
                'disponibilidade_pct': round(ativos / total_veic * 100, 1) if total_veic else 0,
            },
            'manutencoes': {
                'os_abertas': os_abertas,
                'os_aguardando_aprovacao': os_aguardando,
                'custo_mes': custo_manut,
                'total_os_mes': len(manut_rows),
            },
            'combustivel': {
                'litros_mes': round(litros_mes, 2),
                'gasto_mes': round(gasto_combustivel, 2),
                'media_preco_litro': round(gasto_combustivel / litros_mes, 3) if litros_mes else 0,
                'abastecimentos_count': len(abastec_rows),
            },
            'pneus': {
                'total_montados': len(pneu_rows),
                'para_trocar': pneus_trocar,
                'para_rodiziar': pneus_rodiziar,
            },
        })

    @app.get('/api/dashboard')
    @require_auth
    def dashboard(profile):
        filial_id = request.args.get('filial_id', type=int)
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para esta base.'}), 403

        token_counts = []
        for label, table, filial_field in [
            ('Filiais',         'filiais',                 None),
            ('Colaboradores',   'colaboradores',            'filial_id'),
            ('Documentos RH',   'colaborador_documentos',   'filial_id'),
            ('Planejamento RH', 'eventos_rh',               'filial_id'),
            ('Veículos',        'veiculos',                 'filial_id'),
            ('Estoque (itens)', 'estoque_itens',            'filial_id'),
        ]:
            try:
                result = supabase.table(table).select('id', count='exact').limit(1)
                resource_config = RESOURCE_DEFINITIONS.get(table)
                if resource_config:
                    result = apply_filial_scope(result, profile, resource_config)
                if filial_id and filial_field:
                    result = result.eq(filial_field, filial_id)
                result = result.execute()
                count = result.count or 0
            except Exception:
                count = 0
            token_counts.append({
                'label': label,
                'value': count,
                'hint': 'Registros ativos na base operacional',
            })

        workforce_board = build_workforce_board(profile, filial_id)
        loading_summary = build_loading_dashboard_summary(profile, date_class.today())
        collaborator_documents_summary = build_collaborator_documents_alert_summary(profile)

        # Feriados dos próximos 60 dias
        feriados_proximos = []
        try:
            hoje = date_class.today()
            limite = hoje + timedelta(days=60)
            fq = (
                supabase.table('feriados')
                .select('id, nome, data, tipo, uf, tem_expediente, horario_expediente')
                .eq('ativo', True)
                .gte('data', hoje.isoformat())
                .lte('data', limite.isoformat())
                .order('data')
                .limit(8)
            )
            all_feriados = fq.execute().data or []
            # Filtra por UF da filial selecionada quando aplicável
            if filial_id:
                filial_row = next((f for f in (supabase.table('filiais').select('uf').eq('id', filial_id).limit(1).execute().data or [])), None)
                filial_uf = (filial_row or {}).get('uf', '').upper()
                feriados_proximos = [
                    f for f in all_feriados
                    if not f.get('uf') or (filial_uf and f.get('uf', '').upper() == filial_uf)
                ]
            else:
                feriados_proximos = all_feriados
        except Exception:
            feriados_proximos = []

        # Estoque: itens com saldo <= mínimo
        estoque_alertas = []
        try:
            eq = supabase.table('estoque_itens').select('id, nome, estoque_atual, estoque_minimo, filial_id').eq('ativo', True)
            if filial_id:
                eq = eq.eq('filial_id', filial_id)
            else:
                eq = apply_filial_scope(eq, profile, RESOURCE_DEFINITIONS.get('estoque_itens') or {})
            all_itens = eq.execute().data or []
            estoque_alertas = [
                item for item in all_itens
                if item.get('estoque_minimo') is not None
                and float(item.get('estoque_atual') or 0) <= float(item.get('estoque_minimo') or 0)
            ][:10]
        except Exception:
            estoque_alertas = []

        # Pedidos de compra pendentes
        pedidos_pendentes = 0
        try:
            pq = supabase.table('pedidos_compra').select('id', count='exact').eq('ativo', True).eq('status', 'pendente').limit(1)
            if filial_id:
                pq = pq.eq('filial_id', filial_id)
            pedidos_pendentes = pq.execute().count or 0
        except Exception:
            pedidos_pendentes = 0

        with alert_engine_lock:
            raw_alert_items = list(alert_engine_state.get('items') or [])
            last_run_at = alert_engine_state.get('last_run_at')
            last_error = alert_engine_state.get('last_error')
        alert_items = filter_alert_items_for_profile(profile, raw_alert_items)
        alert_summary = summarize_alert_items(alert_items)

        return jsonify({
            'resumo': token_counts,
            'bases': workforce_board['summary_by_filial'],
            'carregamento': loading_summary,
            'rh_documentos': collaborator_documents_summary,
            'feriados_proximos': feriados_proximos,
            'estoque_alertas': estoque_alertas,
            'pedidos_pendentes': pedidos_pendentes,
            'alertas': {
                'last_run_at': last_run_at,
                'last_error': last_error,
                'summary': alert_summary,
                'items': alert_items,
            },
            'profile': profile,
        })

    @app.get('/api/alertas')
    @require_auth
    def alerts_list(profile):
        with alert_engine_lock:
            raw_items = list(alert_engine_state.get('items') or [])
            last_run_at = alert_engine_state.get('last_run_at')
            last_error = alert_engine_state.get('last_error')

        items = filter_alert_items_for_profile(profile, raw_items)
        return jsonify({
            'enabled': alerts_enabled,
            'interval_minutes': alerts_interval_minutes,
            'horizon_days': alerts_horizon_days,
            'last_run_at': last_run_at,
            'last_error': last_error,
            'summary': summarize_alert_items(items),
            'items': items,
        })

    @app.get('/api/quadro-funcionarios')
    @require_auth
    def workforce_board(profile):
        scope_error = require_scope_permission(profile, 'menu.quadro_funcionarios')
        if scope_error:
            return scope_error

        filial_id = request.args.get('filial_id', type=int)
        target_date = parse_iso_date(request.args.get('data')) or date_class.today()
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            return jsonify(build_workforce_board(profile, filial_id, target_date))
        except Exception as exc:
            app.logger.error('Erro ao carregar quadro de funcionários: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/custos-rh')
    @require_auth
    def costs_rh(profile):
        has_cost_scope = profile_has_scope_permission(profile, 'menu.custos_rh')
        has_collab_scope = profile_has_scope_permission(profile, 'menu.colaboradores')
        if not (has_cost_scope or has_collab_scope):
            return jsonify({'error': 'Sem permissão para acessar o painel de custos RH.'}), 403

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        filial_id = request.args.get('filial_id', type=int)
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            return jsonify(build_costs_dashboard(profile, month_reference, filial_id))
        except Exception as exc:
            app.logger.error('Erro ao carregar painel de custos RH: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/<resource_name>')
    @require_auth
    def list_resource(profile, resource_name):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        view_scope_any = config.get('view_scope_any')
        if view_scope_any:
            if not any(profile_has_scope_permission(profile, s) for s in view_scope_any):
                return jsonify({'error': 'Sem permissão para acessar este módulo.'}), 403
        else:
            scope_error = require_scope_permission(profile, config.get('view_scope'))
            if scope_error:
                return scope_error

        try:
            partial_match_fields = set(config.get('partial_match_fields', []))
            _allowed_filter_fields = set(config.get('allowed_fields', []))

            raw_page = request.args.get('page')
            paginated = raw_page is not None
            page = max(1, int(raw_page)) if raw_page and str(raw_page).isdigit() else 1
            per_page = min(200, max(1, request.args.get('per_page', 50, type=int)))
            offset = (page - 1) * per_page

            def execute_list_query_once():
                query = supabase.table(config['table']).select('*', count='exact').order(config['order'])
                query = apply_filial_scope(query, profile, config)

                for key, value in request.args.items():
                    if key in ('page', 'per_page'):
                        continue
                    if key not in _allowed_filter_fields:
                        continue
                    if value == 'true':
                        query = query.eq(key, True)
                    elif value == 'false':
                        query = query.eq(key, False)
                    else:
                        query = query.ilike(key, f'%{value}%') if key in partial_match_fields else query.eq(key, value)

                if paginated:
                    query = query.range(offset, offset + per_page - 1)

                return query.execute()

            response = None
            max_attempts = 3
            for attempt in range(1, max_attempts + 1):
                try:
                    response = execute_list_query_once()
                    break
                except Exception as exc:
                    if not is_transient_disconnect_error(exc) or attempt >= max_attempts:
                        raise

                    backoff_seconds = 0.12 * attempt
                    app.logger.warning(
                        'Retry list_resource %s após desconexão transitória (tentativa %s/%s).',
                        resource_name,
                        attempt + 1,
                        max_attempts,
                    )
                    time.sleep(backoff_seconds)

            rows = response.data or []
            total_count = response.count if paginated else len(rows)
            if resource_name == 'filiais':
                rows = [decorate_filial_row(item) for item in rows]
            if resource_name == 'colaboradores':
                rows = filter_visible_collaborators(rows)
                contract_names_by_collaborator = build_collaborator_contract_names(profile)
                for row in rows:
                    contract_names = contract_names_by_collaborator.get(int(row.get('id')), []) if row.get('id') is not None else []
                    row['contratos_vinculados'] = ', '.join(contract_names) if contract_names else '-'
            if resource_name == 'colaborador_documentos':
                rows = [enrich_collaborator_document(item) for item in rows]
            if resource_name == 'colaborador_beneficios':
                month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
                if not month_reference:
                    today = date_class.today()
                    month_reference = date_class(today.year, today.month, 1)
                raw_filial_id = request.args.get('filial_id')
                scoped_filial_id = int(raw_filial_id) if str(raw_filial_id).isdigit() else None
                rows = enrich_collaborator_benefit_rows(profile, rows, month_reference, scoped_filial_id)
            if resource_name == 'pedidos_compra':
                # Enriquece cada pedido com valor_total calculado pelos itens e colaborador criador
                pedido_ids = [int(row['id']) for row in rows if row.get('id') is not None]
                valor_por_pedido = {}
                colaboradores_por_id = {}
                if pedido_ids:
                    try:
                        itens_rows = (
                            supabase.table('pedidos_compra_itens')
                            .select('pedido_id, valor_total')
                            .in_('pedido_id', pedido_ids)
                            .eq('ativo', True)
                            .execute()
                            .data or []
                        )
                        for item in itens_rows:
                            pid = item.get('pedido_id')
                            if pid is not None:
                                valor_por_pedido[int(pid)] = valor_por_pedido.get(int(pid), 0.0) + parse_float_or_default(item.get('valor_total'), 0.0)
                    except Exception:
                        pass
                    criado_por_ids = sorted({
                        int(row['criado_por']) for row in rows
                        if row.get('criado_por') is not None
                    })
                    if criado_por_ids:
                        try:
                            colabs = (
                                supabase.table('colaboradores')
                                .select('id, nome_completo')
                                .in_('id', criado_por_ids)
                                .execute()
                                .data or []
                            )
                            colaboradores_por_id = {
                                int(c['id']): c.get('nome_completo') or ''
                                for c in colabs if c.get('id') is not None
                            }
                        except Exception:
                            pass
                filiais_por_id = {}
                filial_ids = sorted({
                    int(row['filial_id']) for row in rows
                    if row.get('filial_id') is not None
                })
                if filial_ids:
                    try:
                        filiais_rows = (
                            supabase.table('filiais')
                            .select('id, cidade, uf')
                            .in_('id', filial_ids)
                            .execute()
                            .data or []
                        )
                        filiais_por_id = {
                            int(f['id']): f'{f.get("cidade", "")}/{f.get("uf", "")}'
                            for f in filiais_rows if f.get('id') is not None
                        }
                    except Exception:
                        pass
                for row in rows:
                    pid = int(row['id']) if row.get('id') is not None else None
                    row['valor_total_calculado'] = round(valor_por_pedido.get(pid, 0.0), 2) if pid is not None else 0.0
                    criado_por_id = int(row['criado_por']) if row.get('criado_por') is not None else None
                    row['criado_por_nome'] = colaboradores_por_id.get(criado_por_id, '') if criado_por_id is not None else ''
                    filial_id_row = int(row['filial_id']) if row.get('filial_id') is not None else None
                    row['filial_nome'] = filiais_por_id.get(filial_id_row, '-') if filial_id_row is not None else '-'
            if resource_name == 'estoque_movimentos':
                # Enriquece com nome do item e nome do colaborador
                item_ids = sorted({int(row['item_id']) for row in rows if row.get('item_id') is not None})
                colab_ids = sorted({int(row['colaborador_id']) for row in rows if row.get('colaborador_id') is not None})
                itens_map = {}
                colabs_map = {}
                if item_ids:
                    try:
                        itens_rows = (
                            supabase.table('estoque_itens')
                            .select('id, nome, unidade')
                            .in_('id', item_ids)
                            .execute()
                            .data or []
                        )
                        itens_map = {int(r['id']): r for r in itens_rows if r.get('id') is not None}
                    except Exception:
                        pass
                if colab_ids:
                    try:
                        colabs_rows = (
                            supabase.table('colaboradores')
                            .select('id, nome_completo')
                            .in_('id', colab_ids)
                            .execute()
                            .data or []
                        )
                        colabs_map = {int(r['id']): r.get('nome_completo') or '' for r in colabs_rows if r.get('id') is not None}
                    except Exception:
                        pass
                for row in rows:
                    iid = int(row['item_id']) if row.get('item_id') is not None else None
                    cid = int(row['colaborador_id']) if row.get('colaborador_id') is not None else None
                    row['item_nome'] = itens_map.get(iid, {}).get('nome', '') if iid is not None else ''
                    row['item_unidade'] = itens_map.get(iid, {}).get('unidade', '') if iid is not None else ''
                    row['colaborador_nome'] = colabs_map.get(cid, '') if cid is not None else ''
            if resource_name == 'estoque_itens':
                # Adiciona flag de alerta de estoque baixo
                for row in rows:
                    atual = parse_float_or_default(row.get('estoque_atual'), 0.0)
                    minimo = parse_float_or_default(row.get('estoque_minimo'), 0.0)
                    row['alerta_estoque_baixo'] = minimo > 0 and atual <= minimo
            if paginated:
                total_pages = max(1, -(-total_count // per_page)) if total_count else 1
                return jsonify({
                    'data': rows,
                    'pagination': {
                        'page': page,
                        'per_page': per_page,
                        'total': total_count or 0,
                        'total_pages': total_pages,
                    },
                })
            return jsonify(rows)
        except Exception as exc:
            app.logger.error('Erro ao listar recurso %s: %s', resource_name, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/<resource_name>')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def create_resource(profile, resource_name):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        permission_error = require_create_permission(profile, resource_name)
        if permission_error:
            return permission_error

        raw_payload = request.get_json(silent=True)
        if not isinstance(raw_payload, dict):
            raw_payload = {}

        payload, payload_error = parse_payload(resource_name, partial=False)
        if payload_error:
            return jsonify({'error': payload_error}), 400

        if resource_name == 'colaboradores':
            try:
                auth_user = ensure_auth_user(raw_payload.get('email'), payload.get('nome_completo'))
            except ValueError as exc:
                return jsonify({'error': str(exc)}), 400
            except Exception as exc:
                app.logger.error('Erro ao garantir usuário no Supabase Auth: %s', exc)
                return jsonify({'error': f'Falha ao criar ou localizar usuário no Supabase Auth: {exc}'}), 400

            payload['user_id'] = str(auth_user.id)
            # ativo sempre derivado da data de desligamento
            payload['ativo'] = not bool(payload.get('data_desligamento'))

        if resource_name == 'veiculos':
            payload['created_by'] = profile['user_id']

        filial_scope_field = config.get('filial_scope_field')
        if filial_scope_field and not ensure_profile_can_access_filial(profile, payload.get(filial_scope_field)):
            return jsonify({'error': 'Sem permissão para cadastrar dados nesta base.'}), 403

        # Gera número de solicitação antes do INSERT para incluir diretamente no payload
        _SOLICITACAO_TIPO_MAP = {
            'manutencoes': 'manutencoes',
            'veiculos_abastecimentos': 'abastecimentos',
            'pedidos_compra': 'pedidos_compra',
            'horas_extras': 'horas_extras',
            'veiculos_pneus': 'pneus',
            'diarias_solicitacoes': 'diarias_solicitacoes',
        }
        if resource_name in _SOLICITACAO_TIPO_MAP and 'numero_solicitacao' not in payload:
            _tipo = _SOLICITACAO_TIPO_MAP[resource_name]
            _sb = supabase
            for _attempt in range(3):
                try:
                    _sol = _sb.rpc('gerar_numero_solicitacao', {'p_tipo': _tipo}).execute()
                    if _sol.data:
                        payload['numero_solicitacao'] = _sol.data
                    break
                except Exception as _sol_exc:
                    if is_transient_disconnect_error(_sol_exc) and _attempt < 2:
                        app.logger.warning(
                            'Retry gerar_numero_solicitacao %s (tentativa %s/3): %s',
                            resource_name, _attempt + 2, _sol_exc
                        )
                        time.sleep(0.15 * (_attempt + 1))
                        _sb = create_client(supabase_url, supabase_server_key)
                    else:
                        app.logger.warning(
                            'Falha ao gerar número de solicitação para %s: %s', resource_name, _sol_exc
                        )
                        break

        try:
            removed_columns = []
            if resource_name == 'colaborador_beneficios':
                response, removed_columns = execute_mutation_with_schema_fallback(
                    config['table'],
                    payload,
                    action='insert',
                )
            else:
                response = supabase.table(config['table']).insert(payload).execute()
            created = response.data[0] if response.data else {}

            # Invalida cache do dashboard quando muda algo que afeta cálculo de custo do contrato
            if resource_name in {
                'contratos_operacionais', 'contratos_colaboradores', 'contratos_gastos_extras',
                'colaboradores', 'colaborador_beneficios', 'colaborador_contratos',
            }:
                invalidate_dashboard_cache(created.get('filial_id'))

            write_audit_event(
                profile,
                action='create',
                resource_name=resource_name,
                entity_id=created.get('id'),
                details={
                    'payload_keys': sorted(list(payload.keys())),
                    'removed_columns': removed_columns,
                    'numero_solicitacao': created.get('numero_solicitacao'),
                },
                filial_id=created.get('filial_id'),
            )
            return jsonify(created), 201
        except Exception as exc:
            app.logger.error('Erro ao criar recurso %s: %s', resource_name, exc)
            write_audit_event(profile, 'create', resource_name, status='error', details={'error': str(exc)[:300]}, filial_id=payload.get('filial_id'))
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.patch('/api/<resource_name>/<int:item_id>')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def update_resource(profile, resource_name, item_id):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        scope_error = require_scope_permission(profile, config.get('view_scope'))
        if scope_error:
            return scope_error

        _, item_error = ensure_resource_item_allowed(profile, resource_name, item_id)
        if item_error:
            return item_error

        permission_error = require_edit_permission(profile)
        if permission_error:
            return permission_error

        payload, payload_error = parse_payload(resource_name, partial=True)
        if payload_error:
            return jsonify({'error': payload_error}), 400
        if not payload:
            return jsonify({'error': 'Nenhum campo válido informado para atualização.'}), 400

        if resource_name == 'colaboradores' and 'data_desligamento' in payload:
            payload['ativo'] = not bool(payload.get('data_desligamento'))

        filial_scope_field = config.get('filial_scope_field')
        if filial_scope_field and filial_scope_field in payload and not ensure_profile_can_access_filial(profile, payload.get(filial_scope_field)):
            return jsonify({'error': 'Sem permissão para mover ou alterar dados para uma base não autorizada.'}), 403

        try:
            removed_columns = []
            if resource_name == 'colaborador_beneficios':
                response, removed_columns = execute_mutation_with_schema_fallback(
                    config['table'],
                    payload,
                    action='update',
                    item_id=item_id,
                )
            else:
                response = supabase.table(config['table']).update(payload).eq('id', item_id).execute()
            updated = response.data[0] if response.data else {'id': item_id}

            # Cascata de status: reflete ativo/inativo do colaborador nos vínculos de contrato
            if resource_name == 'colaboradores' and ('data_desligamento' in payload or 'ativo' in payload):
                colaborador_ativo = payload.get('ativo', True)
                cascade_payload = {'ativo': colaborador_ativo}
                data_desligamento = payload.get('data_desligamento')
                if not colaborador_ativo and data_desligamento:
                    cascade_payload['data_fim_vinculo'] = data_desligamento

                # Defesa em profundidade: usuário com escopo de filial só cascateia
                # registros das filiais que ele pode acessar (mesmo que ensure_item
                # já tenha validado o colaborador alvo).
                allowed_filial_ids = (
                    None if not profile_has_filial_scope(profile)
                    else (profile.get('allowed_filial_ids') or [])
                )

                try:
                    q = supabase.table('contratos_colaboradores').update(cascade_payload).eq('colaborador_id', item_id)
                    if allowed_filial_ids is not None:
                        q = q.in_('filial_id', allowed_filial_ids)
                    q.execute()
                except Exception as cascade_exc:
                    app.logger.warning('Falha ao cascatear status para contratos_colaboradores (colaborador %s): %s', item_id, cascade_exc)

                # Quando o colaborador é desligado, encerra também:
                #  1) Vínculos trabalhistas (colaborador_contratos) ainda em curso
                #  2) Documentos RH ativos → status 'nao_se_aplica' (preserva o histórico)
                if not colaborador_ativo and data_desligamento:
                    # 1) Encerra fases de contrato ainda sem data_desligamento
                    try:
                        q = supabase.table('colaborador_contratos').update({
                            'data_desligamento': data_desligamento,
                            'motivo_desligamento': 'Desligamento registrado no cadastro do colaborador.',
                            'ativo': False,
                        }).eq('colaborador_id', item_id).is_('data_desligamento', 'null')
                        if allowed_filial_ids is not None:
                            q = q.in_('filial_id', allowed_filial_ids)
                        q.execute()
                    except Exception as cascade_exc:
                        app.logger.warning('Falha ao encerrar colaborador_contratos (colaborador %s): %s', item_id, cascade_exc)

                    # 2) Marca docs ativos como "não se aplica" (não inativa pra preservar histórico)
                    try:
                        q = supabase.table('colaborador_documentos').update({
                            'status': 'nao_se_aplica',
                        }).eq('colaborador_id', item_id).eq('ativo', True)
                        if allowed_filial_ids is not None:
                            q = q.in_('filial_id', allowed_filial_ids)
                        q.execute()
                    except Exception as cascade_exc:
                        app.logger.warning('Falha ao arquivar colaborador_documentos (colaborador %s): %s', item_id, cascade_exc)

            # Invalida cache do dashboard quando muda algo que afeta cálculo de custo do contrato
            if resource_name in {
                'contratos_operacionais', 'contratos_colaboradores', 'contratos_gastos_extras',
                'colaboradores', 'colaborador_beneficios', 'colaborador_contratos',
            }:
                invalidate_dashboard_cache(updated.get('filial_id') or payload.get('filial_id'))

            write_audit_event(
                profile,
                action='update',
                resource_name=resource_name,
                entity_id=item_id,
                details={
                    'payload_keys': sorted(list(payload.keys())),
                    'removed_columns': removed_columns,
                },
                filial_id=updated.get('filial_id'),
            )
            return jsonify(updated)
        except Exception as exc:
            app.logger.error('Erro ao atualizar recurso %s: %s', resource_name, exc)
            write_audit_event(profile, 'update', resource_name, item_id, status='error', details={'error': str(exc)[:300]}, filial_id=payload.get('filial_id'))
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.delete('/api/<resource_name>/<int:item_id>')
    @rate_limit_endpoint(max_requests=30)
    @require_auth
    def delete_resource(profile, resource_name, item_id):
        config = RESOURCE_DEFINITIONS.get(resource_name)
        if not config:
            return jsonify({'error': 'Recurso não encontrado.'}), 404

        scope_error = require_scope_permission(profile, config.get('view_scope'))
        if scope_error:
            return scope_error

        _, item_error = ensure_resource_item_allowed(profile, resource_name, item_id)
        if item_error:
            return item_error

        if resource_name == 'pedidos_compra':
            # Apenas o autor do pedido pode excluí-lo (não requer permissao_excluir)
            row = (supabase.table('pedidos_compra').select('criado_por').eq('id', item_id).limit(1).execute().data or [{}])[0]
            criado_por = row.get('criado_por')
            is_admin = profile.get('is_super_admin') or any(
                t in (profile.get('cargo') or '').lower() for t in ('admin', 'gestor', 'gerente', 'diretor')
            )
            if not is_admin and (criado_por is None or int(criado_por) != int(profile['id'])):
                return jsonify({'error': 'Apenas o autor do pedido pode excluí-lo.'}), 403
        else:
            permission_error = require_delete_permission(profile)
            if permission_error:
                return permission_error

        try:
            # Busca filial_id antes de deletar para auditoria
            item_response = supabase.table(config['table']).select('filial_id').eq('id', item_id).execute()
            item_filial_id = item_response.data[0].get('filial_id') if item_response.data else None

            supabase.table(config['table']).delete().eq('id', item_id).execute()

            if resource_name in {
                'contratos_operacionais', 'contratos_colaboradores', 'contratos_gastos_extras',
                'colaboradores', 'colaborador_beneficios', 'colaborador_contratos',
            }:
                invalidate_dashboard_cache(item_filial_id)

            write_audit_event(profile, 'delete', resource_name, item_id, filial_id=item_filial_id)
            return jsonify({'status': 'ok'})
        except Exception as exc:
            app.logger.error('Erro ao excluir recurso %s: %s', resource_name, exc)
            write_audit_event(profile, 'delete', resource_name, item_id, status='error', details={'error': str(exc)[:300]})
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.get('/api/cargos/modelos')
    @require_auth
    def cargos_modelos(profile):
        """Retorna todos os cargos com nome e permissoes_padrao para uso no template de permissões."""
        scope_error = require_scope_permission(profile, 'menu.cargos')
        if scope_error:
            # Qualquer usuário autenticado pode ler a lista de cargos para o dropdown
            # (só o menu.cargos bloqueia a tela de gestão, não o dropdown de nomes)
            pass
        try:
            rows = (
                supabase.table('cargos')
                .select('id, nome, descricao, permissoes_padrao, ordem, ativo')
                .eq('ativo', True)
                .order('nome')
                .execute()
                .data or []
            )
            return jsonify(rows)
        except Exception as exc:
            app.logger.error('Erro ao carregar modelos de cargo: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/estoque/resumo')
    @require_auth
    def estoque_resumo(profile):
        """Retorna todos os itens de estoque com saldo atual e flag de alerta por filial autorizada."""
        scope_error = require_scope_permission(profile, 'menu.estoque')
        if scope_error:
            return scope_error

        try:
            q = supabase.table('estoque_itens').select('*').eq('ativo', True).order('nome')
            filial_ids = [int(fid) for fid in (profile.get('filial_ids') or []) if str(fid).isdigit()]
            is_super = profile.get('is_super_admin', False)
            if not is_super and filial_ids:
                q = q.in_('filial_id', filial_ids)
            rows = q.execute().data or []
            for row in rows:
                atual = parse_float_or_default(row.get('estoque_atual'), 0.0)
                minimo = parse_float_or_default(row.get('estoque_minimo'), 0.0)
                row['alerta_estoque_baixo'] = minimo > 0 and atual <= minimo
            return jsonify(rows)
        except Exception as exc:
            app.logger.error('Erro ao carregar resumo de estoque: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/estoque_movimentos')
    @require_auth
    def criar_movimento_estoque(profile):
        """Cria um movimento de estoque e atualiza o saldo do item atomicamente."""
        scope_error = require_scope_permission(profile, 'menu.estoque')
        if scope_error:
            return scope_error

        permission_error = require_scope_permission(profile, 'create.estoque_movimentos')
        if permission_error:
            return permission_error

        raw_payload = request.get_json(silent=True) or {}

        ESTOQUE_TIPOS = {
            'entrada', 'saida_colaborador', 'saida_geral',
            'saida_fornecedor', 'troca', 'ajuste_positivo', 'ajuste_negativo',
        }
        TIPOS_QUE_SAEM = {'saida_colaborador', 'saida_geral', 'saida_fornecedor', 'troca', 'ajuste_negativo'}

        item_id = parse_int_or_default(raw_payload.get('item_id'), None)
        filial_id = parse_int_or_default(raw_payload.get('filial_id'), None)
        tipo = (raw_payload.get('tipo') or '').strip().lower()
        quantidade = max(0.001, parse_float_or_default(raw_payload.get('quantidade'), 0))
        colaborador_id = parse_int_or_default(raw_payload.get('colaborador_id'), None)
        registrado_por = parse_int_or_default(raw_payload.get('registrado_por'), None)
        fornecedor = (raw_payload.get('fornecedor') or '').strip() or None
        numero_nota = (raw_payload.get('numero_nota') or '').strip() or None
        motivo = (raw_payload.get('motivo') or '').strip() or None
        observacoes = (raw_payload.get('observacoes') or '').strip() or None
        data_movimento = (raw_payload.get('data_movimento') or '').strip() or None

        if not item_id:
            return jsonify({'error': 'Informe o item de estoque.'}), 400
        if not filial_id:
            return jsonify({'error': 'Informe a filial.'}), 400
        if not tipo or tipo not in ESTOQUE_TIPOS:
            return jsonify({'error': f'Tipo inválido. Use: {", ".join(sorted(ESTOQUE_TIPOS))}.'}), 400
        if quantidade <= 0:
            return jsonify({'error': 'Quantidade deve ser maior que zero.'}), 400
        if not data_movimento:
            data_movimento = date_class.today().isoformat()

        if not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para lançar movimentos nesta base.'}), 403

        try:
            item_row = (
                supabase.table('estoque_itens')
                .select('id, filial_id, estoque_atual, nome')
                .eq('id', item_id)
                .eq('ativo', True)
                .limit(1)
                .execute()
            ).data or []
            if not item_row:
                return jsonify({'error': 'Item de estoque não encontrado.'}), 404
            item = item_row[0]
            if not ensure_profile_can_access_filial(profile, item.get('filial_id')):
                return jsonify({'error': 'Sem permissão para o item desta base.'}), 403

            estoque_atual = parse_float_or_default(item.get('estoque_atual'), 0.0)
            if tipo in TIPOS_QUE_SAEM:
                novo_saldo = round(estoque_atual - quantidade, 3)
            else:
                novo_saldo = round(estoque_atual + quantidade, 3)

            movimento_payload = {
                'filial_id': filial_id,
                'item_id': item_id,
                'tipo': tipo,
                'quantidade': quantidade,
                'saldo_apos': novo_saldo,
                'data_movimento': data_movimento,
                'ativo': True,
            }
            if colaborador_id:
                movimento_payload['colaborador_id'] = colaborador_id
            if registrado_por:
                movimento_payload['registrado_por'] = registrado_por
            if fornecedor:
                movimento_payload['fornecedor'] = fornecedor
            if numero_nota:
                movimento_payload['numero_nota'] = numero_nota
            if motivo:
                movimento_payload['motivo'] = motivo
            if observacoes:
                movimento_payload['observacoes'] = observacoes

            created_mov = supabase.table('estoque_movimentos').insert(movimento_payload).execute()
            mov = created_mov.data[0] if created_mov.data else {}

            supabase.table('estoque_itens').update({'estoque_atual': novo_saldo}).eq('id', item_id).execute()

            write_audit_event(
                profile, 'create', 'estoque_movimentos', mov.get('id'),
                details={'tipo': tipo, 'quantidade': quantidade, 'item_id': item_id, 'saldo_apos': novo_saldo},
            )
            mov['item_nome'] = item.get('nome', '')
            return jsonify(mov), 201
        except Exception as exc:
            app.logger.error('Erro ao criar movimento de estoque: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.get('/api/estoque_movimentos/historico')
    @require_auth
    def estoque_historico(profile):
        """Retorna movimentos de um item específico com nome do colaborador."""
        scope_error = require_scope_permission(profile, 'menu.estoque')
        if scope_error:
            return scope_error

        item_id = parse_int_or_default(request.args.get('item_id'), None)
        filial_id = parse_int_or_default(request.args.get('filial_id'), None)
        limit = min(200, max(10, parse_int_or_default(request.args.get('limit'), 50)))

        try:
            q = supabase.table('estoque_movimentos').select('*').eq('ativo', True).order('data_movimento', desc=True).order('id', desc=True).limit(limit)
            if item_id:
                q = q.eq('item_id', item_id)
            if filial_id:
                if not ensure_profile_can_access_filial(profile, filial_id):
                    return jsonify({'error': 'Sem permissão para esta base.'}), 403
                q = q.eq('filial_id', filial_id)
            else:
                filial_ids = [int(fid) for fid in (profile.get('filial_ids') or []) if str(fid).isdigit()]
                is_super = profile.get('is_super_admin', False)
                if not is_super and filial_ids:
                    q = q.in_('filial_id', filial_ids)

            rows = q.execute().data or []

            item_ids_needed = sorted({int(r['item_id']) for r in rows if r.get('item_id') is not None})
            colab_ids_needed = sorted({int(r['colaborador_id']) for r in rows if r.get('colaborador_id') is not None})
            itens_map = {}
            colabs_map = {}
            if item_ids_needed:
                try:
                    itens_rows = supabase.table('estoque_itens').select('id, nome, unidade').in_('id', item_ids_needed).execute().data or []
                    itens_map = {int(r['id']): r for r in itens_rows if r.get('id') is not None}
                except Exception:
                    pass
            if colab_ids_needed:
                try:
                    colabs_rows = supabase.table('colaboradores').select('id, nome_completo').in_('id', colab_ids_needed).execute().data or []
                    colabs_map = {int(r['id']): r.get('nome_completo') or '' for r in colabs_rows if r.get('id') is not None}
                except Exception:
                    pass

            for row in rows:
                iid = int(row['item_id']) if row.get('item_id') is not None else None
                cid = int(row['colaborador_id']) if row.get('colaborador_id') is not None else None
                row['item_nome'] = itens_map.get(iid, {}).get('nome', '') if iid is not None else ''
                row['item_unidade'] = itens_map.get(iid, {}).get('unidade', '') if iid is not None else ''
                row['colaborador_nome'] = colabs_map.get(cid, '') if cid is not None else ''

            return jsonify(rows)
        except Exception as exc:
            app.logger.error('Erro ao carregar histórico de estoque: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/pedidos_compra/pre-alocar-numero')
    @require_auth
    def pre_alocar_numero_pedido(profile):
        if not profile_has_scope_permission(profile, 'create.pedidos_compra'):
            return jsonify({'error': 'Sem permissão.'}), 403
        # Try RPC first
        try:
            sol = supabase.rpc('gerar_numero_solicitacao', {'p_tipo': 'pedidos_compra'}).execute()
            if sol.data:
                return jsonify({'numero_solicitacao': sol.data})
        except Exception as exc:
            app.logger.warning('RPC gerar_numero_solicitacao falhou (pre-alocacao): %s', exc)
        # Fallback: derive from max existing number
        try:
            rows = supabase.table('pedidos_compra').select('numero_solicitacao').execute().data or []
            max_num = 0
            for row in rows:
                ns = row.get('numero_solicitacao') or ''
                m = re.match(r'^P(\d+)$', ns)
                if m:
                    max_num = max(max_num, int(m.group(1)))
            return jsonify({'numero_solicitacao': f'P{max_num + 1:05d}'})
        except Exception as exc2:
            app.logger.error('Fallback numero_solicitacao falhou: %s', exc2)
            return jsonify({'error': 'Falha ao gerar número de solicitação.'}), 500

    @app.get('/api/pedidos_compra/<int:pedido_id>/detalhes')
    @require_auth
    def pedido_compra_detalhes(profile, pedido_id):
        """Retorna pedido + itens em JSON estruturado para geração de PDF no frontend."""
        scope_error = require_scope_permission(profile, 'menu.pedidos_compra')
        if scope_error:
            return scope_error

        try:
            pedido_response = (
                supabase.table('pedidos_compra')
                .select('*')
                .eq('id', pedido_id)
                .limit(1)
                .execute()
            )
            if not pedido_response.data:
                return jsonify({'error': 'Pedido não encontrado.'}), 404

            pedido = pedido_response.data[0]
            pedido_filial_id = pedido.get('filial_id')
            if not ensure_profile_can_access_filial(profile, pedido_filial_id):
                return jsonify({'error': 'Sem permissão para acessar este pedido.'}), 403

            # Busca filial
            filial_data = {}
            if pedido_filial_id:
                try:
                    filial_row = (
                        supabase.table('filiais')
                        .select('*')
                        .eq('id', pedido_filial_id)
                        .limit(1)
                        .execute()
                    ).data or []
                    if filial_row:
                        filial_data = decorate_filial_row(filial_row[0])
                except Exception:
                    pass

            # Busca colaborador criador
            criado_por_nome = ''
            criado_por_cargo = ''
            if pedido.get('criado_por'):
                try:
                    colab_row = (
                        supabase.table('colaboradores')
                        .select('nome_completo, cargo')
                        .eq('id', pedido['criado_por'])
                        .limit(1)
                        .execute()
                    ).data or []
                    if colab_row:
                        criado_por_nome = colab_row[0].get('nome_completo') or ''
                        criado_por_cargo = colab_row[0].get('cargo') or ''
                except Exception:
                    pass

            # Resolve nomes dos aprovadores/analisadores
            actor_ids = {
                pedido.get(f) for f in ('aprovado_por', 'em_analise_por', 'reprovado_por')
                if pedido.get(f) and pedido.get(f) != pedido.get('criado_por')
            }
            actors_map = {}
            if actor_ids:
                try:
                    rows = supabase.table('colaboradores').select('id, nome_completo').in_('id', list(actor_ids)).execute().data or []
                    for r in rows:
                        actors_map[r['id']] = r.get('nome_completo') or ''
                except Exception:
                    pass
            if pedido.get('criado_por') and criado_por_nome:
                actors_map[pedido['criado_por']] = criado_por_nome

            # Busca itens
            itens = (
                supabase.table('pedidos_compra_itens')
                .select('*')
                .eq('pedido_id', pedido_id)
                .eq('ativo', True)
                .order('id')
                .execute()
                .data or []
            )

            valor_total = round(sum(parse_float_or_default(i.get('valor_total'), 0.0) for i in itens), 2)

            return jsonify({
                'pedido': pedido,
                'filial': filial_data,
                'criado_por_nome': criado_por_nome,
                'criado_por_cargo': criado_por_cargo,
                'aprovado_por_nome': actors_map.get(pedido.get('aprovado_por'), ''),
                'em_analise_por_nome': actors_map.get(pedido.get('em_analise_por'), ''),
                'reprovado_por_nome': actors_map.get(pedido.get('reprovado_por'), ''),
                'itens': itens,
                'valor_total': valor_total,
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar detalhes do pedido %s: %s', pedido_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.patch('/api/pedidos_compra/<int:pedido_id>/status')
    @require_auth
    def update_pedido_status(profile, pedido_id):
        """Endpoint dedicado para mudança de status do pedido."""
        scope_error = require_scope_permission(profile, 'menu.pedidos_compra')
        if scope_error:
            return scope_error

        PEDIDO_STATUS_OPTIONS = {
            'rascunho', 'pendente', 'analise',
            'pendente_aprovacao', 'em_analise',   # legado
            'aprovado', 'reprovado', 'em_compra', 'recebido', 'cancelado',
        }
        body = request.get_json(silent=True) or {}
        novo_status = (body.get('status') or '').strip().lower()
        if not novo_status or novo_status not in PEDIDO_STATUS_OPTIONS:
            return jsonify({'error': f'Status inválido: {novo_status}. Use: {", ".join(sorted(PEDIDO_STATUS_OPTIONS))}.'}), 400

        try:
            pedido_row = (
                supabase.table('pedidos_compra')
                .select('id, filial_id')
                .eq('id', pedido_id)
                .limit(1)
                .execute()
            ).data or []
            if not pedido_row:
                return jsonify({'error': 'Pedido não encontrado.'}), 404
            if not ensure_profile_can_access_filial(profile, pedido_row[0].get('filial_id')):
                return jsonify({'error': 'Sem permissão para esta base.'}), 403

            supabase.table('pedidos_compra').update({'status': novo_status}).eq('id', pedido_id).execute()
            write_audit_event(profile, 'update', 'pedidos_compra', pedido_id, details={'status': novo_status})
            return jsonify({'status': 'ok', 'novo_status': novo_status})
        except Exception as exc:
            app.logger.error('Erro ao atualizar status do pedido %s: %s', pedido_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/permissoes/config')
    @require_auth
    def permissions_config(profile):
        scope_error = require_scope_permission(profile, 'menu.permissoes')
        if scope_error:
            return scope_error

        try:
            collaborators = (
                supabase.table('colaboradores')
                .select('id, nome_completo, cargo, filial_id, ativo')
                .order('nome_completo')
                .execute()
            )
            filiais = supabase.table('filiais').select('id, cidade, uf').order('cidade').execute()

            return jsonify({
                'scope_groups': PERMISSION_SCOPE_GROUPS,
                'collaborators': filter_visible_collaborators(collaborators.data or []),
                'filiais': filiais.data or [],
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar configuração de permissões: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/permissoes/por-escopo')
    @require_auth
    def permissoes_por_escopo(profile):
        """Retorna IDs dos colaboradores que possuem um determinado escopo."""
        scope_error = require_scope_permission(profile, 'menu.permissoes')
        if scope_error:
            return scope_error
        scope_name = request.args.get('scope', '').strip()
        if not scope_name:
            return jsonify({'error': 'Parâmetro scope é obrigatório.'}), 400
        try:
            rows = (
                supabase.table('permissoes')
                .select('colaborador_id')
                .eq('permissao_nome', scope_name)
                .eq('ativo', True)
                .execute()
                .data or []
            )
            collab_ids = [int(r['colaborador_id']) for r in rows if r.get('colaborador_id')]
            return jsonify({'scope': scope_name, 'collab_ids': collab_ids})
        except Exception as exc:
            app.logger.error('permissoes_por_escopo: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/colaboradores/com-escopo')
    @require_auth
    def colaboradores_com_escopo(profile):
        """Retorna colaboradores ativos que possuem um determinado escopo.
        Acessível a qualquer usuário autenticado com menu.pedidos_compra (ou admin)."""
        scope_error = require_scope_permission(profile, 'menu.pedidos_compra')
        if scope_error:
            return scope_error
        scope_name = request.args.get('scope', '').strip()
        if not scope_name:
            return jsonify({'error': 'Parâmetro scope é obrigatório.'}), 400
        try:
            perm_rows = (
                supabase.table('permissoes')
                .select('colaborador_id')
                .eq('permissao_nome', scope_name)
                .eq('ativo', True)
                .execute()
                .data or []
            )
            collab_ids = [int(r['colaborador_id']) for r in perm_rows if r.get('colaborador_id')]
            if not collab_ids:
                return jsonify([])
            collab_rows = (
                supabase.table('colaboradores')
                .select('id, nome_completo, cargo')
                .in_('id', collab_ids)
                .eq('ativo', True)
                .order('nome_completo')
                .execute()
                .data or []
            )
            return jsonify(collab_rows)
        except Exception as exc:
            app.logger.error('colaboradores_com_escopo: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/filiais/disponiveis')
    @require_auth
    def filiais_disponiveis(profile):
        """Retorna as filiais acessíveis ao usuário logado, sem exigir menu.filiais.
        Usado em formulários (pedidos de compra, etc.) onde o usuário precisa selecionar a filial."""
        try:
            filiais = fetch_accessible_filiais(profile)
            return jsonify([{'id': f['id'], 'cidade': f['cidade'], 'uf': f.get('uf', '')} for f in filiais])
        except Exception as exc:
            app.logger.error('filiais_disponiveis: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/permissoes/toggle-escopo')
    @require_auth
    def toggle_escopo(profile):
        """Ativa ou desativa um único escopo para um colaborador."""
        scope_error = require_scope_permission(profile, 'menu.permissoes')
        if scope_error:
            return scope_error
        body = request.get_json(silent=True) or {}
        colaborador_id = body.get('colaborador_id')
        scope_name = (body.get('scope_name') or '').strip()
        ativo = bool(body.get('ativo', True))
        if not colaborador_id or not scope_name:
            return jsonify({'error': 'colaborador_id e scope_name são obrigatórios.'}), 400
        if scope_name not in PERMISSION_SCOPE_MAP:
            return jsonify({'error': f'Escopo inválido: {scope_name}'}), 400
        try:
            if ativo:
                existing = (
                    supabase.table('permissoes')
                    .select('id')
                    .eq('colaborador_id', colaborador_id)
                    .eq('permissao_nome', scope_name)
                    .execute()
                    .data or []
                )
                if not existing:
                    supabase.table('permissoes').insert({
                        'colaborador_id': colaborador_id,
                        'permissao_nome': scope_name,
                        'ativo': True,
                        'descricao': PERMISSION_SCOPE_MAP[scope_name].get('description', ''),
                    }).execute()
            else:
                supabase.table('permissoes').delete().eq('colaborador_id', colaborador_id).eq('permissao_nome', scope_name).execute()
            write_audit_event(profile, 'update', 'permissoes', entity_id=colaborador_id,
                              details={'scope': scope_name, 'ativo': ativo})
            return jsonify({'ok': True, 'colaborador_id': colaborador_id, 'scope_name': scope_name, 'ativo': ativo})
        except Exception as exc:
            app.logger.error('toggle_escopo: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/permissoes/<int:collaborator_id>')
    @require_auth
    def permissions_detail(profile, collaborator_id):
        scope_error = require_scope_permission(profile, 'menu.permissoes')
        if scope_error:
            return scope_error

        try:
            collaborator_response = (
                supabase.table('colaboradores')
                .select('*')
                .eq('id', collaborator_id)
                .limit(1)
                .execute()
            )
            if not collaborator_response.data:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404

            collaborator = collaborator_response.data[0]
            permissions = fetch_permissions(collaborator_id)
            active_scopes = extract_known_permission_scopes(permissions)

            return jsonify({
                'collaborator': collaborator,
                'permission_flags': {
                    field: bool(collaborator.get(field))
                    for field in MANAGED_COLLABORATOR_PERMISSION_FIELDS
                },
                'active_scopes': active_scopes,
                'active_filial_ids': extract_allowed_filial_ids(permissions),
                'detailed_permissions': permissions,
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar permissões do colaborador %s: %s', collaborator_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.put('/api/permissoes/<int:collaborator_id>')
    @rate_limit_endpoint(max_requests=20)
    @require_auth
    def permissions_update(profile, collaborator_id):
        scope_error = require_scope_permission(profile, 'menu.permissoes', 'Sem permissão para alterar permissões.')
        if scope_error:
            return scope_error

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        raw_flags = payload.get('permission_flags') or {}
        if not isinstance(raw_flags, dict):
            raw_flags = {}

        collaborator_update = {
            field: bool(raw_flags.get(field))
            for field in MANAGED_COLLABORATOR_PERMISSION_FIELDS
            if field in raw_flags
        }

        requested_scopes = payload.get('active_scopes') or []
        if not isinstance(requested_scopes, list):
            return jsonify({'error': 'A lista de escopos informada é inválida.'}), 400

        requested_filial_ids = payload.get('active_filial_ids') or []
        if not isinstance(requested_filial_ids, list):
            return jsonify({'error': 'A lista de filiais informada é inválida.'}), 400

        normalized_scopes = sorted({
            scope_name
            for scope_name in requested_scopes
            if scope_name in PERMISSION_SCOPE_MAP
        })

        if len(normalized_scopes) != len(requested_scopes):
            return jsonify({'error': 'Um ou mais escopos informados são inválidos.'}), 400

        normalized_filial_ids = sorted({
            int(filial_id)
            for filial_id in requested_filial_ids
            if str(filial_id).isdigit()
        })

        if len(normalized_filial_ids) != len(requested_filial_ids):
            return jsonify({'error': 'Uma ou mais filiais informadas são inválidas.'}), 400

        try:
            collaborator_response = (
                supabase.table('colaboradores')
                .select('id')
                .eq('id', collaborator_id)
                .limit(1)
                .execute()
            )
            if not collaborator_response.data:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404

            if collaborator_update:
                supabase.table('colaboradores').update(collaborator_update).eq('id', collaborator_id).execute()

            if ALL_PERMISSION_SCOPES:
                (
                    supabase.table('permissoes')
                    .delete()
                    .eq('colaborador_id', collaborator_id)
                    .in_('permissao_nome', ALL_PERMISSION_SCOPES)
                    .execute()
                )

            (
                supabase.table('permissoes')
                .delete()
                .eq('colaborador_id', collaborator_id)
                .like('permissao_nome', 'filial.%')
                .execute()
            )

            if normalized_scopes:
                permission_rows = [
                    {
                        'colaborador_id': collaborator_id,
                        'permissao_nome': scope_name,
                        'ativo': True,
                        'descricao': PERMISSION_SCOPE_MAP[scope_name]['description'],
                    }
                    for scope_name in normalized_scopes
                ]
                supabase.table('permissoes').insert(permission_rows).execute()

            if normalized_filial_ids:
                filial_permission_rows = [
                    {
                        'colaborador_id': collaborator_id,
                        'permissao_nome': f'filial.{filial_id}',
                        'ativo': True,
                        'descricao': f'Acesso liberado para a filial {filial_id}.',
                    }
                    for filial_id in normalized_filial_ids
                ]
                supabase.table('permissoes').insert(filial_permission_rows).execute()

            write_audit_event(
                profile,
                action='permissions_update',
                resource_name='permissoes',
                entity_id=collaborator_id,
                details={
                    'scopes_count': len(normalized_scopes),
                    'filiais_count': len(normalized_filial_ids),
                },
            )

            return jsonify({'status': 'ok'})
        except Exception as exc:
            app.logger.error('Erro ao salvar permissões do colaborador %s: %s', collaborator_id, exc)
            write_audit_event(profile, 'permissions_update', 'permissoes', collaborator_id, status='error', details={'error': str(exc)[:300]})
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.get('/api/presenca-config')
    @require_auth
    def presence_config(profile):
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        try:
            filiais_query = supabase.table('filiais').select('id, cidade, uf').order('cidade')
            if profile_has_filial_scope(profile):
                filiais_query = filiais_query.in_('id', profile.get('allowed_filial_ids') or [])
            filiais = filiais_query.execute()
            return jsonify({
                'database_ready': attendance_table_ready(),
                'today': date_class.today().isoformat(),
                'status_options': PRESENCE_STATUS_OPTIONS,
                'filiais': filiais.data or [],
                'can_manage': profile_has_scope_permission(profile, 'manage.presenca'),
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar configuração de presença: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca')
    @require_auth
    def presence_list(profile):
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        target_date = parse_iso_date(request.args.get('data')) or date_class.today()
        filial_id = request.args.get('filial_id', type=int)

        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            database_ready = attendance_table_ready()
            stored_rows = []
            rh_events = []

            if database_ready:
                presence_query = (
                    supabase.table('presencas_diarias')
                    .select('*')
                    .eq('data_referencia', target_date.isoformat())
                )
                if filial_id:
                    presence_query = presence_query.eq('filial_id', filial_id)
                stored_rows = presence_query.execute().data or []

            rh_filial_ids = [filial_id] if filial_id else [int(item['id']) for item in fetch_accessible_filiais(profile)]
            rh_events = fetch_active_rh_events_for_date(target_date, rh_filial_ids)
            ensured_collaborator_ids = [
                row.get('colaborador_id')
                for row in stored_rows
                if row.get('colaborador_id') is not None
            ] + [
                event.get('colaborador_id')
                for event in rh_events
                if event.get('colaborador_id') is not None
            ]
            collaborators = list_active_collaborators_for_presence(filial_id, target_date, ensured_collaborator_ids)

            return jsonify({
                'database_ready': database_ready,
                'rh_database_ready': rh_events_table_ready(),
                'date': target_date.isoformat(),
                'items': build_presence_items(target_date, collaborators, stored_rows, rh_events),
            })
        except Exception as exc:
            app.logger.error('Erro ao listar presença diária: %s', exc)
            if is_missing_relation_error(exc):
                collaborators = list_active_collaborators_for_presence(filial_id, target_date)
                return jsonify({
                    'database_ready': False,
                    'rh_database_ready': rh_events_table_ready(),
                    'date': target_date.isoformat(),
                    'items': build_presence_items(target_date, collaborators),
                })
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.put('/api/presenca')
    @rate_limit_endpoint(max_requests=30)
    @require_auth
    def presence_save(profile):
        scope_error = require_scope_permission(profile, 'manage.presenca', 'Sem permissão para alterar o controle de presença.')
        if scope_error:
            return scope_error

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        target_date = parse_iso_date(payload.get('data'))
        if not target_date:
            return jsonify({'error': 'Informe uma data válida para o controle de presença.'}), 400

        entries = payload.get('entries') or []
        if not isinstance(entries, list):
            return jsonify({'error': 'A lista de presenças enviada é inválida.'}), 400

        normalized_entries = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue

            status = (entry.get('status') or 'pendente').strip().lower()
            if status not in PRESENCE_STATUS_OPTIONS:
                return jsonify({'error': f'Status de presença inválido: {status}.'}), 400

            colaborador_id = entry.get('colaborador_id')
            filial_id = entry.get('filial_id')
            if not colaborador_id or not filial_id:
                return jsonify({'error': 'Cada linha de presença precisa de colaborador e filial.'}), 400

            normalized_entries.append({
                'data_referencia': target_date.isoformat(),
                'colaborador_id': colaborador_id,
                'filial_id': filial_id,
                'status': status,
                'observacoes': entry.get('observacoes') or None,
                'origem': entry.get('origem') or 'web',
                'alterado_por': profile['user_id'],
            })

        try:
            if not attendance_table_ready():
                return jsonify({'error': 'A tabela de presença ainda não existe. Rode a migration do módulo primeiro.'}), 400

            if normalized_entries:
                (
                    supabase.table('presencas_diarias')
                    .upsert(normalized_entries, on_conflict='data_referencia,colaborador_id')
                    .execute()
                )

            write_audit_event(
                profile,
                action='presence_save',
                resource_name='presencas_diarias',
                details={'date': target_date.isoformat(), 'entries': len(normalized_entries)},
            )

            return jsonify({'status': 'ok'})
        except Exception as exc:
            app.logger.error('Erro ao salvar presença diária: %s', exc)
            write_audit_event(profile, 'presence_save', 'presencas_diarias', status='error', details={'error': str(exc)[:300]})
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.get('/api/presenca-mes')
    @require_auth
    def presence_month(profile):
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        raw_filial_id = request.args.get('filial_id')
        filial_id = int(raw_filial_id) if str(raw_filial_id).isdigit() else None
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False, 'month_reference': month_reference.isoformat(), 'items': []})

            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            query = (
                supabase.table('presencas_diarias')
                .select('*')
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
            )
            if filial_id:
                query = query.eq('filial_id', filial_id)
            rows = query.execute().data or []

            # Buscar info dos colaboradores envolvidos para enriquecer as linhas
            collaborator_ids = sorted({int(r['colaborador_id']) for r in rows if r.get('colaborador_id') is not None})
            collaborators = []
            if collaborator_ids:
                coll_query = (
                    supabase.table('colaboradores')
                    .select('id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim')
                    .in_('id', collaborator_ids)
                )
                collaborators = coll_query.execute().data or []

            collaborators_by_id = {int(c['id']): c for c in collaborators}

            items = []
            for row in rows:
                cid = row.get('colaborador_id')
                coll = collaborators_by_id.get(int(cid)) if cid is not None else None
                items.append({
                    'id': row.get('id'),
                    'data_referencia': row.get('data_referencia'),
                    'colaborador_id': cid,
                    'filial_id': row.get('filial_id'),
                    'nome_completo': coll.get('nome_completo') if coll else None,
                    'cargo': coll.get('cargo') if coll else None,
                    'turno': coll.get('turno') if coll else None,
                    'escala_servico': coll.get('escala_servico') if coll else None,
                    'horario_padrao_inicio': coll.get('horario_padrao_inicio') if coll else None,
                    'horario_padrao_fim': coll.get('horario_padrao_fim') if coll else None,
                    'status': row.get('status'),
                    'observacoes': row.get('observacoes'),
                    'origem': row.get('origem'),
                })

            return jsonify({'database_ready': True, 'month_reference': month_reference.isoformat(), 'items': items})
        except Exception as exc:
            app.logger.error('Erro ao listar presenças mensais: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca-mes-xlsx')
    @require_auth
    def presence_month_xlsx(profile):
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        raw_filial_id = request.args.get('filial_id')
        filial_id = int(raw_filial_id) if str(raw_filial_id).isdigit() else None
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False, 'month_reference': month_reference.isoformat(), 'items': []})

            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            query = (
                supabase.table('presencas_diarias')
                .select('*')
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
            )
            if filial_id:
                query = query.eq('filial_id', filial_id)
            rows = query.execute().data or []

            collaborator_ids = sorted({int(r['colaborador_id']) for r in rows if r.get('colaborador_id') is not None})
            collaborators = []
            if collaborator_ids:
                coll_query = (
                    supabase.table('colaboradores')
                    .select('id, filial_id, nome_completo, cargo, turno, escala_servico, horario_padrao_inicio, horario_padrao_fim, pis')
                    .in_('id', collaborator_ids)
                )
                collaborators = coll_query.execute().data or []

            collaborators_by_id = {int(c['id']): c for c in collaborators}

            # Montar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Presenças'

            bold = Font(bold=True)
            center = Alignment(horizontal='center', vertical='center')
            left = Alignment(horizontal='left', vertical='center')
            header_fill = PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid')
            thin = Side(border_style='thin', color='FF000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Employer info from ENV or defaults
            employer = os.getenv('EMPREGADOR_NOME', 'GOLD TRANSPORTES')
            employer_cnpj = os.getenv('EMPREGADOR_CNPJ', '')

            # Header rows (merged for nicer layout)
            period_label = f"PERÍODO: {start_date.isoformat()} A {end_date.isoformat()}"
            ws.merge_cells('A1:J1')
            ws['A1'] = period_label
            ws['A1'].font = bold
            ws['A1'].alignment = left

            ws.merge_cells('A2:B2')
            ws['A2'] = f'Empregador: {employer}'
            ws['A2'].font = bold
            ws['A2'].alignment = left
            ws.merge_cells('C2:J2')
            ws['C2'] = f'CNPJ: {employer_cnpj}'
            ws['C2'].alignment = left

            # Spacer
            row_idx = 4

            # For each collaborator, create a block with summary and table
            for coll_id in collaborator_ids:
                coll = collaborators_by_id.get(int(coll_id), {})
                nome = coll.get('nome_completo') or ''
                pis = coll.get('pis') or ''
                setor = ''
                cargo = coll.get('cargo') or ''
                escala = coll.get('escala_servico') or ''

                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
                ws.cell(row=row_idx, column=1, value=f'Empregado: {nome}    Pis: {pis}').font = bold
                row_idx += 1
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
                ws.cell(row=row_idx, column=1, value=f'Setor: {setor}    Cargo: {cargo}    Escala: {escala}').alignment = left
                row_idx += 1

                # Totals placeholder row (user can calculate client-side if needed)
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=10)
                ws.cell(row=row_idx, column=1, value='').alignment = left
                row_idx += 1

                # Table header
                headers = ['DATA','OBSERVAÇÕES','ENTRADA','SAÍDA','ENTRADA','SAÍDA','AD NOTURNO','HORAS TRABALHADAS','HORAS ESCALA','BANCO DE HORAS']
                for col, h in enumerate(headers, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=h)
                    cell.font = bold
                    cell.fill = header_fill
                    cell.alignment = center
                    cell.border = border
                row_idx += 1

                # Rows per day for this collaborator
                # Filter rows for this collaborator and sort by date
                rows_for_coll = [r for r in rows if r.get('colaborador_id') == coll_id]
                rows_for_coll_sorted = sorted(rows_for_coll, key=lambda r: r.get('data_referencia') or '')

                def compute_escala(hstart, hend):
                    if not hstart or not hend:
                        return ''
                    try:
                        t1 = datetime.strptime(hstart, '%H:%M')
                        t2 = datetime.strptime(hend, '%H:%M')
                        diff = (t2 - t1).seconds // 60
                        if diff <= 0:
                            diff += 24*60
                        hours = diff // 60
                        minutes = diff % 60
                        return f"{hours:02d}:{minutes:02d}"
                    except Exception:
                        return ''

                for r in rows_for_coll_sorted:
                    data_ref = r.get('data_referencia')
                    obs = r.get('observacoes') or ''
                    # We don't have entrada/saida no presente payload, leave blank
                    entrada1 = ''
                    saida1 = ''
                    entrada2 = ''
                    saida2 = ''
                    ad_noturno = ''
                    horas_trabalhadas = ''
                    horas_escala = compute_escala(coll.get('horario_padrao_inicio'), coll.get('horario_padrao_fim'))
                    banco = ''

                    row_cells = [data_ref, obs, entrada1, saida1, entrada2, saida2, ad_noturno, horas_trabalhadas, horas_escala, banco]
                    for col, val in enumerate(row_cells, start=1):
                        cell = ws.cell(row=row_idx, column=col, value=val)
                        cell.alignment = left
                        cell.border = border
                    row_idx += 1

                # blank row after collaborator
                row_idx += 1

            # Ajustar largura de colunas básicas
            widths = [12, 40, 10, 10, 10, 10, 10, 12, 12, 12]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

            # Escrever workbook em memória
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            response_stream = make_response(bio.read())
            filename = f'presencas_{month_reference.year}-{month_reference.month:02d}.xlsx'
            response_stream.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response_stream.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
            return response_stream
        except Exception as exc:
            app.logger.error('Erro ao gerar XLSX de presenças mensais: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca-colaborador-calendario')
    @require_auth
    def presence_collaborator_calendar(profile):
        """Gera calendário visual estilo parede para presença de colaborador"""
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        raw_colaborador_id = request.args.get('colaborador_id')
        if not str(raw_colaborador_id).isdigit():
            return jsonify({'error': 'Informe um ID de colaborador válido.'}), 400
        
        colaborador_id = int(raw_colaborador_id)

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False})

            # Buscar colaborador
            coll_response = (
                supabase.table('colaboradores')
                .select('id, filial_id, nome_completo, cargo, turno, horario_padrao_inicio, horario_padrao_fim')
                .eq('id', colaborador_id)
                .execute()
            )
            if not coll_response.data:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404
            
            colaborador = coll_response.data[0]
            filial_id = colaborador.get('filial_id')
            
            # Verificar acesso à filial
            if not ensure_profile_can_access_filial(profile, filial_id):
                return jsonify({'error': 'Sem permissão para consultar este colaborador.'}), 403

            # Calcular período do mês
            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            # Buscar presenças do colaborador no mês
            presence_response = (
                supabase.table('presencas_diarias')
                .select('*')
                .eq('colaborador_id', colaborador_id)
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
                .execute()
            )
            presences = {r['data_referencia']: r for r in (presence_response.data or [])}

            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Calendário'

            # Estilos
            bold = Font(bold=True, color='FFFFFF', size=12)
            bold_dark = Font(bold=True, size=10, color='000000')
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Cores por status
            color_presente = 'FF27AE60'  # Verde
            color_falta = 'FFE74C3C'      # Vermelho
            color_atraso = 'FFF39C12'     # Laranja
            color_folga = 'FF95A5A6'      # Cinza
            color_fim_semana = 'FF34495E' # Azul escuro
            color_header = 'FF2C3E50'     # Azul muito escuro
            
            thin = Side(border_style='thin', color='FF000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Cabeçalho: Nome + Mês
            ws.merge_cells('A1:G1')
            header_cell = ws['A1']
            nome = colaborador.get('nome_completo', '')
            mes_text = month_reference.strftime('%B/%Y').upper()
            header_cell.value = f'{nome} - {mes_text}'
            header_cell.font = Font(bold=True, size=16, color='FFFFFF')
            header_cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
            header_cell.alignment = center
            ws.row_dimensions[1].height = 30

            # Linha de dias da semana
            day_names = ['Do', 'Se', 'Te', 'Qu', 'Qu', 'Se', 'Sa']
            for col, day_name in enumerate(day_names, start=1):
                cell = ws.cell(row=3, column=col, value=day_name)
                cell.font = Font(bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                cell.alignment = center
                cell.border = border
                ws.row_dimensions[3].height = 20

            # Montar grid do calendário
            # Encontrar primeiro dia da semana do mês
            first_day_weekday = start_date.weekday()  # 0=seg, 6=dom
            # Ajustar para começar domingo (0=dom, 1=seg, ...)
            first_day_weekday = (first_day_weekday + 1) % 7
            
            current_row = 4
            current_col = first_day_weekday + 1
            
            # Preencher dias do mês
            current_date = start_date
            while current_date <= end_date:
                date_str = current_date.isoformat()
                presence = presences.get(date_str)
                
                # Determinar status e cor
                is_weekend = current_date.weekday() >= 5
                
                if is_weekend:
                    status_color = color_fim_semana
                    status_symbol = '⚫'
                elif not presence:
                    status_color = color_falta
                    status_symbol = '❌'
                elif presence.get('status') == 'atraso':
                    status_color = color_atraso
                    status_symbol = '🟠'
                elif presence.get('status') == 'folga':
                    status_color = color_folga
                    status_symbol = '🔴'
                else:
                    status_color = color_presente
                    status_symbol = '🟢'
                
                # Célula do dia
                cell = ws.cell(row=current_row, column=current_col)
                cell.value = current_date.day
                cell.font = Font(bold=True, size=18, color='FFFFFF')
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                cell.alignment = center
                cell.border = border
                ws.row_dimensions[current_row].height = 40
                ws.column_dimensions[chr(64 + current_col)].width = 18

                # Observação/duração (linha abaixo)
                obs_row = current_row + 1
                cell_obs = ws.cell(row=obs_row, column=current_col)
                
                if presence and presence.get('observacoes'):
                    cell_obs.value = presence.get('observacoes', '-')
                elif presence and presence.get('status') == 'atraso':
                    # Mostrar duração do atraso se disponível
                    entrada = presence.get('entrada', '')
                    if entrada:
                        cell_obs.value = entrada
                    else:
                        cell_obs.value = '-'
                else:
                    cell_obs.value = '-'
                
                cell_obs.font = Font(size=9, bold=False, color='000000')
                cell_obs.alignment = center
                cell_obs.border = border
                ws.row_dimensions[obs_row].height = 18

                # Próximo dia
                current_col += 1
                if current_col > 7:
                    current_col = 1
                    current_row += 2  # Pula linha de observações

                current_date += timedelta(days=1)

            # Ajustar altura de linha geral
            for row in ws.row_dimensions.values():
                if not row.height:
                    row.height = 20

            # Escrever em memória
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            response_stream = make_response(bio.read())
            filename = f'calendario_{nome.replace(" ", "_")}_{month_reference.year}-{month_reference.month:02d}.xlsx'
            response_stream.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response_stream.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
            return response_stream
        except Exception as exc:
            app.logger.error('Erro ao gerar calendário de presença por colaborador: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca-calendario-massa-xlsx')
    @require_auth
    def presence_calendar_mass_xlsx(profile):
        """Gera Excel com múltiplas abas - um calendário por colaborador + SUMÁRIO"""
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        filial_id_param = request.args.get('filial_id')
        incluir_desligados = request.args.get('incluir_desligados', 'false').lower() == 'true'

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False})

            # Buscar filiais que o usuário pode acessar
            allowed_filial_ids = profile.get('allowed_filial_ids', [])
            if not allowed_filial_ids:
                return jsonify({'error': 'Usuário sem acesso a filiais.'}), 403

            # Se especificou filial, verificar permissão
            if filial_id_param:
                if int(filial_id_param) not in allowed_filial_ids:
                    return jsonify({'error': 'Sem permissão para acessar esta filial.'}), 403
                filiais_filter = [int(filial_id_param)]
            else:
                filiais_filter = allowed_filial_ids

            # Calcular período do mês
            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            # Buscar todos os colaboradores da(s) filial(is)
            colls_response = (
                supabase.table('colaboradores')
                .select('id, filial_id, nome_completo, cargo, turno, horario_padrao_inicio, horario_padrao_fim, ativo, data_admissao')
                .in_('filial_id', filiais_filter)
            )
            if not incluir_desligados:
                colls_response = colls_response.eq('ativo', True)

            colls_response = colls_response.order('nome_completo').execute()
            colaboradores = colls_response.data or []

            if not colaboradores:
                return jsonify({'error': 'Nenhum colaborador encontrado.'}), 404

            # Buscar todas as presenças do mês para todos os colaboradores
            presence_response = (
                supabase.table('presencas_diarias')
                .select('*')
                .in_('colaborador_id', [c['id'] for c in colaboradores])
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
                .execute()
            )
            presences_all = presence_response.data or []

            # Mapa filial_id -> "cidade/uf"
            filiais_resp = supabase.table('filiais').select('id, cidade, uf').in_('id', filiais_filter).execute()
            filiais_map = {item['id']: f"{item['cidade']}/{item['uf']}" for item in (filiais_resp.data or [])}

            # Organizar presenças por colaborador
            presences_by_colab = {}
            for p in presences_all:
                cid = p['colaborador_id']
                if cid not in presences_by_colab:
                    presences_by_colab[cid] = {}
                presences_by_colab[cid][p['data_referencia']] = p

            today = date_class.today()

            # Calcular estatísticas gerais
            total_presentes = len([p for p in presences_all if p.get('status') == 'presente'])
            total_faltas = len([p for p in presences_all if p.get('status') == 'falta'])
            total_atrasos = len([p for p in presences_all if p.get('status') == 'atraso'])

            # Calcular assiduidade por colaborador
            colab_stats = []
            for colab in colaboradores:
                colab_id = colab['id']
                colab_presences = presences_by_colab.get(colab_id, {})

                # Período válido: desde a admissão até hoje (não contar dias futuros)
                raw_admissao = colab.get('data_admissao')
                try:
                    admissao_date = date_class.fromisoformat(str(raw_admissao)) if raw_admissao else start_date
                except (ValueError, TypeError):
                    admissao_date = start_date
                colab_start = max(start_date, admissao_date)
                colab_end = min(end_date, today)

                # Contar apenas dias úteis (seg-sex) dentro do período válido
                dias_uteis = 0
                dias_presentes = 0
                dias_faltas = 0
                dias_atrasos = 0

                current_date = colab_start
                while current_date <= colab_end:
                    if current_date.weekday() < 5:
                        dias_uteis += 1
                        date_str = current_date.isoformat()
                        if date_str in colab_presences:
                            p = colab_presences[date_str]
                            if p.get('status') == 'atraso':
                                dias_atrasos += 1
                            elif p.get('status') in ('presente', 'folga', 'ferias', 'atestado', 'afastado'):
                                dias_presentes += 1
                            elif p.get('status') == 'falta':
                                dias_faltas += 1
                        else:
                            dias_faltas += 1
                    current_date += timedelta(days=1)

                assiduidade = (dias_presentes / dias_uteis * 100) if dias_uteis > 0 else 0
                colab_stats.append({
                    'colab': colab,
                    'presences': colab_presences,
                    'assiduidade': assiduidade,
                    'dias_presentes': dias_presentes,
                    'dias_faltas': dias_faltas,
                    'dias_atrasos': dias_atrasos,
                    'colab_start': colab_start,
                    'colab_end': colab_end,
                })

            # Criar workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove sheet padrão

            # Estilos
            bold = Font(bold=True, color='FFFFFF', size=12)
            bold_dark = Font(bold=True, size=10, color='000000')
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left = Alignment(horizontal='left', vertical='center')
            
            color_header = 'FF2C3E50'
            color_presente = 'FF27AE60'
            color_falta = 'FFE74C3C'
            color_atraso = 'FFF39C12'
            color_folga = 'FF95A5A6'
            
            thin = Side(border_style='thin', color='FF000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ===== ABA 1: SUMÁRIO =====
            ws_sumario = wb.create_sheet('SUMÁRIO', 0)
            
            employer = os.getenv('EMPREGADOR_NOME', 'GOLD TRANSPORTES')
            employer_cnpj = os.getenv('EMPREGADOR_CNPJ', '')
            
            # Cabeçalho
            ws_sumario.merge_cells('A1:D1')
            cell = ws_sumario['A1']
            cell.value = 'RELATÓRIO DE PRESENÇA'
            cell.font = Font(bold=True, size=14, color='FFFFFF')
            cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
            cell.alignment = center
            ws_sumario.row_dimensions[1].height = 25

            row_idx = 3
            ws_sumario[f'A{row_idx}'].value = 'Empresa:'
            ws_sumario[f'A{row_idx}'].font = bold_dark
            ws_sumario[f'B{row_idx}'].value = employer
            row_idx += 1

            ws_sumario[f'A{row_idx}'].value = 'CNPJ:'
            ws_sumario[f'A{row_idx}'].font = bold_dark
            ws_sumario[f'B{row_idx}'].value = employer_cnpj
            row_idx += 1

            ws_sumario[f'A{row_idx}'].value = 'Período:'
            ws_sumario[f'A{row_idx}'].font = bold_dark
            ws_sumario[f'B{row_idx}'].value = f'{start_date.strftime("%d/%m/%Y")} a {end_date.strftime("%d/%m/%Y")}'
            row_idx += 1

            ws_sumario[f'A{row_idx}'].value = 'Data de Geração:'
            ws_sumario[f'A{row_idx}'].font = bold_dark
            ws_sumario[f'B{row_idx}'].value = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
            row_idx += 2

            # Estatísticas
            ws_sumario[f'A{row_idx}'].value = 'ESTATÍSTICAS CONSOLIDADAS'
            ws_sumario[f'A{row_idx}'].font = Font(bold=True, size=11)
            row_idx += 1

            stats_data = [
                ('Total de Colaboradores', len(colaboradores)),
                ('Presenças Totais', total_presentes),
                ('Faltas Totais', total_faltas),
                ('Atrasos Totais', total_atrasos),
            ]
            
            for label, value in stats_data:
                ws_sumario[f'A{row_idx}'].value = label
                ws_sumario[f'A{row_idx}'].font = bold_dark
                ws_sumario[f'B{row_idx}'].value = value
                ws_sumario[f'B{row_idx}'].font = Font(bold=True, size=10)
                row_idx += 1

            row_idx += 1

            # Assiduidade média
            assiduidade_media = sum(s['assiduidade'] for s in colab_stats) / len(colab_stats) if colab_stats else 0
            ws_sumario[f'A{row_idx}'].value = 'Assiduidade Média'
            ws_sumario[f'A{row_idx}'].font = bold_dark
            ws_sumario[f'B{row_idx}'].value = f'{assiduidade_media:.1f}%'
            ws_sumario[f'B{row_idx}'].font = Font(bold=True, size=10, color='27AE60')
            row_idx += 2

            # Distribuição por faixa
            ws_sumario[f'A{row_idx}'].value = 'DISTRIBUIÇÃO POR FAIXA DE ASSIDUIDADE'
            ws_sumario[f'A{row_idx}'].font = Font(bold=True, size=10)
            row_idx += 1

            faixas = [
                ('≥95%', 95),
                ('90-94%', 90),
                ('80-89%', 80),
                ('<80%', 0),
            ]

            for label, threshold in faixas:
                if label == '<80%':
                    count = len([s for s in colab_stats if s['assiduidade'] < threshold])
                else:
                    count = len([s for s in colab_stats if 80 <= s['assiduidade'] < 95]) if label == '80-89%' else len([s for s in colab_stats if 90 <= s['assiduidade'] < 95]) if label == '90-94%' else len([s for s in colab_stats if s['assiduidade'] >= threshold])
                
                ws_sumario[f'A{row_idx}'].value = f'{label}'
                ws_sumario[f'B{row_idx}'].value = count
                row_idx += 1

            row_idx += 1

            # Top 5
            ws_sumario[f'A{row_idx}'].value = 'TOP 5 MELHORES ASSIDUIDADES'
            ws_sumario[f'A{row_idx}'].font = Font(bold=True, size=10)
            row_idx += 1

            sorted_stats = sorted(colab_stats, key=lambda x: x['assiduidade'], reverse=True)[:5]
            for idx, stat in enumerate(sorted_stats, 1):
                ws_sumario[f'A{row_idx}'].value = f'{idx}. {stat["colab"]["nome_completo"]}'
                ws_sumario[f'B{row_idx}'].value = f'{stat["assiduidade"]:.1f}%'
                row_idx += 1

            # Ajustar largura
            ws_sumario.column_dimensions['A'].width = 30
            ws_sumario.column_dimensions['B'].width = 15

            # ===== ABA: TABELA =====
            ws_tab = wb.create_sheet('TABELA', 1)

            status_labels_tab = {
                'presente': 'Presente', 'falta': 'Falta', 'folga': 'Folga',
                'atestado': 'Atestado', 'ferias': 'Férias', 'afastado': 'Afastado',
                'pendente': 'Pendente', 'atraso': 'Atraso',
            }

            tab_headers = ['Data', 'Colaborador', 'Filial', 'Cargo', 'Escala', 'Status', 'Observações']
            for ci, h in enumerate(tab_headers, 1):
                cell = ws_tab.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                cell.alignment = center
                cell.border = border

            tab_row = 2
            cur_date = start_date
            colabs_sorted = sorted(colaboradores, key=lambda c: c.get('nome_completo', ''))
            while cur_date <= end_date:
                date_str = cur_date.isoformat()
                date_fmt = cur_date.strftime('%d/%m/%Y')
                for colab in colabs_sorted:
                    presence = presences_by_colab.get(colab['id'], {}).get(date_str)
                    if not presence:
                        continue
                    raw_status = (presence.get('status') or '').lower()
                    status_label = status_labels_tab.get(raw_status, raw_status.capitalize())
                    row_vals = [
                        date_fmt,
                        colab.get('nome_completo', ''),
                        filiais_map.get(colab.get('filial_id'), '-'),
                        colab.get('cargo', ''),
                        presence.get('escala_servico') or colab.get('turno') or '',
                        status_label,
                        presence.get('observacoes') or '',
                    ]
                    for ci, val in enumerate(row_vals, 1):
                        cell = ws_tab.cell(row=tab_row, column=ci, value=val)
                        cell.border = border
                        cell.alignment = left
                        if ci == 6:
                            if raw_status == 'presente':
                                cell.fill = PatternFill(start_color='FFD5F5E3', end_color='FFD5F5E3', fill_type='solid')
                            elif raw_status == 'falta':
                                cell.fill = PatternFill(start_color='FFFADBD8', end_color='FFFADBD8', fill_type='solid')
                            elif raw_status in ('folga', 'ferias', 'afastado'):
                                cell.fill = PatternFill(start_color='FFEAF0FB', end_color='FFEAF0FB', fill_type='solid')
                    tab_row += 1
                cur_date += timedelta(days=1)

            for ci, w in enumerate([12, 35, 22, 25, 10, 12, 45], 1):
                ws_tab.column_dimensions[chr(64 + ci)].width = w
            ws_tab.freeze_panes = 'A2'

            color_sem_dado = 'FFD0D0D0'

            # ===== ABAS INDIVIDUAIS =====
            for stat_info in colab_stats:
                colab = stat_info['colab']
                colab_presences = stat_info['presences']
                nome = colab.get('nome_completo', 'Sem Nome')
                cal_start = stat_info['colab_start']
                cal_end = stat_info['colab_end']
                
                ws = wb.create_sheet(nome[:31])  # Limite de 31 caracteres no Excel

                # Cabeçalho
                ws.merge_cells('A1:G1')
                header_cell = ws['A1']
                mes_text = month_reference.strftime('%B/%Y').upper()
                header_cell.value = f'{nome} - {mes_text}'
                header_cell.font = Font(bold=True, size=14, color='FFFFFF')
                header_cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                header_cell.alignment = center
                ws.row_dimensions[1].height = 25

                # Linha de dias da semana
                day_names = ['Do', 'Se', 'Te', 'Qu', 'Qu', 'Se', 'Sa']
                for col, day_name in enumerate(day_names, start=1):
                    cell = ws.cell(row=3, column=col, value=day_name)
                    cell.font = Font(bold=True, size=11, color='FFFFFF')
                    cell.fill = PatternFill(start_color=color_header, end_color=color_header, fill_type='solid')
                    cell.alignment = center
                    cell.border = border
                    ws.row_dimensions[3].height = 20

                # Grid do calendário
                first_day_weekday = start_date.weekday()
                first_day_weekday = (first_day_weekday + 1) % 7
                
                current_row = 4
                current_col = first_day_weekday + 1
                current_date = start_date
                
                while current_date <= end_date:
                    date_str = current_date.isoformat()
                    presence = colab_presences.get(date_str)
                    
                    is_weekend = current_date.weekday() >= 5
                    before_admission = current_date < cal_start
                    is_future = current_date > cal_end

                    if before_admission or is_future:
                        status_color = color_sem_dado
                        status_symbol = ''
                    elif is_weekend:
                        status_color = color_folga
                        status_symbol = '⚫'
                    elif not presence:
                        status_color = color_falta
                        status_symbol = '❌'
                    elif presence.get('status') == 'atraso':
                        status_color = color_atraso
                        status_symbol = '🟠'
                    elif presence.get('status') == 'folga':
                        status_color = color_folga
                        status_symbol = '🔴'
                    else:
                        status_color = color_presente
                        status_symbol = '🟢'
                    
                    # Célula do dia
                    cell = ws.cell(row=current_row, column=current_col)
                    cell.value = current_date.day
                    cell.font = Font(bold=True, size=18, color='FFFFFF')
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                    cell.alignment = center
                    cell.border = border
                    ws.row_dimensions[current_row].height = 40
                    ws.column_dimensions[chr(64 + current_col)].width = 18

                    # Observação
                    obs_row = current_row + 1
                    cell_obs = ws.cell(row=obs_row, column=current_col)
                    
                    if presence and presence.get('observacoes'):
                        cell_obs.value = presence.get('observacoes', '-')
                    elif presence and presence.get('status') == 'atraso':
                        entrada = presence.get('entrada', '')
                        cell_obs.value = entrada if entrada else '-'
                    else:
                        cell_obs.value = '-'
                    
                    cell_obs.font = Font(size=9, bold=False, color='000000')
                    cell_obs.alignment = center
                    cell_obs.border = border
                    ws.row_dimensions[obs_row].height = 18

                    # Próximo dia
                    current_col += 1
                    if current_col > 7:
                        current_col = 1
                        current_row += 2

                    current_date += timedelta(days=1)

            # Salvar em memória
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            response_stream = make_response(bio.read())
            filename = f'presencas_massa_{month_reference.year}-{month_reference.month:02d}.xlsx'
            response_stream.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response_stream.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
            return response_stream

        except Exception as exc:
            app.logger.error('Erro ao gerar calendário em massa: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca-calendario-pdf')
    @require_auth
    def presence_calendar_mass_pdf(profile):
        """Gera PDF com página de sumário + uma página de calendário por colaborador"""
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        filial_id_param = request.args.get('filial_id')
        incluir_desligados = request.args.get('incluir_desligados', 'false').lower() == 'true'

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False})

            allowed_filial_ids = profile.get('allowed_filial_ids', [])
            if not allowed_filial_ids:
                return jsonify({'error': 'Usuário sem acesso a filiais.'}), 403

            if filial_id_param:
                if int(filial_id_param) not in allowed_filial_ids:
                    return jsonify({'error': 'Sem permissão para acessar esta filial.'}), 403
                filiais_filter = [int(filial_id_param)]
            else:
                filiais_filter = allowed_filial_ids

            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            colls_response = (
                supabase.table('colaboradores')
                .select('id, filial_id, nome_completo, cargo, turno, horario_padrao_inicio, horario_padrao_fim, ativo')
                .in_('filial_id', filiais_filter)
            )
            if not incluir_desligados:
                colls_response = colls_response.eq('ativo', True)
            colls_response = colls_response.order('nome_completo').execute()
            colaboradores = colls_response.data or []

            if not colaboradores:
                return jsonify({'error': 'Nenhum colaborador encontrado.'}), 404

            presence_response = (
                supabase.table('presencas_diarias')
                .select('*')
                .in_('colaborador_id', [c['id'] for c in colaboradores])
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
                .execute()
            )
            presences_all = presence_response.data or []

            presences_by_colab = {}
            for p in presences_all:
                cid = p['colaborador_id']
                if cid not in presences_by_colab:
                    presences_by_colab[cid] = {}
                presences_by_colab[cid][p['data_referencia']] = p

            colab_stats = []
            for colab in colaboradores:
                colab_id = colab['id']
                colab_presences = presences_by_colab.get(colab_id, {})
                dias_uteis = dias_presentes = dias_faltas = dias_atrasos = 0
                current_date = start_date
                while current_date <= end_date:
                    if current_date.weekday() < 5:
                        dias_uteis += 1
                        date_str = current_date.isoformat()
                        if date_str in colab_presences:
                            p_status = colab_presences[date_str].get('status')
                            if p_status == 'atraso':
                                dias_atrasos += 1
                            elif p_status in ('presente', 'folga'):
                                dias_presentes += 1
                        else:
                            dias_faltas += 1
                    current_date += timedelta(days=1)
                assiduidade = (dias_presentes / dias_uteis * 100) if dias_uteis > 0 else 0
                colab_stats.append({
                    'colab': colab,
                    'presences': colab_presences,
                    'assiduidade': assiduidade,
                    'dias_presentes': dias_presentes,
                    'dias_faltas': dias_faltas,
                    'dias_atrasos': dias_atrasos,
                    'dias_uteis': dias_uteis,
                })

            employer = os.getenv('EMPREGADOR_NOME', 'GOLD TRANSPORTES')
            _month_names_pt = ['', 'Janeiro', 'Fevereiro', 'Marco', 'Abril', 'Maio', 'Junho',
                               'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            mes_text = f'{_month_names_pt[month_reference.month]}/{month_reference.year}'.upper()

            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.set_auto_page_break(auto=True, margin=12)
            pdf.set_margins(10, 10, 10)

            C_DARK = (44, 62, 80)
            C_WHITE = (255, 255, 255)
            C_GREEN = (39, 174, 96)
            C_RED = (231, 76, 60)
            C_ORANGE = (243, 156, 18)
            C_GRAY = (149, 165, 166)
            C_LIGHT = (220, 220, 220)

            def pdf_set_fill(color):
                pdf.set_fill_color(*color)

            def pdf_set_text(color):
                pdf.set_text_color(*color)

            # ===== PÁGINA 1: SUMÁRIO =====
            pdf.add_page()

            pdf_set_fill(C_DARK)
            pdf_set_text(C_WHITE)
            pdf.set_font('Helvetica', 'B', 15)
            pdf.cell(0, 12, f'RELATORIO DE PRESENCA - {employer}', fill=True, align='C')
            pdf.ln(12)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, f'Periodo: {start_date.strftime("%d/%m/%Y")} a {end_date.strftime("%d/%m/%Y")}   |   Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', fill=True, align='C')
            pdf.ln(10)

            total_presentes = len([p for p in presences_all if p.get('status') == 'presente'])
            total_atrasos = len([p for p in presences_all if p.get('status') == 'atraso'])
            total_faltas_sum = sum(s['dias_faltas'] for s in colab_stats)
            assiduidade_media = sum(s['assiduidade'] for s in colab_stats) / len(colab_stats) if colab_stats else 0

            # Caixas de resumo
            pdf_set_text((0, 0, 0))
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 7, 'RESUMO GERAL', align='L')
            pdf.ln(9)

            boxes = [
                ('Colaboradores', str(len(colaboradores)), C_DARK),
                ('Presencas', str(total_presentes), C_GREEN),
                ('Faltas', str(total_faltas_sum), C_RED),
                ('Atrasos', str(total_atrasos), C_ORANGE),
                (f'Assiduidade Media', f'{assiduidade_media:.1f}%', (41, 128, 185)),
            ]
            box_w = 50
            for label, value, color in boxes:
                pdf_set_fill(color)
                pdf_set_text(C_WHITE)
                pdf.set_font('Helvetica', 'B', 16)
                pdf.cell(box_w, 13, value, fill=True, align='C')
            pdf.ln(13)
            for label, _, color in boxes:
                pdf_set_fill((240, 240, 240))
                pdf_set_text((80, 80, 80))
                pdf.set_font('Helvetica', '', 8)
                pdf.cell(box_w, 6, label, fill=True, align='C')
            pdf.ln(12)

            # Tabela de assiduidade
            pdf_set_text((0, 0, 0))
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(0, 7, 'ASSIDUIDADE POR COLABORADOR', align='L')
            pdf.ln(9)

            col_widths_sum = [85, 50, 24, 22, 22, 30]
            headers_sum = ['Colaborador', 'Cargo', 'Presentes', 'Faltas', 'Atrasos', 'Assiduidade']
            pdf_set_fill(C_DARK)
            pdf_set_text(C_WHITE)
            pdf.set_font('Helvetica', 'B', 8)
            for h, w in zip(headers_sum, col_widths_sum):
                pdf.cell(w, 6, h, border=1, fill=True, align='C')
            pdf.ln(6)

            sorted_stats = sorted(colab_stats, key=lambda x: x['assiduidade'], reverse=True)
            for i, stat in enumerate(sorted_stats):
                row_fill = (248, 249, 250) if i % 2 == 0 else (255, 255, 255)
                ass = stat['assiduidade']
                ass_color = C_GREEN if ass >= 90 else C_ORANGE if ass >= 75 else C_RED
                nome = (stat['colab'].get('nome_completo') or '-')[:36]
                cargo = (stat['colab'].get('cargo') or '-')[:24]
                pdf_set_fill(row_fill)
                pdf_set_text((0, 0, 0))
                pdf.set_font('Helvetica', '', 8)
                pdf.cell(85, 5, nome, border=1, fill=True)
                pdf.cell(50, 5, cargo, border=1, fill=True)
                pdf.cell(24, 5, str(stat['dias_presentes']), border=1, fill=True, align='C')
                pdf.cell(22, 5, str(stat['dias_faltas']), border=1, fill=True, align='C')
                pdf.cell(22, 5, str(stat['dias_atrasos']), border=1, fill=True, align='C')
                pdf_set_fill(ass_color)
                pdf_set_text(C_WHITE)
                pdf.cell(30, 5, f'{ass:.1f}%', border=1, fill=True, align='C')
                pdf.ln(5)

            # ===== PÁGINAS INDIVIDUAIS =====
            day_names_pt = ['Dom', 'Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sab']
            cell_w = 277.0 / 7
            cell_h = 19

            for stat_info in colab_stats:
                colab = stat_info['colab']
                colab_presences = stat_info['presences']
                nome = colab.get('nome_completo') or '-'
                cargo = colab.get('cargo') or '-'
                ass = stat_info['assiduidade']
                bar_color = C_GREEN if ass >= 90 else C_ORANGE if ass >= 75 else C_RED

                pdf.add_page()

                pdf_set_fill(C_DARK)
                pdf_set_text(C_WHITE)
                pdf.set_font('Helvetica', 'B', 13)
                pdf.cell(0, 11, f'{nome} - {mes_text}', fill=True, align='C')
                pdf.ln(11)

                pdf_set_fill(bar_color)
                pdf.set_font('Helvetica', '', 9)
                pdf.cell(0, 7, f'Cargo: {cargo}   |   Assiduidade: {ass:.1f}%   |   Presentes: {stat_info["dias_presentes"]}   |   Faltas: {stat_info["dias_faltas"]}   |   Atrasos: {stat_info["dias_atrasos"]}   |   Dias Uteis: {stat_info["dias_uteis"]}', fill=True, align='C')
                pdf.ln(10)

                # Legenda
                pdf.set_font('Helvetica', '', 7)
                legend = [(C_GREEN, 'Presente'), (C_RED, 'Falta'), (C_ORANGE, 'Atraso'), (C_GRAY, 'Folga/FDS')]
                for color, label in legend:
                    pdf_set_fill(color)
                    pdf.cell(4, 4, '', fill=True)
                    pdf_set_text((60, 60, 60))
                    pdf.cell(24, 4, ' ' + label)
                pdf.ln(8)

                # Cabeçalho dos dias da semana
                pdf_set_fill(C_DARK)
                pdf_set_text(C_WHITE)
                pdf.set_font('Helvetica', 'B', 9)
                for day_name in day_names_pt:
                    pdf.cell(cell_w, 7, day_name, border=1, fill=True, align='C')
                pdf.ln(7)

                # Células vazias antes do primeiro dia
                first_day_weekday = (start_date.weekday() + 1) % 7  # Monday=0 → Sunday=0
                col = 0
                for _ in range(first_day_weekday):
                    pdf_set_fill((250, 250, 250))
                    pdf_set_text((200, 200, 200))
                    pdf.cell(cell_w, cell_h, '', border=1, fill=True)
                    col += 1

                # Dias do mês
                current_date = start_date
                while current_date <= end_date:
                    date_str = current_date.isoformat()
                    presence = colab_presences.get(date_str)
                    is_weekend = current_date.weekday() >= 5

                    if is_weekend:
                        pdf_set_fill(C_LIGHT)
                        pdf_set_text((100, 100, 100))
                    elif not presence:
                        pdf_set_fill(C_RED)
                        pdf_set_text(C_WHITE)
                    elif presence.get('status') == 'atraso':
                        pdf_set_fill(C_ORANGE)
                        pdf_set_text(C_WHITE)
                    elif presence.get('status') == 'folga':
                        pdf_set_fill(C_GRAY)
                        pdf_set_text(C_WHITE)
                    else:
                        pdf_set_fill(C_GREEN)
                        pdf_set_text(C_WHITE)

                    pdf.set_font('Helvetica', 'B', 15)
                    pdf.cell(cell_w, cell_h, str(current_date.day), border=1, fill=True, align='C')
                    col += 1
                    if col >= 7:
                        col = 0
                        pdf.ln(cell_h)
                    current_date += timedelta(days=1)

                while 0 < col < 7:
                    pdf_set_fill((250, 250, 250))
                    pdf.cell(cell_w, cell_h, '', border=1, fill=True)
                    col += 1
                if col > 0:
                    pdf.ln(cell_h)

            pdf_bytes = bytes(pdf.output())
            response_stream = make_response(pdf_bytes)
            filename = f'presencas_massa_{month_reference.year}-{month_reference.month:02d}.pdf'
            response_stream.headers.set('Content-Type', 'application/pdf')
            response_stream.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
            return response_stream

        except Exception as exc:
            app.logger.error('Erro ao gerar calendário PDF: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/presenca-colaborador-xlsx')
    @require_auth
    def presence_collaborator_xlsx(profile):
        scope_error = require_scope_permission(profile, 'menu.presenca')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        raw_colaborador_id = request.args.get('colaborador_id')
        if not str(raw_colaborador_id).isdigit():
            return jsonify({'error': 'Informe um ID de colaborador válido.'}), 400
        
        colaborador_id = int(raw_colaborador_id)

        try:
            if not attendance_table_ready():
                return jsonify({'database_ready': False})

            # Buscar colaborador
            coll_response = (
                supabase.table('colaboradores')
                .select('id, filial_id, nome_completo, cargo, turno, horario_padrao_inicio, horario_padrao_fim')
                .eq('id', colaborador_id)
                .execute()
            )
            if not coll_response.data:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404
            
            colaborador = coll_response.data[0]
            filial_id = colaborador.get('filial_id')
            
            # Verificar acesso à filial
            if not ensure_profile_can_access_filial(profile, filial_id):
                return jsonify({'error': 'Sem permissão para consultar este colaborador.'}), 403

            # Calcular período do mês
            start_date = month_reference
            if start_date.month == 12:
                next_month = date_class(start_date.year + 1, 1, 1)
            else:
                next_month = date_class(start_date.year, start_date.month + 1, 1)
            end_date = next_month - timedelta(days=1)

            # Buscar presenças do colaborador no mês
            presence_response = (
                supabase.table('presencas_diarias')
                .select('*')
                .eq('colaborador_id', colaborador_id)
                .gte('data_referencia', start_date.isoformat())
                .lte('data_referencia', end_date.isoformat())
                .execute()
            )
            presences = {r['data_referencia']: r for r in (presence_response.data or [])}

            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Presença'

            bold = Font(bold=True, color='FFFFFF', size=12)
            bold_dark = Font(bold=True, size=11, color='000000')
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left = Alignment(horizontal='left', vertical='center')
            header_fill = PatternFill(start_color='FF2C3E50', end_color='FF2C3E50', fill_type='solid')
            status_presente_fill = PatternFill(start_color='FF27AE60', end_color='FF27AE60', fill_type='solid')
            status_falta_fill = PatternFill(start_color='FFE74C3C', end_color='FFE74C3C', fill_type='solid')
            status_atraso_fill = PatternFill(start_color='FFF39C12', end_color='FFF39C12', fill_type='solid')
            status_folga_fill = PatternFill(start_color='FF95A5A6', end_color='FF95A5A6', fill_type='solid')
            thin = Side(border_style='thin', color='FF000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            employer = os.getenv('EMPREGADOR_NOME', 'GOLD TRANSPORTES')
            employer_cnpj = os.getenv('EMPREGADOR_CNPJ', '')

            # Cabeçalho
            ws.merge_cells('A1:F1')
            title_cell = ws['A1']
            title_cell.value = 'RELATÓRIO DE PRESENÇA - COLABORADOR'
            title_cell.font = Font(bold=True, size=14, color='FFFFFF')
            title_cell.fill = header_fill
            title_cell.alignment = center
            ws.row_dimensions[1].height = 25

            ws.merge_cells('A2:F2')
            ws['A2'].value = f'Empresa: {employer}'
            ws['A2'].font = bold_dark
            ws['A2'].alignment = left

            ws.merge_cells('A3:F3')
            ws['A3'].value = f'CNPJ: {employer_cnpj}'
            ws['A3'].font = bold_dark
            ws['A3'].alignment = left

            # Info colaborador
            nome = colaborador.get('nome_completo', '')
            cargo = colaborador.get('cargo', '')
            turno = colaborador.get('turno', '')
            horario_inicio = colaborador.get('horario_padrao_inicio', '')
            horario_fim = colaborador.get('horario_padrao_fim', '')

            row_idx = 5
            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            ws[f'A{row_idx}'].value = f'Colaborador: {nome}'
            ws[f'A{row_idx}'].font = bold_dark
            ws[f'A{row_idx}'].alignment = left
            row_idx += 1

            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            ws[f'A{row_idx}'].value = f'Cargo: {cargo}  |  Turno: {turno}  |  Horário: {horario_inicio} às {horario_fim}'
            ws[f'A{row_idx}'].font = bold_dark
            ws[f'A{row_idx}'].alignment = left
            row_idx += 1

            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            period_label = f'Período: {start_date.strftime("%d/%m/%Y")} a {end_date.strftime("%d/%m/%Y")}'
            ws[f'A{row_idx}'].value = period_label
            ws[f'A{row_idx}'].font = bold_dark
            ws[f'A{row_idx}'].alignment = left
            row_idx += 2

            # Headers da tabela
            headers = ['Data', 'Dia Semana', 'Status', 'Entrada', 'Saída', 'Observações']
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                cell.font = bold
                cell.fill = header_fill
                cell.alignment = center
                cell.border = border
            
            ws.row_dimensions[row_idx].height = 20
            row_idx += 1

            # Dados por dia
            current_date = start_date
            presente_count = 0
            falta_count = 0
            atraso_count = 0

            while current_date <= end_date:
                date_str = current_date.isoformat()
                day_name = current_date.strftime('%a')
                day_names_pt = {'Mon': 'Segunda', 'Tue': 'Terça', 'Wed': 'Quarta', 'Thu': 'Quinta', 'Fri': 'Sexta', 'Sat': 'Sábado', 'Sun': 'Domingo'}
                day_name_pt = day_names_pt.get(day_name, day_name)

                presence = presences.get(date_str)
                status = presence.get('status') if presence else 'ausente'
                entrada = presence.get('entrada') if presence else ''
                saida = presence.get('saida') if presence else ''
                observacoes = presence.get('observacoes') if presence else ''

                # Determinar status visual
                is_weekend = current_date.weekday() >= 5
                
                if is_weekend:
                    status_display = 'Fim de semana'
                    status_fill = status_folga_fill
                    status_color = Font(color='FFFFFF', bold=True)
                elif not presence:
                    status_display = 'Falta'
                    status_fill = status_falta_fill
                    status_color = Font(color='FFFFFF', bold=True)
                    falta_count += 1
                elif status == 'atraso':
                    status_display = 'Atraso'
                    status_fill = status_atraso_fill
                    status_color = Font(color='FFFFFF', bold=True)
                    atraso_count += 1
                elif status == 'folga':
                    status_display = 'Folga'
                    status_fill = status_folga_fill
                    status_color = Font(color='FFFFFF', bold=True)
                else:
                    status_display = 'Presente'
                    status_fill = status_presente_fill
                    status_color = Font(color='FFFFFF', bold=True)
                    presente_count += 1

                # Preencher linha
                row_cells = [
                    current_date.strftime('%d/%m/%Y'),
                    day_name_pt,
                    status_display,
                    entrada or '',
                    saida or '',
                    observacoes or ''
                ]

                for col, val in enumerate(row_cells, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.border = border
                    cell.alignment = center if col == 3 else left
                    
                    # Cor especial para coluna de status
                    if col == 3:
                        cell.fill = status_fill
                        cell.font = status_color
                
                current_date += timedelta(days=1)
                row_idx += 1

            # Resumo
            row_idx += 1
            ws.merge_cells(f'A{row_idx}:F{row_idx}')
            ws[f'A{row_idx}'].value = 'RESUMO'
            ws[f'A{row_idx}'].font = Font(bold=True, size=11)
            ws[f'A{row_idx}'].alignment = left
            row_idx += 1

            resumo_data = [
                ('Presenças', presente_count),
                ('Faltas', falta_count),
                ('Atrasos', atraso_count),
            ]

            for label, value in resumo_data:
                ws[f'A{row_idx}'].value = label
                ws[f'A{row_idx}'].font = bold_dark
                ws[f'B{row_idx}'].value = value
                ws[f'B{row_idx}'].font = Font(bold=True, size=11, color='000000')
                ws[f'B{row_idx}'].fill = PatternFill(start_color='FFECF0F1', end_color='FFECF0F1', fill_type='solid')
                ws[f'A{row_idx}'].border = border
                ws[f'B{row_idx}'].border = border
                row_idx += 1

            # Ajustar largura
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 30

            # Escrever em memória
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)

            response_stream = make_response(bio.read())
            filename = f'presenca_{nome.replace(" ", "_")}_{month_reference.year}-{month_reference.month:02d}.xlsx'
            response_stream.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response_stream.headers.set('Content-Disposition', f'attachment; filename="{filename}"')
            return response_stream
        except Exception as exc:
            app.logger.error('Erro ao gerar XLSX de presença por colaborador: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/bonificacao')
    @require_auth
    def bonificacao_list(profile):
        scope_error = require_scope_permission(profile, 'menu.bonificacao')
        if scope_error:
            return scope_error

        month_reference = parse_month_reference(request.args.get('mes') or request.args.get('month'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        filial_id = request.args.get('filial_id', type=int)
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403

        try:
            if not bonificacao_tables_ready():
                return jsonify({
                    'database_ready': False,
                    'month_reference': month_reference.isoformat(),
                    'month': month_reference.month,
                    'year': month_reference.year,
                    'month_label': f"{month_reference.month:02d}/{month_reference.year}",
                    'can_manage': profile_has_scope_permission(profile, 'manage.bonificacao'),
                    'filiais': fetch_accessible_filiais(profile),
                    'metricas': [],
                    'colaboradores': [],
                    'indicators_by_collaborator': {},
                    'lancamentos': [],
                    'summary': {
                        'monthly_total_paid': 0,
                        'annual_total_paid': 0,
                        'collaborator_totals': [],
                        'metric_totals': [],
                    },
                })

            return jsonify(build_bonificacao_board(profile, month_reference, filial_id))
        except Exception as exc:
            app.logger.error('Erro ao carregar controle de bonificação: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.put('/api/bonificacao')
    @rate_limit_endpoint(max_requests=10)
    @require_auth
    def bonificacao_save(profile):
        scope_error = require_scope_permission(profile, 'manage.bonificacao', 'Sem permissão para alterar bonificação.')
        if scope_error:
            return scope_error

        if not bonificacao_tables_ready():
            return jsonify({'error': 'As tabelas de bonificação ainda não existem. Rode a migration do módulo primeiro.'}), 400

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        month_reference = parse_month_reference(payload.get('mes') or payload.get('month') or payload.get('month_reference'))
        if not month_reference:
            return jsonify({'error': 'Informe um mês válido no formato YYYY-MM.'}), 400

        raw_filial_id = payload.get('filial_id')
        filial_id = int(raw_filial_id) if str(raw_filial_id).isdigit() else None
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para alterar dados desta base.'}), 403

        entries = payload.get('entries') or []
        if not isinstance(entries, list):
            return jsonify({'error': 'A lista de lançamentos enviada é inválida.'}), 400

        colaboradores = fetch_bonificacao_colaboradores(profile, filial_id)
        colaboradores_by_id = {int(item['id']): item for item in colaboradores}
        metricas = fetch_bonificacao_metricas()
        metricas_by_id = {int(item['id']): item for item in metricas}

        normalized_entries = []
        for entry in entries:
            if not isinstance(entry, dict):
                continue

            colaborador_id = entry.get('colaborador_id')
            metrica_id = entry.get('metrica_id')
            atingiu = bool(entry.get('atingiu'))

            if not str(colaborador_id).isdigit() or not str(metrica_id).isdigit():
                return jsonify({'error': 'Cada lançamento precisa de colaborador e métrica válidos.'}), 400

            colaborador_id = int(colaborador_id)
            metrica_id = int(metrica_id)

            if colaborador_id not in colaboradores_by_id:
                # Ignora entrada fora do escopo atual para evitar falha por estado antigo da tela.
                continue

            if metrica_id not in metricas_by_id:
                # Ignora métrica inativa/removida que ainda esteja em estado local do cliente.
                continue

            if not atingiu:
                continue

            valor_base = float(metricas_by_id[metrica_id].get('valor') or 0)
            percentual_aplicado = parse_float_or_default(entry.get('percentual_aplicado'), 100.0)
            percentual_aplicado = max(0.0, min(200.0, percentual_aplicado))
            valor_aplicado = entry.get('valor_aplicado')
            if valor_aplicado in (None, ''):
                valor_aplicado = valor_base * (percentual_aplicado / 100.0)
            else:
                valor_aplicado = max(0.0, parse_float_or_default(valor_aplicado, 0.0))

            normalized_entries.append({
                'mes_referencia': month_reference.isoformat(),
                'colaborador_id': colaborador_id,
                'filial_id': colaboradores_by_id[colaborador_id].get('filial_id'),
                'metrica_id': metrica_id,
                'atingiu': True,
                'valor_aplicado': round(valor_aplicado, 2),
                'observacoes': entry.get('observacoes') or None,
                'atualizado_por': profile.get('user_id'),
            })

        try:
            delete_query = (
                supabase.table('bonificacao_lancamentos')
                .delete()
                .eq('mes_referencia', month_reference.isoformat())
            )

            if filial_id:
                delete_query = delete_query.eq('filial_id', filial_id)
            elif profile_has_filial_scope(profile):
                delete_query = delete_query.in_('filial_id', profile.get('allowed_filial_ids') or [])

            delete_query.execute()

            if normalized_entries:
                (
                    supabase.table('bonificacao_lancamentos')
                    .insert(normalized_entries)
                    .execute()
                )

            write_audit_event(
                profile,
                action='bonus_save',
                resource_name='bonificacao_lancamentos',
                details={'month': month_reference.isoformat(), 'entries': len(normalized_entries)},
            )

            return jsonify({'status': 'ok', 'saved_rows': len(normalized_entries)})
        except Exception as exc:
            app.logger.error('Erro ao salvar bonificação mensal: %s', exc)
            write_audit_event(profile, 'bonus_save', 'bonificacao_lancamentos', status='error', details={'error': str(exc)[:300]})
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.get('/api/auditoria/config')
    @require_auth
    def auditoria_config(profile):
        scope_error = require_scope_permission(profile, 'menu.auditoria')
        if scope_error:
            return scope_error

        return jsonify({
            'database_ready': audit_table_ready(),
            'filiais': fetch_accessible_filiais(profile),
            'status_options': ['ok', 'error'],
        })

    @app.get('/api/auditoria')
    @require_auth
    def auditoria_list(profile):
        scope_error = require_scope_permission(profile, 'menu.auditoria')
        if scope_error:
            return scope_error

        if not audit_table_ready():
            return jsonify({'database_ready': False, 'items': []})

        filters = {
            'resource': (request.args.get('resource') or '').strip() or None,
            'action': (request.args.get('action') or '').strip() or None,
            'status': (request.args.get('status') or '').strip() or None,
            'filial_id': request.args.get('filial_id', type=int),
            'date_from': (request.args.get('date_from') or '').strip() or None,
            'date_to': (request.args.get('date_to') or '').strip() or None,
            'limit': request.args.get('limit', default=200, type=int),
        }

        try:
            items = list_audit_events(profile, filters)
            return jsonify({'database_ready': True, 'items': items})
        except Exception as exc:
            app.logger.error('Erro ao listar auditoria: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/colaboradores/importar')
    @rate_limit_endpoint(max_requests=5)
    @require_auth
    def colaboradores_importar(profile):
        scope_error = require_scope_permission(profile, 'menu.colaboradores')
        if scope_error:
            return scope_error

        permission_error = require_create_permission(profile, 'colaboradores')
        if permission_error:
            return permission_error

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        rows = payload.get('rows') or []
        if not isinstance(rows, list) or not rows:
            return jsonify({'error': 'Envie uma lista de linhas para importar.'}), 400

        if len(rows) > 500:
            return jsonify({'error': 'Importação limitada a 500 linhas por envio.'}), 400

        import unicodedata as _ud
        import re as _re
        def _strip_accents(s):
            return ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')

        def _is_valid_email(email):
            if not email or '@' not in email:
                return False
            local, _, domain = email.partition('@')
            return bool(local) and not local.endswith('.') and bool(domain) and '.' in domain

        def _email_from_name(nome, domain='gold.com'):
            parts = _strip_accents((nome or '').strip().lower()).split()
            parts = [_re.sub(r'[^a-z0-9]', '', p) for p in parts if p]
            parts = [p for p in parts if p]
            if not parts:
                return None
            local = f"{parts[0]}.{parts[-1]}" if len(parts) > 1 else parts[0]
            return f"{local}@{domain}"

        def _normalize_email(raw_email, nome):
            e = _strip_accents((raw_email or '').strip().lower())
            if _is_valid_email(e):
                return e
            domain = e.partition('@')[2].strip() if '@' in e else 'gold.com'
            return _email_from_name(nome, domain or 'gold.com') or e

        filiais_rows = fetch_accessible_filiais(profile)
        filiais_by_id = {int(item['id']): item for item in filiais_rows if item.get('id') is not None}
        filiais_by_city = {
            (item.get('cidade') or '').strip().lower(): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        filiais_by_city_uf = {
            f"{(item.get('cidade') or '').strip().lower()}/{(item.get('uf') or '').strip().lower()}": int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        # Accent-stripped variants so "SÃO LUIS/MA" matches "SÃO LUÍS/MA" etc.
        filiais_by_city_norm = {
            _strip_accents((item.get('cidade') or '').strip().lower()): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        filiais_by_city_uf_norm = {
            _strip_accents(f"{(item.get('cidade') or '').strip().lower()}/{(item.get('uf') or '').strip().lower()}"): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        filiais_by_partner = {
            (item.get('parceira') or '').strip().lower(): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None and item.get('parceira')
        }
        # "cidade - uf" (dash separator common in Excel exports)
        filiais_by_city_dash_uf = {
            f"{(item.get('cidade') or '').strip().lower()} - {(item.get('uf') or '').strip().lower()}": int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        filiais_by_city_dash_uf_norm = {
            _strip_accents(f"{(item.get('cidade') or '').strip().lower()} - {(item.get('uf') or '').strip().lower()}"): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None
        }
        # computed code (e.g. "BELE", "SAPL") from decorate_filial_row
        filiais_by_codigo = {
            (item.get('codigo') or '').strip().lower(): int(item['id'])
            for item in filiais_rows
            if item.get('id') is not None and item.get('codigo')
        }

        # Normaliza e-mail e deduplica — mantém a primeira ocorrência de cada e-mail
        _seen_emails = set()
        _deduped = []
        skipped_duplicates = 0
        for _r in rows:
            if not isinstance(_r, dict):
                _deduped.append(_r)
                continue
            _nome = (_r.get('nome_completo') or _r.get('nome') or '').strip()
            _email_key = _normalize_email(_r.get('email'), _nome)
            if _email_key and _email_key in _seen_emails:
                skipped_duplicates += 1
                continue
            if _email_key:
                _seen_emails.add(_email_key)
            _deduped.append({**_r, '_email_norm': _email_key})
        rows = _deduped

        imported = 0
        updated = 0
        errors = []
        schema_warnings = []

        for index, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                errors.append({'line': index, 'error': 'Linha inválida, esperado objeto com colunas.'})
                continue

            raw_filial = row.get('filial') or row.get('filial_id')
            filial_id = None
            if str(raw_filial).isdigit():
                filial_id = int(raw_filial)
            elif raw_filial:
                normalized_filial = str(raw_filial).strip().lower()
                normalized_filial_norm = _strip_accents(normalized_filial)
                filial_id = (
                    filiais_by_city.get(normalized_filial)
                    or filiais_by_city_uf.get(normalized_filial)
                    or filiais_by_city_dash_uf.get(normalized_filial)
                    or filiais_by_city_norm.get(normalized_filial_norm)
                    or filiais_by_city_uf_norm.get(normalized_filial_norm)
                    or filiais_by_city_dash_uf_norm.get(normalized_filial_norm)
                    or filiais_by_partner.get(normalized_filial)
                    or filiais_by_codigo.get(normalized_filial)
                )

            if not filial_id and len(filiais_rows) == 1:
                filial_id = int(filiais_rows[0]['id'])

            if not filial_id or filial_id not in filiais_by_id:
                attempted = str(raw_filial or '').strip() or '(vazio)'
                errors.append({'line': index, 'error': f'Filial não encontrada: "{attempted}". Use cidade, cidade/UF ou nome parceira.'})
                continue

            normalized_payload = {
                'filial_id': filial_id,
                'nome_completo': (row.get('nome_completo') or row.get('nome') or '').strip(),
                'cpf': normalize_cpf(row.get('cpf')),
                'telefone': (row.get('telefone') or '').strip() or None,
                'cargo': (row.get('cargo') or '').strip(),
                'turno': ((row.get('turno') or '').strip().lower() or None),
                'escala_servico': (row.get('escala_servico') or '').strip() or None,
                'horario_padrao_inicio': (row.get('horario_padrao_inicio') or '').strip() or None,
                'horario_padrao_fim': (row.get('horario_padrao_fim') or '').strip() or None,
                'tipo_acesso': ((row.get('tipo_acesso') or 'app').strip().lower() or 'app'),
                'data_admissao': normalize_import_date(row.get('data_admissao')),
                'permissao_app': parse_import_bool(row.get('permissao_app'), True),
                'permissao_desktop': parse_import_bool(row.get('permissao_desktop'), False),
                'permissao_editar': parse_import_bool(row.get('permissao_editar'), False),
                'permissao_excluir': parse_import_bool(row.get('permissao_excluir'), False),
                'permissao_aprovar_he': parse_import_bool(row.get('permissao_aprovar_he'), False),
                'ativo': parse_import_bool(row.get('ativo'), True),
            }

            _nome_for_email = (row.get('nome_completo') or row.get('nome') or '').strip()
            raw_email = row.get('_email_norm') or _normalize_email(row.get('email'), _nome_for_email)

            missing = [
                field_name
                for field_name in ['nome_completo', 'cargo', 'data_admissao']
                if not normalized_payload.get(field_name)
            ]
            if missing:
                errors.append({'line': index, 'error': f'Campos obrigatórios ausentes: {", ".join(missing)}.'})
                continue

            if not raw_email:
                errors.append({'line': index, 'error': 'Informe o e-mail para vincular ao Supabase Auth.'})
                continue

            try:
                # Busca por nome e CPF ANTES de criar usuário Auth — evita duplicatas
                existing_collaborator = fetch_collaborator_by_name_and_filial(
                    normalized_payload.get('nome_completo'),
                    filial_id,
                )
                if not existing_collaborator:
                    existing_collaborator = fetch_collaborator_by_cpf_and_filial(
                        normalized_payload.get('cpf'),
                        filial_id,
                    )

                auth_user = ensure_auth_user(raw_email, normalized_payload.get('nome_completo'))
                normalized_payload['user_id'] = str(auth_user.id)

                if not existing_collaborator:
                    existing_collaborator = fetch_collaborator_by_user_id(auth_user.id)

                if existing_collaborator:
                    removed_columns = insert_or_update_collaborator_with_schema_fallback(
                        normalized_payload,
                        existing_id=existing_collaborator['id'],
                    )
                    updated += 1
                else:
                    removed_columns = insert_or_update_collaborator_with_schema_fallback(normalized_payload)
                    imported += 1

                if removed_columns:
                    schema_warnings.append({
                        'line': index,
                        'columns': removed_columns,
                    })
            except Exception as exc:
                translated_error = translate_database_error(exc)
                if translated_error == 'Estrutura de banco ausente para este módulo. Rode a migration correspondente.':
                    errors.append({
                        'line': index,
                        'error': f"{translated_error} Detalhe técnico: {str(exc)[:220]}",
                        'debug': str(exc)[:600],
                    })
                else:
                    errors.append({
                        'line': index,
                        'error': translated_error,
                        'debug': str(exc)[:600],
                    })

        write_audit_event(
            profile,
            action='collaborators_import',
            resource_name='colaboradores',
            details={
                'imported': imported,
                'updated': updated,
                'errors': len(errors),
                'schema_warnings': len(schema_warnings),
            },
        )

        filiais_disponiveis = sorted([
            f"{item.get('cidade', '')}/{item.get('uf', '')}"
            for item in filiais_rows
            if item.get('cidade')
        ])

        return jsonify({
            'status': 'ok',
            'imported': imported,
            'updated': updated,
            'skipped_duplicates': skipped_duplicates,
            'errors': errors,
            'schema_warnings': schema_warnings,
            'filiais_disponiveis': filiais_disponiveis,
        })

    @app.post('/api/veiculos/importar')
    @rate_limit_endpoint(max_requests=5)
    @require_auth
    def veiculos_importar(profile):
        scope_error = require_scope_permission(profile, 'menu.veiculos')
        if scope_error:
            return scope_error

        permission_error = require_create_permission(profile, 'veiculos')
        if permission_error:
            return permission_error

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        rows = payload.get('rows') or []
        if not isinstance(rows, list) or not rows:
            return jsonify({'error': 'Envie uma lista de linhas para importar.'}), 400

        if len(rows) > 500:
            return jsonify({'error': 'Importação limitada a 500 linhas por envio.'}), 400

        import unicodedata as _ud
        def _strip_accents(s):
            return ''.join(c for c in _ud.normalize('NFD', s) if _ud.category(c) != 'Mn')

        filiais_rows = fetch_accessible_filiais(profile)
        filiais_by_id = {int(item['id']): item for item in filiais_rows if item.get('id') is not None}
        filiais_by_city = {
            (item.get('cidade') or '').strip().lower(): int(item['id'])
            for item in filiais_rows if item.get('id') is not None
        }
        filiais_by_city_uf = {
            f"{(item.get('cidade') or '').strip().lower()}/{(item.get('uf') or '').strip().lower()}": int(item['id'])
            for item in filiais_rows if item.get('id') is not None
        }
        filiais_by_city_norm = {
            _strip_accents((item.get('cidade') or '').strip().lower()): int(item['id'])
            for item in filiais_rows if item.get('id') is not None
        }
        filiais_by_city_uf_norm = {
            _strip_accents(f"{(item.get('cidade') or '').strip().lower()}/{(item.get('uf') or '').strip().lower()}"): int(item['id'])
            for item in filiais_rows if item.get('id') is not None
        }
        filiais_by_codigo = {
            (item.get('codigo') or '').strip().lower(): int(item['id'])
            for item in filiais_rows if item.get('id') is not None and item.get('codigo')
        }

        TIPO_VALID = {'rota', 'transferencia', 'diaria', 'multidia'}
        TIPO_ALIASES = {
            'rota': 'rota',
            'transferência': 'transferencia',
            'transferencia': 'transferencia',
            'diária': 'diaria',
            'diaria': 'diaria',
            'multidia': 'multidia',
            'multi-dia': 'multidia',
            'multi dia': 'multidia',
        }
        TIPO_VEICULO_ALIASES = {
            'utilitario': 'utilitario',
            'utilitário': 'utilitario',
            'caminhao leve': 'caminhao_leve',
            'caminhão leve': 'caminhao_leve',
            'caminhao medio': 'caminhao_medio',
            'caminhão médio': 'caminhao_medio',
            'caminhao pesado': 'caminhao_pesado',
            'caminhão pesado': 'caminhao_pesado',
            'van': 'van',
            'furgao': 'van',
            'furgão': 'van',
            'passeio': 'passeio',
            'moto': 'moto',
            'outro': 'outro',
        }
        STATUS_ALIASES = {
            'ativo': 'ativo',
            'manutencao': 'manutencao',
            'manutenção': 'manutencao',
            'em manutencao': 'manutencao',
            'em manutenção': 'manutencao',
            'inativo': 'inativo',
        }
        COMBUSTIVEL_VALID = {'diesel', 'gasolina', 'flex', 'etanol', 'gnv', 'eletrico'}

        def _norm_token(value):
            return _strip_accents(str(value or '').strip().lower())

        imported = 0
        updated = 0
        errors = []

        for index, row in enumerate(rows, start=1):
            if not isinstance(row, dict):
                errors.append({'line': index, 'error': 'Linha inválida.'})
                continue

            raw_filial = row.get('filial') or row.get('filial_id')
            filial_id = None
            if str(raw_filial).isdigit():
                filial_id = int(raw_filial)
            elif raw_filial:
                k = str(raw_filial).strip().lower()
                kn = _strip_accents(k)
                filial_id = (
                    filiais_by_city.get(k)
                    or filiais_by_city_uf.get(k)
                    or filiais_by_city_norm.get(kn)
                    or filiais_by_city_uf_norm.get(kn)
                    or filiais_by_codigo.get(k)
                )

            if not filial_id and len(filiais_rows) == 1:
                filial_id = int(filiais_rows[0]['id'])

            if not filial_id or filial_id not in filiais_by_id:
                attempted = str(raw_filial or '').strip() or '(vazio)'
                errors.append({'line': index, 'error': f'Filial não encontrada: "{attempted}".'})
                continue

            def _as_str(v):
                if v is None or v is False:
                    return ''
                return str(v).strip()

            placa_raw = _as_str(row.get('placa')).upper()
            # Aceita placa combo cavalo/carreta: "ABC1D23/XYZ4W56" ou "ABC1D23,XYZ4W56".
            # Normaliza separadores e espaços para "PLACA1/PLACA2".
            placa = re.sub(r'\s*[/,;]\s*', '/', placa_raw).strip('/').strip()
            if not placa:
                errors.append({'line': index, 'error': 'Placa é obrigatória.'})
                continue
            # marca/modelo aceitos vazios — usa placeholder quando ausente
            # (colunas NOT NULL no banco). Usuário pode editar depois.
            marca = _as_str(row.get('marca')) or '—'
            modelo_raw = _as_str(row.get('modelo'))
            modelo = modelo_raw if modelo_raw and modelo_raw != '0' else '—'

            tipo_raw = _norm_token(row.get('tipo'))
            tipo = TIPO_ALIASES.get(tipo_raw) if tipo_raw else None
            if tipo_raw and tipo not in TIPO_VALID:
                errors.append({'line': index, 'error': f'Tipo inválido: "{row.get("tipo")}". Use rota, transferencia, diaria ou multidia.'})
                continue

            tipo_veiculo_raw = _norm_token(row.get('tipo_veiculo'))
            tipo_veiculo = TIPO_VEICULO_ALIASES.get(tipo_veiculo_raw) or (tipo_veiculo_raw or None)

            status_raw = _norm_token(row.get('status')) or 'ativo'
            status = STATUS_ALIASES.get(status_raw, 'ativo')

            combustivel_raw = _norm_token(row.get('combustivel'))
            combustivel = combustivel_raw if combustivel_raw in COMBUSTIVEL_VALID else (combustivel_raw or None)

            def _to_int(value):
                try:
                    s = _as_str(value)
                    return int(s) if s else None
                except Exception:
                    return None

            def _to_float(value):
                try:
                    s = _as_str(value).replace(',', '.')
                    return float(s) if s else None
                except Exception:
                    return None

            def _to_date(value):
                s = _as_str(value)
                if not s or s in {'0', '0000-00-00'}:
                    return None
                parsed = normalize_import_date(s)
                # normalize_import_date returns raw if no date match — discard non-ISO strings
                if parsed and re.fullmatch(r'\d{4}-\d{2}-\d{2}', parsed):
                    return parsed
                return None

            normalized_payload = {
                'filial_id': filial_id,
                'placa': placa,
                'chassi': _as_str(row.get('chassi')) or None,
                'marca': marca,
                'modelo': modelo,
                'ano_modelo': _to_int(row.get('ano_modelo')),
                'cor': _as_str(row.get('cor')) or None,
                'tipo_veiculo': tipo_veiculo,
                'tipo': tipo,
                'combustivel': combustivel,
                'capacidade_tanque': _to_float(row.get('capacidade_tanque')),
                'status': status,
                'odometro_atual': _to_int(row.get('odometro_atual')) or 0,
                'km_proxima_revisao': _to_int(row.get('km_proxima_revisao')),
                'data_vencimento_crlv': _to_date(row.get('data_vencimento_crlv')),
                'data_vencimento_seguro': _to_date(row.get('data_vencimento_seguro')),
                'data_ultima_revisao': _to_date(row.get('data_ultima_revisao')),
                'descricao': _as_str(row.get('descricao')) or _as_str(row.get('observacoes')) or None,
            }

            try:
                existing = (
                    supabase.table('veiculos')
                    .select('id')
                    .eq('filial_id', filial_id)
                    .eq('placa', placa)
                    .limit(1)
                    .execute()
                )
                existing_row = (existing.data or [None])[0]

                if existing_row:
                    supabase.table('veiculos').update(normalized_payload).eq('id', existing_row['id']).execute()
                    updated += 1
                else:
                    supabase.table('veiculos').insert(normalized_payload).execute()
                    imported += 1
            except Exception as exc:
                errors.append({
                    'line': index,
                    'error': translate_database_error(exc),
                    'debug': str(exc)[:600],
                })

        write_audit_event(
            profile,
            action='veiculos_import',
            resource_name='veiculos',
            details={'imported': imported, 'updated': updated, 'errors': len(errors)},
        )

        filiais_disponiveis = sorted([
            f"{item.get('cidade', '')}/{item.get('uf', '')}"
            for item in filiais_rows if item.get('cidade')
        ])

        return jsonify({
            'status': 'ok',
            'imported': imported,
            'updated': updated,
            'errors': errors,
            'filiais_disponiveis': filiais_disponiveis,
        })

    @app.get('/api/carregamento/config')
    @require_auth
    def loading_config(profile):
        scope_error = require_scope_permission(profile, 'menu.carregamento')
        if scope_error:
            return scope_error

        try:
            database_ready = loading_tables_ready()
            support_rows = fetch_loading_support_rows(profile) if database_ready else {
                'filiais': fetch_accessible_filiais(profile),
                'rotas': [],
                'veiculos': [],
                'motivos': [],
            }

            return jsonify({
                'database_ready': database_ready,
                'today': date_class.today().isoformat(),
                'shift_options': LOADING_SHIFT_OPTIONS,
                'status_options': LOADING_JOURNEY_STATUS_OPTIONS,
                'filiais': support_rows['filiais'],
                'rotas': support_rows['rotas'],
                'veiculos': support_rows['veiculos'],
                'motivos': support_rows['motivos'],
                'current_leader': {
                    'id': profile.get('id'),
                    'nome_completo': profile.get('nome_completo'),
                    'cargo': profile.get('cargo'),
                },
                'can_plan': profile_has_scope_permission(profile, 'manage.programacao_carregamento'),
                'can_operate': profile_has_scope_permission(profile, 'manage.operacao_carregamento'),
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar configuração de carregamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/carregamento/jornadas')
    @require_auth
    def loading_journeys(profile):
        scope_error = require_scope_permission(profile, 'menu.carregamento')
        if scope_error:
            return scope_error

        target_date = parse_iso_date(request.args.get('data')) or date_class.today()
        filial_id = request.args.get('filial_id', type=int)
        turno = (request.args.get('turno') or '').strip().lower() or None

        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para consultar esta base.'}), 403
        if turno and turno not in LOADING_SHIFT_OPTIONS:
            return jsonify({'error': 'Turno informado é inválido.'}), 400

        try:
            if not loading_tables_ready():
                return jsonify({
                    'database_ready': False,
                    'date': target_date.isoformat(),
                    'items': [],
                })

            return jsonify({
                'database_ready': True,
                'date': target_date.isoformat(),
                'items': list_loading_journeys(profile, target_date, filial_id, turno),
            })
        except Exception as exc:
            app.logger.error('Erro ao listar jornadas de carregamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/carregamento/jornadas')
    @require_auth
    def loading_create_journey(profile):
        scope_error = require_scope_permission(profile, 'manage.programacao_carregamento', 'Sem permissão para programar carregamento.')
        if scope_error:
            return scope_error

        if not loading_tables_ready():
            return jsonify({'error': 'As tabelas do módulo de carregamento ainda não existem. Rode a migration primeiro.'}), 400

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        target_date = parse_iso_date(payload.get('data_operacao'))
        filial_id = payload.get('filial_id')
        veiculo_carregamento_id = payload.get('veiculo_carregamento_id')
        veiculo_carregamento_ids = payload.get('veiculo_carregamento_ids') or []
        rota_id = payload.get('rota_id')
        turno = (payload.get('turno') or '').strip().lower()

        normalized_vehicle_ids = []
        if isinstance(veiculo_carregamento_ids, list):
            normalized_vehicle_ids.extend([
                int(vehicle_id)
                for vehicle_id in veiculo_carregamento_ids
                if str(vehicle_id).isdigit()
            ])
        if veiculo_carregamento_id and str(veiculo_carregamento_id).isdigit():
            normalized_vehicle_ids.append(int(veiculo_carregamento_id))
        normalized_vehicle_ids = sorted(set(normalized_vehicle_ids))

        if not target_date:
            return jsonify({'error': 'Informe uma data válida para a jornada.'}), 400
        if not filial_id or not normalized_vehicle_ids:
            return jsonify({'error': 'Filial e pelo menos um caminhão são obrigatórios para programar o turno.'}), 400
        if turno not in LOADING_SHIFT_OPTIONS:
            return jsonify({'error': 'Turno informado é inválido.'}), 400
        if not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para programar esta base.'}), 403

        vehicle_response = (
            supabase.table('veiculos_carregamento')
            .select('*')
            .in_('id', normalized_vehicle_ids)
            .execute()
        )
        vehicles = vehicle_response.data or []
        if len(vehicles) != len(normalized_vehicle_ids):
            return jsonify({'error': 'Um ou mais veículos de carregamento não foram encontrados.'}), 404

        vehicles_by_id = {int(item['id']): item for item in vehicles}
        invalid_filial_vehicle = next(
            (item for item in vehicles if int(item['filial_id']) != int(filial_id)),
            None,
        )
        if invalid_filial_vehicle:
            return jsonify({'error': f"O veículo {invalid_filial_vehicle.get('placa') or invalid_filial_vehicle['id']} pertence a outra base."}), 400

        route_ids_to_load = []
        if rota_id and str(rota_id).isdigit():
            route_ids_to_load.append(int(rota_id))
        route_ids_to_load.extend([
            int(vehicle['rota_id'])
            for vehicle in vehicles
            if vehicle.get('rota_id') and str(vehicle.get('rota_id')).isdigit()
        ])
        route_ids_to_load = sorted(set(route_ids_to_load))

        route_map = {}
        if route_ids_to_load:
            route_response = (
                supabase.table('rotas_carregamento')
                .select('*')
                .in_('id', route_ids_to_load)
                .execute()
            )
            route_map = {int(item['id']): item for item in (route_response.data or [])}

        if rota_id and str(rota_id).isdigit():
            selected_route = route_map.get(int(rota_id))
            if not selected_route:
                return jsonify({'error': 'Referência operacional não encontrada.'}), 404
            if int(selected_route['filial_id']) != int(filial_id):
                return jsonify({'error': 'A referência operacional selecionada pertence a outra base.'}), 400

        existing_rows = (
            supabase.table('jornadas_carregamento')
            .select('veiculo_carregamento_id')
            .eq('data_operacao', target_date.isoformat())
            .eq('turno', turno)
            .in_('veiculo_carregamento_id', normalized_vehicle_ids)
            .execute()
        ).data or []
        existing_vehicle_ids = {
            int(item['veiculo_carregamento_id'])
            for item in existing_rows
            if item.get('veiculo_carregamento_id') is not None
        }

        rows_to_insert = []
        skipped_vehicles = []
        for vehicle_id in normalized_vehicle_ids:
            if vehicle_id in existing_vehicle_ids:
                skipped_vehicles.append(vehicles_by_id[vehicle_id].get('placa') or str(vehicle_id))
                continue

            vehicle = vehicles_by_id[vehicle_id]
            resolved_rota_id = int(rota_id) if rota_id and str(rota_id).isdigit() else vehicle.get('rota_id')
            if not resolved_rota_id:
                return jsonify({'error': f"Defina uma referência para o veículo {vehicle.get('placa') or vehicle_id} antes de abrir a jornada."}), 400

            resolved_route = route_map.get(int(resolved_rota_id))
            if not resolved_route:
                return jsonify({'error': f"A referência do veículo {vehicle.get('placa') or vehicle_id} não foi encontrada."}), 404
            if int(resolved_route['filial_id']) != int(filial_id):
                return jsonify({'error': f"A referência do veículo {vehicle.get('placa') or vehicle_id} pertence a outra base."}), 400

            rows_to_insert.append({
                'data_operacao': target_date.isoformat(),
                'filial_id': filial_id,
                'turno': turno,
                'veiculo_carregamento_id': vehicle_id,
                'rota_id': int(resolved_rota_id),
                'status': 'planejado',
                'observacao_abertura': payload.get('observacao_abertura') or None,
                'lider_colaborador_id': profile.get('id') or None,
                'aberto_por': profile['user_id'],
            })

        try:
            inserted_rows = []
            if rows_to_insert:
                response = (
                    supabase.table('jornadas_carregamento')
                    .insert(rows_to_insert)
                    .execute()
                )
                inserted_rows = response.data or []

            return jsonify({
                'created_count': len(inserted_rows),
                'skipped_count': len(skipped_vehicles),
                'skipped_vehicles': skipped_vehicles,
                'items': inserted_rows,
            }), 201
        except Exception as exc:
            app.logger.error('Erro ao criar jornada de carregamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.post('/api/carregamento/jornadas/<int:journey_id>/eventos')
    @require_auth
    def loading_register_event(profile, journey_id):
        scope_error = require_scope_permission(profile, 'manage.operacao_carregamento', 'Sem permissão para registrar eventos de carregamento.')
        if scope_error:
            return scope_error

        if not loading_tables_ready():
            return jsonify({'error': 'As tabelas do módulo de carregamento ainda não existem. Rode a migration primeiro.'}), 400

        journey, journey_error = ensure_loading_journey_allowed(profile, journey_id)
        if journey_error:
            return journey_error
        if journey.get('status') == 'finalizado':
            return jsonify({'error': 'A jornada já foi finalizada e não aceita novos eventos.'}), 400

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        action = (payload.get('action') or '').strip().lower()
        observacao = (payload.get('observacao') or '').strip()
        timestamp_now = now_iso()
        event_rows = (
            supabase.table('jornada_carregamento_eventos')
            .select('*')
            .eq('jornada_id', journey_id)
            .order('inicio_evento')
            .execute()
        ).data or []
        open_carga = find_open_loading_event(event_rows, 'carga')
        open_parada = find_open_loading_event(event_rows, 'parada')

        try:
            if action == 'iniciar_carga':
                if open_carga:
                    return jsonify({'error': 'Já existe uma carga em andamento para este caminhão.'}), 400
                supabase.table('jornada_carregamento_eventos').insert({
                    'jornada_id': journey_id,
                    'tipo_evento': 'carga',
                    'inicio_evento': timestamp_now,
                    'observacao': observacao or None,
                    'registrado_por': profile['user_id'],
                    'registrado_em': timestamp_now,
                }).execute()
                update_payload = {
                    'status': 'em_operacao',
                    'lider_colaborador_id': profile.get('id') or journey.get('lider_colaborador_id'),
                }
                if not journey.get('iniciado_em'):
                    update_payload['iniciado_em'] = timestamp_now
                supabase.table('jornadas_carregamento').update(update_payload).eq('id', journey_id).execute()

            elif action == 'encerrar_carga':
                if not open_carga:
                    return jsonify({'error': 'Não existe carga em andamento para encerrar.'}), 400
                supabase.table('jornada_carregamento_eventos').update({
                    'fim_evento': timestamp_now,
                    'observacao': observacao or open_carga.get('observacao'),
                }).eq('id', open_carga['id']).execute()

            elif action == 'iniciar_parada':
                if not open_carga:
                    return jsonify({'error': 'Inicie a carga antes de registrar uma parada.'}), 400
                if open_parada:
                    return jsonify({'error': 'Já existe uma parada em andamento.'}), 400

                motivo_parada_id = payload.get('motivo_parada_id')
                if not motivo_parada_id:
                    return jsonify({'error': 'Selecione o motivo da parada.'}), 400

                motive_response = (
                    supabase.table('motivos_parada_carregamento')
                    .select('*')
                    .eq('id', motivo_parada_id)
                    .limit(1)
                    .execute()
                )
                if not motive_response.data:
                    return jsonify({'error': 'Motivo de parada não encontrado.'}), 404
                motive = motive_response.data[0]
                if motive.get('exige_observacao') and not observacao:
                    return jsonify({'error': 'Este motivo exige observação detalhada.'}), 400

                supabase.table('jornada_carregamento_eventos').insert({
                    'jornada_id': journey_id,
                    'tipo_evento': 'parada',
                    'inicio_evento': timestamp_now,
                    'motivo_parada_id': motivo_parada_id,
                    'observacao': observacao or None,
                    'registrado_por': profile['user_id'],
                    'registrado_em': timestamp_now,
                }).execute()

            elif action == 'encerrar_parada':
                if not open_parada:
                    return jsonify({'error': 'Não existe parada em andamento para encerrar.'}), 400
                supabase.table('jornada_carregamento_eventos').update({
                    'fim_evento': timestamp_now,
                    'observacao': observacao or open_parada.get('observacao'),
                }).eq('id', open_parada['id']).execute()

            elif action == 'registrar_ocorrencia':
                if not observacao:
                    return jsonify({'error': 'Informe a ocorrência observada.'}), 400
                supabase.table('jornada_carregamento_eventos').insert({
                    'jornada_id': journey_id,
                    'tipo_evento': 'ocorrencia',
                    'inicio_evento': timestamp_now,
                    'fim_evento': timestamp_now,
                    'observacao': observacao,
                    'registrado_por': profile['user_id'],
                    'registrado_em': timestamp_now,
                }).execute()
            else:
                return jsonify({'error': 'Ação de evento inválida.'}), 400

            refreshed, _ = ensure_loading_journey_allowed(profile, journey_id)
            items = list_loading_journeys(profile, parse_iso_date(refreshed['data_operacao']), refreshed['filial_id'], refreshed['turno'])
            current_item = next((item for item in items if int(item['id']) == int(journey_id)), None)
            return jsonify(current_item or {'status': 'ok'})
        except Exception as exc:
            app.logger.error('Erro ao registrar evento de carregamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 400

    @app.put('/api/carregamento/jornadas/<int:journey_id>/fechamento')
    @require_auth
    def loading_close_journey(profile, journey_id):
        scope_error = require_scope_permission(profile, 'manage.operacao_carregamento', 'Sem permissão para finalizar o carregamento.')
        if scope_error:
            return scope_error

        if not loading_tables_ready():
            return jsonify({'error': 'As tabelas do módulo de carregamento ainda não existem. Rode a migration primeiro.'}), 400

        journey, journey_error = ensure_loading_journey_allowed(profile, journey_id)
        if journey_error:
            return journey_error
        if journey.get('status') == 'finalizado':
            return jsonify({'error': 'A jornada já foi finalizada.'}), 400

        payload = request.get_json(silent=True)
        if not isinstance(payload, dict):
            payload = {}

        quantidade_cilindros = payload.get('quantidade_cilindros')
        if quantidade_cilindros in (None, ''):
            return jsonify({'error': 'Informe a quantidade de cilindros no fechamento.'}), 400

        try:
            quantidade_cilindros = int(quantidade_cilindros)
        except (TypeError, ValueError):
            return jsonify({'error': 'Quantidade de cilindros inválida.'}), 400

        event_rows = (
            supabase.table('jornada_carregamento_eventos')
            .select('*')
            .eq('jornada_id', journey_id)
            .order('inicio_evento')
            .execute()
        ).data or []
        if find_open_loading_event(event_rows, 'carga') or find_open_loading_event(event_rows, 'parada'):
            return jsonify({'error': 'Encerre a carga e as paradas em aberto antes de finalizar a jornada.'}), 400

        timestamp_now = now_iso()
        closure_payload = {
            'jornada_id': journey_id,
            'quantidade_cilindros': quantidade_cilindros,
            'divergencias': payload.get('divergencias') or None,
            'observacao_fechamento': payload.get('observacao_fechamento') or None,
            'finalizado_por': profile['user_id'],
            'finalizado_em': timestamp_now,
        }

        try:
            supabase.table('jornada_carregamento_fechamentos').upsert(closure_payload, on_conflict='jornada_id').execute()
            supabase.table('jornadas_carregamento').update({
                'status': 'finalizado',
                'finalizado_em': timestamp_now,
                'lider_colaborador_id': profile.get('id') or journey.get('lider_colaborador_id'),
            }).eq('id', journey_id).execute()

            items = list_loading_journeys(profile, parse_iso_date(journey['data_operacao']), journey['filial_id'], journey['turno'])
            current_item = next((item for item in items if int(item['id']) == int(journey_id)), None)
            return jsonify(current_item or {'status': 'ok'})
        except Exception as exc:
            app.logger.error('Erro ao finalizar jornada de carregamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 400

    # ─── Gestão de Acessos (somente super admin) ─────────────────────────────

    @app.get('/api/admin/acessos')
    @require_auth
    def admin_listar_acessos(profile):
        """Lista colaboradores com acesso ao sistema + dados do Auth. Apenas super admin."""
        if not profile.get('is_super_admin'):
            return jsonify({'error': 'Apenas o administrador master pode acessar esta área.'}), 403

        try:
            # Busca todos os colaboradores com user_id vinculado
            rows = (
                supabase.table('colaboradores')
                .select('id, nome_completo, cargo, filial_id, user_id, tipo_acesso, permissao_app, permissao_desktop, ativo')
                .not_.is_('user_id', 'null')
                .order('nome_completo')
                .execute()
                .data or []
            )

            # Busca filiais para enriquecer
            filiais_map = {}
            try:
                filiais_rows = supabase.table('filiais').select('id, cidade, uf').execute().data or []
                filiais_map = {f['id']: f for f in filiais_rows}
            except Exception:
                pass

            # Busca dados do Auth por colaborador (get_user_by_id é mais confiável que list_users)
            auth_users_map = {}
            for _row in rows:
                _uid = str(_row.get('user_id') or '').strip()
                if not _uid:
                    continue
                if _uid in auth_users_map:
                    continue
                try:
                    _res = supabase.auth.admin.get_user_by_id(_uid)
                    _u = getattr(_res, 'user', None) or _res
                    if _u and getattr(_u, 'id', None):
                        auth_users_map[_uid] = _u
                except Exception as _e:
                    app.logger.debug('admin_listar_acessos: get_user_by_id(%s) falhou: %s', _uid, _e)

            result = []
            for row in rows:
                uid = str(row.get('user_id') or '')
                auth_user = auth_users_map.get(uid)
                filial = filiais_map.get(row.get('filial_id')) or {}
                result.append({
                    'id': row['id'],
                    'nome_completo': row.get('nome_completo') or '',
                    'cargo': row.get('cargo') or '',
                    'filial_label': f"{filial.get('cidade', '?')}/{filial.get('uf', '?')}" if filial else '—',
                    'filial_id': row.get('filial_id'),
                    'user_id': uid,
                    'email': (getattr(auth_user, 'email', None) if auth_user else None) or row.get('email') or None,
                    'tipo_acesso': row.get('tipo_acesso') or '',
                    'permissao_app': bool(row.get('permissao_app')),
                    'permissao_desktop': bool(row.get('permissao_desktop')),
                    'ativo': bool(row.get('ativo')),
                    'ultimo_login': getattr(auth_user, 'last_sign_in_at', None) if auth_user else None,
                    'criado_em': getattr(auth_user, 'created_at', None) if auth_user else None,
                })

            return jsonify(result)

        except Exception as exc:
            app.logger.error('Erro ao listar acessos: %s', exc)
            return jsonify({'error': 'Falha ao listar acessos.'}), 500

    @app.post('/api/admin/resetar-senha/<int:colaborador_id>')
    @rate_limit_endpoint(max_requests=10)
    @require_auth
    def admin_resetar_senha(profile, colaborador_id):
        """Redefine a senha de um colaborador. Apenas super admin."""
        if not profile.get('is_super_admin'):
            return jsonify({'error': 'Apenas o administrador master pode redefinir senhas.'}), 403

        body = request.get_json(silent=True) or {}
        nova_senha = body.get('nova_senha', '')
        if not nova_senha or len(nova_senha) < 10:
            return jsonify({'error': 'A senha deve ter ao menos 10 caracteres.'}), 400
        if len(nova_senha) > 128:
            return jsonify({'error': 'A senha não pode exceder 128 caracteres.'}), 400

        import re as _re
        if _re.search(r'[\x00-\x1f\x7f]', nova_senha):
            return jsonify({'error': 'A senha contém caracteres inválidos.'}), 400
        if not _re.search(r'[A-Za-z]', nova_senha) or not _re.search(r'[0-9\W_]', nova_senha):
            return jsonify({'error': 'A senha deve conter ao menos uma letra e um número ou símbolo.'}), 400

        try:
            colab_row = (
                supabase.table('colaboradores')
                .select('id, nome_completo, user_id, filial_id')
                .eq('id', colaborador_id)
                .limit(1)
                .execute()
                .data or []
            )
            if not colab_row:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404

            uid = colab_row[0].get('user_id')
            if not uid:
                return jsonify({'error': 'Este colaborador não tem acesso ao sistema. Vincule um e-mail antes.'}), 400

            supabase.auth.admin.update_user_by_id(str(uid), {'password': nova_senha})

            write_audit_event(
                profile, 'update', 'colaboradores', colaborador_id,
                details={'acao': 'reset_senha', 'nome': colab_row[0].get('nome_completo', '')},
                filial_id=colab_row[0].get('filial_id'),
            )
            return jsonify({'status': 'ok', 'mensagem': 'Senha redefinida com sucesso.'})

        except Exception as exc:
            app.logger.error('Erro ao redefinir senha do colaborador %s: %s', colaborador_id, exc)
            return jsonify({'error': 'Falha ao redefinir senha. Verifique a configuração do Supabase.'}), 500

    @app.post('/api/admin/atualizar-email/<int:colaborador_id>')
    @rate_limit_endpoint(max_requests=10)
    @require_auth
    def admin_atualizar_email(profile, colaborador_id):
        """Altera o e-mail de login de um colaborador. Apenas super admin."""
        if not profile.get('is_super_admin'):
            return jsonify({'error': 'Apenas o administrador master pode alterar e-mails.'}), 403

        body = request.get_json(silent=True) or {}
        novo_email = (body.get('novo_email') or '').strip().lower()

        if not novo_email:
            return jsonify({'error': 'Informe o novo e-mail.'}), 400

        import re as _re
        if not _re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', novo_email):
            return jsonify({'error': 'E-mail inválido.'}), 400

        if len(novo_email) > 254:
            return jsonify({'error': 'E-mail muito longo.'}), 400

        try:
            colab_row = (
                supabase.table('colaboradores')
                .select('id, nome_completo, user_id, filial_id')
                .eq('id', colaborador_id)
                .limit(1)
                .execute()
                .data or []
            )
            if not colab_row:
                return jsonify({'error': 'Colaborador não encontrado.'}), 404

            uid = colab_row[0].get('user_id')
            if not uid:
                return jsonify({'error': 'Este colaborador não tem acesso ao sistema.'}), 400

            supabase.auth.admin.update_user_by_id(str(uid), {'email': novo_email})

            write_audit_event(
                profile, 'update', 'colaboradores', colaborador_id,
                details={'acao': 'atualizar_email', 'nome': colab_row[0].get('nome_completo', '')},
                filial_id=colab_row[0].get('filial_id'),
            )
            return jsonify({'status': 'ok', 'mensagem': 'E-mail atualizado com sucesso.'})

        except Exception as exc:
            app.logger.error('Erro ao atualizar e-mail do colaborador %s: %s', colaborador_id, exc)
            return jsonify({'error': 'Falha ao atualizar e-mail. Verifique a configuração do Supabase.'}), 500

    # ─── Rotas de Feriados ────────────────────────────────────────────────────

    @app.get('/api/feriados/calendario')
    @require_auth
    def feriados_calendario(profile):
        """Retorna feriados num intervalo de datas, expandindo recorrentes."""
        scope_error = require_scope_permission(profile, 'menu.feriados')
        if scope_error:
            return scope_error

        ano = request.args.get('ano')
        mes = request.args.get('mes')
        uf = request.args.get('uf', '').strip().upper() or None
        filial_id = request.args.get('filial_id')

        if not ano:
            ano = str(date_class.today().year)
        if not mes:
            mes = str(date_class.today().month)

        try:
            ano_int = int(ano)
            mes_int = int(mes)
            data_ini = date_class(ano_int, mes_int, 1)
            if mes_int == 12:
                data_fim = date_class(ano_int + 1, 1, 1) - timedelta(days=1)
            else:
                data_fim = date_class(ano_int, mes_int + 1, 1) - timedelta(days=1)
        except (ValueError, TypeError):
            return jsonify({'error': 'Ano ou mês inválido.'}), 400

        try:
            query = (
                supabase.table('feriados')
                .select('*')
                .eq('ativo', True)
            )
            rows = query.execute().data or []
        except Exception as exc:
            app.logger.error('Erro ao buscar feriados: %s', exc)
            return jsonify({'error': 'Falha ao consultar feriados.'}), 500

        resultado = []
        for row in rows:
            # Filtra por UF se informado
            row_uf = (row.get('uf') or '').strip().upper()
            row_filial = row.get('filial_id')
            if uf and row_uf and row_uf != uf:
                continue
            if filial_id:
                # Inclui: nacionais (sem filial_id e sem uf), estaduais com UF correspondente, ou da filial exata
                if row_filial and str(row_filial) != str(filial_id):
                    continue

            if row.get('recorrente'):
                # Apenas dia e mês importam
                try:
                    data_base = date_class.fromisoformat(row['data'])
                    data_expandida = date_class(ano_int, data_base.month, data_base.day)
                    if data_ini <= data_expandida <= data_fim:
                        resultado.append({**row, 'data': data_expandida.isoformat(), 'recorrente_expandido': True})
                except (ValueError, KeyError):
                    pass
            else:
                try:
                    data_row = date_class.fromisoformat(row['data'])
                    if data_ini <= data_row <= data_fim:
                        resultado.append(row)
                except (ValueError, KeyError):
                    pass

        resultado.sort(key=lambda r: r.get('data', ''))
        return jsonify(resultado)

    # ─── Rotas de Notas / CT-e ────────────────────────────────────────────────

    @app.get('/api/notas_cte/resumo')
    @require_auth
    def notas_cte_resumo(profile):
        """Retorna totais por status para os cards de resumo."""
        scope_error = require_scope_permission(profile, 'menu.notas_cte')
        if scope_error:
            return scope_error

        filial_id = request.args.get('filial_id')
        try:
            query = supabase.table('notas_cte').select('status, valor_total, data_vencimento, filial_id').eq('ativo', True)
            if filial_id:
                query = query.eq('filial_id', int(filial_id))
            else:
                resource_config = RESOURCE_DEFINITIONS.get('notas_cte')
                if resource_config:
                    query = apply_filial_scope(query, profile, resource_config)
            rows = query.execute().data or []
        except Exception as exc:
            app.logger.error('Erro ao buscar resumo de notas: %s', exc)
            return jsonify({'error': 'Falha ao consultar resumo.'}), 500

        hoje = date_class.today().isoformat()
        totais = {'pendente': 0.0, 'vencido': 0.0, 'pago': 0.0, 'cancelado': 0.0, 'total_pendente_vencido': 0.0}
        contagem = {'pendente': 0, 'vencido': 0, 'pago': 0, 'cancelado': 0}

        for row in rows:
            status = row.get('status', 'pendente')
            valor = parse_float_or_default(row.get('valor_total'), 0.0)
            venc = row.get('data_vencimento') or ''
            # Classifica como vencido se pendente e data_vencimento < hoje
            if status == 'pendente' and venc and venc < hoje:
                status = 'vencido'
            totais[status] = totais.get(status, 0.0) + valor
            contagem[status] = contagem.get(status, 0) + 1

        totais['total_pendente_vencido'] = round(totais['pendente'] + totais['vencido'], 2)
        for k in totais:
            totais[k] = round(totais[k], 2)

        return jsonify({'totais': totais, 'contagem': contagem})

    @app.patch('/api/notas_cte/<int:nota_id>/status')
    @require_auth
    def notas_cte_atualizar_status(profile, nota_id):
        """Endpoint rápido para marcar nota como paga, cancelada etc."""
        scope_error = require_scope_permission(profile, 'menu.notas_cte')
        if scope_error:
            return scope_error

        NOTAS_STATUS_OPTIONS = {'pendente', 'pago', 'vencido', 'cancelado'}
        body = request.get_json(silent=True) or {}
        novo_status = (body.get('status') or '').strip().lower()
        if not novo_status or novo_status not in NOTAS_STATUS_OPTIONS:
            return jsonify({'error': f'Status inválido. Use: {", ".join(sorted(NOTAS_STATUS_OPTIONS))}.'}), 400

        data_pagamento = body.get('data_pagamento') or None
        try:
            nota_row = (
                supabase.table('notas_cte')
                .select('id, filial_id')
                .eq('id', nota_id)
                .limit(1)
                .execute()
            ).data or []
            if not nota_row:
                return jsonify({'error': 'Nota não encontrada.'}), 404
            if not ensure_profile_can_access_filial(profile, nota_row[0].get('filial_id')):
                return jsonify({'error': 'Sem permissão para esta base.'}), 403

            update_payload = {'status': novo_status}
            if novo_status == 'pago' and data_pagamento:
                update_payload['data_pagamento'] = data_pagamento
            elif novo_status == 'pago' and not data_pagamento:
                update_payload['data_pagamento'] = date_class.today().isoformat()

            supabase.table('notas_cte').update(update_payload).eq('id', nota_id).execute()
            write_audit_event(profile, 'update', 'notas_cte', nota_id, details={'status': novo_status})
            return jsonify({'status': 'ok', 'novo_status': novo_status})
        except Exception as exc:
            app.logger.error('Erro ao atualizar status nota %s: %s', nota_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ═══════════════════════════════════════════════════════════════════════════════
    # ENDPOINTS GENÉRICOS DE APROVAÇÃO E ACOMPANHAMENTO
    # Sistema de approval workflow para múltiplos processos de negócio
    # ═══════════════════════════════════════════════════════════════════════════════

    def _get_approval_configs():
        """
        Lê configurações de workflow de aprovação do banco (approval_workflow_configs).
        Fallback para valores hardcoded se a tabela não existir ou falhar.
        Retorna dict keyed por resource_type com chaves no formato esperado pelos endpoints.
        """
        _FALLBACK = {
            'manutencoes': {
                'table': 'manutencoes', 'status_field': 'status',
                'pending_statuses': ['aguardando_aprovacao'],
                'approved_status': 'aprovada', 'rejected_status': 'reprovada',
                'check_scope': 'menu.manutencoes', 'permission': 'aprovar.manutencoes',
                'approval_fields': ['aprovado_por', 'aprovado_em', 'reprovado_por', 'reprovado_em', 'motivo_reprovacao'],
                'approve_update': lambda p: {'aprovado_por': p.get('id'), 'aprovado_em': now_iso()},
                'reject_update': lambda p, m: {'reprovado_por': p.get('id'), 'reprovado_em': now_iso(), 'motivo_reprovacao': m},
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Manutenção de Frota',
            },
            'pedidos_compra': {
                'table': 'pedidos_compra', 'status_field': 'status',
                'pending_statuses': ['pendente', 'analise', 'pendente_aprovacao', 'em_analise'],
                'analise_statuses': ['analise', 'em_analise'],
                'approved_status': 'aprovado', 'rejected_status': 'reprovado',
                'check_scope': 'menu.pedidos_compra', 'permission': 'aprovar.pedidos_compra',
                'analise_permission': 'analisar.pedidos_compra',
                'approval_fields': ['criado_por', 'numero_pedido', 'fornecedor', 'valor_total',
                                    'forma_pagamento', 'tipo_reembolso', 'chave_pix', 'dados_bancarios',
                                    'aprovado_por', 'aprovado_em', 'em_analise_por', 'em_analise_em',
                                    'motivo_reprovacao', 'reprovado_por', 'reprovado_em', 'numero_solicitacao'],
                'approve_update': lambda p: {
                    'aprovado_por': p.get('id'),
                    'aprovado_em': now_iso(),
                },
                'reject_update': lambda p, m: {
                    'reprovado_por': p.get('id'),
                    'reprovado_em': now_iso(),
                    'motivo_reprovacao': m,
                },
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Pedidos de Compra',
            },
            'horas_extras': {
                'table': 'horas_extras', 'status_field': 'status',
                'pending_statuses': ['solicitado'],
                'approved_status': 'aprovado', 'rejected_status': 'reprovado',
                'check_scope': 'menu.horas_extras', 'permission': 'aprovar.horas_extras',
                'approval_fields': ['data_aprovacao', 'aprovado_por', 'justificativa_gestor'],
                'approve_update': lambda p: {'aprovado_por': p.get('id'), 'data_aprovacao': now_iso()},
                'reject_update': lambda p, m: {'reprovado_por': p.get('id'), 'justificativa_gestor': m},
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Horas Extras',
            },
            'abastecimentos': {
                'table': 'veiculos_abastecimentos', 'status_field': 'status',
                'pending_statuses': ['pendente_aprovacao'],
                'approved_status': 'aprovado', 'rejected_status': 'reprovado',
                'check_scope': 'menu.abastecimentos', 'permission': 'aprovar.abastecimentos',
                'approval_fields': [],
                'approve_update': lambda p: {},
                'reject_update': lambda p, m: {},
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Abastecimentos',
            },
            'pneus': {
                'table': 'veiculos_pneus', 'status_field': 'status_aprovacao',
                'pending_statuses': ['pendente_aprovacao'],
                'approved_status': 'aprovado', 'rejected_status': 'reprovado',
                'check_scope': 'menu.pneus', 'permission': 'aprovar.pneus',
                'approval_fields': [],
                'approve_update': lambda p: {},
                'reject_update': lambda p, m: {},
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Controle de Pneus',
            },
            'diarias_solicitacoes': {
                'table': 'diarias_solicitacoes', 'status_field': 'status',
                'pending_statuses': ['pendente'],
                'approved_status': 'aprovado', 'rejected_status': 'reprovado',
                'check_scope': 'menu.diarias', 'permission': 'aprovar.diarias',
                'approval_fields': ['cidade_destino', 'uf_destino', 'data_inicio', 'data_fim', 'rota', 'valor_total', 'banco'],
                'approve_update': lambda p: {},
                'reject_update': lambda p, m: {'observacoes': m},
                'require_comment_on_approve': False, 'require_comment_on_reject': True, 'ativo': True,
                '_label': 'Diárias e Hotelaria',
            },
        }
        try:
            resp = supabase.table('approval_workflow_configs').select('*').execute()
            db_rows = {r['resource_type']: r for r in (resp.data or [])}
            if not db_rows:
                return _FALLBACK
            result = {}
            for rt, fallback in _FALLBACK.items():
                db = db_rows.get(rt, {})
                result[rt] = {
                    **fallback,
                    'table': db.get('tabela', fallback['table']),
                    'status_field': db.get('status_field', fallback['status_field']),
                    'pending_statuses': db.get('pending_statuses') or fallback['pending_statuses'],
                    'approved_status': db.get('approved_status', fallback['approved_status']),
                    'rejected_status': db.get('rejected_status', fallback['rejected_status']),
                    'check_scope': db.get('view_scope', fallback['check_scope']),
                    'permission': db.get('approval_scope', fallback['permission']),
                    'require_comment_on_approve': db.get('require_comment_on_approve', fallback['require_comment_on_approve']),
                    'require_comment_on_reject': db.get('require_comment_on_reject', fallback['require_comment_on_reject']),
                    'ativo': db.get('ativo', fallback['ativo']),
                    '_label': db.get('label', fallback['_label']),
                }
            return result
        except Exception as exc:
            app.logger.warning('approval_workflow_configs lookup failed, using fallback: %s', exc)
            return _FALLBACK

    @app.get('/api/approvals')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def list_approvals(profile):
        """
        Lista solicitações pendentes de aprovação por tipo de processo.
        Suporta: manutencoes, pneus, pedidos_compra, horas_extras, combustivel
        Query params: 
          - resource_type: manutencoes|pneus|pedidos_compra|horas_extras|combustivel (opcional, lista todas se não informado)
          - status: pendente|aprovado|reprovado (opcional)
          - filial_id: int (opcional, respeita escopo do perfil)
          - limit: int (default 100, max 500)
        """
        SUPPORTED_RESOURCES = _get_approval_configs()
        
        resource_type = request.args.get('resource_type', '').strip().lower()
        status_filter = request.args.get('status', '').strip().lower()
        filial_id = request.args.get('filial_id')
        limit = min(int(request.args.get('limit', 100)), 500)
        
        try:
            filial_id = int(filial_id) if filial_id and str(filial_id).isdigit() else None
        except (ValueError, TypeError):
            filial_id = None
        
        if filial_id and not ensure_profile_can_access_filial(profile, filial_id):
            return jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403
        
        results = []
        resources_to_check = (
            [resource_type] if resource_type and resource_type in SUPPORTED_RESOURCES else 
            list(SUPPORTED_RESOURCES.keys())
        )
        
        for res_type in resources_to_check:
            config = SUPPORTED_RESOURCES[res_type]

            # Pular processos desativados
            if not config.get('ativo', True):
                continue

            # Verificar permissão de visualização
            if not profile_has_scope_permission(profile, config['check_scope']):
                continue
            
            try:
                query = supabase.table(config['table']).select('*')

                sf = config['status_field']
                pending = config['pending_statuses']

                def _apply_pending(q):
                    if len(pending) == 1:
                        return q.eq(sf, pending[0])
                    return q.in_(sf, pending)

                if status_filter == 'pendente':
                    query = _apply_pending(query)
                elif status_filter == 'aprovado':
                    query = query.eq(sf, config['approved_status'])
                elif status_filter == 'reprovado':
                    query = query.eq(sf, config['rejected_status'])
                elif status_filter == 'all':
                    pass  # sem filtro de status — retorna tudo
                else:
                    query = _apply_pending(query)
                
                # Aplicar escopo de filial
                if filial_id:
                    query = query.eq('filial_id', filial_id)
                elif profile_has_filial_scope(profile):
                    query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])
                
                query = query.order('id', desc=True).limit(limit)
                response = query.execute()
                
                for item in (response.data or []):
                    if res_type == 'abastecimentos':
                        litros = item.get('litros') or ''
                        combustivel = item.get('tipo_combustivel') or 'combustível'
                        titulo_item = f"{litros}L de {combustivel} – Veículo #{item.get('veiculo_id', '')}"
                    elif res_type == 'pneus':
                        titulo_item = f"Pneu {item.get('posicao', '')} – Veículo #{item.get('veiculo_id', '')}"
                    else:
                        titulo_item = (
                            item.get('titulo') or item.get('numero_pedido') or
                            item.get('motivo') or f'{res_type}#{item.get("id")}'
                        )
                    results.append({
                        'resource_type': res_type,
                        'id': item.get('id'),
                        'status': item.get(config['status_field']),
                        'filial_id': item.get('filial_id'),
                        'data_criacao': (
                            item.get('data_abertura') or item.get('data_pedido') or
                            item.get('data_solicitacao') or item.get('data_abastecimento') or
                            item.get('data_instalacao')
                        ),
                        'titulo': titulo_item,
                        'detalhes': {k: item.get(k) for k in config['approval_fields'] if k in item},
                        'full_item': item,
                    })
            except Exception as exc:
                app.logger.warning('Erro ao listar %s: %s', res_type, exc)
                continue

        # Enriquece pedidos_compra com valor_total_calculado e itens
        pedido_results = [r for r in results if r['resource_type'] == 'pedidos_compra']
        if pedido_results:
            try:
                pedido_ids = [r['id'] for r in pedido_results if r.get('id')]
                if pedido_ids:
                    itens_raw = (
                        supabase.table('pedidos_compra_itens')
                        .select('id, pedido_id, categoria, descricao, quantidade, unidade, valor_unitario, observacoes')
                        .in_('pedido_id', pedido_ids)
                        .eq('ativo', True)
                        .order('id')
                        .execute()
                        .data or []
                    )
                    from collections import defaultdict
                    val_por_pedido   = defaultdict(float)
                    itens_por_pedido = defaultdict(list)
                    for it in itens_raw:
                        pid = it.get('pedido_id')
                        if pid:
                            val = float(it.get('quantidade') or 0) * float(it.get('valor_unitario') or 0)
                            val_por_pedido[pid] += val
                            itens_por_pedido[pid].append({**it, 'total_item': round(val, 2)})
                    for r in pedido_results:
                        pid = r['id']
                        r['full_item']['valor_total_calculado'] = round(val_por_pedido.get(pid, 0.0), 2)
                        r['full_item']['itens'] = itens_por_pedido.get(pid, [])
            except Exception as _exc:
                app.logger.warning('Enriquecimento itens/valor (approvals): %s', _exc)

        return jsonify({
            'total': len(results),
            'items': results,
            'supported_resources': list(SUPPORTED_RESOURCES.keys()),
        })

    @app.post('/api/approvals/<int:item_id>/approve')
    @rate_limit_endpoint(max_requests=30)
    @require_auth
    def approve_request(profile, item_id):
        """
        Aprova uma solicitação.
        Body params:
          - resource_type: manutencoes|pedidos_compra|horas_extras (obrigatório)
          - comentario: string (opcional, salvo em auditoria)
        """
        body = request.get_json(silent=True) or {}
        resource_type = (body.get('resource_type') or '').strip().lower()
        comentario = (body.get('comentario') or '').strip()
        
        SUPPORTED_RESOURCES = _get_approval_configs()

        if not resource_type or resource_type not in SUPPORTED_RESOURCES:
            return jsonify({'error': f'Tipo de recurso inválido: {resource_type}'}), 400

        config = SUPPORTED_RESOURCES[resource_type]

        # Verificar se comentário é obrigatório
        if config.get('require_comment_on_approve') and not comentario:
            return jsonify({'error': 'Comentário é obrigatório para aprovar este tipo de solicitação.'}), 400

        # Verificar permissão
        if not profile_has_scope_permission(profile, config['permission']):
            return jsonify({'error': 'Sem permissão para aprovar este tipo de solicitação.'}), 403
        
        try:
            # Buscar item
            item_query = supabase.table(config['table']).select('*').eq('id', item_id).limit(1)
            item_response = item_query.execute()
            
            if not item_response.data:
                return jsonify({'error': f'{resource_type}: Registro não encontrado.'}), 404
            
            item = item_response.data[0]
            
            # Verificar acesso à filial
            if not ensure_profile_can_access_filial(profile, item.get('filial_id')):
                return jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403

            # Para pedidos_compra: aprovar exige que o pedido esteja em analise
            if resource_type == 'pedidos_compra':
                analise_statuses = config.get('analise_statuses', ['analise', 'em_analise'])
                if item.get(config['status_field']) not in analise_statuses:
                    return jsonify({'error': 'Apenas pedidos em análise podem ser aprovados. Coloque o pedido em análise primeiro.'}), 400

            # Atualizar status (campos base + campos específicos do schema)
            update_data = {config['status_field']: config['approved_status'], **config['approve_update'](profile)}
            supabase.table(config['table']).update(update_data).eq('id', item_id).execute()

            # Usar número já gerado na criação; gerar só se ainda não existe (registros antigos)
            numero_solicitacao = item.get('numero_solicitacao')
            if not numero_solicitacao:
                try:
                    sol_resp = supabase.rpc('gerar_numero_solicitacao', {'p_tipo': resource_type}).execute()
                    numero_solicitacao = sol_resp.data
                    if numero_solicitacao:
                        supabase.table(config['table']).update(
                            {'numero_solicitacao': numero_solicitacao}
                        ).eq('id', item_id).execute()
                except Exception as sol_exc:
                    app.logger.warning(
                        'Falha ao gerar número de solicitação para %s %d: %s',
                        resource_type, item_id, sol_exc
                    )

            # ── Auto-cria contas_a_pagar quando pedido_compra é aprovado ──
            contas_pagar_id = None
            if resource_type == 'pedidos_compra':
                try:
                    valor = float(item.get('valor_total') or 0)
                    filial_id_item = item.get('filial_id')
                    data_pedido = item.get('data_pedido') or date_class.today().isoformat()
                    data_vencimento_pedido = (
                        item.get('data_vencimento') or
                        item.get('data_necessidade') or
                        data_pedido
                    )
                    # Busca nome da filial
                    filial_nome = ''
                    try:
                        f_resp = supabase.table('filiais').select('cidade, uf').eq('id', filial_id_item).limit(1).execute()
                        if f_resp.data:
                            fr = f_resp.data[0]
                            filial_nome = f"{fr.get('cidade','')}/{fr.get('uf','')}"
                    except Exception:
                        pass
                    prazo_info = item.get('prazo_pagamento') or ''
                    descricao_cp = f"Pedido {item.get('numero_pedido') or item_id} – {item.get('fornecedor') or 'Fornecedor não informado'}"
                    if prazo_info:
                        descricao_cp += f" ({prazo_info})"
                    cp_row = {
                        'filial_id': filial_id_item,
                        'filial_nome': filial_nome,
                        'competencia': data_pedido[:7],
                        'tipo_despesa': 'COMPRAS',
                        'descricao': descricao_cp,
                        'fornecedor_nome': item.get('fornecedor') or None,
                        'valor': valor,
                        'data_vencimento': data_vencimento_pedido,
                        'status': 'PENDENTE',
                        'numero_documento': item.get('numero_pedido') or None,
                        'observacoes': f"Gerado automaticamente ao aprovar pedido de compra #{item_id}",
                    }
                    cp_resp = supabase.table('contas_a_pagar').insert(cp_row).execute()
                    if cp_resp.data:
                        contas_pagar_id = cp_resp.data[0].get('id')
                        supabase.table('pedidos_compra').update(
                            {'contas_pagar_id': contas_pagar_id}
                        ).eq('id', item_id).execute()
                except Exception as cp_exc:
                    app.logger.warning('Falha ao criar contas_a_pagar para pedido %d: %s', item_id, cp_exc)

            # Registrar auditoria
            write_audit_event(
                profile,
                'approve',
                resource_type,
                item_id,
                status='ok',
                details={
                    'new_status': config['approved_status'],
                    'comentario': comentario,
                    'approved_by': profile.get('nome_completo'),
                    'numero_solicitacao': numero_solicitacao,
                    'contas_pagar_id': contas_pagar_id,
                }
            )

            return jsonify({
                'status': 'ok',
                'message': 'Solicitação aprovada com sucesso.',
                'resource_type': resource_type,
                'item_id': item_id,
                'numero_solicitacao': numero_solicitacao,
                'contas_pagar_id': contas_pagar_id,
            })
        except Exception as exc:
            app.logger.error('Erro ao aprovar %s %d: %s', resource_type, item_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/approvals/<int:item_id>/reject')
    @rate_limit_endpoint(max_requests=30)
    @require_auth
    def reject_request(profile, item_id):
        """
        Rejeita/reprova uma solicitação.
        Body params:
          - resource_type: manutencoes|pedidos_compra|horas_extras (obrigatório)
          - motivo: string (obrigatório para rejeição)
        """
        body = request.get_json(silent=True) or {}
        resource_type = (body.get('resource_type') or '').strip().lower()
        motivo = (body.get('motivo') or '').strip()

        SUPPORTED_RESOURCES = _get_approval_configs()

        if not resource_type or resource_type not in SUPPORTED_RESOURCES:
            return jsonify({'error': f'Tipo de recurso inválido: {resource_type}'}), 400

        config = SUPPORTED_RESOURCES[resource_type]

        # Verificar obrigatoriedade do motivo (configurável por processo)
        if config.get('require_comment_on_reject', True) and not motivo:
            return jsonify({'error': 'Informe o motivo da rejeição.'}), 400

        # Verificar permissão base (para não-pedidos_compra basta config['permission'])
        if resource_type != 'pedidos_compra':
            if not profile_has_scope_permission(profile, config['permission']):
                return jsonify({'error': 'Sem permissão para rejeitar este tipo de solicitação.'}), 403

        try:
            # Buscar item
            item_query = supabase.table(config['table']).select('*').eq('id', item_id).limit(1)
            item_response = item_query.execute()

            if not item_response.data:
                return jsonify({'error': f'{resource_type}: Registro não encontrado.'}), 404

            item = item_response.data[0]

            # Verificar acesso à filial
            if not ensure_profile_can_access_filial(profile, item.get('filial_id')):
                return jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403

            # Para pedidos_compra: verificar permissão por etapa
            if resource_type == 'pedidos_compra':
                current_status = item.get(config['status_field'], '')
                analise_statuses = config.get('analise_statuses', ['analise', 'em_analise'])
                analise_perm = config.get('analise_permission', 'analisar.pedidos_compra')
                if current_status in analise_statuses:
                    # etapa de aprovação — precisa de aprovar.pedidos_compra
                    if not profile_has_scope_permission(profile, config['permission']):
                        return jsonify({'error': 'Sem permissão para rejeitar pedidos em análise.'}), 403
                else:
                    # etapa de análise (pendente) — precisa de analisar.pedidos_compra
                    if not (profile_has_scope_permission(profile, analise_perm) or
                            profile_has_scope_permission(profile, config['permission'])):
                        return jsonify({'error': 'Sem permissão para rejeitar este pedido.'}), 403

            # Atualizar status (campos base + campos específicos do schema)
            update_data = {config['status_field']: config['rejected_status'], **config['reject_update'](profile, motivo)}
            supabase.table(config['table']).update(update_data).eq('id', item_id).execute()

            # Registrar auditoria
            write_audit_event(
                profile,
                'reject',
                resource_type,
                item_id,
                status='ok',
                details={
                    'new_status': config['rejected_status'],
                    'motivo': motivo,
                    'rejected_by': profile.get('nome_completo'),
                }
            )
            
            return jsonify({
                'status': 'ok',
                'message': f'Solicitação reprovada.',
                'resource_type': resource_type,
                'item_id': item_id,
            })
        except Exception as exc:
            app.logger.error('Erro ao rejeitar %s %d: %s', resource_type, item_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/approvals/history')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def approvals_history(profile):
        """
        Lista histórico de aprovações e rejeições.
        Query params:
          - resource_type: opcional, filtra por tipo
          - status_acao: approve|reject (opcional)
          - days: número de dias para voltar (default 30)
          - limit: max 500
        """
        resource_type = request.args.get('resource_type', '').strip().lower()
        status_acao = request.args.get('status_acao', '').strip().lower()
        days = max(1, min(365, int(request.args.get('days', 30))))
        limit = min(int(request.args.get('limit', 100)), 500)
        
        try:
            from_date = (datetime.now() - timedelta(days=days)).isoformat()
            
            query = (
                supabase.table('auditoria_movimentacoes')
                .select('*')
                .in_('acao', ['approve', 'reject'])
                .gte('criado_em', from_date)
                .order('criado_em', desc=True)
                .limit(limit)
            )
            
            if resource_type:
                query = query.eq('recurso', resource_type)
            
            if status_acao == 'approve':
                query = query.eq('acao', 'approve')
            elif status_acao == 'reject':
                query = query.eq('acao', 'reject')
            
            # Aplicar escopo de filial
            if profile_has_filial_scope(profile):
                query = query.in_('filial_id', profile.get('allowed_filial_ids') or [])
            
            response = query.execute()
            items = response.data or []
            
            return jsonify({
                'total': len(items),
                'period_days': days,
                'items': items,
            })
        except Exception as exc:
            app.logger.error('Erro ao carregar histórico de aprovações: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.post('/api/approvals/<int:item_id>/em-analise')
    @rate_limit_endpoint(max_requests=30)
    @require_auth
    def set_em_analise(profile, item_id):
        """Move pedido_compra de pendente para analise (analista inicia revisão)."""
        body = request.get_json(silent=True) or {}
        resource_type = (body.get('resource_type') or 'pedidos_compra').strip().lower()
        if resource_type != 'pedidos_compra':
            return jsonify({'error': 'Apenas pedidos_compra suportam análise.'}), 400

        SUPPORTED_RESOURCES = _get_approval_configs()
        config = SUPPORTED_RESOURCES.get(resource_type)
        if not config:
            return jsonify({'error': 'Recurso não suportado.'}), 400

        analise_perm = config.get('analise_permission', 'analisar.pedidos_compra')
        if not profile_has_scope_permission(profile, analise_perm):
            return jsonify({'error': 'Sem permissão para colocar pedidos em análise.'}), 403

        try:
            item_resp = supabase.table('pedidos_compra').select('*').eq('id', item_id).limit(1).execute()
            if not item_resp.data:
                return jsonify({'error': 'Pedido não encontrado.'}), 404
            item = item_resp.data[0]
            if not ensure_profile_can_access_filial(profile, item.get('filial_id')):
                return jsonify({'error': 'Sem permissão para acessar dados desta base.'}), 403
            allowed_from = ('pendente', 'pendente_aprovacao', 'analise', 'em_analise')
            if item.get('status') not in allowed_from:
                return jsonify({'error': f"Status atual '{item.get('status')}' não permite mover para análise."}), 400

            supabase.table('pedidos_compra').update({
                'status': 'analise',
                'em_analise_por': profile.get('id'),
                'em_analise_em': now_iso(),
            }).eq('id', item_id).execute()

            write_audit_event(profile, 'em_analise', 'pedidos_compra', item_id, status='ok',
                              details={'analista': profile.get('nome_completo')})

            return jsonify({'status': 'ok', 'message': 'Pedido enviado para análise.'})
        except Exception as exc:
            app.logger.error('set_em_analise %d: %s', item_id, exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.get('/api/pedidos-compra/metricas')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def pedidos_compra_metricas(profile):
        """Métricas agregadas de pedidos de compra para gráficos."""
        scope_error = require_scope_permission(profile, 'menu.pedidos_compra')
        if scope_error:
            return scope_error

        filial_id = request.args.get('filial_id', type=int)

        try:
            query = (
                supabase.table('pedidos_compra')
                .select('id, filial_id, data_pedido, status, forma_pagamento, fornecedor')
                .eq('ativo', True)
                .order('data_pedido', desc=True)
                .limit(3000)
            )
            query = apply_filial_scope(query, profile, RESOURCE_DEFINITIONS['pedidos_compra'])
            if filial_id and ensure_profile_can_access_filial(profile, filial_id):
                query = query.eq('filial_id', filial_id)
            pedidos = query.execute().data or []

            pedido_ids_set = {p['id'] for p in pedidos if p.get('id') is not None}

            # Itens para calcular valores reais e gasto por categoria
            itens = []
            if pedido_ids_set:
                itens_query = (
                    supabase.table('pedidos_compra_itens')
                    .select('pedido_id, categoria, quantidade, valor_unitario')
                    .eq('ativo', True)
                    .limit(10000)
                )
                itens_raw = itens_query.execute().data or []
                itens = [it for it in itens_raw if it.get('pedido_id') in pedido_ids_set]

            # Calcula valor real por pedido a partir dos itens
            from collections import defaultdict
            valor_por_pedido = defaultdict(float)
            for it in itens:
                pid = it.get('pedido_id')
                if pid:
                    valor_por_pedido[pid] += float(it.get('quantidade') or 0) * float(it.get('valor_unitario') or 0)

            # Nomes das filiais
            filial_ids_set = {p['filial_id'] for p in pedidos if p.get('filial_id') is not None}
            filiais_map = {}
            if filial_ids_set:
                try:
                    f_rows = (
                        supabase.table('filiais')
                        .select('id, cidade, uf')
                        .in_('id', list(filial_ids_set))
                        .execute()
                        .data or []
                    )
                    filiais_map = {f['id']: f for f in f_rows}
                except Exception:
                    pass

            # Agrega métricas
            por_mes       = defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0})
            por_status    = defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0})
            por_pagamento = defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0})
            por_fornecedor= defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0})
            por_filial_map= defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0, 'filial_id': None})
            por_categoria = defaultdict(lambda: {'valor_total': 0.0, 'quantidade': 0})

            STATUS_EXCLUIDOS = {'cancelado', 'reprovado', 'reprovada'}

            pedidos_ativos_ids = set()
            for p in pedidos:
                pid  = p['id']
                val  = valor_por_pedido.get(pid, 0.0)
                mes  = (p.get('data_pedido') or '')[:7]
                s    = p.get('status') or 'rascunho'
                fp   = p.get('forma_pagamento') or 'não informado'
                fn   = (p.get('fornecedor') or '').strip() or 'Não informado'
                fid  = p.get('filial_id')
                f    = filiais_map.get(fid, {}) if fid else {}
                fl   = f"{f.get('cidade','?')}/{f.get('uf','?')}" if f else (str(fid) if fid else '—')

                # por_status conta todos, inclusive cancelados/reprovados
                por_status[s]['valor_total']        += val
                por_status[s]['quantidade']         += 1

                # demais agregações excluem cancelados e reprovados
                if s in STATUS_EXCLUIDOS:
                    continue

                pedidos_ativos_ids.add(pid)
                if mes:
                    por_mes[mes]['valor_total']    += val
                    por_mes[mes]['quantidade']      += 1
                por_pagamento[fp]['valor_total']    += val
                por_pagamento[fp]['quantidade']     += 1
                por_fornecedor[fn]['valor_total']   += val
                por_fornecedor[fn]['quantidade']    += 1
                if fid:
                    por_filial_map[fl]['valor_total']  += val
                    por_filial_map[fl]['quantidade']   += 1
                    por_filial_map[fl]['filial_id']    = fid

            for it in itens:
                if it.get('pedido_id') not in pedidos_ativos_ids:
                    continue
                cat = it.get('categoria') or 'outro'
                total_item = float(it.get('quantidade') or 0) * float(it.get('valor_unitario') or 0)
                por_categoria[cat]['valor_total'] += total_item
                por_categoria[cat]['quantidade']  += 1

            top_fornecedores = sorted(
                [{'fornecedor': k, **v} for k, v in por_fornecedor.items()],
                key=lambda x: x['valor_total'], reverse=True
            )[:10]

            top_filiais = sorted(
                [{'filial': k, **v} for k, v in por_filial_map.items()],
                key=lambda x: x['valor_total'], reverse=True
            )

            pedidos_validos = [p for p in pedidos if (p.get('status') or 'rascunho') not in STATUS_EXCLUIDOS]
            valor_total_geral = sum(valor_por_pedido.get(p['id'], 0.0) for p in pedidos_validos)
            pedidos_com_valor = [p for p in pedidos_validos if valor_por_pedido.get(p['id'], 0.0) > 0]

            # Lista de filiais para o filtro no frontend
            filiais_disponiveis = [
                {'id': f['id'], 'label': f"{f.get('cidade','?')}/{f.get('uf','?')}"}
                for f in filiais_map.values()
            ]

            return jsonify({
                'por_mes': [{'mes': k, **v} for k, v in sorted(por_mes.items())],
                'por_status': [{'status': k, **v} for k, v in por_status.items()],
                'por_pagamento': [{'forma_pagamento': k, **v} for k, v in por_pagamento.items()],
                'por_categoria': sorted(
                    [{'categoria': k, **v} for k, v in por_categoria.items()],
                    key=lambda x: x['valor_total'], reverse=True
                ),
                'top_fornecedores': top_fornecedores,
                'top_filiais': top_filiais,
                'filiais_disponiveis': filiais_disponiveis,
                'total_pedidos': len(pedidos_validos),
                'valor_total_geral': round(valor_total_geral, 2),
                'ticket_medio': round(valor_total_geral / len(pedidos_com_valor), 2) if pedidos_com_valor else 0,
            })
        except Exception as exc:
            app.logger.error('pedidos_compra_metricas: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ─── Configuração do workflow de aprovações (tela de Permissões) ─────────────

    @app.get('/api/approval-configs')
    @rate_limit_endpoint(max_requests=60)
    @require_auth
    def get_approval_configs_api(profile):
        """Lista as configurações de workflow de aprovação (para a tela de Permissões)."""
        if not profile_has_scope_permission(profile, 'menu.permissoes'):
            return jsonify({'error': 'Sem permissão.'}), 403
        try:
            resp = supabase.table('approval_workflow_configs').select('*').order('resource_type').execute()
            return jsonify({'items': resp.data or []})
        except Exception as exc:
            app.logger.error('get_approval_configs_api: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.put('/api/approval-configs/<string:resource_type>')
    @rate_limit_endpoint(max_requests=20)
    @require_auth
    def update_approval_config_api(profile, resource_type):
        """Atualiza a configuração de workflow de um tipo de aprovação."""
        if not profile_has_scope_permission(profile, 'menu.permissoes'):
            return jsonify({'error': 'Sem permissão.'}), 403

        body = request.get_json(silent=True) or {}
        allowed_fields = {
            'view_scope', 'approval_scope', 'pending_statuses',
            'approved_status', 'rejected_status',
            'require_comment_on_approve', 'require_comment_on_reject', 'ativo',
        }
        update_data = {k: v for k, v in body.items() if k in allowed_fields}

        if not update_data:
            return jsonify({'error': 'Nenhum campo válido para atualizar.'}), 400

        update_data['atualizado_em'] = now_iso()

        try:
            result = supabase.table('approval_workflow_configs').update(update_data).eq('resource_type', resource_type).execute()
            if not (result.data or []):
                return jsonify({'error': f'Configuração não encontrada: {resource_type}'}), 404

            write_audit_event(
                profile, 'update', 'approval_workflow_configs', resource_type,
                status='ok', details={k: v for k, v in update_data.items() if k != 'atualizado_em'},
            )
            return jsonify({'status': 'ok', 'resource_type': resource_type})
        except Exception as exc:
            app.logger.error('update_approval_config_api: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ============ HORAS EXTRAS RTM — FECHAMENTOS ============

    @app.route('/api/horas-extras-rtm/salvar', methods=['POST'])
    @require_auth
    def salvar_horas_extras_rtm(profile):
        data = request.get_json() or {}
        mes_referencia = data.get('mes_referencia')
        registros = data.get('registros', [])
        if not mes_referencia:
            return jsonify({'error': 'mes_referencia obrigatório.'}), 400
        if not registros:
            return jsonify({'error': 'Nenhum registro para salvar.'}), 400
        try:
            import unicodedata, re as _re
            user_filial_id = profile.get('filial_id')

            # Fetch ALL collaborators for filial resolution and name-based matching
            todos_colaboradores_db = []
            collab_filial_map = {}
            try:
                colab_resp = supabase.table('colaboradores').select('id, nome_completo, filial_id, ativo').execute()
                todos_colaboradores_db = colab_resp.data or []
                for c in todos_colaboradores_db:
                    collab_filial_map[int(c['id'])] = c.get('filial_id')
            except Exception:
                pass

            # Resolve tipo_hora via contratos_colaboradores — same logic as tipo_hora_mapa_rtm
            # Fetch ALL active contracts (no colab_id filter) to guarantee consistency with the calculator display
            tipo_hora_map = {}   # colaborador_id (str) -> 'fixo' | 'extra'
            colab_contrato_map = {}  # colaborador_id (str) -> contrato_operacional_id (fixo only)
            mes_dt = (mes_referencia[:7] + '-01') if len(mes_referencia) >= 7 else mes_referencia
            mes_prefix = mes_referencia[:7]
            cc_resp = supabase.table('contratos_colaboradores').select(
                'colaborador_id, tipo_item, inicio_vigencia, contrato_operacional_id'
            ).execute()
            for cc in (cc_resp.data or []):
                cid_cc = cc.get('colaborador_id')
                if cid_cc is None:
                    continue
                iv = cc.get('inicio_vigencia')
                # Compare by month only — a contract starting mid-month still counts for that month.
                if iv is not None and iv[:7] > mes_prefix:
                    continue
                ti = cc.get('tipo_item', '')
                cid_str = str(cid_cc)
                if ti.lower() == 'colaborador':
                    tipo_hora_map[cid_str] = 'fixo'
                    if cc.get('contrato_operacional_id'):
                        colab_contrato_map[cid_str] = cc['contrato_operacional_id']
                elif cid_str not in tipo_hora_map:
                    tipo_hora_map[cid_str] = 'extra'

            # Build name-keyed map for cases where colaborador_id doesn't align between tables
            import unicodedata as _ud, re as _re2
            def _nrm(n):
                if not n: return ''
                nfd = _ud.normalize('NFD', str(n).upper().strip())
                a = ''.join(c for c in nfd if _ud.category(c) != 'Mn')
                return _re2.sub(r'\s+', ' ', _re2.sub(r'[^A-Z0-9 ]', ' ', a)).strip()
            # Cross-reference: for each collaborator in DB, if their id is in tipo_hora_map, index by name
            tipo_hora_nome_map = {}
            for c in todos_colaboradores_db:
                cid_str2 = str(c.get('id', ''))
                if cid_str2 in tipo_hora_map:
                    tipo_hora_nome_map[_nrm(c.get('nome_completo', ''))] = tipo_hora_map[cid_str2]

            # Fetch cliente_nome for each contrato referenced by fixo collaborators
            contrato_cliente_map = {}  # contrato_id -> {cliente_nome, contrato_nome}
            fixo_contrato_ids = list({v for v in colab_contrato_map.values() if v})
            if fixo_contrato_ids:
                try:
                    cont_resp = supabase.table('contratos_operacionais').select(
                        'id, cliente_nome, nome_contrato'
                    ).in_('id', fixo_contrato_ids).execute()
                    for c in (cont_resp.data or []):
                        contrato_cliente_map[c['id']] = {
                            'cliente_nome': c.get('cliente_nome') or c.get('nome_contrato', ''),
                            'contrato_nome': c.get('nome_contrato', ''),
                        }
                except Exception:
                    pass

            # Build filial name lookup: nome (pasta/arquivo) → filial_id com matching flexível
            filial_nome_to_id_map = {}  # 'JOINVILLE' → filial_id, 'CUIABÁ' → filial_id, etc
            filiais_cadastradas = []  # Guardar todas as filiais para matching fuzzy
            try:
                fil_resp = supabase.table('filiais').select('id, cidade').execute()
                for f in (fil_resp.data or []):
                    filiais_cadastradas.append({'id': f['id'], 'cidade': f.get('cidade', '').upper().strip()})
                    # Map by exact cidade match
                    filial_nome_to_id_map[f.get('cidade', '').upper().strip()] = f['id']
            except Exception:
                pass

            # Função para buscar filial com matching fuzzy
            def find_filial_id(filial_nome_arquivo):
                if not filial_nome_arquivo:
                    return None
                filial_nome_arquivo = filial_nome_arquivo.upper().strip()
                # Primeiro tenta match exato
                if filial_nome_arquivo in filial_nome_to_id_map:
                    return filial_nome_to_id_map[filial_nome_arquivo]
                # Tenta match parcial (contains)
                for f in filiais_cadastradas:
                    if filial_nome_arquivo in f['cidade'] or f['cidade'] in filial_nome_arquivo:
                        return f['id']
                # Tenta remover acentos
                import unicodedata
                filial_norm = ''.join(c for c in unicodedata.normalize('NFD', filial_nome_arquivo) if unicodedata.category(c) != 'Mn')
                for f in filiais_cadastradas:
                    f_norm = ''.join(c for c in unicodedata.normalize('NFD', f['cidade']) if unicodedata.category(c) != 'Mn')
                    if filial_norm == f_norm or filial_norm in f_norm or f_norm in filial_norm:
                        return f['id']
                return None

            # Build filial name lookup: filial_id → cidade (all branches, from the list already fetched)
            filial_nome_map = {f['id']: f['cidade'] for f in filiais_cadastradas}

            # ── Fuzzy name matching (mirrors frontend matchColaborador) ────────────
            _STOP_WORDS = {'DE', 'DA', 'DO', 'DAS', 'DOS', 'E', 'EM', 'OS', 'AS', 'A'}

            def _norm(name):
                if not name:
                    return ''
                nfd = unicodedata.normalize('NFD', name.upper().strip())
                ascii_ = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
                return _re.sub(r'\s+', ' ', _re.sub(r'[^A-Z0-9 ]', ' ', ascii_)).strip()

            def _tokens(name):
                return [w for w in _norm(name).split() if len(w) > 2 and w not in _STOP_WORDS]

            def _overlap(ta, tb):
                if not ta or not tb:
                    return 0.0
                sa, sb = set(ta), set(tb)
                return len(sa & sb) / min(len(sa), len(sb))

            def match_colab_nome(nome_pasta):
                """Return collaborator dict from DB that best matches nome_pasta, or None."""
                n = _norm(nome_pasta)
                if not n:
                    return None
                # Stage 1: exact
                for c in todos_colaboradores_db:
                    if _norm(c.get('nome_completo', '')) == n:
                        return c
                # Stage 2: one contains the other
                for c in todos_colaboradores_db:
                    cn = _norm(c.get('nome_completo', ''))
                    if cn and (n in cn or cn in n):
                        return c
                # Stage 3: all words of shorter name appear in longer
                n_tok = _tokens(n)
                for c in todos_colaboradores_db:
                    cn_tok = _tokens(c.get('nome_completo', ''))
                    if not n_tok or not cn_tok:
                        continue
                    shorter = n_tok if len(n_tok) <= len(cn_tok) else cn_tok
                    longer = cn_tok if shorter is n_tok else n_tok
                    if all(w in longer for w in shorter):
                        return c
                # Stage 4: word overlap >= 60%
                for c in todos_colaboradores_db:
                    if _overlap(n_tok, _tokens(c.get('nome_completo', ''))) >= 0.6:
                        return c
                # Stage 5: first + last word match
                parts = n.split()
                if len(parts) >= 2:
                    first, last = parts[0], parts[-1]
                    for c in todos_colaboradores_db:
                        cp = _norm(c.get('nome_completo', '')).split()
                        if len(cp) >= 2 and cp[0] == first and cp[-1] == last:
                            return c
                return None

            rows = []
            for r in registros:
                cid = r.get('colaborador_id')
                filial_id = None
                filial_nome_arquivo = r.get('filial_nome', '')

                # Resolve tipo_hora: try stored colaborador_id first
                tipo_hora = tipo_hora_map.get(str(cid), None) if cid else None

                # Fallback: if no ID or ID not in contract map, try server-side name matching
                if tipo_hora is None:
                    nome_func = r.get('funcionario_nome', '') or r.get('funcionario', '')
                    matched_c = match_colab_nome(nome_func)
                    if matched_c:
                        alt_cid = matched_c.get('id')
                        tipo_hora = tipo_hora_map.get(str(alt_cid), None)
                        if tipo_hora is None:
                            # Try name-keyed map (handles ID mismatches between tables)
                            tipo_hora = tipo_hora_nome_map.get(_nrm(matched_c.get('nome_completo', '')), None)
                        if alt_cid and not cid:
                            cid = alt_cid
                # Final fallback: try direct name lookup in nome_map
                if tipo_hora is None:
                    nome_func2 = r.get('funcionario_nome', '') or r.get('funcionario', '')
                    tipo_hora = tipo_hora_nome_map.get(_nrm(nome_func2), None)
                tipo_hora = tipo_hora or 'extra'

                # Prioridade 1: tentar resolver filial_id pela filial_nome do arquivo (com matching fuzzy)
                if filial_nome_arquivo:
                    filial_id = find_filial_id(filial_nome_arquivo)

                # Prioridade 2: usar filial_id do colaborador se encontrado
                if filial_id is None:
                    filial_id = collab_filial_map.get(int(cid)) if cid else None

                # Prioridade 3: usar filial_id do usuário logado
                if filial_id is None:
                    filial_id = user_filial_id
                # Prioriza o filial_nome do arquivo; usa o do banco só se não tiver no arquivo
                filial_nome = r.get('filial_nome', '') or filial_nome_map.get(filial_id) or ''
                rows.append({
                    'mes_referencia': mes_referencia,
                    'funcionario_nome': r.get('funcionario_nome', ''),
                    'colaborador_id': cid,
                    'filial_id': filial_id,
                    'filial_nome': filial_nome,
                    'estado': r.get('estado', ''),
                    'horas_normais': float(r.get('horas_normais') or 0),
                    'horas_extra_100': float(r.get('horas_extra_100') or 0),
                    'valor_hora_50': float(r.get('valor_hora_50') or 0),
                    'valor_hora_100': float(r.get('valor_hora_100') or 0),
                    'total_50': float(r.get('total_50') or 0),
                    'total_100': float(r.get('total_100') or 0),
                    'total_geral': float(r.get('total_geral') or 0),
                    'tipo_hora': tipo_hora,
                })

            # Delete only records for filiais present in this import batch (never touch other filiais)
            importing_nomes = list({r.get('filial_nome') for r in rows if r.get('filial_nome')})
            del_query = supabase.table('horas_extras_rtm_registros').delete().eq('mes_referencia', mes_referencia)
            if importing_nomes:
                del_query = del_query.in_('filial_nome', importing_nomes)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'error': 'Sem filiais permitidas para salvar.'}), 403
                del_query = del_query.in_('filial_id', allowed_ids)
            del_query.execute()

            if rows:
                supabase.table('horas_extras_rtm_registros').insert(rows).execute()

            # Registra auditoria com filiais corretas dos registros
            unique_filiais = {r.get('filial_id') for r in rows if r.get('filial_id')}
            for filial_id in unique_filiais:
                write_audit_event(
                    profile,
                    action='create',
                    resource_name='horas_extras_rtm_registros',
                    entity_id=None,
                    details={'mes_referencia': mes_referencia, 'registros_count': len([r for r in rows if r.get('filial_id') == filial_id])},
                    filial_id=filial_id,
                )

            # Auto-create Contas a Receber (fixo only) and Contas a Pagar (extra only)
            try:
                from collections import defaultdict
                # Group by filial_nome (not filial_id) so that each branch gets its own CP/CR
                # even when filial_id resolution falls back to the user's home branch
                fixo_groups = defaultdict(lambda: {
                    'total': 0.0, 'count': 0, 'filial_nome': '', 'filial_id': None,
                    'contrato_id': None, 'cliente_nome': '', 'contrato_nome': '',
                })
                extra_groups = defaultdict(lambda: {'total': 0.0, 'count': 0, 'filial_nome': '', 'filial_id': None})
                for r in rows:
                    fname = r.get('filial_nome') or 'Sem filial'
                    total = float(r.get('total_geral') or 0)
                    cid = r.get('colaborador_id')
                    if r.get('tipo_hora') == 'fixo':
                        contrato_id = colab_contrato_map.get(str(cid)) if cid else None
                        key = (fname, contrato_id)
                        info = fixo_groups[key]
                        info['total'] += total
                        info['count'] += 1
                        info['filial_nome'] = r.get('filial_nome', '')
                        info['filial_id'] = r.get('filial_id')
                        info['contrato_id'] = contrato_id
                        if contrato_id and contrato_id in contrato_cliente_map:
                            info['cliente_nome'] = contrato_cliente_map[contrato_id]['cliente_nome']
                            info['contrato_nome'] = contrato_cliente_map[contrato_id]['contrato_nome']
                    else:
                        info = extra_groups[fname]
                        info['total'] += total
                        info['count'] += 1
                        info['filial_nome'] = r.get('filial_nome', '')
                        info['filial_id'] = r.get('filial_id')

                # Remove auto-generated CR/CP for this month before recreating
                supabase.table('contas_a_receber').delete().eq('horas_extras_rtm_mes', mes_referencia).execute()
                supabase.table('contas_a_pagar').delete().eq('horas_extras_rtm_mes', mes_referencia).execute()

                mes_label = mes_referencia[:7]  # YYYY-MM
                cr_rows, cp_rows = [], []

                # FIXO → Contas a Receber (client reimburses Gold)
                for _key, info in fixo_groups.items():
                    if info['total'] <= 0:
                        continue
                    fid_cr = find_filial_id(info.get('filial_nome', '')) or info['filial_id']
                    cliente = info.get('cliente_nome') or ''
                    cr_rows.append({
                        'filial_id': fid_cr,
                        'filial_nome': info['filial_nome'],
                        'competencia': mes_referencia,
                        'obrigacao': 'HORA EXTRA',
                        'descricao': f"Horas Extras – Fixo Contrato ({mes_label}) – {info['filial_nome']}",
                        'cliente_nome': cliente,
                        'contrato_operacional_id': info.get('contrato_id'),
                        'contrato_nome': info.get('contrato_nome', ''),
                        'valor_gold': round(info['total'], 2),
                        'status_fat': 'NÃO FATURADO',
                        'status': 'FALTA COBRAR',
                        'horas_extras_rtm_mes': mes_referencia,
                        'tipo_hora': 'fixo',
                        'limite_dia': 10,
                        'prazo_envio': 'ATÉ O DIA 10 - SUB',
                    })

                # EXTRA → Contas a Pagar only (Gold absorbs cost, no CR created)
                for _fid, info in extra_groups.items():
                    if info['total'] <= 0:
                        continue
                    fid_cp = find_filial_id(info.get('filial_nome', '')) or info['filial_id']
                    cp_rows.append({
                        'filial_id': fid_cp,
                        'filial_nome': info['filial_nome'],
                        'competencia': mes_referencia,
                        'tipo_despesa': 'HORAS EXTRAS',
                        'descricao': f"Horas Extras Fora Contrato ({mes_label}) – {info['filial_nome']} – {info['count']} funcionários",
                        'valor': round(info['total'], 2),
                        'status': 'PENDENTE',
                        'horas_extras_rtm_mes': mes_referencia,
                        'colaboradores_count': info['count'],
                    })

                if cr_rows:
                    supabase.table('contas_a_receber').insert(cr_rows).execute()
                if cp_rows:
                    supabase.table('contas_a_pagar').insert(cp_rows).execute()
            except Exception as fin_exc:
                app.logger.warning('auto_create_financeiro_rtm: %s', fin_exc)

            return jsonify({'ok': True, 'count': len(rows)})
        except Exception as exc:
            app.logger.error('salvar_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/meses', methods=['GET'])
    @require_auth
    def listar_meses_horas_extras_rtm(profile):
        try:
            query = supabase.table('horas_extras_rtm_registros').select(
                'mes_referencia, horas_normais, horas_extra_100, total_geral'
            ).order('mes_referencia', desc=True)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            result = supabase_retry(query.execute)
            from collections import defaultdict
            meses = defaultdict(lambda: {'funcionarios': 0, 'total_horas_normais': 0.0, 'total_horas_100': 0.0, 'total_geral': 0.0})
            for row in (result.data or []):
                m = row['mes_referencia']
                meses[m]['funcionarios'] += 1
                meses[m]['total_horas_normais'] += float(row.get('horas_normais') or 0)
                meses[m]['total_horas_100'] += float(row.get('horas_extra_100') or 0)
                meses[m]['total_geral'] += float(row.get('total_geral') or 0)
            lista = [{'mes_referencia': k, **v} for k, v in sorted(meses.items(), reverse=True)]
            return jsonify({'data': lista})
        except Exception as exc:
            app.logger.error('listar_meses_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/meses-filiais', methods=['GET'])
    @require_auth
    def listar_meses_filiais_horas_extras_rtm(profile):
        try:
            query = supabase.table('horas_extras_rtm_registros').select(
                'mes_referencia, filial_nome, horas_normais, horas_extra_100, total_geral, tipo_hora'
            ).order('mes_referencia', desc=True)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            result = supabase_retry(query.execute)
            from collections import defaultdict
            groups = defaultdict(lambda: {'funcionarios': 0, 'total_horas_normais': 0.0, 'total_horas_100': 0.0, 'total_geral': 0.0, 'total_fixo': 0.0, 'total_extra': 0.0})
            for row in (result.data or []):
                key = (row['mes_referencia'], row.get('filial_nome') or 'Sem filial')
                g = groups[key]
                g['funcionarios'] += 1
                g['total_horas_normais'] += float(row.get('horas_normais') or 0)
                g['total_horas_100'] += float(row.get('horas_extra_100') or 0)
                g['total_geral'] += float(row.get('total_geral') or 0)
                if row.get('tipo_hora') == 'fixo':
                    g['total_fixo'] += float(row.get('total_geral') or 0)
                elif row.get('tipo_hora') == 'extra':
                    g['total_extra'] += float(row.get('total_geral') or 0)
            lista = [{'mes_referencia': k[0], 'filial_nome': k[1], **v} for k, v in groups.items()]
            lista.sort(key=lambda x: x['filial_nome'])
            lista.sort(key=lambda x: x['mes_referencia'], reverse=True)
            return jsonify({'data': lista})
        except Exception as exc:
            app.logger.error('listar_meses_filiais_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/detalhe', methods=['GET'])
    @require_auth
    def detalhe_horas_extras_rtm(profile):
        mes = request.args.get('mes')
        if not mes:
            return jsonify({'error': 'Parâmetro mes obrigatório.'}), 400
        try:
            query = supabase.table('horas_extras_rtm_registros').select('*').eq('mes_referencia', mes).order('funcionario_nome')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            result = supabase_retry(query.execute)
            return jsonify({'data': result.data or []})
        except Exception as exc:
            app.logger.error('detalhe_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/tipo-hora-mapa', methods=['GET'])
    @require_auth
    def tipo_hora_mapa_rtm(profile):
        """Returns {colaborador_id: {tipo_hora, cliente_nome, contrato_nome, contrato_id}} for a given month."""
        mes = request.args.get('mes')  # YYYY-MM or YYYY-MM-DD
        if not mes:
            return jsonify({'error': 'Parâmetro mes obrigatório.'}), 400
        mes_prefix = mes[:7]  # YYYY-MM
        mes_dt = mes_prefix + '-01'
        try:
            cc_resp = supabase.table('contratos_colaboradores').select(
                'colaborador_id, tipo_item, inicio_vigencia, contrato_operacional_id'
            ).execute()

            tipo_hora_map = {}    # colab_id (str) -> 'fixo' | 'extra'
            colab_contrato = {}   # colab_id (str) -> contrato_operacional_id

            for cc in (cc_resp.data or []):
                cid = cc.get('colaborador_id')
                if cid is None:
                    continue
                iv = cc.get('inicio_vigencia')
                # Compare by month only — a contract starting mid-month still counts for that month.
                if iv is not None and iv[:7] > mes_prefix:
                    continue
                ti = cc.get('tipo_item', '')
                cid_str = str(cid)
                if ti.lower() == 'colaborador':
                    tipo_hora_map[cid_str] = 'fixo'
                    if cc.get('contrato_operacional_id'):
                        colab_contrato[cid_str] = cc['contrato_operacional_id']
                elif cid_str not in tipo_hora_map:
                    tipo_hora_map[cid_str] = 'extra'

            # Fetch client names for all contracts referenced
            contrato_info = {}
            fixo_contrato_ids = list({v for v in colab_contrato.values() if v})
            if fixo_contrato_ids:
                cont_resp = supabase.table('contratos_operacionais').select(
                    'id, cliente_nome, nome_contrato'
                ).in_('id', fixo_contrato_ids).execute()
                for c in (cont_resp.data or []):
                    contrato_info[c['id']] = {
                        'cliente_nome': c.get('cliente_nome') or c.get('nome_contrato', ''),
                        'contrato_nome': c.get('nome_contrato', ''),
                    }

            result = {}
            for cid_str, tipo in tipo_hora_map.items():
                contrato_id = colab_contrato.get(cid_str)
                info = contrato_info.get(contrato_id, {}) if contrato_id else {}
                result[cid_str] = {
                    'tipo_hora': tipo,
                    'cliente_nome': info.get('cliente_nome', ''),
                    'contrato_nome': info.get('contrato_nome', ''),
                    'contrato_id': contrato_id,
                }

            return jsonify({'data': result})
        except Exception as exc:
            app.logger.error('tipo_hora_mapa_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/registro/<registro_id>', methods=['PUT'])
    @require_auth
    def editar_registro_horas_extras_rtm(profile, registro_id):
        data = request.get_json() or {}
        try:
            # Verify filial scope before editing
            if profile_has_filial_scope(profile):
                reg_resp = supabase.table('horas_extras_rtm_registros').select('filial_id').eq('id', registro_id).limit(1).execute()
                reg = (reg_resp.data or [{}])[0]
                if reg.get('filial_id') and int(reg['filial_id']) not in set(profile.get('allowed_filial_ids') or []):
                    return jsonify({'error': 'Sem permissão para editar este registro.'}), 403
            hn = float(data.get('horas_normais') or 0)
            he = float(data.get('horas_extra_100') or 0)
            vh50 = float(data.get('valor_hora_50') or 0)
            vh100 = float(data.get('valor_hora_100') or 0)
            t50 = hn * vh50
            t100 = he * vh100
            update = {
                'horas_normais': hn,
                'horas_extra_100': he,
                'valor_hora_50': vh50,
                'valor_hora_100': vh100,
                'total_50': t50,
                'total_100': t100,
                'total_geral': t50 + t100,
            }
            supabase.table('horas_extras_rtm_registros').update(update).eq('id', registro_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('editar_registro_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/recalcular-tipo-hora', methods=['POST'])
    @require_auth
    def recalcular_tipo_hora_rtm(profile):
        """Re-derives tipo_hora for all existing records of a given month (and optionally filial)."""
        import unicodedata, re as _re
        data = request.get_json() or {}
        mes = data.get('mes')
        filial_nome = data.get('filial_nome')
        if not mes:
            return jsonify({'error': 'Parâmetro mes obrigatório.'}), 400
        try:
            mes_prefix = mes[:7]
            mes_dt = mes_prefix + '-01'

            # Fetch existing records for this month/filial
            q = supabase.table('horas_extras_rtm_registros').select('id, colaborador_id, funcionario_nome').eq('mes_referencia', mes)
            if filial_nome:
                q = q.eq('filial_nome', filial_nome)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if allowed_ids:
                    q = q.in_('filial_id', allowed_ids)
            regs = q.execute().data or []

            # Build tipo_hora_map from all contracts (no ativo filter, date-scoped)
            cc_resp = supabase.table('contratos_colaboradores').select(
                'colaborador_id, tipo_item, inicio_vigencia'
            ).execute()
            tipo_hora_map = {}
            for cc in (cc_resp.data or []):
                cid_cc = cc.get('colaborador_id')
                if cid_cc is None:
                    continue
                iv = cc.get('inicio_vigencia')
                # Compare by month only — a contract starting mid-month still counts for that month.
                if iv is not None and iv[:7] > mes_prefix:
                    continue
                ti = cc.get('tipo_item', '')
                cid_str = str(cid_cc)
                if ti.lower() == 'colaborador':
                    tipo_hora_map[cid_str] = 'fixo'
                elif cid_str not in tipo_hora_map:
                    tipo_hora_map[cid_str] = 'extra'

            # Fetch all collaborators for name-based fallback matching
            todos_db = []
            try:
                todos_db = supabase.table('colaboradores').select('id, nome_completo').execute().data or []
            except Exception:
                pass

            _STOP_W = {'DE', 'DA', 'DO', 'DAS', 'DOS', 'E', 'EM', 'OS', 'AS', 'A'}

            def _nm(n):
                if not n:
                    return ''
                nfd = unicodedata.normalize('NFD', n.upper().strip())
                a = ''.join(c for c in nfd if unicodedata.category(c) != 'Mn')
                return _re.sub(r'\s+', ' ', _re.sub(r'[^A-Z0-9 ]', ' ', a)).strip()

            def _tok(n):
                return [w for w in _nm(n).split() if len(w) > 2 and w not in _STOP_W]

            def _ovl(ta, tb):
                if not ta or not tb:
                    return 0.0
                return len(set(ta) & set(tb)) / min(len(set(ta)), len(set(tb)))

            def _match_nome(nome):
                n = _nm(nome)
                if not n:
                    return None
                for c in todos_db:
                    if _nm(c.get('nome_completo', '')) == n:
                        return c
                for c in todos_db:
                    cn = _nm(c.get('nome_completo', ''))
                    if cn and (n in cn or cn in n):
                        return c
                nt = _tok(n)
                for c in todos_db:
                    ct = _tok(c.get('nome_completo', ''))
                    if not nt or not ct:
                        continue
                    sh = nt if len(nt) <= len(ct) else ct
                    lg = ct if sh is nt else nt
                    if all(w in lg for w in sh):
                        return c
                for c in todos_db:
                    if _ovl(nt, _tok(c.get('nome_completo', ''))) >= 0.6:
                        return c
                parts = n.split()
                if len(parts) >= 2:
                    f, l = parts[0], parts[-1]
                    for c in todos_db:
                        cp = _nm(c.get('nome_completo', '')).split()
                        if len(cp) >= 2 and cp[0] == f and cp[-1] == l:
                            return c
                return None

            # Build name-keyed fallback map (handles ID mismatches between tables)
            tipo_hora_nome_map2 = {}
            for c in todos_db:
                cid_str2 = str(c.get('id', ''))
                if cid_str2 in tipo_hora_map:
                    tipo_hora_nome_map2[_nm(c.get('nome_completo', ''))] = tipo_hora_map[cid_str2]

            # Resolve tipo_hora for each record, then batch-update to minimise DB round-trips
            fixo_ids = []    # record IDs to set tipo_hora='fixo'
            extra_ids = []   # record IDs to set tipo_hora='extra'
            colab_patches = {}  # record id → colaborador_id (only when patching a missing one)
            detalhes_extra = []  # diagnostic: why each 'extra' employee was classified so

            for reg in regs:
                cid = reg.get('colaborador_id')
                nome_func = reg.get('funcionario_nome', '')
                motivo = None
                # 1. Try stored colaborador_id
                tipo_hora = tipo_hora_map.get(str(cid), None) if cid else None
                if tipo_hora is None and cid:
                    motivo = 'colaborador_id_nao_encontrado_em_contratos'
                # 2. Direct name lookup
                if tipo_hora is None:
                    tipo_hora = tipo_hora_nome_map2.get(_nm(nome_func), None)
                # 3. Fuzzy name match → nome_map
                if tipo_hora is None:
                    matched = _match_nome(nome_func)
                    if matched:
                        alt_cid = matched.get('id')
                        tipo_hora = tipo_hora_map.get(str(alt_cid), None)
                        if tipo_hora is None:
                            tipo_hora = tipo_hora_nome_map2.get(_nm(matched.get('nome_completo', '')), None)
                        if tipo_hora is not None and not cid:
                            colab_patches[reg['id']] = alt_cid
                    else:
                        motivo = motivo or ('sem_colaborador_id_e_nome_nao_encontrado' if not cid else 'colaborador_id_nao_encontrado_e_nome_sem_match')
                tipo_hora = tipo_hora or 'extra'
                if tipo_hora == 'fixo':
                    fixo_ids.append(reg['id'])
                else:
                    extra_ids.append(reg['id'])
                    detalhes_extra.append({
                        'nome': nome_func,
                        'colaborador_id': cid,
                        'motivo': motivo or 'nao_vinculado_como_colaborador_em_contrato',
                    })

            # Batch updates — 2 queries instead of N
            if fixo_ids:
                supabase.table('horas_extras_rtm_registros').update({'tipo_hora': 'fixo'}).in_('id', fixo_ids).execute()
            if extra_ids:
                supabase.table('horas_extras_rtm_registros').update({'tipo_hora': 'extra'}).in_('id', extra_ids).execute()
            # Patch missing colaborador_ids individually (usually zero or very few)
            for rec_id, c_id in colab_patches.items():
                supabase.table('horas_extras_rtm_registros').update({'colaborador_id': c_id}).eq('id', rec_id).execute()

            updated = len(fixo_ids) + len(extra_ids)
            return jsonify({'ok': True, 'updated': updated, 'fixo': len(fixo_ids), 'extra': len(extra_ids), 'detalhes_extra': detalhes_extra})
        except Exception as exc:
            app.logger.error('recalcular_tipo_hora_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/mes/<mes>', methods=['DELETE'])
    @require_auth
    def deletar_mes_horas_extras_rtm(profile, mes):
        try:
            filial_nome = request.args.get('filial_nome')
            del_query = supabase.table('horas_extras_rtm_registros').delete().eq('mes_referencia', mes)
            if filial_nome:
                del_query = del_query.eq('filial_nome', filial_nome)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'error': 'Sem filiais permitidas.'}), 403
                del_query = del_query.in_('filial_id', allowed_ids)
            del_query.execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('deletar_mes_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/horas-extras-rtm/metricas', methods=['GET'])
    @require_auth
    def metricas_horas_extras_rtm(profile):
        try:
            query = supabase.table('horas_extras_rtm_registros').select(
                'mes_referencia, funcionario_nome, filial_nome, estado, horas_normais, horas_extra_100, total_50, total_100, total_geral'
            )
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'top_filiais': [], 'top_funcionarios_100': [], 'top_funcionarios_50': [], 'evolucao_mensal': [], 'resumo_por_ano': []})
                query = query.in_('filial_id', allowed_ids)
            result = query.execute()
            rows = result.data or []
            from collections import defaultdict
            filiais = defaultdict(lambda: {'horas_normais': 0.0, 'horas_extra_100': 0.0, 'total_50': 0.0, 'total_100': 0.0, 'total': 0.0, 'funcionarios': 0})
            funcionarios = defaultdict(lambda: {'horas_normais': 0.0, 'horas_extra_100': 0.0, 'total_50': 0.0, 'total_100': 0.0, 'total': 0.0, 'meses': 0})
            meses_map = defaultdict(lambda: {'horas_normais': 0.0, 'horas_extra_100': 0.0, 'total_50': 0.0, 'total_100': 0.0, 'total': 0.0, 'funcionarios': 0})
            anos_map = defaultdict(lambda: {'horas_normais': 0.0, 'horas_extra_100': 0.0, 'total_50': 0.0, 'total_100': 0.0, 'total': 0.0, 'meses': set()})
            for r in rows:
                f = r.get('filial_nome') or 'Sem filial'
                filiais[f]['horas_normais'] += float(r.get('horas_normais') or 0)
                filiais[f]['horas_extra_100'] += float(r.get('horas_extra_100') or 0)
                filiais[f]['total_50'] += float(r.get('total_50') or 0)
                filiais[f]['total_100'] += float(r.get('total_100') or 0)
                filiais[f]['total'] += float(r.get('total_geral') or 0)
                filiais[f]['funcionarios'] += 1
                fn = r.get('funcionario_nome') or 'Desconhecido'
                funcionarios[fn]['horas_normais'] += float(r.get('horas_normais') or 0)
                funcionarios[fn]['horas_extra_100'] += float(r.get('horas_extra_100') or 0)
                funcionarios[fn]['total_50'] += float(r.get('total_50') or 0)
                funcionarios[fn]['total_100'] += float(r.get('total_100') or 0)
                funcionarios[fn]['total'] += float(r.get('total_geral') or 0)
                funcionarios[fn]['meses'] += 1
                m = r.get('mes_referencia') or ''
                meses_map[m]['horas_normais'] += float(r.get('horas_normais') or 0)
                meses_map[m]['horas_extra_100'] += float(r.get('horas_extra_100') or 0)
                meses_map[m]['total_50'] += float(r.get('total_50') or 0)
                meses_map[m]['total_100'] += float(r.get('total_100') or 0)
                meses_map[m]['total'] += float(r.get('total_geral') or 0)
                meses_map[m]['funcionarios'] += 1
                ano = m[:4] if m else ''
                if ano:
                    anos_map[ano]['horas_normais'] += float(r.get('horas_normais') or 0)
                    anos_map[ano]['horas_extra_100'] += float(r.get('horas_extra_100') or 0)
                    anos_map[ano]['total_50'] += float(r.get('total_50') or 0)
                    anos_map[ano]['total_100'] += float(r.get('total_100') or 0)
                    anos_map[ano]['total'] += float(r.get('total_geral') or 0)
                    anos_map[ano]['meses'].add(m)
            top_filiais = sorted([{'filial': k, **v} for k, v in filiais.items()], key=lambda x: x['total'], reverse=True)[:12]
            top_func_100 = sorted([{'funcionario': k, **v} for k, v in funcionarios.items()], key=lambda x: x['horas_extra_100'], reverse=True)[:10]
            top_func_50 = sorted([{'funcionario': k, **v} for k, v in funcionarios.items()], key=lambda x: x['horas_normais'], reverse=True)[:10]
            evolucao = sorted([{'mes': k, **v} for k, v in meses_map.items()], key=lambda x: x['mes'])
            resumo_por_ano = sorted([{'ano': k, 'meses': len(v.pop('meses')), **v} for k, v in anos_map.items()], key=lambda x: x['ano'], reverse=True)
            return jsonify({
                'top_filiais': top_filiais,
                'top_funcionarios_100': top_func_100,
                'top_funcionarios_50': top_func_50,
                'evolucao_mensal': evolucao,
                'resumo_por_ano': resumo_por_ano,
            })
        except Exception as exc:
            app.logger.error('metricas_horas_extras_rtm: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ============ CONTAS A RECEBER ============

    @app.route('/api/contas-receber', methods=['GET'])
    @require_auth
    def listar_contas_receber(profile):
        try:
            query = supabase.table('contas_a_receber').select('*').order('competencia', desc=True).order('filial_nome')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            competencia = request.args.get('competencia')
            obrigacao = request.args.get('obrigacao')
            status_fat = request.args.get('status_fat')
            status = request.args.get('status')
            if competencia:
                query = query.eq('competencia', competencia)
            if obrigacao:
                query = query.eq('obrigacao', obrigacao)
            if status_fat:
                query = query.eq('status_fat', status_fat)
            if status:
                query = query.eq('status', status)
            result = supabase_retry(query.execute)
            return jsonify({'data': result.data or []})
        except Exception as exc:
            app.logger.error('listar_contas_receber: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-receber/alertas', methods=['GET'])
    @require_auth
    def alertas_contas_receber(profile):
        try:
            query = supabase.table('contas_a_receber').select(
                'status_fat, status, cobrado_wm, valor_gold, data_limite, data_vencimento, o_que_falta'
            )
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'total_a_receber': 0, 'nao_faturado': 0, 'falta_cobrar': 0, 'vencidos': 0, 'pendentes_autorizacao': 0})
                query = query.in_('filial_id', allowed_ids)
            result = supabase_retry(query.execute)
            rows = result.data or []
            from datetime import date as _date
            hoje = _date.today().isoformat()
            total_a_receber = sum(float(r.get('cobrado_wm') or 0) for r in rows if r.get('status_fat') != 'FATURADO')
            nao_faturado = sum(1 for r in rows if r.get('status_fat') == 'NÃO FATURADO')
            falta_cobrar = sum(1 for r in rows if r.get('status') == 'FALTA COBRAR')
            vencidos = sum(1 for r in rows if (r.get('data_limite') or r.get('data_vencimento') or '') < hoje and r.get('status_fat') != 'FATURADO')
            pendentes = sum(1 for r in rows if r.get('o_que_falta'))
            return jsonify({
                'total_a_receber': total_a_receber,
                'nao_faturado': nao_faturado,
                'falta_cobrar': falta_cobrar,
                'vencidos': vencidos,
                'pendentes_autorizacao': pendentes,
            })
        except Exception as exc:
            app.logger.error('alertas_contas_receber: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-receber', methods=['POST'])
    @require_auth
    def criar_conta_receber(profile):
        data = request.get_json() or {}
        required = ['filial_id', 'competencia', 'obrigacao']
        for f in required:
            if not data.get(f):
                return jsonify({'error': f'{f} obrigatório.'}), 400
        if profile_has_filial_scope(profile):
            allowed_ids = profile.get('allowed_filial_ids') or []
            if int(data['filial_id']) not in set(allowed_ids):
                return jsonify({'error': 'Sem permissão para esta filial.'}), 403
        allowed = [
            'filial_id', 'filial_nome', 'competencia', 'obrigacao', 'descricao',
            'cliente_nome', 'contrato_nome', 'contrato_operacional_id',
            'limite_dia', 'data_limite', 'ult_dia_competencia', 'prazo_envio',
            'valor_gold', 'data_pagamento_gold', 'cobrado_wm', 'data_envio', 'data_ajuste', 'vlr_ajustado_wm',
            'frete', 'vlr_cte', 'vlr_fixo_icms',
            'emissao', 'nd', 'cte', 'tipo_documento', 'ferramenta', 'prestacao', 'contato',
            'double_check', 'autorizacao', 'o_que_falta', 'motivo_pendencia', 'setor_responsavel', 'previsao',
            'status_fat', 'status', 'data_vencimento', 'sla', 'tipo_hora',
        ]
        row = {k: data[k] for k in allowed if k in data}
        try:
            result = supabase.table('contas_a_receber').insert(row).execute()
            return jsonify({'ok': True, 'data': (result.data or [{}])[0]})
        except Exception as exc:
            app.logger.error('criar_conta_receber: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-receber/<cr_id>', methods=['PUT'])
    @require_auth
    def editar_conta_receber(profile, cr_id):
        data = request.get_json() or {}
        if profile_has_filial_scope(profile):
            reg = supabase.table('contas_a_receber').select('filial_id').eq('id', cr_id).limit(1).execute()
            fid = ((reg.data or [{}])[0]).get('filial_id')
            if fid and int(fid) not in set(profile.get('allowed_filial_ids') or []):
                return jsonify({'error': 'Sem permissão para esta filial.'}), 403
        allowed = [
            'filial_nome', 'competencia', 'obrigacao', 'descricao',
            'cliente_nome', 'contrato_nome', 'contrato_operacional_id',
            'limite_dia', 'data_limite', 'ult_dia_competencia', 'prazo_envio',
            'valor_gold', 'data_pagamento_gold', 'cobrado_wm', 'data_envio', 'data_ajuste', 'vlr_ajustado_wm',
            'frete', 'vlr_cte', 'vlr_fixo_icms',
            'emissao', 'nd', 'cte', 'tipo_documento', 'ferramenta', 'prestacao', 'contato',
            'double_check', 'autorizacao', 'o_que_falta', 'motivo_pendencia', 'setor_responsavel', 'previsao',
            'status_fat', 'status', 'data_vencimento', 'sla', 'tipo_hora',
        ]
        update = {k: data[k] for k in allowed if k in data}
        update['updated_at'] = datetime.utcnow().isoformat()
        try:
            supabase.table('contas_a_receber').update(update).eq('id', cr_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('editar_conta_receber: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-receber/<cr_id>', methods=['DELETE'])
    @require_auth
    def deletar_conta_receber(profile, cr_id):
        if profile_has_filial_scope(profile):
            reg = supabase.table('contas_a_receber').select('filial_id').eq('id', cr_id).limit(1).execute()
            fid = ((reg.data or [{}])[0]).get('filial_id')
            if fid and int(fid) not in set(profile.get('allowed_filial_ids') or []):
                return jsonify({'error': 'Sem permissão.'}), 403
        try:
            supabase.table('contas_a_receber').delete().eq('id', cr_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('deletar_conta_receber: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ============ CONTAS A PAGAR ============

    @app.route('/api/contas-pagar', methods=['GET'])
    @require_auth
    def listar_contas_pagar(profile):
        try:
            query = supabase.table('contas_a_pagar').select('*').order('competencia', desc=True).order('filial_nome')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            competencia = request.args.get('competencia')
            status = request.args.get('status')
            tipo_despesa = request.args.get('tipo_despesa')
            if competencia:
                query = query.eq('competencia', competencia)
            if status:
                query = query.eq('status', status)
            if tipo_despesa:
                query = query.eq('tipo_despesa', tipo_despesa)
            result = supabase_retry(query.execute)
            return jsonify({'data': result.data or []})
        except Exception as exc:
            app.logger.error('listar_contas_pagar: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-pagar/alertas', methods=['GET'])
    @require_auth
    def alertas_contas_pagar(profile):
        try:
            query = supabase.table('contas_a_pagar').select('status, valor, valor_pago, data_vencimento')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'total_a_pagar': 0, 'pendente': 0, 'vencidos': 0, 'pago_mes': 0})
                query = query.in_('filial_id', allowed_ids)
            result = supabase_retry(query.execute)
            rows = result.data or []
            from datetime import date as _date
            hoje = _date.today().isoformat()
            mes_atual = hoje[:7]
            total_a_pagar = sum(float(r.get('valor') or 0) - float(r.get('valor_pago') or 0) for r in rows if r.get('status') not in ('PAGO', 'CANCELADO'))
            pendente = sum(1 for r in rows if r.get('status') == 'PENDENTE')
            vencidos = sum(1 for r in rows if (r.get('data_vencimento') or '') < hoje and r.get('status') not in ('PAGO', 'CANCELADO'))
            pago_mes = sum(float(r.get('valor_pago') or 0) for r in rows if r.get('status') == 'PAGO' and (r.get('data_vencimento') or '')[:7] == mes_atual)
            return jsonify({
                'total_a_pagar': total_a_pagar,
                'pendente': pendente,
                'vencidos': vencidos,
                'pago_mes': pago_mes,
            })
        except Exception as exc:
            app.logger.error('alertas_contas_pagar: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-pagar', methods=['POST'])
    @require_auth
    def criar_conta_pagar(profile):
        data = request.get_json() or {}
        required = ['filial_id', 'competencia', 'tipo_despesa', 'valor']
        for f in required:
            if data.get(f) is None:
                return jsonify({'error': f'{f} obrigatório.'}), 400
        if profile_has_filial_scope(profile):
            allowed_ids = profile.get('allowed_filial_ids') or []
            if int(data['filial_id']) not in set(allowed_ids):
                return jsonify({'error': 'Sem permissão para esta filial.'}), 403
        allowed = [
            'filial_id', 'filial_nome', 'competencia', 'tipo_despesa', 'descricao',
            'fornecedor_nome', 'colaborador_id', 'valor', 'data_vencimento', 'data_pagamento',
            'valor_pago', 'status', 'tipo_documento', 'numero_documento', 'observacoes',
        ]
        row = {k: data[k] for k in allowed if k in data}
        try:
            result = supabase.table('contas_a_pagar').insert(row).execute()
            return jsonify({'ok': True, 'data': (result.data or [{}])[0]})
        except Exception as exc:
            app.logger.error('criar_conta_pagar: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-pagar/<cp_id>', methods=['PUT'])
    @require_auth
    def editar_conta_pagar(profile, cp_id):
        data = request.get_json() or {}
        if profile_has_filial_scope(profile):
            reg = supabase.table('contas_a_pagar').select('filial_id').eq('id', cp_id).limit(1).execute()
            fid = ((reg.data or [{}])[0]).get('filial_id')
            if fid and int(fid) not in set(profile.get('allowed_filial_ids') or []):
                return jsonify({'error': 'Sem permissão.'}), 403
        allowed = [
            'filial_nome', 'competencia', 'tipo_despesa', 'descricao', 'fornecedor_nome',
            'colaborador_id', 'valor', 'data_vencimento', 'data_pagamento', 'valor_pago',
            'status', 'tipo_documento', 'numero_documento', 'observacoes',
        ]
        update = {k: data[k] for k in allowed if k in data}
        update['updated_at'] = datetime.utcnow().isoformat()
        try:
            supabase.table('contas_a_pagar').update(update).eq('id', cp_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('editar_conta_pagar: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/contas-pagar/<cp_id>', methods=['DELETE'])
    @require_auth
    def deletar_conta_pagar(profile, cp_id):
        if profile_has_filial_scope(profile):
            reg = supabase.table('contas_a_pagar').select('filial_id').eq('id', cp_id).limit(1).execute()
            fid = ((reg.data or [{}])[0]).get('filial_id')
            if fid and int(fid) not in set(profile.get('allowed_filial_ids') or []):
                return jsonify({'error': 'Sem permissão.'}), 403
        try:
            supabase.table('contas_a_pagar').delete().eq('id', cp_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('deletar_conta_pagar: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ============ BANCO ============

    @app.route('/api/banco/contas', methods=['GET'])
    @require_auth
    def listar_banco_contas(profile):
        try:
            query = supabase.table('banco_contas').select('*').order('banco_nome')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            result = query.execute()
            contas = result.data or []
            # Calcula saldo atual para cada conta
            for conta in contas:
                lresp = supabase.table('banco_lancamentos').select('tipo, valor').eq('conta_id', conta['id']).execute()
                saldo = float(conta.get('saldo_inicial') or 0)
                for l in (lresp.data or []):
                    v = float(l.get('valor') or 0)
                    saldo += v if l.get('tipo') == 'ENTRADA' else -v
                conta['saldo_atual'] = round(saldo, 2)
            return jsonify({'data': contas})
        except Exception as exc:
            app.logger.error('listar_banco_contas: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/contas', methods=['POST'])
    @require_auth
    def criar_banco_conta(profile):
        data = request.get_json() or {}
        if not data.get('banco_nome'):
            return jsonify({'error': 'banco_nome obrigatório.'}), 400
        if profile_has_filial_scope(profile):
            allowed_ids = profile.get('allowed_filial_ids') or []
            if data.get('filial_id') and int(data['filial_id']) not in set(allowed_ids):
                return jsonify({'error': 'Sem permissão para esta filial.'}), 403
        allowed = ['filial_id', 'filial_nome', 'banco_nome', 'agencia', 'conta', 'tipo', 'saldo_inicial', 'ativo']
        row = {k: data[k] for k in allowed if k in data}
        try:
            result = supabase.table('banco_contas').insert(row).execute()
            return jsonify({'ok': True, 'data': (result.data or [{}])[0]})
        except Exception as exc:
            app.logger.error('criar_banco_conta: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/contas/<conta_id>', methods=['PUT'])
    @require_auth
    def editar_banco_conta(profile, conta_id):
        data = request.get_json() or {}
        allowed = ['banco_nome', 'agencia', 'conta', 'tipo', 'saldo_inicial', 'ativo', 'filial_nome']
        update = {k: data[k] for k in allowed if k in data}
        try:
            supabase.table('banco_contas').update(update).eq('id', conta_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('editar_banco_conta: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/contas/<conta_id>', methods=['DELETE'])
    @require_auth
    def deletar_banco_conta(profile, conta_id):
        try:
            supabase.table('banco_contas').delete().eq('id', conta_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('deletar_banco_conta: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/lancamentos', methods=['GET'])
    @require_auth
    def listar_banco_lancamentos(profile):
        try:
            query = supabase.table('banco_lancamentos').select('*').order('data_lancamento', desc=True)
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'data': []})
                query = query.in_('filial_id', allowed_ids)
            conta_id = request.args.get('conta_id')
            conciliado = request.args.get('conciliado')
            mes = request.args.get('mes')
            if conta_id:
                query = query.eq('conta_id', conta_id)
            if conciliado is not None:
                query = query.eq('conciliado', conciliado == 'true')
            if mes:
                query = query.gte('data_lancamento', mes + '-01').lte('data_lancamento', mes + '-31')
            result = query.execute()
            return jsonify({'data': result.data or []})
        except Exception as exc:
            app.logger.error('listar_banco_lancamentos: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/lancamentos', methods=['POST'])
    @require_auth
    def criar_banco_lancamento(profile):
        data = request.get_json() or {}
        required = ['conta_id', 'data_lancamento', 'tipo', 'descricao', 'valor']
        for f in required:
            if not data.get(f):
                return jsonify({'error': f'{f} obrigatório.'}), 400
        allowed = [
            'conta_id', 'filial_id', 'filial_nome', 'data_lancamento', 'tipo', 'categoria',
            'descricao', 'valor', 'conta_receber_id', 'conta_pagar_id', 'observacoes',
        ]
        row = {k: data[k] for k in allowed if k in data}
        row['conciliado'] = False
        try:
            result = supabase.table('banco_lancamentos').insert(row).execute()
            inserted = (result.data or [{}])[0]
            # Se veio com conta_receber_id ou conta_pagar_id, concilia automaticamente
            if inserted.get('id') and data.get('conta_receber_id'):
                supabase.table('contas_a_receber').update({'banco_lancamento_id': inserted['id'], 'status_fat': 'FATURADO', 'status': 'RECEBIDO'}).eq('id', data['conta_receber_id']).execute()
            if inserted.get('id') and data.get('conta_pagar_id'):
                supabase.table('contas_a_pagar').update({'banco_lancamento_id': inserted['id'], 'status': 'PAGO', 'data_pagamento': data.get('data_lancamento'), 'valor_pago': float(data.get('valor') or 0)}).eq('id', data['conta_pagar_id']).execute()
            return jsonify({'ok': True, 'data': inserted})
        except Exception as exc:
            app.logger.error('criar_banco_lancamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/lancamentos/<lancamento_id>', methods=['PUT'])
    @require_auth
    def editar_banco_lancamento(profile, lancamento_id):
        data = request.get_json() or {}
        allowed = ['data_lancamento', 'tipo', 'categoria', 'descricao', 'valor', 'conciliado', 'conta_receber_id', 'conta_pagar_id', 'observacoes']
        update = {k: data[k] for k in allowed if k in data}
        try:
            supabase.table('banco_lancamentos').update(update).eq('id', lancamento_id).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('editar_banco_lancamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/lancamentos/<lancamento_id>/conciliar', methods=['POST'])
    @require_auth
    def conciliar_lancamento(profile, lancamento_id):
        data = request.get_json() or {}
        try:
            lan_resp = supabase.table('banco_lancamentos').select('*').eq('id', lancamento_id).limit(1).execute()
            lan = (lan_resp.data or [{}])[0]
            if not lan:
                return jsonify({'error': 'Lançamento não encontrado.'}), 404
            conciliado = bool(data.get('conciliado', True))
            supabase.table('banco_lancamentos').update({'conciliado': conciliado}).eq('id', lancamento_id).execute()
            if conciliado:
                if lan.get('conta_receber_id'):
                    supabase.table('contas_a_receber').update({'status_fat': 'FATURADO', 'status': 'RECEBIDO'}).eq('id', lan['conta_receber_id']).execute()
                if lan.get('conta_pagar_id'):
                    supabase.table('contas_a_pagar').update({'status': 'PAGO', 'valor_pago': float(lan.get('valor') or 0)}).eq('id', lan['conta_pagar_id']).execute()
            return jsonify({'ok': True})
        except Exception as exc:
            app.logger.error('conciliar_lancamento: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    @app.route('/api/banco/saldos', methods=['GET'])
    @require_auth
    def saldos_banco(profile):
        try:
            query_contas = supabase.table('banco_contas').select('id, banco_nome, filial_id, filial_nome, saldo_inicial, tipo, ativo')
            if profile_has_filial_scope(profile):
                allowed_ids = profile.get('allowed_filial_ids') or []
                if not allowed_ids:
                    return jsonify({'saldo_total': 0, 'nao_conciliados': 0, 'contas': []})
                query_contas = query_contas.in_('filial_id', allowed_ids)
            contas_resp = query_contas.eq('ativo', True).execute()
            contas = contas_resp.data or []
            lan_resp = supabase.table('banco_lancamentos').select('conta_id, tipo, valor, conciliado').execute()
            lancamentos = lan_resp.data or []
            nao_conciliados = sum(1 for l in lancamentos if not l.get('conciliado'))
            saldo_total = 0
            from collections import defaultdict
            lan_by_conta = defaultdict(list)
            for l in lancamentos:
                lan_by_conta[l['conta_id']].append(l)
            for conta in contas:
                saldo = float(conta.get('saldo_inicial') or 0)
                for l in lan_by_conta.get(conta['id'], []):
                    v = float(l.get('valor') or 0)
                    saldo += v if l.get('tipo') == 'ENTRADA' else -v
                conta['saldo_atual'] = round(saldo, 2)
                saldo_total += saldo
            return jsonify({'saldo_total': round(saldo_total, 2), 'nao_conciliados': nao_conciliados, 'contas': contas})
        except Exception as exc:
            app.logger.error('saldos_banco: %s', exc)
            return jsonify({'error': translate_database_error(exc)}), 500

    # ============ SERVIR FRONTEND REACT ============
    @app.route('/', defaults={'path': ''})
    @app.route('/<path:path>')
    def serve_frontend(path):
        """Serve arquivos estáticos do frontend React compilado"""
        if path != '' and os.path.exists(os.path.join(FRONTEND_DIST_DIR, path)):
            return send_from_directory(FRONTEND_DIST_DIR, path)
        elif path.startswith('api/'):
            return jsonify({'error': 'API endpoint not found'}), 404
        else:
            return send_from_directory(FRONTEND_DIST_DIR, 'index.html')

    # ─── Inicializar motor de alertas ─────────────────────────────────────────
    if alerts_enabled:
        try:
            refresh_alert_engine_snapshot()
        except Exception as exc:
            app.logger.error('Falha na carga inicial do motor de alertas: %s', exc)
            with alert_engine_lock:
                alert_engine_state['last_error'] = str(exc)
                alert_engine_state['last_run_at'] = datetime.now().astimezone().isoformat()

        engine_thread = threading.Thread(target=run_alert_engine_loop, name='seg-alert-engine', daemon=True)
        engine_thread.start()

    return app


app = create_app()


if __name__ == '__main__':
    log_level_name = os.getenv('LOG_LEVEL', 'INFO').upper()
    log_level = getattr(logging, log_level_name, logging.INFO)
    logging.basicConfig(level=log_level)
    app.logger.setLevel(log_level)
    port = int(os.getenv('PORT', '5000'))
    debug_mode = os.getenv('FLASK_DEBUG', '0').strip().lower() in {'1', 'true', 'yes'}
    app.logger.info('Servidor Flask iniciado. debug=%s porta=%s', debug_mode, port)
    app.run(debug=debug_mode, use_reloader=False, host='0.0.0.0', port=port)

